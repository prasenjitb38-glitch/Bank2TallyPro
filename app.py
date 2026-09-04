from __future__ import annotations

import csv
import base64
import hashlib
import hmac
import html
import io
import json
import os
import re
import queue
import shutil
import socket
import sqlite3
import subprocess
import sys
import tempfile
import threading
import traceback
import unicodedata
import uuid
import webbrowser
import urllib.error
import urllib.request
import zipfile
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from generic_mapping import (
    apply_mapping,
    extract_pdf_grid,
    format_fingerprint,
    load_profiles,
    number,
    parse_date,
    save_profiles,
)
from scanned_bank_ocr import parse_scanned_bank_pdf


class LicenseStore:
    """Cloud edition: licensing and credit limits are disabled."""
    def __init__(self, data_dir):
        self.processed_files = set()

    def status(self):
        return {
            "device_id": "cloud",
            "trial_remaining": 0,
            "paid_remaining": 0,
            "yearly_remaining": 0,
            "remaining": 999999999,
            "yearly_expiry": "",
            "plan": "Cloud",
        }

    def already_processed(self, digest):
        return digest in self.processed_files

    def can_charge(self, credits):
        return True

    def charge(self, credits, digest, filename):
        duplicate = digest in self.processed_files
        self.processed_files.add(digest)
        return {"charged": 0, "duplicate": duplicate, **self.status()}

    def activate(self, key, customer_mobile=""):
        return self.status()


def file_digest(raw):
    return hashlib.sha256(raw).hexdigest()


def credit_cost(filename, raw, transaction_count=0, password=""):
    return 0

APP_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
STATIC_DIR = APP_DIR / "static"
if getattr(sys, "frozen", False):
    data_candidates = [
        Path(os.environ.get("LOCALAPPDATA", Path.home())) / "Bank2Tally",
        Path(sys.executable).resolve().parent / "Bank2TallyData",
        Path.cwd() / "Bank2TallyData",
    ]
    for DATA_DIR in data_candidates:
        try:
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            break
        except OSError:
            continue
    else:
        raise RuntimeError("Bank2Tally could not create its local data folder.")
    MAPPINGS_PATH = DATA_DIR / "saved_mappings.json"
    bundled_mappings = APP_DIR / "saved_mappings.json"
    if not MAPPINGS_PATH.exists() and bundled_mappings.exists():
        shutil.copy2(bundled_mappings, MAPPINGS_PATH)
else:
    DATA_DIR = APP_DIR
    MAPPINGS_PATH = APP_DIR / "saved_mappings.json"
# Render (and similar hosts) inject PORT and require binding on all interfaces;
# local Windows use keeps the previous loopback-only default.
HOST = os.environ.get("HOST", "0.0.0.0")
PORT = int(os.environ.get("PORT", "8765"))
LICENSE_STORE = LicenseStore(DATA_DIR)
PENDING_CHARGES = {}
ACCOUNT_PATH = DATA_DIR / "customer_account.json"
SESSION_PATH = DATA_DIR / "login_session.json"
IMPORT_HISTORY_PATH = DATA_DIR / "tally_import_history.json"
HSN_MASTER_PATH = DATA_DIR / "hsn_master.sqlite3"
GST_RECON_DB_PATH = DATA_DIR / "gst_reconciliation.sqlite3"
ACCOUNT_LOGGED_IN = False
TALLY_CACHE = {"connected": False, "company": "", "ledgers": [], "items": [], "stock_groups": [], "voucher_types": [],
               "company_state": "", "company_country": "India", "synced_at": ""}
# Portal import badges for the current app/login session only (cleared on start + login).
GST_SESSION_IMPORTS = {"GSTR-2B": set(), "GSTR-2A": set(), "GSTR-1": set(), "GSTR-3B": set()}

GST_STATE_CODES = {
    "01": "Jammu & Kashmir", "02": "Himachal Pradesh", "03": "Punjab", "04": "Chandigarh",
    "05": "Uttarakhand", "06": "Haryana", "07": "Delhi", "08": "Rajasthan",
    "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim", "12": "Arunachal Pradesh",
    "13": "Nagaland", "14": "Manipur", "15": "Mizoram", "16": "Tripura",
    "17": "Meghalaya", "18": "Assam", "19": "West Bengal", "20": "Jharkhand",
    "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh", "24": "Gujarat",
    "26": "Dadra & Nagar Haveli and Daman & Diu", "27": "Maharashtra",
    "28": "Andhra Pradesh (Before Division)", "29": "Karnataka", "30": "Goa",
    "31": "Lakshadweep", "32": "Kerala", "33": "Tamil Nadu", "34": "Puducherry",
    "35": "Andaman & Nicobar Islands", "36": "Telangana", "37": "Andhra Pradesh",
    "38": "Ladakh", "97": "Other Territory",
}


def hsn_connection():
    connection = sqlite3.connect(HSN_MASTER_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute(
        """CREATE TABLE IF NOT EXISTS hsn_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hsn_code TEXT NOT NULL,
        item_name TEXT NOT NULL,
        description TEXT DEFAULT '',
        gst_rate REAL DEFAULT 0,
        uqc TEXT DEFAULT '',
        category TEXT DEFAULT '',
        effective_date TEXT DEFAULT '',
        active INTEGER DEFAULT 1,
        UNIQUE(hsn_code, item_name)
        )"""
    )
    connection.execute("CREATE INDEX IF NOT EXISTS idx_hsn_code ON hsn_master(hsn_code)")
    connection.commit()
    return connection


def list_hsn_master(query="", limit=1000):
    search = f"%{gst_text(query)}%"
    connection = hsn_connection()
    try:
        rows = connection.execute(
            """SELECT id,hsn_code,item_name,description,gst_rate,uqc,category,effective_date,active
               FROM hsn_master
               WHERE hsn_code LIKE ? OR item_name LIKE ? OR description LIKE ? OR category LIKE ?
               ORDER BY hsn_code,item_name LIMIT ?""",
            (search, search, search, search, max(1, min(int(limit or 1000), 5000))),
        ).fetchall()
    finally:
        connection.close()
    return [dict(row) for row in rows]


def save_hsn_record(record):
    code = re.sub(r"\s+", "", gst_text(record.get("hsn_code") or record.get("hsn")))
    name = gst_text(record.get("item_name") or record.get("item") or record.get("description"))
    if not code or not name:
        raise ValueError("HSN Code and Item Name are required.")
    values = (
        code, name, gst_text(record.get("description")), gst_number(record.get("gst_rate")),
        gst_text(record.get("uqc") or record.get("unit")), gst_text(record.get("category")),
        gst_text(record.get("effective_date")), 1 if record.get("active", True) else 0,
    )
    connection = hsn_connection()
    try:
        connection.execute(
            """INSERT INTO hsn_master
               (hsn_code,item_name,description,gst_rate,uqc,category,effective_date,active)
               VALUES (?,?,?,?,?,?,?,?)
               ON CONFLICT(hsn_code,item_name) DO UPDATE SET
               description=excluded.description,gst_rate=excluded.gst_rate,uqc=excluded.uqc,
               category=excluded.category,effective_date=excluded.effective_date,active=excluded.active""",
            values,
        )
        connection.commit()
    finally:
        connection.close()


def import_hsn_master(name, raw):
    suffix = Path(name).suffix.lower()
    records = []
    if suffix == ".csv":
        text = raw.decode("utf-8-sig", errors="replace")
        records = list(csv.DictReader(io.StringIO(text)))
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            headers = next(values, None)
            if not headers:
                continue
            keys = [gst_text(value).lower() for value in headers]
            records.extend(dict(zip(keys, row)) for row in values if any(value not in (None, "") for value in row))
    else:
        raise ValueError("HSN Master must be an Excel .xlsx or CSV file.")
    aliases = {
        "hsn_code": ("hsn code", "hsn", "hsn/sac", "hsn sac code", "code"),
        "item_name": ("item name", "item", "product", "stock item", "description of goods"),
        "description": ("description", "hsn description", "goods description"),
        "gst_rate": ("gst rate", "rate", "tax rate", "gst%"),
        "uqc": ("uqc", "unit", "unit code"),
        "category": ("category", "chapter", "group"),
        "effective_date": ("effective date", "date"),
    }
    saved, skipped = 0, 0
    for source in records:
        lowered = {gst_text(key).lower(): value for key, value in source.items()}
        normalized = {}
        for target, names in aliases.items():
            normalized[target] = next((lowered[name] for name in names if name in lowered), "")
        try:
            save_hsn_record(normalized)
            saved += 1
        except ValueError:
            skipped += 1
    return saved, skipped


def load_account():
    try:
        return json.loads(ACCOUNT_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None


def save_login_session(logged_in=True):
    """Keep local login across app.py / server restarts (desktop, single-user)."""
    payload = {
        "logged_in": bool(logged_in),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    if logged_in:
        account = load_account() or {}
        payload["mobile"] = account.get("mobile", "")
    try:
        SESSION_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    except OSError:
        pass


def clear_login_session():
    save_login_session(False)


def restore_login_session():
    """Restore ACCOUNT_LOGGED_IN from disk if the registered account still exists."""
    global ACCOUNT_LOGGED_IN
    account = load_account()
    if not account:
        ACCOUNT_LOGGED_IN = False
        clear_login_session()
        return False
    try:
        data = json.loads(SESSION_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        ACCOUNT_LOGGED_IN = False
        return False
    if not data.get("logged_in"):
        ACCOUNT_LOGGED_IN = False
        return False
    saved_mobile = str(data.get("mobile") or "")
    if saved_mobile and saved_mobile != str(account.get("mobile") or ""):
        ACCOUNT_LOGGED_IN = False
        clear_login_session()
        return False
    ACCOUNT_LOGGED_IN = True
    return True


# Restore desktop login after process restart (Connect Tally / APIs).
restore_login_session()


def save_account(customer_name, mobile, business_name, pin):
    if not customer_name.strip() or not mobile.strip() or not business_name.strip():
        raise ValueError("Customer Name, Mobile Number and Business Name are required.")
    if not re.fullmatch(r"(?:\d{4}|\d{6})", pin or ""):
        raise ValueError("Login PIN must be exactly 4 or 6 digits.")
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", pin.encode(), salt, 180000)
    account = {
        "customer_name": customer_name.strip()[:100],
        "mobile": re.sub(r"\D", "", mobile)[:15],
        "business_name": business_name.strip()[:120],
        "pin_salt": salt.hex(),
        "pin_hash": digest.hex(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    ACCOUNT_PATH.write_text(json.dumps(account, indent=2), encoding="utf-8")
    return account


def verify_pin(account, pin):
    try:
        salt = bytes.fromhex(account["pin_salt"])
        expected = bytes.fromhex(account["pin_hash"])
        actual = hashlib.pbkdf2_hmac("sha256", str(pin).encode(), salt, 180000)
        return hmac.compare_digest(actual, expected)
    except (KeyError, ValueError):
        return False


def public_account(account):
    if not account:
        return {}
    return {key: account.get(key, "") for key in ("customer_name", "mobile", "business_name")}


def amount(value):
    if value is None or value == "":
        return 0.0
    cleaned = re.sub(r"[^\d.\-]", "", str(value).replace(",", ""))
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return 0.0


def legacy_xls_to_xlsx(raw):
    with tempfile.TemporaryDirectory(prefix="Bank2Tally_xls_", ignore_cleanup_errors=True) as folder:
        source = Path(folder) / "source.xls"
        target = Path(folder) / "converted.xlsx"
        source.write_bytes(raw)
        source_ps = str(source).replace("'", "''")
        target_ps = str(target).replace("'", "''")
        script = (
            "$excel=New-Object -ComObject Excel.Application;"
            "$excel.Visible=$false;$excel.DisplayAlerts=$false;"
            f"try{{$wb=$excel.Workbooks.Open('{source_ps}',0,$true);"
            f"$wb.SaveAs('{target_ps}',51);$wb.Close($false)}}"
            "finally{$excel.Quit();[Runtime.InteropServices.Marshal]::ReleaseComObject($excel)|Out-Null}"
        )
        completed = subprocess.run(
            ["powershell.exe", "-NoProfile", "-Command", script],
            capture_output=True, text=True, timeout=90,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        if completed.returncode or not target.exists():
            detail = (completed.stderr or completed.stdout or "").strip()
            raise ValueError("Old .xls conversion failed. Microsoft Excel is required." +
                             (f" {detail[:180]}" if detail else ""))
        return target.read_bytes()


def legacy_xls_grid(raw):
    """Read a legacy binary XLS export without requiring Microsoft Excel."""
    import xlrd
    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    return [sheet.row_values(index) for index in range(sheet.nrows)]


TALLY_HTTP_URL = "http://127.0.0.1:9000"
TALLY_HTTP_HOST = "127.0.0.1"
TALLY_HTTP_PORT = 9000
# The Tally computer opens the connection to Render.  Render never needs a
# tunnel URL or an inbound address for that computer.
TALLY_CONNECTOR_ID = os.environ.get("TALLY_CONNECTOR_ID", "primary")
TALLY_CONNECTOR_TOKEN = os.environ.get("TALLY_CONNECTOR_TOKEN", "")
CONNECTOR_COMMANDS = queue.Queue()
CONNECTOR_PENDING = {}
CONNECTOR_LOCK = threading.Lock()


def tally_log(message):
    stamp = datetime.now().isoformat(timespec="seconds")
    line = f"[Tally] {stamp} {message}"
    print(line)
    try:
        log_path = DATA_DIR / "tally_sales_sync.log"
        with log_path.open("a", encoding="utf-8") as handle:
            handle.write(line + "\n")
    except OSError:
        pass


def connector_authorized(headers, connector_id):
    return (
        bool(TALLY_CONNECTOR_TOKEN)
        and connector_id == TALLY_CONNECTOR_ID
        and hmac.compare_digest(headers.get("X-Connector-Token", ""), TALLY_CONNECTOR_TOKEN)
    )


def connector_request(kind, payload, timeout=120):
    """Send work to the Windows Connector and wait for its outbound reply."""
    if not TALLY_CONNECTOR_TOKEN:
        raise ValueError("Tally Connector is not configured on Render. Set TALLY_CONNECTOR_TOKEN.")
    command_id = str(uuid.uuid4())
    pending = {"event": threading.Event(), "result": None}
    with CONNECTOR_LOCK:
        CONNECTOR_PENDING[command_id] = pending
    CONNECTOR_COMMANDS.put({"id": command_id, "kind": kind, **payload})
    if not pending["event"].wait(timeout):
        with CONNECTOR_LOCK:
            CONNECTOR_PENDING.pop(command_id, None)
        raise ValueError("Tally Connector is offline or did not respond in time. Start it on the Tally computer.")
    result = pending["result"] or {}
    if not result.get("ok"):
        raise ValueError(result.get("error") or "Tally Connector request failed.")
    return result


def tally_port_reachable(timeout=2.0):
    """Return (reachable: bool, detail: str) for Tally HTTP port."""
    try:
        with socket.create_connection((TALLY_HTTP_HOST, TALLY_HTTP_PORT), timeout=timeout):
            return True, f"{TALLY_HTTP_HOST}:{TALLY_HTTP_PORT} is reachable"
    except OSError as exc:
        return False, f"{TALLY_HTTP_HOST}:{TALLY_HTTP_PORT} not reachable ({type(exc).__name__}: {exc})"


def tally_post(raw, timeout=25, purpose="Tally export"):
    """
    POST XML to TallyPrime HTTP server.
    Logs URL, request size, reachability, and exact network/response outcome.
    """
    payload = raw if isinstance(raw, (bytes, bytearray)) else gst_text(raw).encode("utf-8")
    if TALLY_CONNECTOR_TOKEN:
        result = connector_request("tally", {"xml": payload.decode("utf-8", errors="replace"), "timeout": timeout}, timeout + 15)
        body = result.get("body", "")
        if not body.strip():
            raise ValueError(f"Tally Connector returned an empty response for {purpose}.")
        return body
    reachable, reach_detail = tally_port_reachable()
    tally_log(f"{purpose} | URL={TALLY_HTTP_URL} | port_reachable={reachable} | {reach_detail}")
    tally_log(f"{purpose} | request_bytes={len(payload)} | timeout={timeout}s")
    try:
        preview = payload.decode("utf-8", errors="replace")
        tally_log(f"{purpose} | request_xml_preview={preview[:500].replace(chr(10), ' ')}")
        (DATA_DIR / "tally_last_sales_request.xml").write_bytes(payload)
    except OSError:
        pass
    if not reachable:
        raise ValueError(
            f"TallyPrime HTTP port is not reachable on {TALLY_HTTP_HOST}:{TALLY_HTTP_PORT}. "
            "Open TallyPrime, load the company, and enable the HTTP Server on port 9000. "
            f"Detail: {reach_detail}"
        )
    request = urllib.request.Request(
        TALLY_HTTP_URL, data=payload,
        headers={"Content-Type": "text/xml; charset=utf-8"}, method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            status = getattr(response, "status", None) or response.getcode()
            body = response.read().decode("utf-8", errors="replace")
        tally_log(f"{purpose} | http_status={status} | response_bytes={len(body.encode('utf-8', errors='ignore'))}")
        if not body.strip():
            tally_log(f"{purpose} | response=EMPTY")
            raise ValueError(
                f"TallyPrime returned an empty response for {purpose}. "
                "Confirm the company is loaded and HTTP Server is enabled on port 9000."
            )
        lowered = body.lower()
        if "<lineerror>" in lowered or "<error>" in body[:2000].lower() and "<envelope>" not in lowered:
            tally_log(f"{purpose} | response=ERROR_XML preview={body[:400].replace(chr(10), ' ')}")
        else:
            tally_log(f"{purpose} | response=XML preview={body[:400].replace(chr(10), ' ')}")
        try:
            (DATA_DIR / "tally_last_sales_response.xml").write_text(body, encoding="utf-8", errors="replace")
        except OSError:
            pass
        return body
    except TimeoutError as exc:
        tally_log(f"{purpose} | network_error=TimeoutError | {exc}")
        raise ValueError(
            f"TallyPrime timed out after {timeout}s while handling {purpose} "
            f"at {TALLY_HTTP_URL}. Port reachable={reachable}. "
            "Keep TallyPrime open with the company loaded; large voucher exports need more time."
        ) from exc
    except urllib.error.HTTPError as exc:
        detail = ""
        try:
            detail = exc.read().decode("utf-8", errors="replace")[:300]
        except Exception:
            detail = str(exc)
        tally_log(f"{purpose} | network_error=HTTPError | code={exc.code} | {detail}")
        raise ValueError(
            f"TallyPrime HTTP error {exc.code} for {purpose} at {TALLY_HTTP_URL}. {detail}"
        ) from exc
    except urllib.error.URLError as exc:
        tally_log(f"{purpose} | network_error=URLError | {exc.reason!r}")
        raise ValueError(
            f"TallyPrime connection failed for {purpose} at {TALLY_HTTP_URL}. "
            f"Port reachable={reachable}. Network error: {exc.reason!r}. "
            "Open TallyPrime, load the company, and enable the HTTP Server on port 9000."
        ) from exc


def tally_test_connection(timeout=10):
    """Probe port 9000 and request a lightweight Company collection before heavy sync."""
    if TALLY_CONNECTOR_TOKEN:
        try:
            result = connector_request("health", {}, timeout)
            reachable = bool(result.get("reachable"))
            reach_detail = result.get("detail", "Tally Connector responded")
        except Exception as exc:
            reachable = False
            reach_detail = str(exc)
    else:
        reachable, reach_detail = tally_port_reachable()
    result = {
        "ok": False,
        "url": "Outbound Connector" if TALLY_CONNECTOR_TOKEN else TALLY_HTTP_URL,
        "port": TALLY_HTTP_PORT,
        "port_reachable": reachable,
        "port_detail": reach_detail,
        "company": "",
        "response_kind": "",
        "error": "",
    }
    tally_log(f"connection_test | {reach_detail}")
    if not reachable:
        result["error"] = reach_detail
        result["response_kind"] = "port_closed"
        return result
    raw = (
        '<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>'
        '<TYPE>Collection</TYPE><ID>Bank2TallySalesPing</ID></HEADER><BODY><DESC>'
        '<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>'
        '<TDL><TDLMESSAGE><COLLECTION NAME="Bank2TallySalesPing" ISMODIFY="No">'
        '<TYPE>Company</TYPE><FETCH>Name</FETCH>'
        '</COLLECTION></TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>'
    ).encode("utf-8")
    try:
        xml = tally_post(raw, timeout=timeout, purpose="sales-connection-test")
    except ValueError as exc:
        result["error"] = str(exc)
        result["response_kind"] = "error"
        return result
    if not gst_text(xml):
        result["error"] = "Empty response from TallyPrime"
        result["response_kind"] = "empty"
        return result
    companies = parse_tally_objects(xml, "COMPANY", {})
    company = companies[0]["name"] if companies else tag_value(xml, "SVCURRENTCOMPANY")
    result["ok"] = True
    result["company"] = company or "Open Tally company"
    result["response_kind"] = "xml"
    tally_log(f"connection_test | ok company={result['company']}")
    return result


def tally_collection_xml(collection, object_type, fetch, timeout=120, purpose="", filter_formula=""):
    # TDL FETCH expects comma-separated method names. A space-separated value
    # returns only object names, which prevents unit/parent/GSTIN matching.
    fields = ", ".join(fetch)
    label = purpose or f"collection:{object_type}"
    filter_name = "Bank2TallyObjectFilter" if gst_text(filter_formula) else ""
    filter_xml = f"<FILTER>{filter_name}</FILTER>" if filter_name else ""
    formula_xml = (
        f'<SYSTEM TYPE="Formulae" NAME="{filter_name}">{xml_escape(filter_formula)}</SYSTEM>'
        if filter_name else ""
    )
    raw = (
        '<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>'
        '<TYPE>Collection</TYPE><ID>Bank2TallySync</ID></HEADER><BODY><DESC>'
        '<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>'
        '<TDL><TDLMESSAGE><COLLECTION NAME="Bank2TallySync" ISMODIFY="No">'
        f'<TYPE>{xml_escape(object_type)}</TYPE><FETCH>{xml_escape(fields)}</FETCH>{filter_xml}'
        f'</COLLECTION>{formula_xml}</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>'
    ).encode("utf-8")
    return tally_post(raw, timeout=timeout, purpose=label)


def tag_value(block, name):
    match = re.search(rf"<{re.escape(name)}(?:\s[^>]*)?>(.*?)</{re.escape(name)}>", block, re.I | re.S)
    return re.sub(r"<[^>]+>", "", match.group(1)).strip() if match else ""


def parse_tally_objects(xml, tag, fields):
    objects = []
    for match in re.finditer(rf"<{tag}\b([^>]*)>(.*?)</{tag}>", xml, re.I | re.S):
        attrs, body = match.groups()
        name_match = re.search(r'\bNAME="([^"]*)"', attrs, re.I)
        raw_name = name_match.group(1) if name_match else tag_value(body, "NAME")
        exact_name = tally_master_name(raw_name)
        item = {
            "name": clean_ledger_name(raw_name),
            "exact_name": exact_name,
        }
        for field, tally_tag in fields.items():
            item[field] = tag_value(body, tally_tag)
        if item["name"]:
            objects.append(item)
    return objects


def sync_tally():
    company_xml = tally_collection_xml(
        "Company", "Company", ["Name", "Guid", "StateName", "CountryName"]
    )
    ledger_xml = tally_collection_xml(
        "Ledger", "Ledger", ["Name", "Parent", "GSTIN", "PartyGSTIN", "ClosingBalance", "GSTDetails"]
    )
    item_xml = tally_collection_xml(
        "StockItem", "Stock Item",
        ["Name", "Parent", "BaseUnits", "GSTApplicable", "HSNCode", "HSNDetails", "GSTDetails"]
    )
    stock_group_xml = tally_collection_xml("StockGroup", "Stock Group", ["Name", "Parent"])
    voucher_type_xml = tally_collection_xml("VoucherType", "Voucher Type", ["Name", "Parent"])
    companies = parse_tally_objects(
        company_xml, "COMPANY",
        {"guid": "GUID", "state": "STATENAME", "country": "COUNTRYNAME"},
    )
    ledgers = parse_tally_objects(
        ledger_xml, "LEDGER",
        {"parent": "PARENT", "gstin": "PARTYGSTIN", "closing_balance": "CLOSINGBALANCE", "taxability": "TAXABILITY"},
    )
    items = parse_tally_objects(
        item_xml, "STOCKITEM",
        {"parent": "PARENT", "unit": "BASEUNITS", "hsn": "HSNCODE"},
    )
    stock_groups = parse_tally_objects(stock_group_xml, "STOCKGROUP", {"parent": "PARENT"})
    voucher_types = parse_tally_objects(voucher_type_xml, "VOUCHERTYPE", {"parent": "PARENT"})
    company = companies[0]["name"] if companies else tag_value(company_xml, "SVCURRENTCOMPANY")
    company_state = companies[0].get("state", "") if companies else ""
    company_country = companies[0].get("country", "") if companies else ""
    TALLY_CACHE.update({
        "connected": True, "company": company or "Open Tally company",
        "ledgers": ledgers, "items": items, "stock_groups": stock_groups, "voucher_types": voucher_types,
        "company_state": company_state, "company_country": company_country or "India",
        "synced_at": datetime.now().isoformat(timespec="seconds"),
    })
    return TALLY_CACHE


def sync_tally_vouchers(party_ledger=""):
    xml = tally_collection_xml(
        "Voucher", "Voucher",
        ["Date", "VoucherTypeName", "VoucherNumber", "Reference", "Narration",
         "PartyLedgerName", "AllLedgerEntries.LedgerName", "AllLedgerEntries.Amount"],
    )
    rows = []
    for match in re.finditer(r"<VOUCHER\b[^>]*>(.*?)</VOUCHER>", xml, re.I | re.S):
        body = match.group(1)
        party = tag_value(body, "PARTYLEDGERNAME")
        entries = []
        for entry in re.finditer(r"<ALLLEDGERENTRIES\.LIST>(.*?)</ALLLEDGERENTRIES\.LIST>", body, re.I | re.S):
            entry_body = entry.group(1)
            entries.append({"ledger": tag_value(entry_body, "LEDGERNAME"),
                            "amount": amount(tag_value(entry_body, "AMOUNT"))})
        selected = next((entry for entry in entries
                         if entry["ledger"].strip().lower() == party_ledger.strip().lower()), None)
        if party_ledger and not selected and party.strip().lower() != party_ledger.strip().lower():
            continue
        if not selected and party:
            selected = next((entry for entry in entries
                             if entry["ledger"].strip().lower() == party.strip().lower()), None)
        value = abs(selected["amount"]) if selected else max((abs(entry["amount"]) for entry in entries), default=0)
        raw_date = tag_value(body, "DATE")
        date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}" if re.fullmatch(r"\d{8}", raw_date) else raw_date
        rows.append({
            "date": date, "particulars": tag_value(body, "NARRATION") or tag_value(body, "VOUCHERTYPENAME"),
            "reference": tag_value(body, "REFERENCE") or tag_value(body, "VOUCHERNUMBER"),
            "amount": value, "voucher_type": tag_value(body, "VOUCHERTYPENAME"),
            "party_ledger": party or party_ledger, "source": "Tally",
        })
    return rows


def sync_tally_gst_monthly():
    """Aggregate the open Tally company's Sales/Purchase and GST ledgers month-wise."""
    cache = sync_tally()
    parents = {gst_text(item.get("name")).lower(): gst_text(item.get("parent")).lower()
               for item in cache.get("ledgers", [])}
    xml = tally_collection_xml("Voucher", "Voucher", ["Date", "VoucherTypeName",
        "AllLedgerEntries.LedgerName", "AllLedgerEntries.Amount"])
    result = {"sales": {}, "purchase": {}}
    for match in re.finditer(r"<VOUCHER\b[^>]*>(.*?)</VOUCHER>", xml, re.I | re.S):
        body = match.group(1); raw_date = tag_value(body, "DATE")
        if not re.fullmatch(r"\d{8}", raw_date):
            continue
        month = raw_date[:6]
        entries = []
        for entry in re.finditer(r"<ALLLEDGERENTRIES\.LIST>(.*?)</ALLLEDGERENTRIES\.LIST>", body, re.I | re.S):
            part = entry.group(1); entries.append((tag_value(part, "LEDGERNAME"), amount(tag_value(part, "AMOUNT"))))
        mode = "sales" if any(parents.get(name.lower()) == "sales accounts" for name, _ in entries) else ("purchase" if any(parents.get(name.lower()) == "purchase accounts" for name, _ in entries) else "")
        if not mode:
            continue
        total = result[mode].setdefault(month, {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0})
        for name, value in entries:
            parent = parents.get(name.lower(), ""); lname = name.lower(); val = abs(value)
            if parent == f"{mode} accounts": total["taxable_value"] += val
            elif "igst" in lname: total["igst"] += val
            elif "cgst" in lname: total["cgst"] += val
            elif "sgst" in lname: total["sgst"] += val
    for groups in result.values():
        for values in groups.values():
            for key in values: values[key] = round(values[key], 2)
    return {"company": cache.get("company", ""), **result}


def normalized_words(value):
    ignored = {"UPI", "NEFT", "IMPS", "RTGS", "BANK", "PAYMENT", "TRANSFER", "DR", "CR", "P2A", "P2M"}
    words = re.findall(r"[A-Z0-9]{3,}", str(value or "").upper())
    return [word for word in words if word not in ignored and not word.isdigit()]


def tally_ledger_lookup(ledgers=None):
    lookup = {}
    for ledger in (ledgers if ledgers is not None else TALLY_CACHE.get("ledgers", [])):
        name = clean_ledger_name(ledger.get("name", ""))
        if name:
            lookup[name.casefold()] = ledger_exact_name(ledger) or name
    return lookup


def match_tally_ledger(text, ledgers=None):
    ledgers = ledgers if ledgers is not None else TALLY_CACHE.get("ledgers", [])
    query = clean_ledger_name(text).casefold()
    if query:
        for ledger in ledgers:
            name = clean_ledger_name(ledger.get("name", ""))
            if name.casefold() == query:
                return ledger, 1.0
    source = set(normalized_words(text))
    if not source:
        return None, 0.0
    best, best_score = None, 0.0
    for ledger in ledgers:
        name = clean_ledger_name(ledger.get("name", ""))
        name_words = set(normalized_words(name))
        if not name_words:
            continue
        overlap_words = source & name_words
        overlap = len(overlap_words)
        # A single shared word like INDIA must not map GOOGLE receipts to
        # unrelated ledgers such as Central Bank of India.
        if overlap < 2 and name.casefold() not in query and query not in name.casefold():
            continue
        score = overlap / len(name_words)
        coverage = overlap / len(source)
        score = min(1.0, score * 0.65 + coverage * 0.35)
        if len(name) >= 6 and name.casefold() in query:
            score = max(score, 0.92)
        if score > best_score:
            best, best_score = ledger, score
    return (best, round(best_score, 3)) if best_score >= 0.55 else (None, 0.0)


def resolve_tally_ledger_name(name, ledgers=None):
    cleaned = clean_ledger_name(name)
    if not cleaned:
        return cleaned
    lookup = tally_ledger_lookup(ledgers)
    exact = lookup.get(cleaned.casefold())
    if exact:
        return exact
    ledger, score = match_tally_ledger(cleaned, ledgers)
    if ledger and score >= 0.5:
        return ledger_exact_name(ledger) or cleaned
    fetched = fetch_tally_ledger_exact(cleaned)
    return fetched or cleaned


def resolve_row_ledgers(rows):
    ledgers = TALLY_CACHE.get("ledgers", [])
    for row in rows:
        row["bank_ledger"] = resolve_tally_ledger_name(row.get("bank_ledger", ""), ledgers)
        row["counter_ledger"] = resolve_tally_ledger_name(row.get("counter_ledger", ""), ledgers)
    return rows


def suggested_party(particulars):
    text = clean_ledger_name(particulars)
    upper = text.upper()
    if "/" in text:
        candidate = text.rstrip("/").split("/")[-1].strip()
        if candidate and not re.fullmatch(r"\d+", candidate):
            return candidate[:80]
    for prefix in ("BY TRF ", "TRF TO ", "BY TRANSFER ", "TO TRANSFER "):
        pos = upper.find(prefix)
        if pos >= 0:
            return text[pos + len(prefix):].strip()[:80]
    return ""


def bank_voucher_number(row):
    """Cheque/NEFT/RTGS/UPI reference for Payment or Receipt No.; blank if none."""
    particulars = str(row.get("particulars") or "")
    upper = particulars.upper()

    instrument = re.sub(r"\s+", "", str(row.get("instrument") or ""))
    if re.fullmatch(r"\d{5,}", instrument):
        return instrument

    for pattern in (
        r"NEFT[^A-Z0-9]*(?:IN|OUT)?[^A-Z0-9]*(?:null//)?([A-Z0-9]{10,})",
        r"RTGS[^A-Z0-9]*(?:IN|OUT)?[^A-Z0-9]*(?:null//)?([A-Z0-9]{10,})",
        r"IMPS[^A-Z0-9]*(?:IN|OUT)?[^A-Z0-9]*(?:null//)?([A-Z0-9]{10,})",
        r"UPI/\d+/(\d{10,})",
        r"UPI[^A-Z0-9/]*(\d{10,})",
        r"//([A-Z]{4}[A-Z0-9]{10,})/",
        r"UTR[\s.:]*([A-Z0-9]{10,})",
        r"CHQ(?:\s*NO)?[\s.:]*(\d{5,})",
        r"CHEQUE[\s.:]*(\d{5,})",
    ):
        match = re.search(pattern, particulars, re.I)
        if match:
            return match.group(1).upper()

    reference = str(row.get("reference") or "").strip().lstrip(":")
    if reference:
        if reference.upper().startswith("UPI/"):
            for part in reference.split("/")[1:]:
                digits = re.sub(r"\D", "", part)
                if len(digits) >= 10:
                    return digits
        elif any(token in upper for token in ("NEFT", "RTGS", "CHQ", "CHEQUE", "IMPS", "UPI")):
            cleaned = reference.upper()
            if cleaned.startswith("UPI/"):
                for part in cleaned.split("/")[1:]:
                    digits = re.sub(r"\D", "", part)
                    if len(digits) >= 10:
                        return digits
            else:
                return cleaned
        elif re.fullmatch(r"\d{5,}", reference):
            return reference
    return ""


def classify(row, bank_ledger):
    particulars = row["particulars"].strip()
    upper = particulars.upper()
    debit, credit = row["debit"], row["credit"]
    if "BY CASH" in upper or "CASH DEPOSIT" in upper:
        voucher, ledger, confidence = "Contra", "Cash", "High"
    elif "CASH WITHDRAW" in upper or "TO CASH" in upper:
        voucher, ledger, confidence = "Contra", "Cash", "High"
    elif "CHARGE" in upper or "CHRG" in upper:
        voucher, ledger, confidence = "Payment", "Bank Charges", "High"
    elif "GOOGLE INDIA DIGITAL" in upper.replace(" ", "") or re.search(r"GOOGLE\s*INDIA\s*DIGITAL", upper):
        voucher = "Receipt" if credit > 0 else "Payment"
        ledger, confidence = "Google India Digital", "High"
    elif credit > 0:
        voucher, ledger, confidence = "Receipt", "Suspense", "High"
    else:
        voucher, ledger, confidence = "Payment", "Suspense", "High"
    ref_match = re.search(
        r"(PUNB[RN]\d+|UPI/\d+|:\d{9,}|//([A-Z0-9]{10,})/|NEFT[^/]*//([A-Z0-9]{10,})|RTGS[^/]*//([A-Z0-9]{10,}))",
        particulars,
        re.I,
    )
    if ref_match:
        reference = next((group.lstrip(":") for group in ref_match.groups()[1:] if group), ref_match.group(1).lstrip(":"))
    else:
        reference = row.get("reference", "")
    if voucher == "Contra":
        narration = f"Being cash transaction - {particulars}."
    elif ledger == "Bank Charges":
        narration = f"Being bank charges debited - {particulars}."
    else:
        direction = "received" if credit else "paid"
        narration = f"Being amount {direction} - {particulars}"
        if reference:
            narration += f", Ref: {reference}"
        narration += "."
    return {
        **row, "reference": reference, "voucher_type": voucher, "bank_ledger": clean_ledger_name(bank_ledger),
        "counter_ledger": clean_ledger_name(ledger), "narration": narration, "confidence": confidence,
        "approval": "Ready" if confidence == "High" else "Pending",
    }


def load_import_history():
    try:
        data = json.loads(IMPORT_HISTORY_PATH.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except (OSError, ValueError):
        return []


def save_import_history(history):
    IMPORT_HISTORY_PATH.write_text(json.dumps(history, indent=2), encoding="utf-8")


def parse_pnb_text(text, bank_ledger):
    opening_match = re.search(r"Opening Balance\s*:\s*([\d,]+\.\d{2})(Cr|Dr)", text)
    if not opening_match:
        raise ValueError("PNB opening balance was not found.")
    opening = amount(opening_match.group(1)) * (1 if opening_match.group(2) == "Cr" else -1)
    normalized = re.sub(r"\s+", " ", text)
    pattern = re.compile(
        r"(\d{2}-\d{2}-\d{4})\s+(\d{2}-\d{2}-\d{4})\s+(.+?)\s+"
        r"([\d,]+\.\d{2})(Cr|Dr)\s+(\S+)(?:\s+(\S+))?"
        r"(?=\s+\d{2}-\d{2}-\d{4}\s+\d{2}-\d{2}-\d{4}|\s*-{10,}|\s+Page\s+\d+\s+of\s+\d+|$)"
    )
    parsed = []
    for index, match in enumerate(pattern.finditer(normalized)):
        gl_date, value_date, body, balance_text, side, _, _ = match.groups()
        balance = amount(balance_text) * (1 if side == "Cr" else -1)
        tokens = list(re.finditer(r"[\d,]+\.\d{2}", body))
        if tokens:
            token = tokens[-1]
            body = (body[: token.start()] + body[token.end() :]).strip()
        instrument = ""
        inst = re.match(r"^(\d{5,})\s+(.*)$", body)
        if inst:
            instrument, body = inst.groups()
        parsed.append({
            "index": index,
            "date": datetime.strptime(gl_date, "%d-%m-%Y").strftime("%Y-%m-%d"),
            "value_date": datetime.strptime(value_date, "%d-%m-%Y").strftime("%Y-%m-%d"),
            "instrument": instrument,
            "particulars": body.strip(),
            "reference": "",
            "balance": balance,
            "balance_available": True,
        })
    if not parsed:
        raise ValueError("No PNB transactions were detected.")
    newest_first = parsed[0]["date"] > parsed[-1]["date"]
    parsed.sort(key=lambda row: (row["date"], -row["index"] if newest_first else row["index"]))
    rows = []
    previous = opening
    for row in parsed:
        change = round(row["balance"] - previous, 2)
        if not change:
            continue
        base = {
            **row,
            "debit": abs(change) if change < 0 else 0,
            "credit": change if change > 0 else 0,
        }
        rows.append(classify(base, bank_ledger))
        previous = row["balance"]
    if not rows:
        raise ValueError("No PNB transactions were detected.")
    return rows, opening, previous


def parse_pnb_cc_statement(text, bank_ledger):
    """Parse PNB credit-card style 'Statement of Account' PDFs (Amount/Type/Balance rows).

    Use the statement Amount + DR/CR columns directly. Balance-delta derivation
    merges distinct rows when several closure/folio entries land on 0.00 (for
    example 65,774 + 217.12 becomes 65,991.12).
    """
    pattern = re.compile(
        r"(\d{2}/\d{2}/\d{4})\s+([\d,]+(?:\.\d+)?)\s+(DR|CR)\s+(-?[\d,]+(?:\.\d+)?)\s*"
        r"(.*?)(?=\n\d{2}/\d{2}/\d{4}\s|\nDate:\s|\Z)",
        re.I | re.S,
    )
    parsed = []
    for index, match in enumerate(pattern.finditer(text)):
        date, txn_amount, side, balance_text, remarks = match.groups()
        value = amount(txn_amount)
        if not value:
            continue
        side = side.upper()
        parsed.append({
            "index": index,
            "date": datetime.strptime(date, "%d/%m/%Y").strftime("%Y-%m-%d"),
            "txn_amount": value,
            "side": side,
            "balance": amount(balance_text),
            "particulars": re.sub(r"\s+", " ", remarks).strip(),
            "instrument": "",
            "reference": "",
            "balance_available": True,
            "debit": value if side == "DR" else 0,
            "credit": value if side == "CR" else 0,
        })
    if not parsed:
        raise ValueError("No PNB CC transactions were detected.")
    newest_first = parsed[0]["date"] > parsed[-1]["date"]
    parsed.sort(key=lambda row: (row["date"], -row["index"] if newest_first else row["index"]))
    rows = [classify(row, bank_ledger) for row in parsed]
    opening = round(rows[0]["balance"] + rows[0]["debit"] - rows[0]["credit"], 2)
    closing = round(rows[-1]["balance"], 2)
    return rows, opening, closing


def parse_sbi_pdf(reader, bank_ledger):
    pages = [page.extract_text() or "" for page in reader.pages]
    full_text = "\n".join(pages)
    opening_match = re.search(r"Balance\s+as\s+on\s+.+?\s*:\s*([\d,]+\.\d{2})", full_text, re.I)
    if not opening_match:
        raise ValueError("SBI opening balance was not found.")
    opening = amount(opening_match.group(1))
    previous = opening
    rows = []
    pattern = re.compile(r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(.*?)(?=\n\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+|\Z)", re.S)
    for page_text in pages:
        if "Txn Date" not in page_text or "Balance" not in page_text:
            continue
        for match in pattern.finditer(page_text):
            txn_date, value_date, body = match.groups()
            body = re.sub(r"\s+", " ", body).strip()
            money_tokens = list(re.finditer(r"[\d,]+\.\d{2}", body))
            if len(money_tokens) < 2:
                continue
            txn_token, balance_token = money_tokens[-2], money_tokens[-1]
            balance = amount(balance_token.group())
            change = round(balance - previous, 2)
            if not change:
                continue
            particulars = re.sub(r"\s+\d{4,6}\s*$", "", body[:txn_token.start()].strip()).strip()
            ref_match = re.search(r"(SBIN\d+|CT[A-Z0-9]+|NEFT\s+INB:\s*[A-Z0-9]+)", particulars, re.I)
            base = {
                "date": datetime.strptime(txn_date, "%d/%m/%Y").strftime("%Y-%m-%d"),
                "value_date": datetime.strptime(value_date, "%d/%m/%Y").strftime("%Y-%m-%d"),
                "instrument": "",
                "particulars": particulars,
                "reference": ref_match.group(1) if ref_match else "",
                "debit": abs(change) if change < 0 else 0,
                "credit": change if change > 0 else 0,
                "balance": balance,
                "balance_available": True,
            }
            rows.append(classify(base, bank_ledger))
            previous = balance
    if not rows:
        raise ValueError("No SBI transactions were detected.")
    return rows, opening, previous

def rows_from_grid(grid, bank_ledger):
    if not grid:
        raise ValueError("The file has no rows.")
    headers = [str(x or "").strip().lower() for x in grid[0]]
    aliases = {
        "date": ["date", "transaction date", "txn date", "gl date", "booking date", "booking dt", "posting date", "posting dt"],
        "value_date": ["value date", "valuedate"],
        "particulars": ["particulars", "description", "narration", "remarks", "transaction details", "memo", "memo text"],
        "debit": ["debit", "withdrawal", "withdrawal amount", "dr amount", "paid out", "payment"],
        "credit": ["credit", "deposit", "deposit amount", "cr amount", "money in", "receipt"],
        "balance": ["balance", "closing balance", "running balance", "closing total"],
        "reference": ["reference", "ref no", "utr", "cheque no", "transaction id"],
    }
    indexes = {key: next((i for i, h in enumerate(headers) if h in names), None) for key, names in aliases.items()}
    if indexes["date"] is None or indexes["particulars"] is None:
        raise ValueError("Date and Particulars/Description columns are required.")
    output = []
    for source in grid[1:]:
        if not any(x not in (None, "") for x in source):
            continue
        def get(key, default=""):
            idx = indexes[key]
            return source[idx] if idx is not None and idx < len(source) else default
        raw_date = get("date")
        if isinstance(raw_date, datetime):
            date_text = raw_date.strftime("%Y-%m-%d")
        else:
            date_text = str(raw_date).strip()
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    date_text = datetime.strptime(date_text, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    pass
        base = {
            "date": date_text, "value_date": str(get("value_date", date_text))[:10],
            "instrument": "", "particulars": str(get("particulars")).strip(),
            "reference": str(get("reference")).strip(), "debit": amount(get("debit")),
            "credit": amount(get("credit")), "balance": amount(get("balance")),
            "balance_available": indexes["balance"] is not None,
        }
        output.append(classify(base, bank_ledger))
    return output


def serializable_grid(grid):
    output = []
    width = max((len(row) for row in grid), default=0)
    for row in grid:
        converted = []
        for value in row:
            if isinstance(value, datetime):
                converted.append(value.strftime("%Y-%m-%d"))
            elif value is None:
                converted.append("")
            else:
                converted.append(value)
        output.append(converted + [""] * (width - len(converted)))
    return output


def guessed_header_row(grid):
    keywords = {"date", "txn date", "transaction date", "booking date", "booking dt", "posting date", "posting dt", "description", "particulars", "narration", "memo", "memo text", "debit", "withdrawal amount", "paid out", "credit", "deposit amount", "money in", "balance", "closing total", "running balance"}
    best_row, best_score = 0, -1
    for index, row in enumerate(grid[:30]):
        labels = {str(value or "").strip().lower() for value in row}
        score = len(labels & keywords)
        if score > best_score:
            best_row, best_score = index, score
    return best_row


def icici_statement_grid(grid):
    """Collapse ICICI PDF line-wrapped remarks into one row per transaction."""
    if not grid:
        return None
    header = next((i for i, row in enumerate(grid[:20])
                   if "s no." in " ".join(str(v or "").lower() for v in row)
                   and "transaction date" in " ".join(str(v or "").lower() for v in row)), None)
    if header is None:
        return None
    out = [list(grid[header])]
    current = None
    date_re = re.compile(r"^\d{2}[./-]\d{2}[./-]\d{4}$")
    def finish(row):
        if not row:
            return
        # pdfplumber often places ICICI's remarks on a separate, non-tabular
        # line.  A non-empty particulars value is still required by the
        # review/mapping pipeline, so retain a stable fallback when the bank
        # did not expose the remark in the table extraction.
        if not str(row[3] or "").strip():
            row[3] = f"ICICI Bank transaction {row[0]}"
        out.append(row)

    for source in grid[header + 1:]:
        cells = [str(v or "").strip() for v in source]
        raw_date = cells[1] if len(cells) > 1 else ""
        is_txn = bool(date_re.fullmatch(raw_date)) and bool(re.fullmatch(r"\d+", cells[0] or ""))
        if is_txn:
            finish(current)
            current = cells[:7] + [""] * max(0, 7 - len(cells))
            current[3] = ""
            continue
        if current:
            text = " ".join(c for c in cells[:4] if c)
            if text and not re.search(
                r"transaction remarks|www\.icici\.bank\.in|never share your otp|"
                r"please call from your registered mobile|this is a system generated statement|"
                r"team icici bank",
                text,
                re.I,
            ):
                current[3] = (current[3] + " " + text).strip()
    finish(current)
    return out if len(out) > 1 else None


def icici_text_remarks(text):
    """Extract ICICI remarks from the PDF text layer (the table grid omits them)."""
    if not text:
        return []
    start = text.find("S No. Transaction")
    body = text[start:] if start >= 0 else text
    pattern = re.compile(
        r"(?ms)^\s*(\d+)\s+(\d{2}[./-]\d{2}[./-]\d{4})\s+(.*?)"
        r"(?=^\s*\d+\s+\d{2}[./-]\d{2}[./-]\d{4}\s+|\Z)"
    )
    remarks = []
    for match in pattern.finditer(body):
        value = re.sub(r"\s+", " ", match.group(3)).strip()
        money = list(re.finditer(r"[\d,]+\.\d{2}", value))
        if len(money) >= 2:
            value = value[: money[-2].start()].strip()
        value = re.sub(r"\s+", " ", value).strip(" -")
        if value and not re.search(r"never share your otp|transaction remarks|www\.icici", value, re.I):
            remarks.append(value)
    return remarks


def effective_mapping_header_row(grid, mapping, header_row):
    """Include a selected/saved row when it is actually the first transaction.

    Some statement PDFs do not expose a usable column-heading row. In that case
    the mapping screen allows the first transaction row to be selected. Older
    code always treated the selected row as headings and consequently dropped
    that first transaction (and used its closing balance as the opening balance).
    """
    header_row = int(header_row)
    if not (0 <= header_row < len(grid)):
        return header_row

    row = grid[header_row]

    def mapped_value(key):
        index = mapping.get(key)
        if index in (None, "", -1):
            return ""
        index = int(index)
        return row[index] if index < len(row) else ""

    is_transaction = (
        bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", parse_date(mapped_value("date"))))
        and bool(str(mapped_value("particulars") or "").strip())
        and any(
            abs(number(mapped_value(key))) > 0
            for key in ("debit", "credit", "amount", "balance")
        )
    )
    return header_row - 1 if is_transaction else header_row


def prepare_grid(name, grid, bank_ledger):
    grid = serializable_grid(grid)
    icici_grid = icici_statement_grid(grid)
    if icici_grid:
        rows = apply_mapping(
            icici_grid,
            {"date": 1, "particulars": 3, "debit": 4, "credit": 5,
             "balance": 6, "reference": 2},
            0, bank_ledger, classify,
        )
        return rows, {"format": "ICICI PDF (automatic mapping)"}
    profiles = load_profiles(MAPPINGS_PATH)
    suffix = Path(name).suffix.lower()
    for profile in profiles.values():
        if profile.get("suffix") != suffix:
            continue
        saved_headers = profile.get("headers", [])
        for header_row, row in enumerate(grid[:30]):
            headers = [str(value or "").strip().lower() for value in row]
            if headers == saved_headers:
                effective_header = effective_mapping_header_row(
                    grid, profile["mapping"], header_row
                )
                rows = apply_mapping(
                    grid, profile["mapping"], effective_header, bank_ledger, classify
                )
                return rows, {"format": "Saved mapping", "mapping_name": profile.get("name", "Saved")}
    header_row = guessed_header_row(grid)
    return [], {"mapping_required": True, "filename": name, "grid": grid, "header_row": header_row}


def bridge_scanned_ocr(raw, name="statement.pdf"):
    if TALLY_CONNECTOR_TOKEN:
        return connector_request("ocr", {
            "name": name, "data": base64.b64encode(raw).decode("ascii"),
        }, timeout=180)
    # Kept as a clear failure for deployments that have not yet configured
    # their outbound Connector.
    raise ValueError("Windows OCR requires the outbound Tally Connector.")
    """Legacy tunnel implementation retained below for migration reference."""
    request = urllib.request.Request(
        TALLY_BRIDGE_URL + "/ocr",
        data=json.dumps({"name": name, "data": base64.b64encode(raw).decode("ascii")}).encode("utf-8"),
        headers={"Content-Type": "application/json", "X-Bridge-Token": TALLY_BRIDGE_TOKEN},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            result = json.loads(response.read().decode("utf-8"))
        if result.get("error"):
            raise ValueError(result["error"])
        return result
    except urllib.error.URLError as exc:
        raise ValueError(
            "Local Tally Connector is not reachable for scanned-PDF OCR. "
            "Start the connector and check its tunnel URL and token."
        ) from exc


def parse_bank_of_baroda_pdf(text, bank_ledger):
    """Read Bank of Baroda's two-line statement table without column mapping."""
    record_start = re.compile(
        r"(?m)^(?P<balance>[\d,]+\.\d{2})(?P<serial>\d+)\s+"
        r"(?P<date>\d{2}-\d{2}-\d{4})\s+(?P<value_date>\d{2}-\d{2}-\d{4})\s+"
    )
    matches = list(record_start.finditer(text))
    rows = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        content = re.sub(r"\s+", " ", text[match.end():end]).split("This is a computer-generated statement", 1)[0].strip()
        amounts = re.match(r"^(?P<particulars>.*?)\s+(?P<debit>[\d,]+\.\d{2}|-)\s+(?P<credit>[\d,]+\.\d{2}|-)$", content)
        if not amounts:
            continue
        debit = 0 if amounts["debit"] == "-" else number(amounts["debit"])
        credit = 0 if amounts["credit"] == "-" else number(amounts["credit"])
        if not (debit or credit):
            continue
        rows.append(classify({
            "date": match["date"],
            "value_date": match["value_date"],
            "particulars": amounts["particulars"].strip(),
            "debit": debit,
            "credit": credit,
            "balance": number(match["balance"]),
        }, bank_ledger))
    if not rows:
        raise ValueError("Bank of Baroda transaction rows were not found.")
    opening_match = re.search(r"(?m)^(?P<opening>[\d,]+\.\d{2})1\s+\d{2}-\d{2}-\d{4}\s+Opening Balance", text)
    return rows, number(opening_match["opening"]) if opening_match else 0, rows[-1].get("balance", 0)


def parse_hdfc_pdf(text, bank_ledger):
    """Parse HDFC statements with Withdrawal/Deposit/Closing columns."""
    rows = []
    pattern = re.compile(r"(?ms)^(?P<date>\d{2}/\d{2}/\d{4})\s+(?P<body>.*?)\s+(?P<value_date>\d{2}/\d{2}/\d{4})\s+(?P<debit>[\d,]+\.\d{2}|-)\s+(?P<credit>[\d,]+\.\d{2}|-)\s+(?P<balance>[\d,]+\.\d{2})\s*$")
    for match in pattern.finditer(text):
        withdrawal, deposit, balance = number(match["debit"]), number(match["credit"]), number(match["balance"])
        body = re.sub(r"\s+", " ", match["body"]).strip()
        ref_match = re.search(r"\b[A-Z0-9]{8,}\b", body)
        rows.append(classify({
            "date": datetime.strptime(match["date"], "%d/%m/%Y").strftime("%Y-%m-%d"),
            "value_date": match["value_date"], "particulars": body,
            "reference": ref_match.group() if ref_match else "",
            "debit": withdrawal, "credit": deposit, "balance": balance,
        }, bank_ledger))
    if not rows:
        raise ValueError("HDFC transaction rows were not found.")
    opening = round(rows[0]["balance"] + rows[0]["debit"] - rows[0]["credit"], 2)
    closing = rows[-1]["balance"]
    return rows, opening, closing


def parse_file(name, raw, bank_ledger, password=""):
    suffix = Path(name).suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(io.BytesIO(raw))
        if reader.is_encrypted:
            if not password:
                return [], {"password_required": True, "filename": name}
            try:
                unlocked = reader.decrypt(password)
            except Exception:
                unlocked = 0
            if not unlocked:
                raise ValueError("Incorrect PDF password.")
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        if "Txn Date" in text and "Ref No./Cheque" in text and "Branch" in text:
            rows, opening, closing = parse_sbi_pdf(reader, bank_ledger)
            return rows, {"opening": opening, "closing": closing, "format": "SBI PDF"}
        if "Opening Balance" in text and (
            "Customer Account Ledger Report" in text or "Account Ledger Report" in text
        ):
            rows, opening, closing = parse_pnb_text(text, bank_ledger)
            return rows, {"opening": opening, "closing": closing, "format": "PNB PDF"}
        if re.search(r"Statement of Account:\d+", text, re.I) and re.search(
            r"Amount\s*\(\s*INR\s*\)", text, re.I
        ):
            rows, opening, closing = parse_pnb_cc_statement(text, bank_ledger)
            return rows, {"opening": opening, "closing": closing, "format": "PNB CC PDF"}
        if "Account Statement from" in text and "Description Cheque" in text and "Sr.No Transaction" in text:
            rows, opening, closing = parse_bank_of_baroda_pdf(text, bank_ledger)
            return rows, {"opening": opening, "closing": closing, "format": "Bank of Baroda PDF"}
        if "S No." in text and "Transaction Remarks" in text and "icici" in text.lower():
            rows, meta = prepare_grid(name, extract_pdf_grid(raw, password), bank_ledger)
            remarks = icici_text_remarks(text)
            for index, row in enumerate(rows):
                if index < len(remarks):
                    row["particulars"] = remarks[index]
                    row["narration"] = re.sub(
                        r"Being amount (?:paid|received|transferred|cash transaction)\s*-\s*.*?\.$",
                        lambda m: m.group(0).split(" - ")[0] + " - " + remarks[index] + ".",
                        row.get("narration", ""),
                    )
            meta["format"] = "ICICI PDF (automatic mapping)"
            return rows, meta
        if not text.strip() and TALLY_CONNECTOR_TOKEN:
            remote = bridge_scanned_ocr(raw, name)
            rows = remote.get("rows", [])
            return [classify(row, bank_ledger) for row in rows], {
                "opening": remote.get("opening", 0),
                "closing": remote.get("closing", 0),
                "format": "Scanned PDF (Local Bridge OCR)",
            }
        if not text.strip():
            rows, opening, closing = parse_scanned_bank_pdf(
                raw, APP_DIR / "windows_ocr.ps1"
            )
            return [classify(row, bank_ledger) for row in rows], {
                "opening": opening,
                "closing": closing,
                "format": "Scanned PDF (Windows OCR)",
            }
        if "Withdrawal Amount" in text and "Deposit Amount" in text and "Closing Balance" in text:
            rows, opening, closing = parse_hdfc_pdf(text, bank_ledger)
            return rows, {"opening": opening, "closing": closing, "format": "HDFC PDF (automatic mapping)"}
        grid = extract_pdf_grid(raw, password)
        return prepare_grid(name, grid, bank_ledger)
    if suffix == ".xls":
        grid = legacy_xls_grid(raw)
        try:
            return rows_from_grid(grid, bank_ledger), {"format": "Excel automatic"}
        except ValueError:
            return prepare_grid(name, grid, bank_ledger)
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        grid = list(wb.active.iter_rows(values_only=True))
        try:
            return rows_from_grid(grid, bank_ledger), {"format": "Excel automatic"}
        except ValueError:
            return prepare_grid(name, grid, bank_ledger)
    text = raw.decode("utf-8-sig", errors="replace")
    if "Customer Account Ledger Report" in text and "Opening Balance" in text:
        rows, opening, closing = parse_pnb_text(text, bank_ledger)
        return rows, {"opening": opening, "closing": closing, "format": "PNB Text"}
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
        grid = list(csv.reader(io.StringIO(text), dialect))
    except csv.Error:
        grid = [[line] for line in text.splitlines() if line.strip()]
    try:
        return rows_from_grid(grid, bank_ledger), {"format": "Text automatic"}
    except ValueError:
        return prepare_grid(name, grid, bank_ledger)

HEADERS = ["Date", "Value Date", "Instrument No.", "Bank Particulars", "Reference/UTR", "Debit", "Credit",
           "Balance", "Voucher Type", "Bank Ledger", "Counter Ledger", "Narration", "Confidence", "Approval"]


def make_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Voucher Review"
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get("date"), row.get("value_date"), row.get("instrument"), row.get("particulars"),
                   row.get("reference"), amount(row.get("debit")), amount(row.get("credit")),
                   amount(row.get("balance")), row.get("voucher_type"), row.get("bank_ledger"),
                   row.get("counter_ledger"), row.get("narration"), row.get("confidence"), row.get("approval")])
    fill = PatternFill("solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = fill, Font(bold=True, color="FFFFFF"), Alignment(wrap_text=True)
    for col in ("F", "G", "H"):
        for cell in ws[col][1:]:
            cell.number_format = "#,##0.00"
    for i, width in enumerate([13, 13, 16, 48, 28, 15, 15, 15, 15, 24, 28, 70, 12, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes, ws.auto_filter.ref = "A2", ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def make_gst_preview_xlsx(rows, title="GSTR-2 Preview"):
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\\/*?:\[\]]", " ", gst_text(title))[:31] or "GSTR-2 Preview"
    headers = ["GSTIN", "Party / Tally Ledger", "Item / Tally Stock Item", "Document Type",
               "Invoice No.", "Original Invoice Date", "GSTR-2B Period", "Tally Entry Date",
               "Invoice Value", "Taxable", "GST Rate", "IGST", "CGST", "SGST", "CESS", "Status"]
    ws.append(headers)
    for row in rows:
        taxable = gst_number(row.get("taxable_value"))
        tax = sum(gst_number(row.get(key)) for key in ("igst", "cgst", "sgst", "cess"))
        actual_rate = (tax * 100 / taxable) if taxable else 0
        gst_rate = min((0, 5, 12, 18, 28), key=lambda rate: abs(rate - actual_rate))
        allocations = row.get("sales_allocations") or []
        item_name = row.get("expense_ledger") or ", ".join(
            gst_text(item.get("item_name")) for item in allocations if gst_text(item.get("item_name"))
        )
        ws.append([row.get("gstin"), row.get("party_ledger") or row.get("party_name"), item_name,
                   row.get("document_type"), row.get("invoice_no"),
                   row.get("original_invoice_date") or row.get("invoice_date"), row.get("gstr2b_period"),
                   row.get("tally_entry_date"), gst_number(row.get("invoice_value")), taxable, gst_rate,
                   gst_number(row.get("igst")), gst_number(row.get("cgst")), gst_number(row.get("sgst")),
                   gst_number(row.get("cess")), row.get("status")])
    fill = PatternFill("solid", fgColor="176B4A")
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = fill, Font(bold=True, color="FFFFFF"), Alignment(wrap_text=True)
    for column in range(9, 16):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            cell[0].number_format = "#,##0.00"
    widths = [18, 28, 28, 22, 18, 20, 15, 18, 15, 15, 11, 13, 13, 13, 13, 22]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def make_gst_summary_xlsx(rows, title="GST Summary"):
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\\/*?:\[\]]", " ", gst_text(title))[:31] or "GST Summary"
    headers = list((rows or [{}])[0].keys()) if rows else ["Particulars"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    fill = PatternFill("solid", fgColor="493184")
    for cell in ws[1]:
        cell.fill, cell.font = fill, Font(bold=True, color="FFFFFF")
    for index, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(index)].width = max(14, min(36, len(str(header)) + 6))
    ws.freeze_panes, ws.auto_filter.ref = "A2", ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def tally_master_name(value):
    """Unescape Tally XML name entities but keep CR/LF that are part of the master identity."""
    return html.unescape(str(value or ""))


def clean_ledger_name(value):
    text = tally_master_name(value)
    text = text.replace("&#13;&#10;", " ").replace("&#10;", " ").replace("&#13;", " ")
    text = re.sub(r"[\r\n\t\x04\x0b\f\u200b-\u200d\ufeff]+", " ", text)
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", " ", text).strip()


def ledger_exact_name(ledger):
    if not isinstance(ledger, dict):
        return tally_master_name(ledger)
    return ledger.get("exact_name") or tally_master_name(ledger.get("name", ""))


def register_tally_ledger(name, parent="", exact_name=""):
    cleaned = clean_ledger_name(name)
    if not cleaned:
        return cleaned
    exact = tally_master_name(exact_name or name) or cleaned
    ledgers = TALLY_CACHE.setdefault("ledgers", [])
    lookup = tally_ledger_lookup(ledgers)
    if cleaned.casefold() not in lookup:
        ledgers.append({
            "name": cleaned,
            "exact_name": exact,
            "parent": clean_ledger_name(parent),
        })
    else:
        for item in ledgers:
            if clean_ledger_name(item.get("name", "")).casefold() == cleaned.casefold():
                item["exact_name"] = exact
                break
    return cleaned


def fetch_tally_ledger_exact(name):
    cleaned = clean_ledger_name(name)
    if not cleaned:
        return ""
    # Try the cleaned name first, then common hidden trailing CR/LF variants that
    # appear in some Tally masters (visible in UI, invisible to plain XML lookup).
    candidates = [cleaned, cleaned + "\r\n", cleaned + "\n", cleaned + "\r"]
    for candidate in candidates:
        raw = (
            '<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>'
            '<TYPE>Object</TYPE><SUBTYPE>Ledger</SUBTYPE>'
            f'<ID TYPE="Name">{xml_escape_tally_master(candidate)}</ID>'
            '</HEADER><BODY></BODY></ENVELOPE>'
        ).encode("utf-8")
        try:
            response = tally_post(raw, timeout=30)
        except ValueError:
            return ""
        if re.search(r"Could not find Ledger", response, re.I):
            continue
        match = re.search(r'<LEDGER\b[^>]*\bNAME="([^"]*)"', response, re.I)
        if match:
            exact = tally_master_name(match.group(1))
            display = clean_ledger_name(exact)
            if display:
                parent = tag_value(response, "PARENT")
                register_tally_ledger(display, parent, exact_name=exact)
                return exact
        alt_exact = tally_master_name(tag_value(response, "NAME"))
        alt = clean_ledger_name(alt_exact)
        if alt:
            register_tally_ledger(alt, exact_name=alt_exact)
            return alt_exact or alt
    return ""


def xml_escape_tally_master(value):
    """Escape a Tally master name for XML, preserving CR/LF as numeric entities."""
    text = tally_master_name(value)
    return (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;")
            .replace("\r", "&#13;").replace("\n", "&#10;"))


def xml_escape(value, *, ledger=False):
    text = clean_ledger_name(value) if ledger else str(value or "")
    return (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;"))


def existing_tally_ledgers():
    return {
        clean_ledger_name(item.get("name", "")).casefold()
        for item in TALLY_CACHE.get("ledgers", [])
        if item.get("name")
    }


def missing_bank_counter_ledgers(rows):
    existing = existing_tally_ledgers()
    # Cash is built in and "Select Ledger" is only a placeholder. Suspense and
    # Bank Charges are ordinary ledger masters and must be created when absent.
    skip = {"cash", "select ledger"}
    missing = {}
    for row in rows:
        if str(row.get("approval", row.get("status", "Ready"))).lower() != "ready":
            continue
        counter = clean_ledger_name(row.get("counter_ledger", ""))
        bank = clean_ledger_name(row.get("bank_ledger", ""))
        if not counter or counter.casefold() in skip or counter.casefold() == bank.casefold():
            continue
        if counter.casefold() in existing:
            continue
        selected_voucher = row.get("voucher_type", "Payment")
        credit = amount(row.get("credit"))
        debit = amount(row.get("debit"))
        if selected_voucher == "Contra":
            continue
        counter_key = counter.casefold()
        if counter_key == "suspense":
            parent = "Suspense A/c"
        elif counter_key == "bank charges":
            parent = "Indirect Expenses"
        elif selected_voucher == "Receipt" or credit:
            parent = "Sundry Debtors"
        else:
            parent = "Sundry Creditors"
        missing[counter] = parent
    return missing


def tally_parent_candidates(prefer_creditors=False):
    counts = {}
    for ledger in TALLY_CACHE.get("ledgers", []):
        parent = clean_ledger_name(ledger.get("parent", ""))
        if parent:
            counts[parent] = counts.get(parent, 0) + 1
    ordered = sorted(counts.keys(), key=lambda name: (-counts[name], name))
    preferred = []
    if prefer_creditors:
        preferred = [name for name in ordered if "creditor" in name.casefold() or "liabilit" in name.casefold()]
    else:
        preferred = [name for name in ordered if "debtor" in name.casefold() or "asset" in name.casefold()]
    fallback = (
        ["Sundry Creditors", "Current Liabilities", "Sundry Debtors", "Current Assets"]
        if prefer_creditors else
        ["Sundry Debtors", "Current Assets", "Sundry Creditors", "Current Liabilities"]
    )
    result = []
    for name in preferred + fallback + ordered:
        if name and name not in result:
            result.append(name)
    return result


def extract_missing_ledgers_from_details(details):
    names = []
    for detail in details or []:
        for match in re.finditer(r"Ledger '([^']+)' does not exist", detail, re.I):
            name = clean_ledger_name(match.group(1))
            if name and name not in names:
                names.append(name)
    return names


def make_single_ledger_create_xml(name, parent):
    message = (
        f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(name, ledger=True)}" ACTION="Create">'
        f'<NAME>{xml_escape(name, ledger=True)}</NAME>'
        f'<PARENT>{xml_escape(parent)}</PARENT>'
        f'<ISBILLWISEON>Yes</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
        f'</LEDGER></TALLYMESSAGE>'
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>'
        '</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>All Masters</REPORTNAME></REQUESTDESC>'
        f'<REQUESTDATA>{message}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>'
    ).encode("utf-8")


def make_ledger_rename_xml(old_name, new_name):
    """Alter a Tally ledger master name (used to strip hidden CR/LF from names)."""
    message = (
        f'<TALLYMESSAGE xmlns:UDF="TallyUDF">'
        f'<LEDGER NAME="{xml_escape_tally_master(old_name)}" ACTION="Alter">'
        f'<NAME>{xml_escape(new_name, ledger=True)}</NAME>'
        f'<LANGUAGENAME.LIST><NAME.LIST TYPE="String">'
        f'<NAME>{xml_escape(new_name, ledger=True)}</NAME>'
        f'</NAME.LIST><LANGUAGEID>1033</LANGUAGEID></LANGUAGENAME.LIST>'
        f'</LEDGER></TALLYMESSAGE>'
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>'
        '</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>All Masters</REPORTNAME></REQUESTDESC>'
        f'<REQUESTDATA>{message}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>'
    ).encode("utf-8")


def repair_dirty_tally_ledger_names(rows):
    """Rename ledgers whose Tally master name contains hidden CR/LF so voucher import matches."""
    sync_tally()
    needed = set()
    for row in rows:
        for key in ("counter_ledger", "bank_ledger", "party_ledger"):
            cleaned = clean_ledger_name(row.get(key, ""))
            if cleaned:
                needed.add(cleaned.casefold())
    notes = []
    for ledger in list(TALLY_CACHE.get("ledgers", [])):
        cleaned = clean_ledger_name(ledger.get("name", ""))
        exact = ledger_exact_name(ledger)
        if not cleaned or cleaned.casefold() not in needed:
            continue
        if exact == cleaned or not re.search(r"[\r\n\x04]", exact):
            continue
        raw = make_ledger_rename_xml(exact, cleaned)
        tally_response = tally_post(raw, timeout=75)
        (DATA_DIR / "tally_last_ledger_repair_request.xml").write_bytes(raw)
        (DATA_DIR / "tally_last_ledger_repair_response.xml").write_text(
            tally_response, encoding="utf-8", errors="replace"
        )
        result = tally_import_result(tally_response)
        if result["errors"]:
            detail = result["details"][0] if result["details"] else "rename failed"
            notes.append(f"Could not fix hidden characters in ledger '{cleaned}': {detail}")
            continue
        register_tally_ledger(cleaned, ledger.get("parent", ""), exact_name=cleaned)
        notes.append(f"Fixed ledger name '{cleaned}' (removed hidden line break).")
    if notes:
        sync_tally()
    return notes


def ensure_bank_counter_ledgers(rows):
    sync_tally()
    notes = repair_dirty_tally_ledger_names(rows)
    missing = missing_bank_counter_ledgers(rows)
    if not missing:
        return notes
    for name, default_parent in sorted(missing.items()):
        fetched = fetch_tally_ledger_exact(name)
        if fetched:
            display = clean_ledger_name(fetched)
            if fetched != display and re.search(r"[\r\n\x04]", fetched):
                repair_notes = repair_dirty_tally_ledger_names(
                    [{"counter_ledger": display, "bank_ledger": "", "party_ledger": ""}]
                )
                notes.extend(repair_notes)
                notes.append(f"Matched existing Tally ledger '{display}'.")
            else:
                notes.append(f"Matched existing Tally ledger '{display or fetched}'.")
            continue
        prefer_creditors = default_parent == "Sundry Creditors"
        parents = tally_parent_candidates(prefer_creditors=prefer_creditors)
        if default_parent not in parents:
            parents.insert(0, default_parent)
        created = False
        for parent in parents:
            raw = make_single_ledger_create_xml(name, parent)
            tally_response = tally_post(raw, timeout=75)
            (DATA_DIR / "tally_last_ledger_request.xml").write_bytes(raw)
            (DATA_DIR / "tally_last_ledger_response.xml").write_text(
                tally_response, encoding="utf-8", errors="replace"
            )
            result = tally_import_result(tally_response)
            sync_tally()
            if name.casefold() in existing_tally_ledgers():
                notes.append(f"Created ledger '{name}' under '{parent}'.")
                created = True
                break
            if result["details"]:
                notes.append(f"Tally rejected '{name}' under '{parent}': {result['details'][0]}")
        if not created:
            notes.append(f"Could not create ledger '{name}' in Tally. Create it manually, then send again.")
    return notes


def make_tally_xml(rows, batch_id="", return_records=False):
    non_party = {"cash", "bank charges", "suspense", "select ledger"}

    def entry_xml(name, value, positive, is_party=False):
        party_flag = "<ISPARTYLEDGER>Yes</ISPARTYLEDGER>" if is_party else ""
        return (
            f"<ALLLEDGERENTRIES.LIST><LEDGERNAME>{xml_escape_tally_master(name)}</LEDGERNAME>"
            f"<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE>{party_flag}"
            f"<AMOUNT>{value:.2f}</AMOUNT></ALLLEDGERENTRIES.LIST>"
        )

    vouchers = []
    records = []
    for row in rows:
        if str(row.get("approval", row.get("status", "Ready"))).lower() != "ready":
            continue
        selected_voucher = row.get("voucher_type", "Payment")
        bank, counter = (
            resolve_tally_ledger_name(row.get("bank_ledger", "")),
            resolve_tally_ledger_name(row.get("counter_ledger", "")),
        )
        debit, credit = amount(row.get("debit")), amount(row.get("credit"))
        total = credit or debit
        if not bank or not counter or "Select Ledger" in clean_ledger_name(counter) or not total:
            continue
        # The running balance is the source of truth for bank direction. Some
        # narrations contain misleading words (for example "DR" inside a
        # recharge reference), so a stale Receipt/Payment label must never flip
        # the bank ledger side during export.
        voucher = selected_voucher if selected_voucher == "Contra" else ("Receipt" if credit else "Payment")
        party_name = counter if voucher in ("Payment", "Receipt") and clean_ledger_name(counter).casefold() not in non_party else ""
        if credit:
            entries = [(bank, -total, "Yes", False), (counter, total, "No", bool(party_name))]
        else:
            entries = [(counter, -total, "Yes", bool(party_name)), (bank, total, "No", False)]
        ledger_xml = "".join(entry_xml(name, value, positive, is_party) for name, value, positive, is_party in entries)
        party_xml = (
            f"<PARTYLEDGERNAME>{xml_escape_tally_master(party_name)}</PARTYLEDGERNAME>"
            if party_name else ""
        )
        # Tally expects voucher dates as YYYYMMDD. Bank statements commonly
        # provide DD-MM-YYYY; passing that through as DDMMYYYY makes Tally
        # report the voucher date as missing/invalid.
        raw_date = str(row.get("date", "")).strip().replace("/", "-")
        date_match = re.fullmatch(r"(\d{2})-(\d{2})-(\d{4})", raw_date)
        if date_match:
            date = f"{date_match.group(3)}{date_match.group(2)}{date_match.group(1)}"
        else:
            date = raw_date.replace("-", "")
        index = len(records) + 1
        tracking_number = f"B2T-{batch_id[:6].upper()}-{index:05d}" if batch_id else ""
        tally_number = bank_voucher_number(row)
        remote_id = str(uuid.uuid5(
            uuid.NAMESPACE_URL,
            f"bank2tally:{batch_id}:{index}:{date}:{clean_ledger_name(bank)}:{clean_ledger_name(counter)}:{total:.2f}",
        )) if batch_id else ""
        identity = (
            f' REMOTEID="{remote_id}"'
            if remote_id else ""
        )
        number_xml = f"<VOUCHERNUMBER>{xml_escape(tally_number)}</VOUCHERNUMBER>" if tally_number else ""
        vouchers.append(
            f"<TALLYMESSAGE xmlns:UDF=\"TallyUDF\"><VOUCHER{identity} VCHTYPE=\"{xml_escape(voucher)}\" ACTION=\"Create\">"
            f"<DATE>{date}</DATE><VOUCHERTYPENAME>{xml_escape(voucher)}</VOUCHERTYPENAME>"
            f"{number_xml}{party_xml}<PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW><NARRATION>{xml_escape(row.get('narration'))}</NARRATION>"
            f"{ledger_xml}</VOUCHER></TALLYMESSAGE>")
        records.append({
            "remote_id": remote_id, "voucher_number": tally_number or tracking_number,
            "date": date, "voucher_type": voucher, "counter_ledger": clean_ledger_name(counter),
        })
    raw = ("<?xml version=\"1.0\" encoding=\"UTF-8\"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>"
            "</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>"
            f"<REQUESTDATA>{''.join(vouchers)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>").encode("utf-8")
    return (raw, records) if return_records else raw


def make_tally_delete_xml(records):
    vouchers = []
    for record in records:
        vouchers.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER REMOTEID="{xml_escape(record.get("remote_id"))}" '
            f'DATE="{xml_escape(record.get("date"))}" TAGNAME="Voucher Number" '
            f'TAGVALUE="{xml_escape(record.get("voucher_number"))}" '
            f'VCHTYPE="{xml_escape(record.get("voucher_type"))}" ACTION="Delete">'
            f'<VOUCHERNUMBER>{xml_escape(record.get("voucher_number"))}</VOUCHERNUMBER>'
            f'</VOUCHER></TALLYMESSAGE>'
        )
    return ("<?xml version=\"1.0\" encoding=\"UTF-8\"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>"
            "</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>"
            f"<REQUESTDATA>{''.join(vouchers)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>").encode("utf-8")


def gst_text(value):
    return "" if value is None else str(value).strip()


def gst_party_ledger(row):
    party = gst_text(row.get("party_ledger") or row.get("party_name"))
    if not party or party.lower() in {
        "cash", "cash sales and purchase", "cash sales & purchase", "cash sale", "cash sales",
    }:
        return "Cash"
    return party


def gst_number(value):
    try:
        return round(float(str(value).replace(",", "").strip()), 2) if value not in (None, "") else 0.0
    except ValueError:
        return 0.0


def infer_gst_period(name):
    """Infer MMYYYY return period from common GST Portal/export file names."""
    text = Path(gst_text(name)).name.upper()
    matches = re.findall(r"(?<!\d)(0[1-9]|1[0-2])[-_]?((?:20)?\d{2})(?!\d)", text)
    if matches:
        month, year = matches[-1]
        return f"{month}{year if len(year) == 4 else '20' + year}"
    matches = re.findall(r"(?<!\d)(20\d{2})[-_]?(0[1-9]|1[0-2])(?!\d)", text)
    if matches:
        year, month = matches[-1]
        return f"{month}{year}"
    return ""


def tally_date_for_return_period(invoice_date, return_period):
    """
    Tally voucher DATE for invoices available in GSTR-2B:
    - invoice month equals the GSTR-2B month: original invoice date;
    - invoice belongs to another/earlier month: first day of GSTR-2B month.

    Supplier invoice date (REFERENCEDATE) stays on the original invoice date
    and is handled separately by make_gst_purchase_xml.
    """
    period = re.fullmatch(r"(0[1-9]|1[0-2])(20\d{2})", gst_text(return_period))
    if not period:
        return gst_text(invoice_date)
    month, year = int(period.group(1)), int(period.group(2))
    invoice_text = gst_text(invoice_date).replace("/", "-")
    invoice_parsed = None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%b-%Y", "%d-%b-%y"):
        try:
            invoice_parsed = datetime.strptime(invoice_text, fmt)
            break
        except ValueError:
            pass
    if invoice_parsed and (invoice_parsed.month, invoice_parsed.year) == (month, year):
        return invoice_parsed.strftime("%d-%m-%Y")
    return datetime(year, month, 1).strftime("%d-%m-%Y")


def gst_pick(record, *names):
    if not isinstance(record, dict):
        return ""
    lowered = {str(key).strip().lower(): value for key, value in record.items()}
    return next((lowered[name.lower()] for name in names if lowered.get(name.lower()) not in (None, "")), "")


def _gst_missing_field_warn(field, row=None):
    invoice = gst_text((row or {}).get("invoice_no") or (row or {}).get("inum") or "?")
    print(f"[GST] warning: missing '{field}' on invoice '{invoice}'; using safe default.")


def ensure_gst_invoice_fields(row):
    """
    Normalize GST invoice amount/identity fields without assuming keys exist.
    Never raises KeyError for missing portal/JSON variants.
    """
    if not isinstance(row, dict):
        return {
            "gstin": "", "invoice_no": "", "invoice_date": "",
            "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
            "invoice_value": 0.0,
        }
    out = dict(row)

    def tax_amount(*names):
        for name in names:
            if name in out and out.get(name) not in (None, ""):
                return gst_number(out.get(name))
        return gst_number(gst_pick(out, *names))

    taxable = tax_amount("taxable_value", "txval", "taxable value", "taxable amount")
    igst = tax_amount("igst", "iamt", "integrated tax", "igst amount")
    cgst = tax_amount("cgst", "camt", "central tax", "cgst amount")
    sgst = tax_amount("sgst", "samt", "state/ut tax", "sgst amount", "utgst")
    cess = tax_amount("cess", "csamt", "cess amount")

    had_invoice_value = "invoice_value" in out and out.get("invoice_value") not in (None, "")
    invoice_value = gst_number(
        out.get("invoice_value") if had_invoice_value
        else gst_pick(out, "val", "invoice value", "invoice amount", "total invoice value", "total_value", "note value")
    )
    if not invoice_value:
        invoice_value = round(taxable + igst + cgst + sgst + cess, 2)
        if not had_invoice_value:
            _gst_missing_field_warn("invoice_value", out)

    out["gstin"] = gst_text(
        out.get("gstin") or gst_pick(out, "ctin", "gstin", "supplier gstin", "recipient gstin", "gstin/uin")
    ).upper()
    out["invoice_no"] = gst_text(
        out.get("invoice_no")
        or gst_pick(out, "inum", "invoice number", "invoice no", "invoice_no", "inv no", "nt_num", "note number", "document number")
    )
    out["invoice_date"] = gst_text(
        out.get("invoice_date")
        or gst_pick(out, "idt", "invoice date", "date", "nt_dt", "note date", "document date")
    )
    out["taxable_value"] = round(gst_number(taxable), 2)
    out["igst"] = round(gst_number(igst), 2)
    out["cgst"] = round(gst_number(cgst), 2)
    out["sgst"] = round(gst_number(sgst), 2)
    out["cess"] = round(gst_number(cess), 2)
    out["invoice_value"] = round(gst_number(invoice_value), 2)
    if not out.get("invoice_no_norm") and out.get("invoice_no"):
        out["invoice_no_norm"] = normalize_invoice_number(out.get("invoice_no"))
    return out


def normalize_gst_invoice(record, context=None):
    context = context or {}
    invoice_no = gst_pick(record, "inum", "invoice number", "invoice no", "invoice_no", "inv no", "document number", "doc no", "nt_num", "note number", "note no")
    if not invoice_no:
        return None
    gstin = gst_pick(record, "ctin", "gstin", "supplier gstin", "recipient gstin", "gstin/uin") or context.get("gstin", "")
    date = gst_pick(record, "idt", "invoice date", "date", "document date", "doc date", "nt_dt", "note date")
    value = gst_number(gst_pick(record, "val", "invoice value", "invoice amount", "total invoice value", "total_value", "note value"))
    items = record.get("itms") or record.get("items")
    # Portal-authoritative: when itms exist, store exact line-item tax sums only.
    # Never rewrite camt/samt to force balance with val (portal may report asymmetric taxes).
    if isinstance(items, list) and items:
        totals = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        for item in items:
            detail = item.get("itm_det", item) if isinstance(item, dict) else {}
            if isinstance(detail, dict):
                totals["taxable_value"] += gst_number(gst_pick(detail, "txval", "taxable value"))
                totals["igst"] += gst_number(gst_pick(detail, "iamt", "igst"))
                totals["cgst"] += gst_number(gst_pick(detail, "camt", "cgst"))
                totals["sgst"] += gst_number(gst_pick(detail, "samt", "sgst"))
                totals["cess"] += gst_number(gst_pick(detail, "csamt", "cess"))
    else:
        totals = {
            "taxable_value": gst_number(gst_pick(record, "txval", "taxable value", "taxable amount")),
            "igst": gst_number(gst_pick(record, "iamt", "integrated tax", "igst", "igst amount")),
            "cgst": gst_number(gst_pick(record, "camt", "central tax", "cgst", "cgst amount")),
            "sgst": gst_number(gst_pick(record, "samt", "state/ut tax", "sgst", "sgst amount", "utgst")),
            "cess": gst_number(gst_pick(record, "csamt", "cess", "cess amount")),
        }
    if not value:
        value = round(sum(totals.values()), 2)
    return ensure_gst_invoice_fields({
        "gstin": gst_text(gstin).upper(), "party_name": gst_text(context.get("party_name", "")),
        "invoice_no": gst_text(invoice_no), "invoice_date": gst_text(date),
        "invoice_value": value,
        "taxable_value": round(gst_number(totals["taxable_value"]), 2),
        "igst": round(gst_number(totals["igst"]), 2),
        "cgst": round(gst_number(totals["cgst"]), 2),
        "sgst": round(gst_number(totals["sgst"]), 2),
        "cess": round(gst_number(totals["cess"]), 2),
        "source": context.get("source", ""), "section": context.get("section", ""),
        "document_type": context.get("document_type", "Invoice"),
        "itc_availability": gst_text(context.get("itc_availability", "")),
    })


def gst_rows_from_json(data, source=""):
    rows = []
    def walk(node, context):
        if isinstance(node, dict):
            next_context = dict(context)
            gstin = gst_pick(node, "ctin", "gstin", "supplier gstin", "recipient gstin")
            if gstin:
                next_context["gstin"] = gst_text(gstin)
            candidate = normalize_gst_invoice(node, next_context)
            if candidate:
                rows.append(candidate)
            for value in node.values():
                walk(value, next_context)
        elif isinstance(node, list):
            for value in node:
                walk(value, context)
    walk(data, {"source": source})
    unique = {}
    for row in rows:
        safe = ensure_gst_invoice_fields(row)
        unique[(
            safe.get("gstin"),
            safe.get("invoice_no"),
            safe.get("invoice_date"),
            safe.get("invoice_value"),
        )] = safe
    return list(unique.values())


def gst_header_col(parent, child, terms, prefer_last=False):
    matches = []
    for index in range(max(len(parent), len(child))):
        text = f"{gst_text(parent[index] if index < len(parent) else '')} {gst_text(child[index] if index < len(child) else '')}".lower()
        if all(term in text for term in terms):
            matches.append(index)
    if not matches:
        return None
    return matches[-1] if prefer_last else matches[0]


def gst_rows_from_excel(raw, source=""):
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        if ws.title.lower() in {"read me", "itc available", "itc not available", "itc reversal", "itc rejected"}:
            continue
        # Some GST Portal workbooks contain a stale worksheet dimension
        # (for example A1:X30 although the B2B sheet actually has 900+ rows).
        # Read-only openpyxl otherwise stops at that false boundary and only
        # the first supplier is imported. Recalculate the real used range.
        if hasattr(ws, "reset_dimensions"):
            ws.reset_dimensions()
        grid = list(ws.iter_rows(values_only=True))
        # HSN-wise sales register exports (for example ``HSN Wise Sales.xlsx``)
        # are invoice-line data without a GST Portal-style invoice total row.
        # Keep the existing Return Type choices unchanged, but recognise this
        # format when it is imported through MARG Backup / Sales Register.
        hsn_header = next((i for i, line in enumerate(grid[:20])
                           if {gst_text(v).strip().lower() for v in line if gst_text(v).strip()}
                           >= {"date", "vno", "account", "gst %", "hsn/sac", "taxable value"}), None)
        if hsn_header is not None:
            h = [gst_text(v).strip().lower() for v in grid[hsn_header]]
            def hcol(*names):
                return next((h.index(name) for name in names if name in h), None)
            date_col = hcol("date")
            inv_col = hcol("vno")
            party_col = hcol("account")
            gstin_col = hcol("partygstin")
            reg_col = hcol("gst registrationtype")
            rate_col = hcol("gst %")
            hsn_col = hcol("hsn/sac")
            qty_col = hcol("total qty")
            taxable_col = hcol("taxable value")
            igst_col = hcol("igst")
            cgst_col = hcol("cgst")
            sgst_col = hcol("sgst/utgst")
            cess_col = hcol("cess")
            grouped_hsn = {}
            for values in grid[hsn_header + 1:]:
                def hcell(index):
                    return values[index] if index is not None and index < len(values) else ""
                invoice_no = gst_text(hcell(inv_col))
                if not invoice_no or invoice_no.upper() in {"TOTAL", "GRAND TOTAL"}:
                    continue
                raw_date = hcell(date_col)
                if hasattr(raw_date, "strftime") and not isinstance(raw_date, str):
                    date = raw_date.strftime("%d-%m-%Y")
                else:
                    date = gst_text(raw_date)
                party = gst_text(hcell(party_col))
                gstin = gst_text(hcell(gstin_col)).upper()
                key = (invoice_no, date, party, gstin)
                rate_raw = gst_number(hcell(rate_col))
                rate = rate_raw * 100 if 0 < abs(rate_raw) <= 1 else rate_raw
                taxable = gst_number(hcell(taxable_col))
                taxes = {
                    "igst": gst_number(hcell(igst_col)),
                    "cgst": gst_number(hcell(cgst_col)),
                    "sgst": gst_number(hcell(sgst_col)),
                    "cess": gst_number(hcell(cess_col)),
                }
                row = grouped_hsn.setdefault(key, {
                    "gstin": gstin, "party_name": party, "party_ledger": party,
                    "invoice_no": invoice_no, "invoice_date": date,
                    "invoice_value": 0.0, "taxable_value": 0.0,
                    "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
                    "source": source, "section": ws.title,
                    "document_type": "Sales Invoice", "items": [],
                    "registration_type": gst_text(hcell(reg_col)),
                })
                row["taxable_value"] = round(row["taxable_value"] + taxable, 2)
                for field, amount in taxes.items():
                    row[field] = round(row[field] + amount, 2)
                hsn = gst_text(hcell(hsn_col))
                row["items"].append({
                    "name": f"{rate:g}% Items", "hsn": hsn,
                    "gst_rate": rate, "quantity": gst_number(hcell(qty_col)),
                    "taxable_value": taxable, **taxes,
                })
            for row in grouped_hsn.values():
                row["invoice_value"] = round(row["taxable_value"] + row["igst"] + row["cgst"] + row["sgst"] + row["cess"], 2)
                rows.append(ensure_gst_invoice_fields(row))
            continue
        register_header = next((i for i, line in enumerate(grid[:40])
                                if any(gst_text(v).lower() in {"inv no", "invoice no.", "invoice no", "bill no", "note no", "credit note no"} for v in line)
                                and any(gst_text(v).lower() in {"taxable amount", "taxable value"} for v in line)), None)
        if register_header is not None:
            rows.extend(gst_rows_from_register_grid(
                grid, register_header, source, ws.title,
                "Credit Note" if "credit note" in " ".join(
                    gst_text(v).lower() for line in grid[:10] for v in line
                ) else ("Purchase" if "purchase" in " ".join(
                    gst_text(v).lower() for line in grid[:10] for v in line
                ) else "Sales Invoice")
            ))
            continue
        header_index = next((i for i, line in enumerate(grid[:40])
                             if any(gst_text(v).lower() in
                                    {"invoice number", "note number", "note number ", "document number"}
                                    for v in line)), None)
        if header_index is None:
            continue
        parent = list(grid[max(0, header_index - 1)])
        child = list(grid[header_index])
        amended = ws.title.upper().endswith("A") or "AMEND" in " ".join(gst_text(v).upper() for v in grid[3] if v)
        data_start = header_index + 1
        # GST Portal amendment sheets (for example B2BA/CDNRA) use a
        # three-level header: Original details, Revised details, then the
        # revised invoice/tax sub-headings.  Use the last two header rows so
        # revised invoice value and tax columns are selected, not the original
        # invoice number columns.
        if amended and header_index + 1 < len(grid):
            next_header = list(grid[header_index + 1])
            next_text = " ".join(gst_text(v).lower() for v in next_header if v)
            if any(term in next_text for term in ("invoice value", "integrated tax", "central tax", "note value")):
                parent = list(grid[header_index])
                child = next_header
                data_start = header_index + 2
        number_col = gst_header_col(parent, child, ("invoice", "number"), amended)
        date_col = gst_header_col(parent, child, ("invoice", "date"), amended)
        value_col = gst_header_col(parent, child, ("invoice", "value"), amended)
        document_type = "Invoice Amendment" if amended else "Invoice"
        if number_col is None:
            number_col = gst_header_col(parent, child, ("note", "number"), amended)
            date_col = gst_header_col(parent, child, ("note", "date"), amended)
            value_col = gst_header_col(parent, child, ("note", "value"), amended)
            document_type = "Credit/Debit Note Amendment" if amended else "Credit/Debit Note"
            note_type_col = gst_header_col(parent, child, ("note", "type"), amended)
        else:
            note_type_col = None
        if number_col is None:
            number_col = gst_header_col(parent, child, ("document", "number"), amended)
            date_col = gst_header_col(parent, child, ("document", "date"), amended)
            value_col = gst_header_col(parent, child, ("document", "value"), amended)
            document_type = "Document"
            note_type_col = None
        gstin_col = gst_header_col(parent, child, ("gstin",))
        party_col = gst_header_col(parent, child, ("trade/legal",))
        taxable_col = gst_header_col(parent, child, ("taxable", "value"))
        rate_col = gst_header_col(parent, child, ("rate",))
        igst_col = gst_header_col(parent, child, ("integrated", "tax"))
        cgst_col = gst_header_col(parent, child, ("central", "tax"))
        sgst_col = gst_header_col(parent, child, ("state", "tax"))
        if sgst_col is None:
            sgst_col = gst_header_col(parent, child, ("state/ut", "tax"))
        cess_col = gst_header_col(parent, child, ("cess",))
        itc_col = gst_header_col(parent, child, ("itc", "availability"))
        filing_date_col = gst_header_col(parent, child, ("filing", "date"))
        if gstin_col is None or number_col is None:
            continue
        grouped = {}
        for values in grid[data_start:]:
            if number_col >= len(values):
                continue
            raw_number = gst_text(values[number_col])
            if not raw_number:
                continue
            is_total = bool(re.search(r"(?i)[-\s]*total\s*$", raw_number))
            invoice_no = re.sub(r"(?i)[-\s]*total\s*$", "", raw_number).strip()
            gstin = gst_text(values[gstin_col] if gstin_col < len(values) else "").upper()
            if not gstin and document_type != "Document":
                continue
            def cell(index):
                return values[index] if index is not None and index < len(values) else ""
            row_document_type = document_type
            note_type = gst_text(cell(note_type_col)) if note_type_col is not None else ""
            if note_type:
                lowered = note_type.lower()
                if "credit" in lowered:
                    row_document_type = "Credit Note Amendment" if amended else "Credit Note"
                elif "debit" in lowered:
                    row_document_type = "Debit Note Amendment" if amended else "Debit Note"
            row = {
                "gstin": gstin, "party_name": gst_text(cell(party_col)),
                "invoice_no": invoice_no, "invoice_date": gst_text(cell(date_col)),
                "invoice_value": gst_number(cell(value_col)), "taxable_value": gst_number(cell(taxable_col)),
                "igst": gst_number(cell(igst_col)), "cgst": gst_number(cell(cgst_col)),
                "sgst": gst_number(cell(sgst_col)), "cess": gst_number(cell(cess_col)),
                "source": source, "section": ws.title, "document_type": row_document_type,
                "note_type": note_type,
                "itc_availability": gst_text(cell(itc_col)),
                "invoice_upload_date": gst_text(cell(filing_date_col)),
                "items": [],
            }
            if amended:
                row["amendment_direction"] = "increase"
                # B2BA keeps original invoice identity in its first columns.
                if ws.title.upper() == "B2BA":
                    row["original_invoice_no"] = gst_text(values[0] if len(values) > 0 else "")
                    row["original_invoice_date"] = gst_text(values[1] if len(values) > 1 else "")
            key = (gstin, re.sub(r"[^A-Z0-9]", "", invoice_no.upper()), row["invoice_date"], row_document_type)
            is_note_doc = any(token in row_document_type.lower() for token in ("note", "refund", "return"))
            if is_total:
                # The GST portal's bold ``Invoice/Note No.-Total`` row is the
                # authoritative document total. For Credit/Debit Notes, keep the
                # Excel total amounts exactly and rebuild items from that total
                # so rate-line aggregation cannot inflate CN values.
                rate_items = grouped.get(key, {}).get("items", [])
                if is_note_doc:
                    rate = gst_rate_for_values(
                        row["taxable_value"], row["igst"], row["cgst"], row["sgst"], row["cess"]
                    )
                    hsn = gst_text((rate_items[0] or {}).get("hsn")) if rate_items else ""
                    name = gst_text((rate_items[0] or {}).get("name")) if rate_items else f"{rate:g}% Items"
                    grouped[key] = {
                        **row,
                        "items": [{
                            "name": name or f"{rate:g}% Items",
                            "hsn": hsn,
                            "gst_rate": rate,
                            "taxable_value": row["taxable_value"],
                            "igst": row["igst"], "cgst": row["cgst"],
                            "sgst": row["sgst"], "cess": row["cess"],
                        }],
                        "_total": True,
                    }
                else:
                    grouped[key] = {**row, "items": rate_items, "_total": True}
            elif key not in grouped:
                rate = gst_number(cell(rate_col))
                rate_item = {
                    "name": f"{rate:g}% Items",
                    "gst_rate": rate,
                    "taxable_value": row["taxable_value"],
                    "igst": row["igst"], "cgst": row["cgst"],
                    "sgst": row["sgst"], "cess": row["cess"],
                }
                grouped[key] = {**row, "items": [rate_item], "_total": False}
            elif not grouped[key].get("_total"):
                for field in ("taxable_value", "igst", "cgst", "sgst", "cess"):
                    grouped[key][field] = round(grouped[key][field] + row[field], 2)
                rate = gst_number(cell(rate_col))
                grouped[key].setdefault("items", []).append({
                    "name": f"{rate:g}% Items",
                    "gst_rate": rate,
                    "taxable_value": row["taxable_value"],
                    "igst": row["igst"], "cgst": row["cgst"],
                    "sgst": row["sgst"], "cess": row["cess"],
                })
        for row in grouped.values():
            row.pop("_total", None)
            rows.append(reconcile_gstr1_note_document_amounts(row))
    return rows


def reconcile_gstr1_note_document_amounts(row):
    """
    Credit/Debit Note / Sales Return: keep Excel document totals authoritative.
    If rate/item lines disagree with Note Value / Taxable / tax columns, rebuild
    items from the document row so downstream prep cannot inflate CN amounts.
    """
    if not isinstance(row, dict):
        return row
    doc = gst_text(row.get("document_type")).lower()
    if not any(token in doc for token in ("note", "refund", "return")):
        return row
    taxable = gst_number(row.get("taxable_value"))
    igst = gst_number(row.get("igst"))
    cgst = gst_number(row.get("cgst"))
    sgst = gst_number(row.get("sgst"))
    cess = gst_number(row.get("cess"))
    invoice_value = gst_number(row.get("invoice_value"))
    items = list(row.get("items") or [])
    item_taxable = round(sum(gst_number(item.get("taxable_value")) for item in items), 2)
    item_cgst = round(sum(gst_number(item.get("cgst")) for item in items), 2)
    item_sgst = round(sum(gst_number(item.get("sgst")) for item in items), 2)
    mismatch = (
        items
        and abs(taxable) > 0.005
        and (
            abs(abs(item_taxable) - abs(taxable)) > 0.05
            or abs(abs(item_cgst) - abs(cgst)) > 0.05
            or abs(abs(item_sgst) - abs(sgst)) > 0.05
        )
    )
    if mismatch or (abs(taxable) > 0.005 and not items):
        rate = gst_rate_for_values(abs(taxable), abs(igst), abs(cgst), abs(sgst), abs(cess))
        seed = items[0] if items else {}
        row["items"] = [{
            "name": gst_text(seed.get("name") or seed.get("item_name")) or f"{rate:g}% Items",
            "hsn": gst_text(seed.get("hsn")),
            "gst_rate": rate,
            "quantity": gst_number(seed.get("quantity")) or 1,
            "unit": gst_text(seed.get("unit") or seed.get("uqc")) or "Pcs",
            "taxable_value": taxable,
            "igst": igst, "cgst": cgst, "sgst": sgst, "cess": cess,
        }]
    if not invoice_value and abs(taxable) > 0.005:
        row["invoice_value"] = round(taxable + igst + cgst + sgst + cess, 2)
    return row


def gst_rows_from_register_grid(grid, header_index, source, section, document_type):
    header_values = [gst_text(value).lower() for value in grid[header_index]]
    headers = {value: index for index, value in enumerate(header_values) if value}
    def col(*names):
        return next((headers[name] for name in names if name in headers), None)
    def after(label):
        index = next((i for i, value in enumerate(header_values) if value == label), None)
        return index + 1 if index is not None and index + 1 < len(header_values) else None
    inv_col = col("inv no", "invoice no.", "invoice no", "bill no", "note no", "credit note no")
    date_col = col("inv date", "invoice date", "bill date", "v.date", "note date")
    gstin_col = col("gstr no", "depo gst no", "gstin")
    party_col = col("party name", "supplier name", "desc")
    value_col = col("net amount", "invoice value", "invoice/note value", "note value")
    taxable_col, hsn_col = col("taxable amount", "taxable value"), col("hsn code", "hsn")
    sku_col, sku_name_col = col("sku code"), col("sku name")
    qty_col, rate_col = col("qty", "quantity"), col("rate")
    tax_cols = {"igst": col("igst amt"), "cgst": col("cgst amt"),
                "sgst": col("sgst amt"), "cess": col("cess amt")}
    if all(index is None for index in tax_cols.values()):
        tax_cols = {"igst": after("igst"), "cgst": after("cgst"),
                    "sgst": after("sgst"), "cess": col("cess")}
    ref_col, ref_date_col = col("ref.invno"), col("ref.invdate")
    grouped = {}
    carry = {}
    current_document_type = document_type
    for values in grid[header_index + 1:]:
        def cell(index):
            return values[index] if index is not None and index < len(values) else ""
        row_text = " ".join(gst_text(value) for value in values if value not in (None, "")).lower()
        if "credit/debit note" in row_text or ("credit note" in row_text and "less" in row_text):
            current_document_type = "Credit/Debit Note & Refund"
            carry = {}
            continue
        if "export invoices" in row_text or "gross total" in row_text:
            continue
        fresh_invoice = gst_text(cell(inv_col))
        item_identity = gst_text(cell(sku_col)) or gst_text(cell(sku_name_col)) or gst_text(cell(hsn_col))
        if "TOTAL" in item_identity.upper():
            continue
        if not fresh_invoice and not item_identity:
            continue
        invoice_no = fresh_invoice or carry.get("invoice_no", "")
        if not invoice_no:
            continue
        if gst_text(cell(inv_col)):
            carry = {
                "invoice_no": invoice_no, "invoice_date": gst_text(cell(date_col)),
                "gstin": gst_text(cell(gstin_col)).upper(), "party_name": gst_text(cell(party_col)),
                "invoice_value": gst_number(cell(value_col)),
                "reference_invoice": gst_text(cell(ref_col)),
                "reference_date": gst_text(cell(ref_date_col)),
            }
        row_document_type = (
            "Credit Note" if current_document_type == "Credit/Debit Note & Refund"
            and invoice_no.upper().startswith(("CN", "CR")) else current_document_type
        )
        is_note_doc = any(token in row_document_type.lower() for token in ("note", "refund", "return"))
        key = (carry.get("gstin", ""), re.sub(r"[^A-Z0-9]", "", invoice_no.upper()), row_document_type)
        row = grouped.setdefault(key, {
            **carry, "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
            "source": source, "section": section, "document_type": row_document_type, "items": [],
        })
        # Credit Note header row (invoice/note no present, no HSN/SKU): lock Excel
        # document totals so later item lines cannot inflate taxable/tax.
        header_only = bool(fresh_invoice) and not item_identity and is_note_doc
        line_taxable = gst_number(cell(taxable_col))
        line_taxes = {field: gst_number(cell(index)) for field, index in tax_cols.items()}
        if header_only and abs(line_taxable) > 0.005:
            row["invoice_value"] = gst_number(cell(value_col)) or gst_number(carry.get("invoice_value"))
            row["taxable_value"] = line_taxable
            for field, amount in line_taxes.items():
                row[field] = amount
            row["_locked_doc_totals"] = True
            continue
        if row.get("_locked_doc_totals"):
            if item_identity:
                row["items"].append({
                    "code": gst_text(cell(sku_col)), "name": gst_text(cell(sku_name_col)),
                    "hsn": gst_text(cell(hsn_col)), "quantity": gst_number(cell(qty_col)),
                    "rate": gst_number(cell(rate_col)), "taxable_value": line_taxable,
                    **line_taxes,
                })
            continue
        row["invoice_value"] = max(gst_number(row.get("invoice_value")), gst_number(carry.get("invoice_value")))
        row["taxable_value"] = round(gst_number(row.get("taxable_value")) + line_taxable, 2)
        for field, amount in line_taxes.items():
            row[field] = round(gst_number(row.get(field)) + amount, 2)
        if gst_text(cell(hsn_col)) or gst_text(cell(sku_name_col)):
            row["items"].append({
                "code": gst_text(cell(sku_col)), "name": gst_text(cell(sku_name_col)),
                "hsn": gst_text(cell(hsn_col)), "quantity": gst_number(cell(qty_col)),
                "rate": gst_number(cell(rate_col)), "taxable_value": line_taxable,
                **line_taxes,
            })
    return [reconcile_gstr1_note_document_amounts({k: v for k, v in row.items() if not str(k).startswith("_")})
            for row in grouped.values()]


def parse_gst_file(name, raw):
    suffix = Path(name).suffix.lower()
    source_period = infer_gst_period(name)
    if suffix == ".zip":
        rows = []
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            for member in archive.infolist():
                if not member.is_dir() and Path(member.filename).suffix.lower() in {".json", ".xlsx", ".xlsm"}:
                    child_rows = parse_gst_file(member.filename, archive.read(member))
                    for row in child_rows:
                        if source_period:
                            row["source_period"] = source_period
                        row["source_container"] = name
                    rows.extend(child_rows)
        return apply_gstr2b_portal_signs(rows)
    if suffix == ".json":
        rows = gst_rows_from_json(json.loads(raw.decode("utf-8-sig")), name)
        for row in rows:
            row["source_period"] = source_period
        return apply_gstr2b_portal_signs(rows)
    if suffix in {".xlsx", ".xlsm"}:
        rows = gst_rows_from_excel(raw, name)
        for row in rows:
            row["source_period"] = source_period
        return apply_gstr2b_portal_signs(rows)
    if suffix == ".xls":
        rows = gst_rows_from_excel(legacy_xls_to_xlsx(raw), name)
        for row in rows:
            row["source_period"] = source_period
        return apply_gstr2b_portal_signs(rows)
    raise ValueError("GST import supports JSON, Excel or ZIP files.")


def parse_gst_financial_file(name, raw, report_type=""):
    report_type = gst_text(report_type)
    suffix = Path(name).suffix.lower()
    if suffix == ".csv":
        text = raw.decode("utf-8-sig", errors="replace")
        grid = list(csv.reader(io.StringIO(text)))
        if "Payment" in report_type:
            header = next((i for i, row in enumerate(grid) if "CPIN" in row), None)
            if header is None:
                raise ValueError("GST Payment List header was not found.")
            columns = [gst_text(value) for value in grid[header]]
            records = [dict(zip(columns, row)) for row in grid[header + 1:] if any(gst_text(v) for v in row)]
            return {"kind": "GST Payment List", "records": records,
                    "total": round(sum(gst_number(row.get("Amount")) for row in records), 2)}
        header = next((i for i, row in enumerate(grid)
                       if any("Transaction Type" in gst_text(value) for value in row)), None)
        if header is None:
            raise ValueError("GST electronic ledger header was not found.")
        sub = grid[header + 1] if header + 1 < len(grid) else []
        width = max(len(grid[header]), len(sub))
        parents, last = [], ""
        for index in range(width):
            value = gst_text(grid[header][index] if index < len(grid[header]) else "")
            if value:
                last = value
            parents.append(last)
        columns = []
        seen = {}
        for index in range(width):
            child = gst_text(sub[index] if index < len(sub) else "")
            label = f"{parents[index]} {child}".strip()
            seen[label] = seen.get(label, 0) + 1
            columns.append(f"{label} #{seen[label]}" if seen[label] > 1 else label)
        records = []
        for values in grid[header + 2:]:
            description = gst_text(values[6] if len(values) > 6 else "")
            if not description or description in {"Opening Balance", "Closing Balance"}:
                continue
            record = {columns[i]: values[i] if i < len(values) else "" for i in range(width)}
            record["description"] = description
            record["transaction_type"] = gst_text(values[7] if len(values) > 7 else "")
            record["date"] = gst_text(values[1] if len(values) > 1 else "")
            record["reference"] = gst_text(values[4] if len(values) > 4 else "")
            record["tax_period"] = gst_text(values[5] if len(values) > 5 else "")
            records.append(record)
        return {"kind": "Electronic Cash Ledger" if "Cash" in report_type else "Electronic Credit Ledger",
                "records": records, "rcm_records": [row for row in records
                                                    if "reverse charge" in row["description"].lower()
                                                    and "other than reverse charge" not in row["description"].lower()]}
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        periods = {}
        for ws in wb.worksheets:
            grid = list(ws.iter_rows(values_only=True))
            for row_index, values in enumerate(grid):
                period_index = next((i for i, value in enumerate(values[:8])
                                     if re.fullmatch(r"[A-Z][a-z]{2}-\d{2}", gst_text(value))), None)
                if period_index is None:
                    continue
                period = gst_text(values[period_index])
                numeric = [gst_number(value) for value in values[period_index + 1:]]
                periods.setdefault(period, {})[ws.title] = numeric
        if not periods:
            raise ValueError("Tax liability/ITC period rows were not found.")
        return {"kind": "Tax Liability and ITC Comparison", "periods": periods,
                "sheets": [ws.title for ws in wb.worksheets]}
    raise ValueError("GST financial import supports CSV or Excel files.")


def parse_gstr3b_totals(name, raw):
    """Read GSTR-3B eligible ITC and reverse-charge totals from PDF/JSON/ZIP/Excel."""
    suffix = Path(name).suffix.lower()
    if suffix == ".zip":
        totals = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        net_itc = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        reverse_charge = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            for member in archive.infolist():
                if member.is_dir():
                    continue
                child = parse_gstr3b_totals(member.filename, archive.read(member))
                for key in totals:
                    totals[key] += gst_number(child.get(key))
                    net_itc[key] += gst_number(child.get("net_itc", child).get(key))
                    reverse_charge[key] += gst_number(child.get("reverse_charge", {}).get(key))
        result = {key: round(value, 2) for key, value in totals.items()}
        result["net_itc"] = {key: round(value, 2) for key, value in net_itc.items()}
        result["reverse_charge"] = {key: round(value, 2) for key, value in reverse_charge.items()}
        return result
    if suffix == ".json":
        data = json.loads(raw.decode("utf-8-sig"))
        available = []
        reverse_charge_rows = []
        def walk(value):
            if isinstance(value, dict):
                for key, child in value.items():
                    if key.lower() in {"itc_avl", "itc_available"} and isinstance(child, list):
                        available.extend(item for item in child if isinstance(item, dict))
                    if key.lower() in {"isup_rev", "inward_supplies_liable_to_reverse_charge"}:
                        if isinstance(child, dict):
                            reverse_charge_rows.append(child)
                        elif isinstance(child, list):
                            reverse_charge_rows.extend(item for item in child if isinstance(item, dict))
                    walk(child)
            elif isinstance(value, list):
                for child in value:
                    walk(child)
        walk(data)
        if not available:
            raise ValueError("GSTR-3B eligible ITC section was not found in this JSON file.")
        result = {
            "taxable_value": 0.0,
            "igst": round(sum(gst_number(item.get("iamt")) for item in available), 2),
            "cgst": round(sum(gst_number(item.get("camt")) for item in available), 2),
            "sgst": round(sum(gst_number(item.get("samt")) for item in available), 2),
            "cess": round(sum(gst_number(item.get("csamt")) for item in available), 2),
        }
        result["reverse_charge"] = {
            "taxable_value": round(sum(gst_number(item.get("txval")) for item in reverse_charge_rows), 2),
            "igst": round(sum(gst_number(item.get("iamt")) for item in reverse_charge_rows), 2),
            "cgst": round(sum(gst_number(item.get("camt")) for item in reverse_charge_rows), 2),
            "sgst": round(sum(gst_number(item.get("samt")) for item in reverse_charge_rows), 2),
            "cess": round(sum(gst_number(item.get("csamt")) for item in reverse_charge_rows), 2),
        }
        return result
    if suffix == ".pdf":
        try:
            import pdfplumber
        except ImportError as exc:
            raise ValueError("PDF support is not installed. Run Install_Requirements.bat once.") from exc

        itc = None
        gross_itc = None
        reversed_itc = None
        reverse_charge = None
        outward_31a = None
        outward_31c = None
        interest = None
        late_fee = None
        interest_61 = {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        late_fee_61 = {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        pdf_gstin = ""
        pdf_period = ""

        def pdf_cell_money(cell):
            """Parse a GSTR-3B PDF amount cell; ignore '-', blanks, and watermark letters (D/E/L/I/F)."""
            text = gst_text(cell).replace("₹", "").replace("\n", " ").strip()
            if not text or text in {"-", "—", "–"}:
                return None
            text = re.sub(r"^[A-Za-z]+\s*", "", text).strip()
            text = re.sub(r"\s+[A-Za-z]+$", "", text).strip()
            if not text or text in {"-", "—", "–"}:
                return None
            if re.fullmatch(r"\(?-?[\d,]+(?:\.\d+)?\)?", text):
                negative = text.startswith("(") and text.endswith(")")
                value = gst_number(text.strip("()"))
                return -value if negative else value
            return None

        def values_after_label(cells, label_index, count):
            values = []
            for cell in cells[label_index + 1:]:
                value = pdf_cell_money(cell)
                if value is not None:
                    values.append(value)
            return ([0.0] * count + values)[-count:] if values else []

        def values_after_label_aligned(cells, label_index, count):
            """Keep column positions; treat '-' / blank as 0 (needed for Table 3.1(c) / Late fee)."""
            values = []
            for cell in cells[label_index + 1:]:
                text = gst_text(cell).replace("₹", "").replace("\n", " ").strip()
                if not text or text in {"-", "—", "–"}:
                    values.append(0.0)
                    continue
                value = pdf_cell_money(cell)
                if value is not None:
                    values.append(value)
            return (values + [0.0] * count)[:count]

        def tax4_from_numbers(numbers):
            padded = ([0.0] * 4 + list(numbers or []))[-4:]
            return dict(zip(("igst", "cgst", "sgst", "cess"), [round(gst_number(v), 2) for v in padded]))

        def add_tax4(left, right):
            left = left or {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
            right = right or {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
            return {
                key: round(gst_number(left.get(key)) + gst_number(right.get(key)), 2)
                for key in ("igst", "cgst", "sgst", "cess")
            }

        def is_table_31a_row(joined):
            text = gst_text(joined).lower()
            if "(a)" not in text or "outward taxable" not in text:
                return False
            # Prefer official 3.1(a) wording; accept short label if not clearly (b)/(c)/(d)/(e).
            if "other than" in text or ("zero rated" in text and "nil" in text):
                return True
            return "zero rated" not in text and "nil rated" not in text and "reverse charge" not in text and "non-gst" not in text

        def is_table_31c_row(joined):
            text = gst_text(joined).lower()
            if "(c" not in text:
                return False
            return "nil" in text or "exempt" in text

        def infer_pdf_period_and_gstin(lines):
            """Read Year / Period / GSTIN from GST Portal GSTR-3B PDF header text."""
            year = ""
            fy_end = ""
            month_name = ""
            gstin = ""
            month_map = {
                "january": "01", "february": "02", "march": "03", "april": "04",
                "may": "05", "june": "06", "july": "07", "august": "08",
                "september": "09", "october": "10", "november": "11", "december": "12",
            }
            for line in lines:
                text = gst_text(line)
                lowered = text.lower()
                if not gstin:
                    match = re.search(r"\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][A-Z0-9]Z[A-Z0-9])\b", text.upper())
                    if match:
                        gstin = match.group(1)
                if not year:
                    match = re.search(r"\byear\s*(20\d{2})\s*[-–/]\s*(\d{2})\b", lowered)
                    if match:
                        year = match.group(1)
                        fy_end = match.group(2)
                if not month_name:
                    match = re.search(
                        r"\bperiod\s+(january|february|march|april|may|june|july|august|september|october|november|december)\b",
                        lowered,
                    )
                    if match:
                        month_name = match.group(1)
            period = ""
            if year and month_name:
                month = month_map[month_name]
                # Portal FY 2025-26: Apr-Dec use 2025, Jan-Mar use 2026.
                cal_year = year
                if fy_end and month in {"01", "02", "03"}:
                    cal_year = f"20{fy_end}" if len(fy_end) == 2 else fy_end
                period = f"{month}{cal_year}"
            return gstin, period

        def pdf_line_amounts(text):
            """Extract money tokens; require a digit so lone commas in labels are ignored."""
            return [
                gst_number(value)
                for value in re.findall(r"-?\d[\d,]*(?:\.\d+)?", gst_text(text))
            ]

        # Pass 1: fast text (pypdf). Pass 2: pdfplumber tables only if Net ITC / 3.1(a) still missing.
        text_lines = []
        available_from_rows = {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        reversed_from_rows = {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        try:
            pdf_reader = PdfReader(io.BytesIO(raw))
            for page in pdf_reader.pages:
                text_lines.extend((page.extract_text() or "").splitlines())
        except Exception:
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages:
                    text_lines.extend((page.extract_text() or "").splitlines())

        pdf_gstin, pdf_period = infer_pdf_period_and_gstin(text_lines)

        # Text path for GST Portal PDFs / split tables across pages.
        available_values = [0.0, 0.0, 0.0, 0.0]
        reversed_values = [0.0, 0.0, 0.0, 0.0]
        in_available_section = False
        in_reversed_section = False
        for index, line in enumerate(text_lines):
            lowered = line.lower()
            nearby = " ".join(text_lines[index:index + 3])
            line_numbers = pdf_line_amounts(line)
            nearby_numbers = pdf_line_amounts(nearby)

            if "a. itc available" in lowered:
                in_available_section = True
                in_reversed_section = False
                continue
            if "b. itc reversed" in lowered:
                in_available_section = False
                in_reversed_section = True
                continue
            if "c. net itc available" in lowered or "net itc available (a-b)" in lowered:
                in_available_section = False
                in_reversed_section = False
                if itc is None and len(line_numbers) >= 4:
                    itc = tax4_from_numbers(line_numbers[-4:])
                elif itc is None and len(nearby_numbers) >= 4:
                    itc = tax4_from_numbers(nearby_numbers[-4:])
                continue
            if in_available_section and re.match(r"^\s*\([1-5]\)", line) and len(line_numbers) >= 4:
                available_values = [left + right for left, right in zip(available_values, line_numbers[-4:])]
            if in_reversed_section and re.match(r"^\s*\([12]\)", line) and len(line_numbers) >= 4:
                reversed_values = [left + right for left, right in zip(reversed_values, line_numbers[-4:])]

            if interest is None and "interest paid" in lowered:
                numbers = line_numbers if len(line_numbers) >= 4 else nearby_numbers
                if len(numbers) >= 4:
                    interest = tax4_from_numbers(numbers[-4:])
            if late_fee is None and lowered.strip().startswith("late fee"):
                amounts = []
                for token in re.findall(r"(?:(?<=\s)|^)(-|\d[\d,]*(?:\.\d+)?)(?=\s|$)", line):
                    if token in {"-", "—", "–"}:
                        amounts.append(0.0)
                    else:
                        amounts.append(gst_number(token))
                if len(amounts) >= 4:
                    late_fee = tax4_from_numbers(amounts[-4:])
                elif len(line_numbers) >= 2:
                    # Common portal layout: Late fee - 475.00 475.00 -
                    late_fee = {
                        "igst": 0.0,
                        "cgst": round(line_numbers[0], 2),
                        "sgst": round(line_numbers[1] if len(line_numbers) > 1 else 0.0, 2),
                        "cess": 0.0,
                    }

            if reverse_charge is None and "inward supplies" in lowered and "reverse charge" in nearby.lower():
                numbers = line_numbers if len(line_numbers) >= 5 else nearby_numbers
                if len(numbers) >= 5:
                    reverse_charge = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), numbers[-5:]))
            # Table 3.1(a): portal PDFs often wrap the label; amounts are on the next line.
            if outward_31a is None and is_table_31a_row(lowered):
                numbers = list(line_numbers)
                if len(numbers) < 5:
                    for look in range(1, 3):
                        if index + look >= len(text_lines):
                            break
                        candidate = pdf_line_amounts(text_lines[index + look])
                        # Prefer a pure amount row (5 values) over mixed label rows.
                        if len(candidate) >= 5 and not re.search(r"\([b-e]\)", text_lines[index + look].lower()):
                            numbers = candidate
                            break
                if len(numbers) >= 5:
                    outward_31a = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), numbers[-5:]))
            if outward_31c is None and is_table_31c_row(lowered):
                # Prefer 5-value row; else taxable-only (common when tax columns are '-').
                if len(line_numbers) >= 5:
                    outward_31c = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), line_numbers[-5:]))
                elif line_numbers:
                    outward_31c = {
                        "taxable_value": line_numbers[0],
                        "igst": 0.0,
                        "cgst": 0.0,
                        "sgst": 0.0,
                        "cess": 0.0,
                    }
            # Table 3.1(e) Non-GST outward — fold into nil/exempt bucket for 3.1(c)-style turnover.
            if outward_31c is None and ("(e)" in lowered or "(e " in lowered) and "non-gst" in lowered:
                if len(line_numbers) >= 1:
                    outward_31c = {
                        "taxable_value": line_numbers[0],
                        "igst": 0.0,
                        "cgst": 0.0,
                        "sgst": 0.0,
                        "cess": 0.0,
                    }

        if any(available_values):
            gross_itc = tax4_from_numbers(available_values)
        if any(reversed_values):
            reversed_itc = tax4_from_numbers(reversed_values)

        # Slow table scrape only when text path missed required Net ITC / 3.1(a).
        if itc is None or outward_31a is None:
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages:
                    for table in page.extract_tables() or []:
                        for row in table or []:
                            cells = [gst_text(cell).replace("\n", " ").strip() for cell in (row or [])]
                            joined = " ".join(cells).lower()
                            label = gst_text(cells[0] if cells else "").lower()

                            # Table 3.1(a) Output Liability.
                            if outward_31a is None and is_table_31a_row(joined):
                                label_index = next(
                                    (i for i, cell in enumerate(cells) if cell and "(a)" in cell.lower() and "outward" in cell.lower()),
                                    0,
                                )
                                numbers = values_after_label(cells, label_index, 5)
                                if len(numbers) >= 5:
                                    outward_31a = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), numbers))
                            # Table 3.1(c) Nil/exempt.
                            if outward_31c is None and is_table_31c_row(joined):
                                label_index = next(
                                    (i for i, cell in enumerate(cells) if cell and "(c" in cell.lower()),
                                    0,
                                )
                                numbers = values_after_label_aligned(cells, label_index, 5)
                                if any(abs(gst_number(v)) > 0.005 for v in numbers):
                                    outward_31c = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), numbers))
                            # Table 3.1(e) Non-GST supplies (turnover only) — keep with nil/exempt bucket.
                            if outward_31c is None and "(e" in joined and "non-gst" in joined:
                                label_index = next(
                                    (i for i, cell in enumerate(cells) if cell and "(e" in cell.lower()),
                                    0,
                                )
                                numbers = values_after_label_aligned(cells, label_index, 5)
                                if any(abs(gst_number(v)) > 0.005 for v in numbers):
                                    outward_31c = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), numbers))
                            # Table 3.1(d) reverse charge.
                            if reverse_charge is None and "inward supplies" in joined and "reverse charge" in joined and "other than" not in joined:
                                label_index = next((i for i, cell in enumerate(cells) if "inward supplies" in cell.lower()), 0)
                                numbers = values_after_label_aligned(cells, label_index, 5)
                                if numbers:
                                    reverse_charge = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), numbers))

                            # Table 4A rows (1)-(5)
                            if re.match(r"^\(?[1-5]\)?", label) and (
                                "import" in label or "inward" in label or "all other itc" in label or "isd" in label
                            ):
                                numbers = values_after_label_aligned(cells, 0, 4)
                                if len(numbers) == 4:
                                    available_from_rows = add_tax4(available_from_rows, tax4_from_numbers(numbers))
                            # Table 4B rows (1)-(2) under ITC Reversed (rules / others)
                            if re.match(r"^\(?[12]\)?", label) and (
                                "rules 38" in label or "section 17" in label or label.startswith("(2) others") or label == "(2) others"
                                or ("others" in label and "ineligible" not in label and "reclaimed" not in label and "import" not in label)
                            ):
                                numbers = values_after_label_aligned(cells, 0, 4)
                                if len(numbers) == 4 and ("rules" in label or "others" in label or "17(5)" in label):
                                    reversed_from_rows = add_tax4(reversed_from_rows, tax4_from_numbers(numbers))
                            # Table 4C Net ITC
                            if "net itc available" in joined or ("c. net itc" in joined):
                                label_index = next(
                                    (i for i, cell in enumerate(cells) if cell and "net itc" in cell.lower()),
                                    0,
                                )
                                numbers = values_after_label_aligned(cells, label_index, 4)
                                if numbers and (any(numbers) or itc is None):
                                    itc = tax4_from_numbers(numbers)

                            # Section 5.1 Interest Paid / Late fee (primary source).
                            if "interest paid" in joined:
                                label_index = next(
                                    (i for i, cell in enumerate(cells) if cell and "interest paid" in cell.lower()),
                                    0,
                                )
                                numbers = values_after_label_aligned(cells, label_index, 4)
                                if numbers:
                                    interest = tax4_from_numbers(numbers)
                            if label.startswith("late fee") or joined.strip().startswith("late fee"):
                                label_index = next(
                                    (i for i, cell in enumerate(cells) if cell and "late fee" in cell.lower()),
                                    0,
                                )
                                numbers = values_after_label_aligned(cells, label_index, 4)
                                if numbers:
                                    late_fee = tax4_from_numbers(numbers)

                            # Section 6.1 Payment of tax — Interest / Late fee columns (fallback).
                            # Row shape: Description | Tax payable | Adj | Net | ITC... | Cash | Interest | Late fee
                            if any(h in label for h in ("integrated tax", "central tax", "state/ut tax", "cess")) and len(cells) >= 8:
                                head = "igst" if "integrated" in label else (
                                    "cgst" if "central" in label else (
                                        "sgst" if "state" in label else ("cess" if label.startswith("cess") else "")
                                    )
                                )
                                if head:
                                    trailing = []
                                    for cell in cells[1:]:
                                        text = gst_text(cell).replace("₹", "").replace("\n", " ").strip()
                                        if not text or text in {"-", "—", "–"}:
                                            trailing.append(None)
                                            continue
                                        value = pdf_cell_money(cell)
                                        trailing.append(value)
                                    # Interest / Late fee are the last two numeric-ish columns on 6.1 rows.
                                    money_tail = [v for v in trailing if v is not None]
                                    if len(money_tail) >= 2 and "(a)" not in joined and "(b)" not in joined:
                                        # Prefer explicit last two cells when present.
                                        interest_val = trailing[-2] if len(trailing) >= 2 and trailing[-2] is not None else money_tail[-2]
                                        late_val = trailing[-1] if trailing[-1] is not None else money_tail[-1]
                                        # Guard: tax payable rows have many amounts; only accept when row looks like payment line.
                                        if "payable" in joined or "tax" in label:
                                            if interest_val is not None:
                                                interest_61[head] = round(gst_number(interest_61[head]) + gst_number(interest_val), 2)
                                            if late_val is not None:
                                                late_fee_61[head] = round(gst_number(late_fee_61[head]) + gst_number(late_val), 2)

            if not gross_itc and any(available_from_rows.values()):
                gross_itc = available_from_rows
            if not reversed_itc and any(reversed_from_rows.values()):
                reversed_itc = reversed_from_rows

        if itc is None:
            raise ValueError("GSTR-3B Net ITC Available was not found in this PDF. Upload the GST Portal GSTR-3B PDF.")
        if outward_31a is None:
            raise ValueError(
                "GSTR-3B Table 3.1(a) Outward taxable supplies was not found in this PDF. "
                "Upload the official GST Portal GSTR-3B PDF (Returns Dashboard → Download GSTR-3B)."
            )
        # Prefer Section 5.1; fall back to Section 6.1 payment columns when 5.1 missing.
        if interest is None and any(interest_61.values()):
            interest = dict(interest_61)
        if late_fee is None and any(late_fee_61.values()):
            late_fee = dict(late_fee_61)
        interest = interest or {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        late_fee = late_fee or {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        reversed_itc = reversed_itc or {"igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
        gross_itc = gross_itc or itc

        def bucket4(data):
            out = {
                "taxable_value": 0.0,
                "igst": round(gst_number((data or {}).get("igst")), 2),
                "cgst": round(gst_number((data or {}).get("cgst")), 2),
                "sgst": round(gst_number((data or {}).get("sgst")), 2),
                "cess": round(gst_number((data or {}).get("cess")), 2),
            }
            out["output_gst"] = round(out["igst"] + out["cgst"] + out["sgst"] + out["cess"], 2)
            return out

        result = bucket4(gross_itc)
        result["itc_available_gross"] = bucket4(gross_itc)
        result["itc_reversed"] = bucket4(reversed_itc)
        result["net_itc"] = bucket4(itc)
        result["itc_claimed"] = bucket4(itc)
        result["interest"] = bucket4(interest)
        result["late_fee"] = bucket4(late_fee)
        result["reverse_charge"] = reverse_charge or {
            "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0, "output_gst": 0.0,
        }
        if "output_gst" not in result["reverse_charge"]:
            rc = result["reverse_charge"]
            rc["output_gst"] = round(sum(gst_number(rc.get(k)) for k in ("igst", "cgst", "sgst", "cess")), 2)
        outward = {
            "taxable_value": round(gst_number(outward_31a.get("taxable_value")), 2),
            "igst": round(gst_number(outward_31a.get("igst")), 2),
            "cgst": round(gst_number(outward_31a.get("cgst")), 2),
            "sgst": round(gst_number(outward_31a.get("sgst")), 2),
            "cess": round(gst_number(outward_31a.get("cess")), 2),
        }
        outward["output_gst"] = round(sum(outward[key] for key in ("igst", "cgst", "sgst", "cess")), 2)
        result["outward_supplies"] = outward
        if outward_31c:
            nil_exempt = {
                "taxable_value": round(gst_number(outward_31c.get("taxable_value")), 2),
                "igst": round(gst_number(outward_31c.get("igst")), 2),
                "cgst": round(gst_number(outward_31c.get("cgst")), 2),
                "sgst": round(gst_number(outward_31c.get("sgst")), 2),
                "cess": round(gst_number(outward_31c.get("cess")), 2),
            }
            nil_exempt["output_gst"] = round(sum(nil_exempt[key] for key in ("igst", "cgst", "sgst", "cess")), 2)
            result["outward_nil_exempt"] = nil_exempt
        if pdf_gstin:
            result["gstin"] = pdf_gstin
        if pdf_period:
            result["return_period"] = pdf_period
        return result
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for ws in wb.worksheets:
            grid = list(ws.iter_rows(values_only=True))
            reverse_charge = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
            for index, values in enumerate(grid):
                label = " ".join(gst_text(value) for value in values[:8]).lower()
                if "reverse charge" in label and "other than reverse charge" not in label:
                    numbers = [gst_number(value) for value in values if isinstance(value, (int, float))]
                    if numbers:
                        padded = ([0.0] * 5 + numbers)[-5:]
                        reverse_charge = dict(zip(("taxable_value", "igst", "cgst", "sgst", "cess"), padded))
                if "net itc available" not in label and "itc available" not in label:
                    continue
                headers = [gst_text(value).lower() for value in (grid[index - 1] if index else [])]
                def value_for(*tokens):
                    column = next((i for i, header in enumerate(headers) if all(token in header for token in tokens)), None)
                    return gst_number(values[column]) if column is not None and column < len(values) else 0.0
                totals = {"taxable_value": 0.0, "igst": value_for("integrated"),
                          "cgst": value_for("central"), "sgst": value_for("state"),
                          "cess": value_for("cess")}
                if any(totals[key] for key in ("igst", "cgst", "sgst", "cess")):
                    totals["reverse_charge"] = reverse_charge
                    return totals
        raise ValueError("GSTR-3B ITC Available/Net ITC Available row was not found in this Excel file.")
    raise ValueError("GSTR-3B import supports GST Portal PDF, JSON, ZIP or Excel files.")


def company_date(value):
    text = gst_text(value).replace("/", "-").replace(".", "-")
    for fmt in ("%d-%m-%Y", "%d-%m-%y", "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %b %y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def company_row(date, particulars, reference, debit=0, credit=0, balance=0, source="", confidence="High"):
    debit, credit = amount(debit), amount(credit)
    if debit and credit and abs(debit - credit) < .01:
        voucher = "Journal"
        status = "Review Required"
    elif debit:
        voucher, status = "Purchase", "Ready"
    elif credit:
        voucher, status = "Payment", "Ready"
    else:
        voucher, status = "Journal", "Review Required"
    return {
        "date": date, "particulars": gst_text(particulars), "reference": gst_text(reference),
        "debit": debit, "credit": credit, "balance": amount(balance),
        "voucher_type": voucher, "status": status, "confidence": confidence,
        "source": source,
    }


def parse_company_pdf(name, raw, password=""):
    reader = PdfReader(io.BytesIO(raw))
    if reader.is_encrypted:
        if not password or not reader.decrypt(password):
            raise ValueError("Incorrect PDF password.")
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    rows, source = [], name
    if "Dabur India Ltd." in text and "STATEMENT OF ACCOUNT" in text:
        pattern = re.compile(
            r"(?m)^(\d{2}\.\d{2}\.\d{4})\s+([A-Z]{2})\s+(.+?)\s+(\d{7,12})\s+"
            r"([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)\s+(-?[\d,]+(?:\.\d+)?)"
        )
        for match in pattern.finditer(text):
            date, doc_type, particulars, reference, debit, credit, balance = match.groups()
            rows.append(company_row(company_date(date), f"{doc_type} {particulars}", reference,
                                    debit, credit, balance, source))
        format_name = "Dabur Statement"
    elif "CHAUDHARY MARKETING" in text and "Ledger Report" in text:
        pattern = re.compile(
            r"(?m)^(\d{2}\s+[A-Z]{3}\s+\d{2})\s+(.+?)\s+(\S+)\s+"
            r"([\d,]+(?:\.\d+)?)\s+(?:(?:([\d,]+(?:\.\d+)?)\s+)?([\d,]+(?:\.\d+)?)\s+(Dr|Cr))$"
        )
        for match in pattern.finditer(text):
            date, vtype, reference, first, second, balance, side = match.groups()
            # Receipt rows carry the amount in Credit; invoice rows in Debit.
            is_credit = "RECEIPT" in vtype.upper() or "CREDIT NOTE" in vtype.upper()
            rows.append(company_row(company_date(date), vtype, reference,
                                    0 if is_credit else first, first if is_credit else 0,
                                    balance, source))
        format_name = "Chaudhary Ledger"
    elif "TRISHUL" in text and "|Particulars" in text:
        pattern = re.compile(
            r"(?m)^([A-Z]+)\s+\|(\d{2}/[A-Za-z]{3}/\d{4})\|([^|]+)\|"
            r"\s*([\d,.]*)\|\s*([\d,.]*)\|\s*([\d,.]+)\|(Dr|Cr)"
        )
        for match in pattern.finditer(text):
            book, date, particulars, debit, credit, balance, side = match.groups()
            rows.append(company_row(company_date(date), f"{book} {particulars}", "",
                                    debit, credit, balance, source))
        format_name = "Trishul Ledger"
    elif "Manoj Kumar Baid" in text and "Marico" in text:
        compact = re.sub(r"\s+", " ", text)
        pattern = re.compile(
            r"(\d{1,2}-\d{1,2}-\d{4})(To|By)(.+?)"
            r"([\d,]+\.\d{2})(\d+)(Marico Tax \(GST\)|Receipt|Credit Note Journal|Journal)"
        )
        for match in pattern.finditer(compact):
            date, side, particulars, value, reference, voucher = match.groups()
            is_credit = side == "By"
            rows.append(company_row(company_date(date), f"{side} {particulars.strip()}",
                                    reference, 0 if is_credit else value,
                                    value if is_credit else 0, 0, source, "Review"))
            rows[-1]["voucher_type"] = "Credit Note" if "Credit Note" in voucher else (
                "Payment" if voucher == "Receipt" else "Purchase"
            )
            rows[-1]["status"] = "Review Required"
        format_name = "Tally-style Marico Ledger"
    else:
        # Generic statement: date + particulars + debit + credit + optional balance.
        pattern = re.compile(
            r"(?m)^(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})\s+(.+?)\s+"
            r"([\d,]+(?:\.\d{2}))\s+([\d,]+(?:\.\d{2}))(?:\s+(-?[\d,]+(?:\.\d{2})))?\s*$"
        )
        for match in pattern.finditer(text):
            date, particulars, debit, credit, balance = match.groups()
            reference = next(iter(re.findall(r"\b[A-Z0-9/-]{5,}\b", particulars)), "")
            rows.append(company_row(company_date(date), particulars, reference,
                                    debit, credit, balance or 0, source, "Review"))
        format_name = "Generic Company Statement"
    rows = [row for row in rows if row["date"] and (row["debit"] or row["credit"])]
    if not rows:
        raise ValueError("No company ledger transaction rows were detected. Use Column Mapping or upload a clearer file.")
    return rows, {"format": format_name, "pages": len(reader.pages), "rows": len(rows)}


def parse_company_image(name, raw):
    tesseract = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    if not tesseract.exists():
        raise ValueError("Image OCR requires Tesseract OCR. Use a PDF/Excel file or install Tesseract.")
    suffix = Path(name).suffix.lower() or ".png"
    with tempfile.TemporaryDirectory(prefix="Bank2Tally_ocr_", ignore_cleanup_errors=True) as folder:
        image_path = Path(folder) / f"statement{suffix}"
        image_path.write_bytes(raw)
        completed = subprocess.run(
            [str(tesseract), str(image_path), "stdout", "--psm", "6", "-l", "eng"],
            capture_output=True, text=True, timeout=120,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    if completed.returncode:
        raise ValueError("The statement image could not be read by OCR.")
    text = completed.stdout
    rows = []
    pattern = re.compile(
        r"(?m)^(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4})\s+(.+?)\s+"
        r"([\d,]+(?:\.\d{2}))\s+([\d,]+(?:\.\d{2}))(?:\s+(-?[\d,]+(?:\.\d{2})))?\s*$"
    )
    for match in pattern.finditer(text):
        date, particulars, debit, credit, balance = match.groups()
        rows.append(company_row(company_date(date), particulars, "", debit, credit, balance or 0,
                                name, "OCR Review"))
        rows[-1]["status"] = "Review Required"
    if not rows:
        raise ValueError(
            "OCR read the image but no reliable dated ledger rows were found. "
            "Upload a straight, clear table image or use Column Mapping from Excel."
        )
    return rows, {"format": "Image OCR - Review Required", "rows": len(rows)}


def company_match_key(row):
    reference = re.sub(r"[^A-Z0-9]", "", gst_text(row.get("reference")).upper())
    return company_date(row.get("date")) or gst_text(row.get("date")), reference, round(
        max(amount(row.get("debit")), amount(row.get("credit")), amount(row.get("amount"))), 2
    )


def reconcile_company_rows(statement_rows, tally_rows, tolerance=1):
    tolerance = max(0, amount(tolerance))
    used, results = set(), []
    for source_row in statement_rows:
        source_date, source_ref, source_amount = company_match_key(source_row)
        best = None
        for index, tally_row in enumerate(tally_rows):
            if index in used:
                continue
            tally_date, tally_ref, tally_amount = company_match_key(tally_row)
            score = 0
            if source_ref and tally_ref and source_ref == tally_ref:
                score += 4
            if source_date == tally_date:
                score += 2
            if abs(source_amount - tally_amount) <= tolerance:
                score += 3
            if score >= 5 and (best is None or score > best[0]):
                best = (score, index, tally_row)
        if best:
            used.add(best[1])
            results.append({**source_row, "match_status": "Matched", "tally_row": best[2], "selected": False})
        else:
            results.append({**source_row, "match_status": "Only in Statement",
                            "selected": source_row.get("status") == "Ready"})
    for index, tally_row in enumerate(tally_rows):
        if index not in used:
            results.append({**tally_row, "match_status": "Only in Tally", "selected": False})
    return results


def make_company_statement_xml(rows, party_ledger, counter_ledger):
    vouchers = []
    for index, row in enumerate(rows, 1):
        if not row.get("selected") or row.get("match_status") != "Only in Statement":
            continue
        value = amount(row.get("debit")) or amount(row.get("credit"))
        if not value:
            continue
        voucher = gst_text(row.get("voucher_type")) or ("Purchase" if amount(row.get("debit")) else "Payment")
        if voucher not in {"Purchase", "Payment", "Receipt", "Journal", "Credit Note", "Debit Note"}:
            voucher = "Journal"
        if amount(row.get("debit")):
            entries = [(party_ledger, value, "No"), (counter_ledger, -value, "Yes")]
        else:
            entries = [(party_ledger, -value, "Yes"), (counter_ledger, value, "No")]
        ledger_xml = "".join(
            f"<ALLLEDGERENTRIES.LIST><LEDGERNAME>{xml_escape(name)}</LEDGERNAME>"
            f"<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE><AMOUNT>{entry_amount:.2f}</AMOUNT>"
            f"</ALLLEDGERENTRIES.LIST>" for name, entry_amount, positive in entries
        )
        date = gst_text(row.get("date")).replace("-", "")
        reference = gst_text(row.get("reference"))
        remote_id = str(uuid.uuid5(
            uuid.NAMESPACE_URL, f"bank2tally-company:{party_ledger}:{date}:{reference}:{value:.2f}:{voucher}"
        ))
        narration = f"Imported after company ledger reconciliation - {gst_text(row.get('particulars'))}"
        vouchers.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER REMOTEID="{remote_id}" VCHTYPE="{xml_escape(voucher)}" ACTION="Create">'
            f'<DATE>{date}</DATE><VOUCHERTYPENAME>{xml_escape(voucher)}</VOUCHERTYPENAME>'
            f'<REFERENCE>{xml_escape(reference)}</REFERENCE><PARTYLEDGERNAME>{xml_escape(party_ledger)}</PARTYLEDGERNAME>'
            f'<PERSISTEDVIEW>Accounting Voucher View</PERSISTEDVIEW><NARRATION>{xml_escape(narration)}</NARRATION>'
            f'{ledger_xml}</VOUCHER></TALLYMESSAGE>'
        )
    if not vouchers:
        raise ValueError("Select at least one reviewed row that exists only in the statement.")
    return ("<?xml version=\"1.0\" encoding=\"UTF-8\"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>"
            "</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>"
            f"<REQUESTDATA>{''.join(vouchers)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>").encode("utf-8")


def gst_summary(rows):
    safe_rows = [ensure_gst_invoice_fields(row) for row in rows or []]
    return {
        "invoices": len(safe_rows),
        **{key: round(sum(gst_number(row.get(key)) for row in safe_rows), 2)
           for key in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")},
    }


def reconcile_gst_rows(rows_2a, rows_2b, tolerance=1.0):
    """
    Match GSTR-2A detail with GSTR-2B posting eligibility.

    Date rules:
    - 2A+2B matched / 2B-present: same-month invoice keeps its original date;
      an invoice from another/earlier month posts on day 1 of GSTR-2B month
    - Supplier invoice date = original invoice date (from 2A when available)
    - 2A-only: Tally voucher date = original GSTR-2A invoice date; ITC pending
    """
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    left, right = {}, {}
    for row in rows_2a:
        left.setdefault(gst_invoice_key(row), []).append(row)
    for row in rows_2b:
        right.setdefault(gst_invoice_key(row), []).append(row)
    results = []
    for key in sorted(set(left) | set(right)):
        left_rows, right_rows = list(left.get(key, [])), list(right.get(key, []))
        used_right = set()
        for row_2a in left_rows:
            candidates = [(index, row) for index, row in enumerate(right_rows) if index not in used_right]
            if candidates:
                def candidate_score(pair):
                    _, candidate = pair
                    date_penalty = 0 if gst_text(row_2a.get("invoice_date")).replace("/", "-") == gst_text(candidate.get("invoice_date")).replace("/", "-") else 1000000
                    amount_penalty = sum(abs(gst_number(candidate.get(field)) - gst_number(row_2a.get(field))) for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess"))
                    return date_penalty + amount_penalty
                right_index, row_2b = min(candidates, key=candidate_score)
                used_right.add(right_index)
            else:
                row_2b = None
            if row_2b:
                differences = {
                    field: round(gst_number(row_2b.get(field)) - gst_number(row_2a.get(field)), 2)
                    for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
                }
                date_match = gst_text(row_2a.get("invoice_date")).replace("/", "-") == gst_text(row_2b.get("invoice_date")).replace("/", "-")
                value_match = all(abs(value) <= tolerance for value in differences.values())
                if date_match and value_match:
                    status = "2A + 2B Matched"
                    category = "matched_2a_2b"
                    review_required = False
                    ready = True
                    itc_status = "ITC Eligible via GSTR-2B"
                elif value_match:
                    status = "Date Mismatch"
                    category = "mismatch"
                    review_required = True
                    ready = True
                    itc_status = "Review Required"
                else:
                    status = "Amount/Tax Mismatch"
                    category = "mismatch"
                    review_required = True
                    ready = True
                    itc_status = "Review Required"
                # Accounting entry carries original GSTR-2A invoice details;
                # GSTR-2B controls eligible return period / voucher DATE only.
                base = row_2a
                period = gst_text(row_2b.get("source_period") or row_2b.get("gstr2b_period"))
                original_date = gst_text(row_2a.get("invoice_date") or row_2b.get("invoice_date"))
                tally_date = tally_date_for_return_period(original_date, period)
                in_2a, in_2b = True, True
            else:
                differences, status = {}, "GSTR-2A Only / Not in GSTR-2B"
                category = "only_2a"
                review_required = False
                ready = True
                itc_status = "ITC Pending / Not Available in GSTR-2B"
                base = row_2a
                period = ""
                original_date = gst_text(row_2a.get("invoice_date"))
                # RULE 5: 2A-only voucher date = original invoice date (not a fake 2B month).
                tally_date = original_date
                in_2a, in_2b = True, False
            is_invoice = gst_text(base.get("document_type") or "Invoice").lower() in {"", "invoice", "b2b", "inv"}
            is_note = "note" in gst_text(base.get("document_type")).lower() or "amend" in gst_text(base.get("document_type")).lower()
            if is_note:
                is_invoice = False
            results.append({
                **base,
                "status": status,
                "category": category,
                "gstr2a": row_2a,
                "gstr2b": row_2b,
                "differences": differences,
                "gstr2b_period": period,
                "gstr2a_period": gst_text(row_2a.get("source_period") or row_2a.get("gstr2a_period") or ""),
                "original_invoice_no": gst_text(row_2a.get("invoice_no") or base.get("invoice_no")),
                "original_invoice_date": original_date,
                "tally_entry_date": tally_date,
                "available_in_gstr2a": in_2a,
                "available_in_gstr2b": in_2b,
                "itc_status": itc_status,
                "review_required": review_required,
                "tally_status": "Missing in Tally",
                "purchase_booked": False,
                "ready_for_tally": bool(ready and is_invoice),
                "ready_for_purchase_note": bool(ready and is_note),
            })
        for index, row_2b in enumerate(right_rows):
            if index in used_right:
                continue
            period = gst_text(row_2b.get("source_period") or row_2b.get("gstr2b_period"))
            original_date = gst_text(row_2b.get("invoice_date"))
            document_text = gst_text(row_2b.get("document_type") or "Invoice").lower()
            is_note = "note" in document_text or "amend" in document_text
            # RULE 7: 2B without 2A details → Review Required (do not invent 2A detail).
            results.append({
                **row_2b,
                "status": "GSTR-2B Found / GSTR-2A Details Missing",
                "category": "only_2b_review",
                "gstr2a": None,
                "gstr2b": row_2b,
                "differences": {},
                "gstr2b_period": period,
                "gstr2a_period": "",
                "original_invoice_no": gst_text(row_2b.get("invoice_no")),
                "original_invoice_date": original_date,
                "tally_entry_date": tally_date_for_return_period(original_date, period),
                "available_in_gstr2a": False,
                "available_in_gstr2b": True,
                "itc_status": "Review Required — 2A details missing",
                "review_required": True,
                "tally_status": "Missing in Tally",
                "purchase_booked": False,
                "ready_for_tally": not is_note,
                "ready_for_purchase_note": is_note,
            })
    # Annotate against existing Tally purchase rows when available (non-fatal).
    try:
        tally_rows = [ensure_gst_invoice_fields(row) for row in gst_recon_load_rows("TALLY_PURCHASE")]
    except Exception:
        tally_rows = []
    if tally_rows:
        tally_by_key = {}
        for row in tally_rows:
            tally_by_key.setdefault(gst_invoice_key(row), []).append(row)
        for item in results:
            matches = tally_by_key.get(gst_invoice_key(item)) or []
            if not matches:
                continue
            # Prefer same original invoice date + amount when multiple.
            best = None
            for candidate in matches:
                same_date = gst_text(candidate.get("invoice_date")).replace("/", "-") == gst_text(
                    item.get("original_invoice_date") or item.get("invoice_date")
                ).replace("/", "-")
                amount_ok = abs(
                    gst_number(candidate.get("invoice_value") or candidate.get("taxable_value"))
                    - gst_number(item.get("invoice_value") or item.get("taxable_value"))
                ) <= max(tolerance, 1.0)
                if same_date and amount_ok:
                    best = candidate
                    break
            if best is None:
                best = matches[0]
            item["tally"] = best
            item["purchase_booked"] = True
            item["tally_status"] = "Already in Tally"
            item["tally_voucher_no"] = gst_text(best.get("voucher_no") or best.get("voucher_number") or best.get("invoice_no"))
            if item.get("category") == "only_2a":
                item["status"] = "Booked in Tally / Not Available in GSTR-2B"
                item["itc_status"] = "ITC Pending / Not Available in GSTR-2B"
            elif item.get("category") == "matched_2a_2b":
                item["status"] = "2A + 2B + Tally Matched"
                item["category"] = "matched_2a_2b_tally"
            item["ready_for_tally"] = False

    counts = {}
    category_counts = {}
    for row in results:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
        category_counts[row.get("category") or ""] = category_counts.get(row.get("category") or "", 0) + 1
    dashboard = build_purchase_2a2b_dashboard(results)
    return results, counts, category_counts, dashboard


def build_purchase_2a2b_dashboard(results):
    """Summary cards for Purchase / GSTR-2 (count + taxable)."""
    def pack(predicate):
        rows = [row for row in results if predicate(row)]
        return {
            "count": len(rows),
            "taxable_value": round(sum(gst_number(row.get("taxable_value")) for row in rows), 2),
            "invoice_value": round(sum(gst_number(row.get("invoice_value")) for row in rows), 2),
        }

    return {
        "total_2b": pack(lambda r: r.get("available_in_gstr2b")),
        "total_2a": pack(lambda r: r.get("available_in_gstr2a")),
        "matched_2a_2b": pack(lambda r: r.get("category") in {"matched_2a_2b", "matched_2a_2b_tally"}),
        "matched_2a_2b_tally": pack(lambda r: r.get("category") == "matched_2a_2b_tally"),
        "matched_missing_tally": pack(
            lambda r: r.get("category") == "matched_2a_2b" and not r.get("purchase_booked")
        ),
        "only_2b_review": pack(lambda r: r.get("category") == "only_2b_review"),
        "only_2a": pack(lambda r: r.get("category") == "only_2a" or (
            r.get("available_in_gstr2a") and not r.get("available_in_gstr2b")
        )),
        "mismatch": pack(lambda r: r.get("category") == "mismatch"),
        "already_in_tally": pack(lambda r: r.get("purchase_booked")),
        "ready_to_send": pack(lambda r: r.get("ready_for_tally")),
        "sent_to_tally": pack(lambda r: gst_text(r.get("tally_status")).lower() in {"sent to tally", "already in tally"}),
    }


def purchase_row_already_in_tally(row, tally_rows=None, tolerance=1.0):
    """Return matching Tally purchase row if duplicate would be created."""
    if tally_rows is None:
        try:
            tally_rows = [ensure_gst_invoice_fields(item) for item in gst_recon_load_rows("TALLY_PURCHASE")]
        except Exception:
            tally_rows = []
    key = gst_invoice_key(row)
    original_date = gst_text(row.get("original_invoice_date") or row.get("invoice_date")).replace("/", "-")
    amount = gst_number(row.get("invoice_value") or row.get("taxable_value"))
    for candidate in tally_rows:
        if gst_invoice_key(candidate) != key:
            continue
        cand_date = gst_text(candidate.get("invoice_date") or candidate.get("original_invoice_date")).replace("/", "-")
        cand_amount = gst_number(candidate.get("invoice_value") or candidate.get("taxable_value"))
        if original_date and cand_date and original_date != cand_date:
            continue
        if abs(cand_amount - amount) > max(tolerance, 1.0):
            continue
        return candidate
    return None


def gst_recon_connection():
    connection = sqlite3.connect(GST_RECON_DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_recon_meta (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL DEFAULT '',
        updated_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_recon_rows (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dataset_key TEXT NOT NULL,
        row_json TEXT NOT NULL,
        gstin TEXT DEFAULT '',
        invoice_no TEXT DEFAULT '',
        source_period TEXT DEFAULT '',
        imported_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_gst_recon_dataset ON gst_recon_rows(dataset_key)"
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_recon_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        recon_type TEXT NOT NULL DEFAULT '2b_tally',
        row_json TEXT NOT NULL,
        status TEXT DEFAULT '',
        reconciled_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    # Phase 2: GSTR-1 import batches / invoices / sales reconciliation
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_1_import_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        gstin TEXT DEFAULT '',
        financial_year TEXT DEFAULT '',
        return_period TEXT DEFAULT '',
        import_date TEXT NOT NULL DEFAULT '',
        file_name TEXT NOT NULL DEFAULT '',
        record_count INTEGER DEFAULT 0,
        return_type TEXT NOT NULL DEFAULT 'GSTR1',
        file_digest TEXT DEFAULT '',
        UNIQUE(file_digest, return_period, gstin)
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_1_invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id INTEGER,
        row_json TEXT NOT NULL,
        gstin TEXT DEFAULT '',
        invoice_no TEXT DEFAULT '',
        invoice_no_norm TEXT DEFAULT '',
        invoice_date TEXT DEFAULT '',
        section TEXT DEFAULT '',
        document_type TEXT DEFAULT '',
        source_period TEXT DEFAULT '',
        taxable_value REAL DEFAULT 0,
        igst REAL DEFAULT 0,
        cgst REAL DEFAULT 0,
        sgst REAL DEFAULT 0,
        cess REAL DEFAULT 0,
        invoice_value REAL DEFAULT 0,
        imported_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_gst1_invoice_period ON gst_1_invoices(source_period)"
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_1_reconciliation (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        return_period TEXT DEFAULT '',
        row_json TEXT NOT NULL,
        status TEXT DEFAULT '',
        review_action TEXT DEFAULT '',
        reconciled_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    # Phase 3: GSTR-3B / liability / ITC claim summaries
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_3b_import_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        gstin TEXT DEFAULT '',
        financial_year TEXT DEFAULT '',
        return_period TEXT DEFAULT '',
        import_date TEXT NOT NULL DEFAULT '',
        file_name TEXT NOT NULL DEFAULT '',
        record_count INTEGER DEFAULT 0,
        return_type TEXT NOT NULL DEFAULT 'GSTR3B',
        file_digest TEXT DEFAULT '',
        UNIQUE(file_digest, return_period, gstin)
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_3b_summary (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id INTEGER,
        return_period TEXT DEFAULT '',
        summary_json TEXT NOT NULL,
        imported_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_liability_summary (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        return_period TEXT DEFAULT '',
        source TEXT DEFAULT '',
        summary_json TEXT NOT NULL,
        calculated_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_itc_claim_summary (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        return_period TEXT DEFAULT '',
        summary_json TEXT NOT NULL,
        calculated_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    # GST Payment & Ledger datasets (isolated from GSTR-1/2B/3B).
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_payment_import_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        gstin TEXT DEFAULT '',
        financial_year TEXT DEFAULT '',
        data_type TEXT NOT NULL,
        import_date TEXT NOT NULL DEFAULT '',
        file_name TEXT NOT NULL DEFAULT '',
        record_count INTEGER DEFAULT 0,
        file_digest TEXT DEFAULT '',
        meta_json TEXT DEFAULT '',
        UNIQUE(file_digest, data_type, gstin, financial_year)
        )"""
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_payment_rows (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        batch_id INTEGER,
        data_type TEXT NOT NULL,
        gstin TEXT DEFAULT '',
        financial_year TEXT DEFAULT '',
        source_period TEXT DEFAULT '',
        row_json TEXT NOT NULL,
        imported_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_gst_payment_scope "
        "ON gst_payment_rows(data_type, gstin, financial_year, source_period)"
    )
    connection.execute(
        """CREATE TABLE IF NOT EXISTS gst_payment_recon_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        gstin TEXT DEFAULT '',
        financial_year TEXT DEFAULT '',
        recon_type TEXT NOT NULL DEFAULT '',
        row_json TEXT NOT NULL,
        status TEXT DEFAULT '',
        reconciled_at TEXT NOT NULL DEFAULT ''
        )"""
    )
    # Portal isolation columns (GSTIN + FY). Safe to re-run.
    for ddl in (
        "ALTER TABLE gst_recon_rows ADD COLUMN taxpayer_gstin TEXT DEFAULT ''",
        "ALTER TABLE gst_recon_rows ADD COLUMN financial_year TEXT DEFAULT ''",
        "ALTER TABLE gst_1_invoices ADD COLUMN taxpayer_gstin TEXT DEFAULT ''",
        "ALTER TABLE gst_1_invoices ADD COLUMN financial_year TEXT DEFAULT ''",
    ):
        try:
            connection.execute(ddl)
        except sqlite3.OperationalError:
            pass
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_gst_recon_portal_scope "
        "ON gst_recon_rows(dataset_key, taxpayer_gstin, financial_year, source_period)"
    )
    connection.commit()
    return connection


GST_PORTAL_CONTEXT_KEY = "gst_portal_context"
GST_PORTAL_DATASETS = {"GSTR-2B", "GSTR-2A", "GSTR-1", "GSTR-3B"}


def infer_taxpayer_gstin_from_text(*values):
    """Extract a GSTIN from filenames / headers (filing taxpayer)."""
    for value in values:
        text = gst_text(value).upper()
        match = re.search(r"\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][A-Z0-9]Z[A-Z0-9])\b", text)
        if match:
            return match.group(1)
    return ""


def gst_portal_default_fy():
    return "2025-26"


def gst_portal_get_context():
    meta = gst_recon_get_meta(GST_PORTAL_CONTEXT_KEY, {}) or {}
    gstin = gst_text(meta.get("gstin")).upper()
    fy = gst_text(meta.get("financial_year")) or gst_portal_default_fy()
    return {"gstin": gstin, "financial_year": fy}


def gst_portal_set_context(gstin="", financial_year=""):
    gstin = gst_text(gstin).upper()
    fy = gst_text(financial_year) or gst_portal_default_fy()
    if not gstin:
        return gst_portal_get_context()
    context = {"gstin": gstin, "financial_year": fy, "updated_at": gst_recon_now()}
    gst_recon_set_meta(GST_PORTAL_CONTEXT_KEY, context)
    return context


def gst_portal_infer_gstin_from_db():
    """Best-effort filing GSTIN from stored portal imports."""
    connection = gst_recon_connection()
    try:
        for query in (
            "SELECT gstin FROM gst_3b_import_batches WHERE IFNULL(gstin,'')!='' ORDER BY id DESC LIMIT 1",
            "SELECT gstin FROM gst_1_import_batches WHERE IFNULL(gstin,'')!='' ORDER BY id DESC LIMIT 1",
            "SELECT taxpayer_gstin FROM gst_recon_rows WHERE dataset_key='GSTR-2B' AND IFNULL(taxpayer_gstin,'')!='' ORDER BY id DESC LIMIT 1",
        ):
            row = connection.execute(query).fetchone()
            if row and gst_text(row[0]):
                return gst_text(row[0]).upper()
        # Filename fallback (older GSTR-1/3B imports).
        for query in (
            "SELECT file_name FROM gst_3b_import_batches ORDER BY id DESC LIMIT 5",
            "SELECT file_name FROM gst_1_import_batches ORDER BY id DESC LIMIT 5",
        ):
            for row in connection.execute(query).fetchall():
                gstin = infer_taxpayer_gstin_from_text(row[0])
                if gstin:
                    return gstin
    finally:
        connection.close()
    return ""


def gst_portal_resolve_context(gstin="", financial_year=""):
    """
    Resolve active portal scope.
    Explicit args win; else saved context; else infer from DB / default FY.
    Never silently replace a saved GSTIN with a different inferred GSTIN.
    """
    explicit_gstin = gst_text(gstin).upper()
    fy = gst_text(financial_year) or ""
    saved = gst_portal_get_context()
    if not explicit_gstin:
        explicit_gstin = gst_text(saved.get("gstin")).upper()
    if not explicit_gstin:
        explicit_gstin = gst_portal_infer_gstin_from_db()
        if explicit_gstin:
            fy = fy or saved.get("financial_year") or gst_portal_default_fy()
            gst_portal_set_context(explicit_gstin, fy)
            return {"gstin": explicit_gstin, "financial_year": fy}
    if not fy:
        fy = saved.get("financial_year") or gst_portal_default_fy()
    if explicit_gstin and explicit_gstin != saved.get("gstin"):
        gst_portal_set_context(explicit_gstin, fy)
    elif explicit_gstin and not saved.get("gstin"):
        gst_portal_set_context(explicit_gstin, fy)
    return {"gstin": explicit_gstin, "financial_year": fy}


def gst_row_financial_year(row):
    row = row or {}
    fy = gst_text(row.get("financial_year"))
    if fy:
        return fy
    return gstr1_financial_year(row.get("source_period") or row.get("gstr2b_period") or row.get("return_period"))


def gst_stamp_portal_row(row, taxpayer_gstin="", financial_year="", return_type=""):
    """Attach filing GSTIN + FY markers used for persistence isolation."""
    item = dict(row or {})
    gstin = gst_text(taxpayer_gstin or item.get("taxpayer_gstin") or item.get("filing_gstin")).upper()
    fy = gst_text(financial_year or item.get("financial_year")) or gst_row_financial_year(item) or gst_portal_default_fy()
    if gstin:
        item["taxpayer_gstin"] = gstin
        item["filing_gstin"] = gstin
    item["financial_year"] = fy
    if return_type:
        item["return_type"] = gst_text(item.get("return_type") or return_type)
    return item


def gst_portal_row_in_scope(row, gstin="", financial_year="", context=None):
    ctx = context or gst_portal_resolve_context(gstin, financial_year)
    want_gstin = gst_text(ctx.get("gstin")).upper()
    want_fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()
    item = row or {}
    row_gstin = gst_text(item.get("taxpayer_gstin") or item.get("filing_gstin")).upper()
    row_fy = gst_text(item.get("financial_year")) or gst_row_financial_year(item) or ""
    if want_gstin and row_gstin and row_gstin != want_gstin:
        return False
    if want_gstin and not row_gstin:
        # Legacy unscoped rows belong only to the active filing GSTIN.
        return True
    if not want_gstin and row_gstin:
        # No active company — do not surface GSTIN-scoped portal rows.
        return False
    if want_fy and row_fy and row_fy not in {want_fy, "", "ALL"}:
        period = normalize_gst_recon_period(item.get("source_period") or item.get("return_period"))
        if period and period not in set(gst_fy_period_values(want_fy)):
            return False
    elif want_fy and not row_fy:
        period = normalize_gst_recon_period(item.get("source_period") or item.get("return_period"))
        if period and period not in set(gst_fy_period_values(want_fy)):
            return False
    return True


def gst_recon_now():
    return datetime.now().isoformat(timespec="seconds")


def gst_recon_set_meta(key, value):
    connection = gst_recon_connection()
    try:
        connection.execute(
            "INSERT INTO gst_recon_meta(key,value,updated_at) VALUES(?,?,?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at",
            (key, json.dumps(value, ensure_ascii=False), gst_recon_now()),
        )
        connection.commit()
    finally:
        connection.close()


def gst_recon_get_meta(key, default=None):
    connection = gst_recon_connection()
    try:
        row = connection.execute("SELECT value FROM gst_recon_meta WHERE key=?", (key,)).fetchone()
    finally:
        connection.close()
    if not row:
        return default
    try:
        return json.loads(row["value"])
    except json.JSONDecodeError:
        return default


def gst_portal_backfill_scope_columns(connection=None):
    """Stamp legacy portal rows with taxpayer_gstin + financial_year when missing."""
    owns = connection is None
    connection = connection or gst_recon_connection()
    try:
        marker = connection.execute(
            "SELECT value FROM gst_recon_meta WHERE key=?",
            ("gst_portal_scope_backfilled_v1",),
        ).fetchone()
        # Re-run when any portal row still lacks taxpayer_gstin.
        pending = connection.execute(
            "SELECT COUNT(*) AS n FROM gst_recon_rows "
            "WHERE dataset_key IN ('GSTR-2B','GSTR-1') AND IFNULL(taxpayer_gstin,'')=''"
        ).fetchone()
        pending_n = int((pending["n"] if pending else 0) or 0)
        if marker and pending_n == 0:
            return
        filing = ""
        for query in (
            "SELECT gstin FROM gst_3b_import_batches WHERE IFNULL(gstin,'')!='' ORDER BY id DESC LIMIT 1",
            "SELECT gstin FROM gst_1_import_batches WHERE IFNULL(gstin,'')!='' ORDER BY id DESC LIMIT 1",
        ):
            row = connection.execute(query).fetchone()
            if row and gst_text(row[0]):
                filing = gst_text(row[0]).upper()
                break
        if not filing:
            for query in (
                "SELECT file_name FROM gst_3b_import_batches ORDER BY id DESC LIMIT 8",
                "SELECT file_name FROM gst_1_import_batches ORDER BY id DESC LIMIT 8",
            ):
                for row in connection.execute(query).fetchall():
                    filing = infer_taxpayer_gstin_from_text(row[0])
                    if filing:
                        break
                if filing:
                    break
        records = connection.execute(
            "SELECT id, row_json, source_period, taxpayer_gstin, financial_year "
            "FROM gst_recon_rows WHERE dataset_key IN ('GSTR-2B','GSTR-1')"
        ).fetchall()
        for record in records:
            try:
                payload = json.loads(record["row_json"] or "{}")
            except (TypeError, json.JSONDecodeError, ValueError):
                payload = {}
            taxpayer = gst_text(
                record["taxpayer_gstin"] or payload.get("taxpayer_gstin") or payload.get("filing_gstin") or filing
            ).upper()
            fy = gst_text(record["financial_year"] or payload.get("financial_year")) or gstr1_financial_year(
                record["source_period"] or payload.get("source_period") or payload.get("gstr2b_period")
            ) or gst_portal_default_fy()
            if taxpayer:
                payload["taxpayer_gstin"] = taxpayer
                payload["filing_gstin"] = taxpayer
            payload["financial_year"] = fy
            connection.execute(
                "UPDATE gst_recon_rows SET taxpayer_gstin=?, financial_year=?, row_json=? WHERE id=?",
                (taxpayer, fy, json.dumps(payload, ensure_ascii=False), record["id"]),
            )
        inv_rows = connection.execute(
            "SELECT id, row_json, source_period, taxpayer_gstin, financial_year FROM gst_1_invoices"
        ).fetchall()
        for record in inv_rows:
            try:
                payload = json.loads(record["row_json"] or "{}")
            except (TypeError, json.JSONDecodeError, ValueError):
                payload = {}
            taxpayer = gst_text(
                record["taxpayer_gstin"] or payload.get("taxpayer_gstin") or payload.get("filing_gstin") or filing
            ).upper()
            fy = gst_text(record["financial_year"] or payload.get("financial_year")) or gstr1_financial_year(
                record["source_period"] or payload.get("source_period")
            ) or gst_portal_default_fy()
            if taxpayer:
                payload["taxpayer_gstin"] = taxpayer
                payload["filing_gstin"] = taxpayer
            payload["financial_year"] = fy
            connection.execute(
                "UPDATE gst_1_invoices SET taxpayer_gstin=?, financial_year=?, row_json=? WHERE id=?",
                (taxpayer, fy, json.dumps(payload, ensure_ascii=False), record["id"]),
            )
        # Fill blank batch FY/GSTIN from filenames where possible.
        for table in ("gst_1_import_batches", "gst_3b_import_batches"):
            batches = connection.execute(
                f"SELECT id, gstin, financial_year, return_period, file_name FROM {table}"
            ).fetchall()
            for batch in batches:
                gstin = gst_text(batch["gstin"]).upper() or infer_taxpayer_gstin_from_text(batch["file_name"]) or filing
                fy = gst_text(batch["financial_year"]) or gstr1_financial_year(batch["return_period"]) or gst_portal_default_fy()
                if gstin != gst_text(batch["gstin"]).upper() or fy != gst_text(batch["financial_year"]):
                    connection.execute(
                        f"UPDATE {table} SET gstin=?, financial_year=? WHERE id=?",
                        (gstin, fy, batch["id"]),
                    )
        if filing:
            existing = connection.execute(
                "SELECT value FROM gst_recon_meta WHERE key=?",
                (GST_PORTAL_CONTEXT_KEY,),
            ).fetchone()
            if not existing:
                connection.execute(
                    "INSERT INTO gst_recon_meta(key,value,updated_at) VALUES(?,?,?)",
                    (
                        GST_PORTAL_CONTEXT_KEY,
                        json.dumps(
                            {"gstin": filing, "financial_year": gst_portal_default_fy(), "updated_at": gst_recon_now()},
                            ensure_ascii=False,
                        ),
                        gst_recon_now(),
                    ),
                )
        connection.execute(
            "INSERT INTO gst_recon_meta(key,value,updated_at) VALUES(?,?,?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at",
            ("gst_portal_scope_backfilled_v1", json.dumps({"ok": True}, ensure_ascii=False), gst_recon_now()),
        )
        connection.commit()
    finally:
        if owns:
            connection.close()


def gst_recon_delete_portal_scope_rows(connection, dataset_key, gstin="", financial_year=""):
    """Delete portal rows for one GSTIN + FY scope only (keeps other companies/years)."""
    ctx = gst_portal_resolve_context(gstin, financial_year)
    want_gstin = ctx.get("gstin")
    want_fy = ctx.get("financial_year") or gst_portal_default_fy()
    period_set = set(gst_fy_period_values(want_fy))
    records = connection.execute(
        "SELECT id, row_json, source_period, taxpayer_gstin, financial_year FROM gst_recon_rows WHERE dataset_key=?",
        (dataset_key,),
    ).fetchall()
    deleted = 0
    for record in records:
        try:
            payload = json.loads(record["row_json"] or "{}")
        except (TypeError, json.JSONDecodeError, ValueError):
            payload = {}
        row_gstin = gst_text(
            record["taxpayer_gstin"] or payload.get("taxpayer_gstin") or payload.get("filing_gstin")
        ).upper()
        row_fy = gst_text(record["financial_year"] or payload.get("financial_year")) or gstr1_financial_year(
            record["source_period"] or payload.get("source_period") or payload.get("gstr2b_period")
        )
        period = normalize_gst_recon_period(record["source_period"] or payload.get("source_period") or payload.get("gstr2b_period"))
        if want_gstin and row_gstin and row_gstin != want_gstin:
            continue
        if want_gstin and not row_gstin:
            # Legacy unscoped rows belong to the active filing GSTIN only.
            pass
        elif not want_gstin and row_gstin:
            continue
        if period and period not in period_set:
            continue
        if row_fy and row_fy not in {want_fy, ""}:
            continue
        connection.execute("DELETE FROM gst_recon_rows WHERE id=?", (record["id"],))
        deleted += 1
    return deleted


def gst_recon_save_rows(dataset_key, rows, gstin="", financial_year=""):
    dataset_key = gst_text(dataset_key)
    if not dataset_key:
        raise ValueError("Dataset key is required.")
    stamp = gst_recon_now()
    portal = dataset_key in GST_PORTAL_DATASETS
    ctx = gst_portal_resolve_context(gstin, financial_year) if portal else {"gstin": "", "financial_year": ""}
    if portal and ctx.get("gstin"):
        gst_portal_set_context(ctx["gstin"], ctx.get("financial_year") or gst_portal_default_fy())
    stamped_rows = []
    for row in rows or []:
        item = dict(row or {})
        if portal:
            item = gst_stamp_portal_row(
                item,
                taxpayer_gstin=ctx.get("gstin") or item.get("taxpayer_gstin") or item.get("filing_gstin"),
                financial_year=ctx.get("financial_year") or item.get("financial_year"),
                return_type=dataset_key,
            )
            if not ctx.get("gstin") and item.get("taxpayer_gstin"):
                ctx = gst_portal_resolve_context(item.get("taxpayer_gstin"), item.get("financial_year"))
                gst_portal_set_context(ctx["gstin"], ctx.get("financial_year") or gst_portal_default_fy())
        stamped_rows.append(item)
    connection = gst_recon_connection()
    try:
        if portal:
            gst_portal_backfill_scope_columns(connection)
            gst_recon_delete_portal_scope_rows(
                connection,
                dataset_key,
                gstin=ctx.get("gstin") or gstin,
                financial_year=ctx.get("financial_year") or financial_year or gst_portal_default_fy(),
            )
        else:
            connection.execute("DELETE FROM gst_recon_rows WHERE dataset_key=?", (dataset_key,))
        for row in stamped_rows:
            connection.execute(
                """INSERT INTO gst_recon_rows
                   (dataset_key,row_json,gstin,invoice_no,source_period,imported_at,taxpayer_gstin,financial_year)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (
                    dataset_key,
                    json.dumps(row, ensure_ascii=False),
                    gst_text(row.get("gstin")).upper(),
                    gst_text(row.get("invoice_no")),
                    gst_text(row.get("source_period") or row.get("gstr2b_period")),
                    stamp,
                    gst_text(row.get("taxpayer_gstin") or row.get("filing_gstin")).upper(),
                    gst_text(row.get("financial_year")) or gst_portal_default_fy(),
                ),
            )
        connection.commit()
    finally:
        connection.close()
    scoped_count = len(stamped_rows)
    gst_recon_set_meta(f"dataset_count:{dataset_key}", scoped_count if portal else scoped_count)
    if dataset_key == "GSTR-2B" and stamped_rows:
        periods = {
            normalize_gst_recon_period(row.get("source_period") or row.get("gstr2b_period"))
            for row in stamped_rows
        }
        for period in periods:
            if period:
                gst_session_mark_imported("GSTR-2B", period)


def gst_recon_load_rows(dataset_key, gstin="", financial_year=""):
    dataset_key = gst_text(dataset_key)
    connection = gst_recon_connection()
    try:
        if dataset_key in GST_PORTAL_DATASETS:
            gst_portal_backfill_scope_columns(connection)
        records = connection.execute(
            "SELECT row_json, taxpayer_gstin, financial_year, source_period FROM gst_recon_rows "
            "WHERE dataset_key=? ORDER BY id",
            (dataset_key,),
        ).fetchall()
    finally:
        connection.close()
    rows = []
    context = None
    if dataset_key in GST_PORTAL_DATASETS:
        context = gst_portal_resolve_context(gstin, financial_year)
    for record in records:
        try:
            item = json.loads(record["row_json"])
        except (TypeError, json.JSONDecodeError, ValueError):
            continue
        if dataset_key in GST_PORTAL_DATASETS:
            if not item.get("taxpayer_gstin"):
                item["taxpayer_gstin"] = gst_text(record["taxpayer_gstin"]).upper()
            if not item.get("financial_year"):
                item["financial_year"] = gst_text(record["financial_year"]) or gstr1_financial_year(
                    record["source_period"]
                )
            if not gst_portal_row_in_scope(item, gstin=gstin, financial_year=financial_year, context=context):
                continue
        rows.append(item)
    return rows


def gst_recon_save_results(results, recon_type="2b_tally"):
    stamp = gst_recon_now()
    connection = gst_recon_connection()
    try:
        connection.execute("DELETE FROM gst_recon_results WHERE recon_type=?", (recon_type,))
        for row in results or []:
            connection.execute(
                "INSERT INTO gst_recon_results(recon_type,row_json,status,reconciled_at) VALUES (?,?,?,?)",
                (recon_type, json.dumps(row, ensure_ascii=False), gst_text(row.get("status")), stamp),
            )
        connection.commit()
    finally:
        connection.close()


def gst_recon_load_results(recon_type="2b_tally"):
    connection = gst_recon_connection()
    try:
        records = connection.execute(
            "SELECT row_json FROM gst_recon_results WHERE recon_type=? ORDER BY id",
            (recon_type,),
        ).fetchall()
    finally:
        connection.close()
    rows = []
    for record in records:
        try:
            rows.append(json.loads(record["row_json"]))
        except json.JSONDecodeError:
            continue
    return rows


def gst_recon_count_rows(dataset_key, gstin="", financial_year=""):
    """Fast row count without parsing invoice JSON blobs."""
    dataset_key = gst_text(dataset_key)
    if not dataset_key:
        return 0
    connection = gst_recon_connection()
    try:
        if dataset_key in GST_PORTAL_DATASETS:
            gst_portal_backfill_scope_columns(connection)
            ctx = gst_portal_resolve_context(gstin, financial_year)
            want_gstin = gst_text(ctx.get("gstin")).upper()
            want_fy = gst_text(ctx.get("financial_year") or gst_portal_default_fy())
            if want_gstin:
                row = connection.execute(
                    "SELECT COUNT(*) AS c FROM gst_recon_rows WHERE dataset_key=? AND "
                    "(IFNULL(taxpayer_gstin,'')='' OR UPPER(taxpayer_gstin)=?) AND "
                    "(IFNULL(financial_year,'')='' OR financial_year=?)",
                    (dataset_key, want_gstin, want_fy),
                ).fetchone()
            else:
                row = connection.execute(
                    "SELECT COUNT(*) AS c FROM gst_recon_rows WHERE dataset_key=?",
                    (dataset_key,),
                ).fetchone()
            return int((row["c"] if row else 0) or 0)
        row = connection.execute(
            "SELECT COUNT(*) AS c FROM gst_recon_rows WHERE dataset_key=?",
            (dataset_key,),
        ).fetchone()
        return int((row["c"] if row else 0) or 0)
    finally:
        connection.close()


def gst_recon_count_results(recon_type="2b_tally"):
    connection = gst_recon_connection()
    try:
        row = connection.execute(
            "SELECT COUNT(*) AS c FROM gst_recon_results WHERE recon_type=?",
            (gst_text(recon_type) or "2b_tally",),
        ).fetchone()
        return int((row["c"] if row else 0) or 0)
    finally:
        connection.close()


def gstr1_count_invoices(gstin="", financial_year=""):
    """Count GSTR-1 invoices without loading full JSON rows."""
    connection = gst_recon_connection()
    try:
        gst_portal_backfill_scope_columns(connection)
        ctx = gst_portal_resolve_context(gstin, financial_year)
        want_gstin = gst_text(ctx.get("gstin")).upper()
        want_fy = gst_text(ctx.get("financial_year") or gst_portal_default_fy())
        # Prefer invoice table (same source as gstr1_load_invoices).
        try:
            if want_gstin:
                row = connection.execute(
                    "SELECT COUNT(*) AS c FROM gst_1_invoices WHERE "
                    "(IFNULL(taxpayer_gstin,'')='' OR UPPER(taxpayer_gstin)=?) AND "
                    "(IFNULL(financial_year,'')='' OR financial_year=?)",
                    (want_gstin, want_fy),
                ).fetchone()
            else:
                row = connection.execute("SELECT COUNT(*) AS c FROM gst_1_invoices").fetchone()
            count = int((row["c"] if row else 0) or 0)
            if count:
                return count
        except Exception:
            pass
        return gst_recon_count_rows("GSTR-1", gstin=gstin, financial_year=financial_year)
    finally:
        connection.close()


def gstr1_count_reconciliation(return_period=""):
    period = normalize_gst_recon_period(return_period)
    connection = gst_recon_connection()
    try:
        if period:
            row = connection.execute(
                "SELECT COUNT(*) AS c FROM gst_1_reconciliation WHERE return_period=?",
                (period,),
            ).fetchone()
        else:
            row = connection.execute("SELECT COUNT(*) AS c FROM gst_1_reconciliation").fetchone()
        count = int((row["c"] if row else 0) or 0)
        if count:
            return count
        return gst_recon_count_results("gstr1_tally")
    finally:
        connection.close()


def gst_strip_dashboard_rows(dashboard):
    """Return dashboard cards/meta without embedding full voucher row lists."""
    if not isinstance(dashboard, dict) or not dashboard:
        return {}
    out = {key: value for key, value in dashboard.items() if key != "rows"}
    rows = dashboard.get("rows")
    if isinstance(rows, list):
        out["row_count"] = len(rows)
    elif "row_count" not in out:
        out["row_count"] = int(gst_number(dashboard.get("row_count")))
    return out


def gst_strip_gstr3b_dashboard_wire(dashboard, include_drilldown=False):
    """
    Keep GSTR-3B card/compare payload; drop heavy voucher drilldown lists by default.
    Accounting figures are unchanged — only the HTTP payload is trimmed.
    """
    if not isinstance(dashboard, dict) or not dashboard:
        return {}
    skip = {"books_itc_lines", "gstr3b"}
    if not include_drilldown:
        skip.add("outward_classification_drilldown")
    out = {key: value for key, value in dashboard.items() if key not in skip}
    g3 = dashboard.get("gstr3b")
    if isinstance(g3, dict):
        out["gstr3b"] = {
            "return_period": g3.get("return_period"),
            "imported_periods": g3.get("imported_periods"),
            "gstin": g3.get("gstin"),
        }
    drill = dashboard.get("outward_classification_drilldown")
    if isinstance(drill, dict):
        out["has_outward_drilldown"] = True
        if include_drilldown:
            out["outward_classification_drilldown"] = drill
    return out


def gst_recon_overview_wire(overview):
    """Overview cards only — omit nested multi-MB dashboard copies."""
    if not isinstance(overview, dict) or not overview:
        return {}
    return {
        "return_period": overview.get("return_period") or "ALL",
        "import_status": overview.get("import_status") or {},
        "purchase": overview.get("purchase") or {},
        "sales": overview.get("sales") or {},
        "gstr3b": overview.get("gstr3b") or {},
    }


def gst_recon_load_dataset_bundle(include, gstin="", financial_year=""):
    """Load only requested recon datasets (full row arrays)."""
    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = ctx.get("gstin")
    fy = ctx.get("financial_year") or gst_portal_default_fy()
    wanted = {gst_text(item).lower() for item in (include or []) if gst_text(item)}
    if not wanted or "all" in wanted:
        wanted = {
            "gstr2b", "gstr2a", "tally_purchase", "gstr1", "tally_sales",
            "results", "gstr1_results", "gstr3b",
        }
    out = {"portal_context": ctx}
    if "gstr2b" in wanted:
        out["gstr2b"] = gst_recon_load_rows("GSTR-2B", gstin=gstin, financial_year=fy)
    if "gstr2a" in wanted:
        out["gstr2a"] = gst_recon_load_rows("GSTR-2A", gstin=gstin, financial_year=fy)
    if "tally_purchase" in wanted:
        out["tally_purchase"] = gst_recon_load_rows("TALLY_PURCHASE")
    if "gstr1" in wanted:
        out["gstr1"] = gstr1_load_invoices(gstin=gstin, financial_year=fy)
    if "tally_sales" in wanted:
        out["tally_sales"] = [ensure_gst_invoice_fields(row) for row in gst_recon_load_rows("TALLY_SALES")]
    if "results" in wanted:
        out["results"] = gst_recon_load_results("2b_tally")
    if "gstr1_results" in wanted:
        out["gstr1_results"] = gstr1_load_reconciliation() or gst_recon_load_results("gstr1_tally")
    if "gstr3b" in wanted:
        out["gstr3b"] = gstr3b_load_summary() or {}
    return out


GST_PORTAL_META_KEYS = {
    "GSTR-2B": (
        "itc_dashboard",
        "itc_difference_recon",
        "gst_recon_overview",
        "dataset_count:GSTR-2B",
    ),
    "GSTR-2A": (
        "dataset_count:GSTR-2A",
        "gst_recon_overview",
    ),
    "GSTR-1": (
        "sales_dashboard",
        "gstr1_last_import",
        "gstr1_difference_recon",
        "gst_recon_overview",
        "dataset_count:GSTR-1",
    ),
    "GSTR-3B": (
        "GSTR-3B",
        "gstr3b_last_import",
        "gstr3b_dashboard",
        "gst_recon_overview",
        "itc_dashboard",
        "itc_difference_recon",
    ),
}


def normalize_portal_return_type(value):
    """Map UI / API return-type labels to GSTR-1 | GSTR-2A | GSTR-2B | GSTR-3B."""
    text = gst_text(value).upper().replace("_", " ").strip()
    compact = re.sub(r"[\s\-]", "", text)
    if "3B" in text or compact in {"GSTR3B", "GSTR03B"}:
        return "GSTR-3B"
    # Check 2A before 2B so labels containing neither are unambiguous.
    if "2A" in text or compact in {"GSTR2A", "GSTR02A"}:
        return "GSTR-2A"
    if "2B" in text or compact in {"GSTR2B", "GSTR02B"}:
        return "GSTR-2B"
    if (
        compact.startswith("GSTR1")
        or text.startswith("GSTR-1")
        or "SALES REGISTER" in text
        or "SALES JSON" in text
    ):
        return "GSTR-1"
    return ""


def gst_recon_delete_meta_keys(keys):
    """Hard-delete meta keys so empty {} caches cannot revive portal totals."""
    key_list = [gst_text(key) for key in (keys or []) if gst_text(key)]
    if not key_list:
        return
    connection = gst_recon_connection()
    try:
        for key in key_list:
            connection.execute("DELETE FROM gst_recon_meta WHERE key=?", (key,))
        connection.commit()
    finally:
        connection.close()


def gst_fy_period_values(financial_year=""):
    """Return MMYYYY periods for the selected FY (currently FY 2025-26)."""
    fy = gst_text(financial_year) or "2025-26"
    if fy in {"2025-26", "2526", "FY2025-26", "FY 2025-26"}:
        return [period for period, _label in gst_fy_2025_26_periods()]
    return [period for period, _label in gst_fy_2025_26_periods()]


def gst_session_clear_imports_for(kind, periods=None):
    kind = normalize_portal_return_type(kind) or gst_text(kind).upper()
    if kind not in {"GSTR-2B", "GSTR-2A", "GSTR-1", "GSTR-3B"}:
        return
    if periods is None:
        GST_SESSION_IMPORTS[kind] = set()
        return
    drop = {normalize_gst_recon_period(item) for item in periods if normalize_gst_recon_period(item)}
    GST_SESSION_IMPORTS[kind] = {
        period for period in (GST_SESSION_IMPORTS.get(kind) or set()) if period not in drop
    }


def gst_recon_clear_portal_return(return_type, gstin="", financial_year=""):
    """
    Completely remove one portal return type for the selected GSTIN + FY.
    Keeps Tally sync rows/meta and the other GST return types.
    """
    kind = normalize_portal_return_type(return_type)
    if kind not in {"GSTR-1", "GSTR-2A", "GSTR-2B", "GSTR-3B"}:
        raise ValueError("returnType must be GSTR-1, GSTR-2A, GSTR-2B, or GSTR-3B.")
    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = gst_text(ctx.get("gstin")).upper()
    fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()
    periods = gst_fy_period_values(fy)
    period_set = set(periods)

    connection = gst_recon_connection()
    try:
        gst_portal_backfill_scope_columns(connection)
        if kind == "GSTR-2B":
            gst_recon_delete_portal_scope_rows(connection, "GSTR-2B", gstin=gstin, financial_year=fy)
            connection.execute("DELETE FROM gst_recon_results WHERE recon_type=?", ("2b_tally",))
            connection.execute("DELETE FROM gst_recon_results WHERE recon_type=?", ("2a_2b",))

        elif kind == "GSTR-2A":
            gst_recon_delete_portal_scope_rows(connection, "GSTR-2A", gstin=gstin, financial_year=fy)
            connection.execute("DELETE FROM gst_recon_results WHERE recon_type=?", ("2a_2b",))

        elif kind == "GSTR-1":
            batch_ids = []
            for batch in connection.execute(
                "SELECT id, gstin, financial_year, return_period, file_name FROM gst_1_import_batches"
            ).fetchall():
                batch_gstin = gst_text(batch["gstin"]).upper() or infer_taxpayer_gstin_from_text(batch["file_name"])
                batch_fy = gst_text(batch["financial_year"]) or gstr1_financial_year(batch["return_period"]) or fy
                period = normalize_gst_recon_period(batch["return_period"])
                if gstin and batch_gstin and batch_gstin != gstin:
                    continue
                if batch_fy and batch_fy not in {fy, ""}:
                    continue
                if period and period not in period_set:
                    continue
                batch_ids.append(batch["id"])
            # Invoice rows scoped by taxpayer_gstin / FY / period.
            for record in connection.execute(
                "SELECT id, taxpayer_gstin, financial_year, source_period, row_json FROM gst_1_invoices"
            ).fetchall():
                try:
                    payload = json.loads(record["row_json"] or "{}")
                except (TypeError, json.JSONDecodeError, ValueError):
                    payload = {}
                row_gstin = gst_text(
                    record["taxpayer_gstin"] or payload.get("taxpayer_gstin") or payload.get("filing_gstin")
                ).upper()
                row_fy = gst_text(record["financial_year"] or payload.get("financial_year")) or gstr1_financial_year(
                    record["source_period"]
                )
                period = normalize_gst_recon_period(record["source_period"] or payload.get("source_period"))
                if gstin and row_gstin and row_gstin != gstin:
                    continue
                if period and period not in period_set:
                    continue
                if row_fy and row_fy not in {fy, ""}:
                    continue
                connection.execute("DELETE FROM gst_1_invoices WHERE id=?", (record["id"],))
            for period in periods:
                connection.execute("DELETE FROM gst_1_reconciliation WHERE return_period=?", (period,))
            connection.execute(
                "DELETE FROM gst_1_reconciliation WHERE IFNULL(return_period,'')=''"
            )
            if batch_ids:
                placeholders = ",".join("?" * len(batch_ids))
                connection.execute(
                    f"DELETE FROM gst_1_import_batches WHERE id IN ({placeholders})",
                    batch_ids,
                )
            gst_recon_delete_portal_scope_rows(connection, "GSTR-1", gstin=gstin, financial_year=fy)
            connection.execute("DELETE FROM gst_recon_results WHERE recon_type=?", ("gstr1_tally",))

        elif kind == "GSTR-3B":
            batch_ids = []
            for batch in connection.execute(
                "SELECT id, gstin, financial_year, return_period, file_name FROM gst_3b_import_batches"
            ).fetchall():
                batch_gstin = gst_text(batch["gstin"]).upper() or infer_taxpayer_gstin_from_text(batch["file_name"])
                batch_fy = gst_text(batch["financial_year"]) or gstr1_financial_year(batch["return_period"]) or fy
                period = normalize_gst_recon_period(batch["return_period"])
                if gstin and batch_gstin and batch_gstin != gstin:
                    continue
                if batch_fy and batch_fy not in {fy, ""}:
                    continue
                if period and period not in period_set:
                    continue
                batch_ids.append(batch["id"])
            if batch_ids:
                placeholders = ",".join("?" * len(batch_ids))
                connection.execute(
                    f"DELETE FROM gst_3b_summary WHERE batch_id IN ({placeholders})",
                    batch_ids,
                )
                connection.execute(
                    f"DELETE FROM gst_3b_import_batches WHERE id IN ({placeholders})",
                    batch_ids,
                )
            # Other GSTINs still present?
            other_gstin = connection.execute(
                "SELECT COUNT(*) AS n FROM gst_3b_import_batches WHERE IFNULL(gstin,'') NOT IN ('', ?)",
                (gstin or "__none__",),
            ).fetchone()
            other_count = int((other_gstin["n"] if other_gstin else 0) or 0)
            # Always drop FY derived caches for this company scope.
            for period in periods:
                connection.execute(
                    "DELETE FROM gst_liability_summary WHERE return_period=?",
                    (period,),
                )
                connection.execute(
                    "DELETE FROM gst_itc_claim_summary WHERE return_period=?",
                    (period,),
                )
            connection.execute(
                "DELETE FROM gst_liability_summary WHERE IFNULL(return_period,'') IN ('', 'ALL')"
            )
            connection.execute(
                "DELETE FROM gst_itc_claim_summary WHERE IFNULL(return_period,'') IN ('', 'ALL')"
            )
            if other_count == 0:
                # Single-company / no remaining batches: wipe every FY summary row
                # including orphans with blank/mismatched batch_id.
                for period in periods:
                    connection.execute("DELETE FROM gst_3b_summary WHERE return_period=?", (period,))
                connection.execute(
                    "DELETE FROM gst_3b_summary WHERE IFNULL(return_period,'')=''"
                )
                connection.execute("DELETE FROM gst_3b_import_batches")
                connection.execute("DELETE FROM gst_3b_summary")
            else:
                # Delete orphan summaries whose period was cleared for this GSTIN.
                for period in periods:
                    still = connection.execute(
                        """SELECT COUNT(*) AS n FROM gst_3b_summary s
                           LEFT JOIN gst_3b_import_batches b ON b.id = s.batch_id
                           WHERE s.return_period=? AND IFNULL(b.gstin,'') NOT IN ('', ?)""",
                        (period, gstin or "__none__"),
                    ).fetchone()
                    if not still or int(still["n"] or 0) == 0:
                        connection.execute(
                            "DELETE FROM gst_3b_summary WHERE return_period=? "
                            "AND (batch_id IS NULL OR batch_id NOT IN "
                            "(SELECT id FROM gst_3b_import_batches))",
                            (period,),
                        )

        connection.commit()
    finally:
        connection.close()

    gst_recon_delete_meta_keys(GST_PORTAL_META_KEYS.get(kind, ()))
    gst_session_clear_imports_for(kind, periods)

    if kind == "GSTR-1":
        # Recreate GSTR-1 recon mirror from any surviving invoice rows only.
        sync_gstr1_recon_rows_from_invoices()
    elif kind == "GSTR-3B":
        gstr3b_rebuild_period_index()
        # Rebuild overview cache without 3B figures so restart cannot revive them.
        overview = build_gst_recon_overview("", 1, gstin=gstin, financial_year=fy)
        gst_recon_set_meta("gst_recon_overview", overview)
    elif kind == "GSTR-2B":
        gst_recon_set_meta("dataset_count:GSTR-2B", len(gst_recon_load_rows("GSTR-2B", gstin=gstin, financial_year=fy)))
    elif kind == "GSTR-2A":
        gst_recon_set_meta("dataset_count:GSTR-2A", len(gst_recon_load_rows("GSTR-2A", gstin=gstin, financial_year=fy)))

    return {
        "cleared": True,
        "return_type": kind,
        "gstin": gstin,
        "financial_year": fy,
        "periods_cleared": periods,
    }


def gst_recon_clear_all():
    """Clear all portal imports + derived caches. Keep Tally sync rows/meta."""
    for kind in ("GSTR-2B", "GSTR-1", "GSTR-3B"):
        gst_recon_clear_portal_return(kind, gstin="", financial_year="2025-26")
    gst_session_set_imports({"GSTR-2B": [], "GSTR-1": [], "GSTR-3B": []})
    return {"cleared": True, "return_type": "ALL"}


# ---------------------------------------------------------------------------
# GST Payment & Ledger — challan / cash / credit parsers + reconciliation
# ---------------------------------------------------------------------------

GST_PAYMENT_DATA_TYPES = (
    "challan_history",
    "cash_ledger",
    "credit_ledger",
)

GST_PAYMENT_TYPE_LABELS = {
    "challan_history": "GST Payment List / Challan History",
    "cash_ledger": "Electronic Cash Ledger",
    "credit_ledger": "Electronic Credit / ITC Ledger",
}

# UI / legacy aliases → canonical storage keys
GST_PAYMENT_TYPE_ALIASES = {
    "GST_PAYMENT_LIST": "challan_history",
    "GST_CASH_LEDGER": "cash_ledger",
    "GST_ITC_LEDGER": "credit_ledger",
    "challan_history": "challan_history",
    "cash_ledger": "cash_ledger",
    "credit_ledger": "credit_ledger",
    "itc_ledger": "credit_ledger",
    "payment_list": "challan_history",
}

# Reverse map for status chips that still expect legacy keys
GST_PAYMENT_LEGACY_KEYS = {
    "challan_history": "GST_PAYMENT_LIST",
    "cash_ledger": "GST_CASH_LEDGER",
    "credit_ledger": "GST_ITC_LEDGER",
}


def normalize_gst_payment_data_type(value):
    """Return canonical data type, or '' if unknown."""
    try:
        import gst_payment_ledger as gpl
        kind = gpl.normalize_payment_data_type(value)
        if kind in GST_PAYMENT_DATA_TYPES:
            return kind
    except Exception:
        pass
    key = gst_text(value).strip()
    if not key:
        return ""
    mapped = GST_PAYMENT_TYPE_ALIASES.get(key) or GST_PAYMENT_TYPE_ALIASES.get(key.upper())
    if mapped in GST_PAYMENT_DATA_TYPES:
        return mapped
    lower = key.lower().replace("-", "_").replace(" ", "_")
    return lower if lower in GST_PAYMENT_DATA_TYPES else ""


def gst_payment_type_aliases(data_type):
    canon = normalize_gst_payment_data_type(data_type)
    aliases = [canon] if canon else []
    for alias, target in GST_PAYMENT_TYPE_ALIASES.items():
        if target == canon and alias not in aliases:
            aliases.append(alias)
    return aliases


def gst_payment_fy_bounds(financial_year):
    text = gst_text(financial_year) or gst_portal_default_fy()
    match = re.fullmatch(r"(20\d{2})\s*[-–]\s*(\d{2}|\d{4})", text)
    if not match:
        return None, None
    start_year = int(match.group(1))
    end_raw = match.group(2)
    end_year = int(end_raw) if len(end_raw) == 4 else 2000 + int(end_raw)
    try:
        return datetime(start_year, 4, 1).date(), datetime(end_year, 3, 31).date()
    except ValueError:
        return None, None


def gst_payment_parse_portal_date(value):
    text = gst_text(value)
    if not text:
        return None
    # Strip Excel wrapper if still present
    m = re.fullmatch(r'="(.*)"', text)
    if m:
        text = m.group(1)
    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def gst_payment_delete_dataset(connection, gstin, fy, data_type):
    """Permanently clear one Payment & Ledger dataset for GSTIN+FY only."""
    gstin = gst_text(gstin).upper()
    fy = gst_text(fy) or gst_portal_default_fy()
    aliases = gst_payment_type_aliases(data_type)
    if not gstin or not fy or not aliases:
        return 0
    placeholders = ",".join("?" for _ in aliases)
    params = [gstin, fy, *aliases]
    batch_ids = [
        int(r["id"])
        for r in connection.execute(
            f"""
            SELECT id FROM gst_payment_import_batches
            WHERE IFNULL(gstin,'')=? AND IFNULL(financial_year,'')=? AND data_type IN ({placeholders})
            """,
            params,
        ).fetchall()
    ]
    deleted_rows = 0
    if batch_ids:
        ph = ",".join("?" for _ in batch_ids)
        deleted_rows = connection.execute(
            f"DELETE FROM gst_payment_rows WHERE batch_id IN ({ph})",
            batch_ids,
        ).rowcount
        # Also delete by scope in case batch_id is null on older rows
        connection.execute(
            f"""
            DELETE FROM gst_payment_rows
            WHERE IFNULL(gstin,'')=? AND IFNULL(financial_year,'')=? AND data_type IN ({placeholders})
            """,
            params,
        )
        connection.execute(
            f"DELETE FROM gst_payment_import_batches WHERE id IN ({ph})",
            batch_ids,
        )
    connection.execute(
        f"""
        DELETE FROM gst_payment_recon_results
        WHERE IFNULL(gstin,'')=? AND IFNULL(financial_year,'')=?
          AND (recon_type IN ({placeholders}) OR IFNULL(recon_type,'') LIKE ?)
        """,
        [*params, f"{normalize_gst_payment_data_type(data_type)}%"],
    )
    return int(deleted_rows or 0)


def gst_payment_clear_dataset(data_type, gstin="", financial_year=""):
    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = gst_text(ctx.get("gstin")).upper()
    fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()
    kind = normalize_gst_payment_data_type(data_type)
    if kind not in GST_PAYMENT_DATA_TYPES:
        raise ValueError("dataType must be challan_history, cash_ledger, credit_ledger (or GST_PAYMENT_LIST / GST_CASH_LEDGER / GST_ITC_LEDGER).")
    connection = gst_recon_connection()
    try:
        deleted = gst_payment_delete_dataset(connection, gstin, fy, kind)
        connection.commit()
    finally:
        connection.close()
    status = gst_payment_build_status(gstin=gstin, financial_year=fy)
    return {
        "ok": True,
        "cleared": True,
        "data_type": kind,
        "legacy_data_type": GST_PAYMENT_LEGACY_KEYS.get(kind, kind),
        "deleted_rows": deleted,
        "gstin": gstin,
        "financial_year": fy,
        "status": status,
        "message": (
            f"Cleared {GST_PAYMENT_TYPE_LABELS.get(kind, kind)} for {gstin} / {fy}. "
            "GSTR-1, GSTR-2B, GSTR-3B, Tally and other Payment & Ledger datasets were not touched."
        ),
    }


def gst_payment_load_rows(data_type="", gstin="", financial_year=""):
    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = gst_text(ctx.get("gstin")).upper()
    fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()
    connection = gst_recon_connection()
    try:
        sql = """
            SELECT id, batch_id, data_type, gstin, financial_year, source_period, row_json, imported_at
            FROM gst_payment_rows
            WHERE IFNULL(gstin,'')=? AND IFNULL(financial_year,'')=?
        """
        params = [gstin, fy]
        if data_type:
            aliases = gst_payment_type_aliases(data_type)
            placeholders = ",".join("?" for _ in aliases)
            sql += f" AND data_type IN ({placeholders})"
            params.extend(aliases)
        sql += " ORDER BY data_type, id"
        out = []
        for record in connection.execute(sql, params).fetchall():
            try:
                payload = json.loads(record["row_json"] or "{}")
            except Exception:
                payload = {}
            if not isinstance(payload, dict):
                payload = {"value": payload}
            item = dict(payload)
            item["_row_id"] = record["id"]
            item["_data_type"] = normalize_gst_payment_data_type(record["data_type"])
            item["_source_period"] = record["source_period"] or ""
            # Legacy scaffold compatibility
            if "CPIN" not in item and item.get("cpin"):
                item["CPIN"] = item.get("cpin")
            if "Deposit Date" not in item and item.get("deposit_date"):
                item["Deposit Date"] = item.get("deposit_date")
            if "Amount" not in item and item.get("amount") is not None:
                item["Amount"] = item.get("amount")
            out.append(item)
        return out
    finally:
        connection.close()


def gst_payment_load_batches(gstin="", financial_year=""):
    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = gst_text(ctx.get("gstin")).upper()
    fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()
    connection = gst_recon_connection()
    try:
        rows = connection.execute(
            """
            SELECT * FROM gst_payment_import_batches
            WHERE IFNULL(gstin,'')=? AND IFNULL(financial_year,'')=?
            ORDER BY id DESC
            """,
            (gstin, fy),
        ).fetchall()
        out = []
        for row in rows:
            item = dict(row)
            try:
                item["meta"] = json.loads(item.get("meta_json") or "{}")
            except Exception:
                item["meta"] = {}
            item["data_type"] = normalize_gst_payment_data_type(item.get("data_type"))
            out.append(item)
        return out
    finally:
        connection.close()


def gst_payment_dataset_status(data_type, gstin="", financial_year=""):
    kind = normalize_gst_payment_data_type(data_type)
    rows = gst_payment_load_rows(kind, gstin=gstin, financial_year=financial_year)
    batches = [
        b for b in gst_payment_load_batches(gstin, financial_year)
        if b.get("data_type") == kind
    ]
    latest = batches[0] if batches else {}
    meta = latest.get("meta") or {}
    return {
        "data_type": kind,
        "legacy_data_type": GST_PAYMENT_LEGACY_KEYS.get(kind, kind),
        "imported": bool(rows),
        "record_count": len(rows),
        "row_count": len(rows),
        "file_name": latest.get("file_name") or meta.get("source_filename") or "",
        "import_date": latest.get("import_date") or "",
        "validation": meta.get("validation") or {},
        "stats": meta.get("stats") or {},
        "meta": meta,
    }


def gst_payment_save_import(data_type, file_name, raw, gstin="", financial_year=""):
    """Parse and permanently store one Payment & Ledger dataset for GSTIN+FY."""
    return gst_payment_save_imports(
        data_type,
        [{"name": file_name, "raw": raw or b""}],
        gstin=gstin,
        financial_year=financial_year,
    )


def gst_payment_save_imports(data_type, file_items, gstin="", financial_year=""):
    """
    Store one Payment & Ledger dataset from one or more files.
    Challan History: merge + CPIN de-dupe across files.
    Cash / Credit: last-file-wins (single-file expected).
    """
    import gst_payment_ledger as gpl

    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = gst_text(ctx.get("gstin") or gstin).upper()
    fy = gst_text(ctx.get("financial_year") or financial_year) or gst_portal_default_fy()
    kind = normalize_gst_payment_data_type(data_type)
    if kind not in GST_PAYMENT_DATA_TYPES:
        raise ValueError("dataType must be challan_history, cash_ledger, credit_ledger (or GST_PAYMENT_LIST / GST_CASH_LEDGER / GST_ITC_LEDGER).")
    items = [item for item in (file_items or []) if item is not None]
    if not items:
        raise ValueError("Select a GST Payment / Ledger file to import.")
    if not gstin:
        # Infer from first filename when context empty.
        gstin = infer_taxpayer_gstin_from_text(items[0].get("name") or "")
    if not gstin:
        raise ValueError("Company GSTIN is required before importing Payment & Ledger files.")

    parsed_list = []
    source_names = []
    digests = []
    for item in items:
        name = item.get("name") or "payment.csv"
        raw = item.get("raw") or b""
        source_names.append(Path(name).name)
        digests.append(hashlib.sha256(raw).hexdigest())
        parsed = gpl.parse_payment_ledger_file(kind, name, raw)
        # Attach source filename into meta for merge stats.
        meta = dict(parsed.get("meta") or {})
        meta["source_filename"] = Path(name).name
        parsed["meta"] = meta
        for record in parsed.get("records") or []:
            if not record.get("source_file"):
                record["source_file"] = Path(name).name
        # GSTIN / FY validation per file when metadata present.
        detected_gstin = gst_text(meta.get("gstin")).upper()
        if not detected_gstin:
            detected_gstin = infer_taxpayer_gstin_from_text(name)
        if detected_gstin and detected_gstin != gstin:
            raise ValueError(
                f"GSTIN mismatch in {Path(name).name}: file has {detected_gstin}, "
                f"current company is {gstin}. Import rejected."
            )
        from_date = meta.get("from_date") or ""
        to_date = meta.get("to_date") or ""
        fy_start, fy_end = gst_payment_fy_bounds(fy)
        d_from = gst_payment_parse_portal_date(from_date)
        d_to = gst_payment_parse_portal_date(to_date)
        if fy_start and fy_end and d_from and d_to and (d_to < fy_start or d_from > fy_end):
            warnings = list((parsed.get("validation") or {}).get("warnings") or [])
            warnings.append(
                f"{Path(name).name}: date range {from_date}–{to_date} does not overlap selected FY {fy}."
            )
            validation = dict(parsed.get("validation") or {})
            validation["warnings"] = warnings
            if validation.get("status") == "OK" or validation.get("ok"):
                validation["status"] = "WARN"
            parsed["validation"] = validation
        # Soft FY check from challan deposit dates when no metadata range.
        if kind == "challan_history" and fy_start and fy_end and not (d_from and d_to):
            dates = []
            for record in parsed.get("records") or []:
                dt = gst_payment_parse_portal_date(record.get("deposit_date") or record.get("created_on") or "")
                if dt:
                    dates.append(dt)
            if dates and (max(dates) < fy_start or min(dates) > fy_end):
                warnings = list((parsed.get("validation") or {}).get("warnings") or [])
                warnings.append(
                    f"{Path(name).name}: challan dates do not overlap selected FY {fy}."
                )
                validation = dict(parsed.get("validation") or {})
                validation["warnings"] = warnings
                validation["status"] = "WARN"
                parsed["validation"] = validation
        parsed_list.append(parsed)

    merge_stats = None
    if kind == "challan_history":
        merged = gpl.merge_challan_history_parses(parsed_list)
        records = list(merged.get("records") or [])
        meta_in = dict(merged.get("meta") or {})
        validation = dict(merged.get("validation") or {})
        merge_stats = dict(merged.get("merge_stats") or {})
        detected_gstin = gstin
        from_date = meta_in.get("from_date") or ""
        to_date = meta_in.get("to_date") or ""
        # Prefer earliest/latest deposit dates for display.
        deposit_dates = [
            gst_payment_parse_portal_date(r.get("deposit_date") or r.get("created_on") or "")
            for r in records
        ]
        deposit_dates = [d for d in deposit_dates if d]
        if deposit_dates:
            from_date = min(deposit_dates).strftime("%d/%m/%Y")
            to_date = max(deposit_dates).strftime("%d/%m/%Y")
            meta_in["from_date"] = from_date
            meta_in["to_date"] = to_date
    else:
        # Cash / Credit remain single-file oriented; use the last file if several are sent.
        parsed = parsed_list[-1]
        records = list(parsed.get("records") or [])
        meta_in = dict(parsed.get("meta") or {})
        validation = dict(parsed.get("validation") or {})
        detected_gstin = gst_text(meta_in.get("gstin")).upper() or gstin
        from_date = meta_in.get("from_date") or ""
        to_date = meta_in.get("to_date") or ""

    errors = list(validation.get("errors") or [])
    warnings = list(validation.get("warnings") or [])
    validation["warnings"] = warnings
    validation["errors"] = errors
    if "status" not in validation:
        validation["status"] = "OK" if validation.get("ok", len(errors) == 0) else "WARN"
    if warnings and validation.get("status") == "OK":
        validation["status"] = "WARN"
    validation["detected_gstin"] = detected_gstin or ""
    validation["company_gstin"] = gstin
    validation["from_date"] = from_date
    validation["to_date"] = to_date
    validation["rows_imported"] = len(records)
    validation["files_processed"] = len(source_names)

    stats = {
        "rows_imported": len(records),
        "files_processed": len(source_names),
        "credit_rows": meta_in.get("credit_rows"),
        "debit_rows": meta_in.get("debit_rows"),
        "opening_balance": meta_in.get("opening_balance_total"),
        "closing_balance": meta_in.get("closing_balance_total"),
        "paid_challans": meta_in.get("paid_challans"),
        "failed_challans": meta_in.get("failed_challans"),
        "total_paid_amount": meta_in.get("total_paid_amount"),
        "rows_read": (merge_stats or {}).get("rows_read", len(records)),
        "unique_challans": (merge_stats or {}).get("unique_challans", len(records) if kind == "challan_history" else None),
        "duplicates_skipped": (merge_stats or {}).get("duplicates_skipped", 0 if kind == "challan_history" else None),
    }
    if merge_stats:
        stats.update({
            "files_processed": merge_stats.get("files_processed", len(source_names)),
            "rows_read": merge_stats.get("rows_read"),
            "unique_challans": merge_stats.get("unique_challans"),
            "duplicates_skipped": merge_stats.get("duplicates_skipped"),
            "paid_challans": merge_stats.get("paid_challans"),
            "failed_challans": merge_stats.get("failed_challans"),
            "total_paid_amount": merge_stats.get("total_paid_amount"),
        })

    stamp = gst_recon_now()
    digest = hashlib.sha256("|".join(digests).encode("utf-8")).hexdigest()
    display_name = source_names[0] if len(source_names) == 1 else f"{len(source_names)} files"
    meta = {
        "source_filename": display_name,
        "source_files": source_names,
        "detected_gstin": detected_gstin or "",
        "legal_name": meta_in.get("legal_name") or "",
        "from_date": from_date,
        "to_date": to_date,
        "validation": validation,
        "stats": stats,
        "merge_stats": merge_stats or {},
        "parser": "gst_payment_ledger",
        "kind": kind,
        "label": GST_PAYMENT_TYPE_LABELS.get(kind, kind),
    }

    connection = gst_recon_connection()
    try:
        gst_payment_delete_dataset(connection, gstin, fy, kind)
        cur = connection.execute(
            """
            INSERT INTO gst_payment_import_batches(
                gstin, financial_year, data_type, import_date, file_name,
                record_count, file_digest, meta_json
            ) VALUES (?,?,?,?,?,?,?,?)
            """,
            (
                gstin,
                fy,
                kind,
                stamp,
                display_name,
                len(records),
                digest,
                json.dumps(meta, ensure_ascii=False),
            ),
        )
        batch_id = int(cur.lastrowid)
        for record in records:
            period = (
                gst_text(record.get("tax_period_norm"))
                or gst_text(record.get("tax_period"))
                or gst_text(record.get("period_key"))
                or ""
            )
            connection.execute(
                """
                INSERT INTO gst_payment_rows(
                    batch_id, data_type, gstin, financial_year, source_period, row_json, imported_at
                ) VALUES (?,?,?,?,?,?,?)
                """,
                (
                    batch_id,
                    kind,
                    gstin,
                    fy,
                    period,
                    json.dumps(record, ensure_ascii=False),
                    stamp,
                ),
            )
        connection.commit()
    finally:
        connection.close()

    if gstin:
        gst_portal_set_context(gstin, fy)

    status = gst_payment_build_status(gstin=gstin, financial_year=fy)
    if kind == "challan_history" and merge_stats:
        message = (
            f"Challan History: {merge_stats.get('files_processed', 0)} file(s) · "
            f"rows read {merge_stats.get('rows_read', 0)} · "
            f"unique {merge_stats.get('unique_challans', 0)} · "
            f"duplicates skipped {merge_stats.get('duplicates_skipped', 0)} · "
            f"PAID {merge_stats.get('paid_challans', 0)} · "
            f"FAILED {merge_stats.get('failed_challans', 0)} · "
            f"PAID amount {merge_stats.get('total_paid_amount', 0)}"
        )
    else:
        message = (
            f"Imported {len(records)} {GST_PAYMENT_TYPE_LABELS.get(kind, kind)} row(s). "
            f"GSTIN={detected_gstin or gstin}; From={from_date or '—'}; To={to_date or '—'}; "
            f"Validation={validation.get('status') or ('OK' if validation.get('ok') else 'WARN')}."
        )
    return {
        "ok": True,
        "data_type": kind,
        "legacy_data_type": GST_PAYMENT_LEGACY_KEYS.get(kind, kind),
        "record_count": len(records),
        "row_count": len(records),
        "rows_imported": len(records),
        "file_name": display_name,
        "source_files": source_names,
        "gstin": gstin,
        "financial_year": fy,
        "from_date": from_date,
        "to_date": to_date,
        "validation": validation,
        "stats": stats,
        "merge_stats": merge_stats or {},
        "meta": meta,
        "status": status,
        "message": message,
        "format_pending": False,
        "parse_error": "",
    }


def gst_payment_num(value, default=0.0):
    """Coerce GSTR-3B / ledger values that may be numbers or head dicts."""
    if value is None or value == "":
        return float(default)
    if isinstance(value, dict):
        for key in ("output_gst", "total", "amount", "value"):
            if value.get(key) is not None and not isinstance(value.get(key), dict):
                try:
                    return float(value.get(key) or 0)
                except (TypeError, ValueError):
                    continue
        total = 0.0
        for key in ("igst", "cgst", "sgst", "cess"):
            try:
                total += float(value.get(key) or 0)
            except (TypeError, ValueError):
                pass
        return round(total, 2)
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def gst_payment_gstr3b_by_period(gstin, fy):
    """Read-only GSTR-3B liability map for payment comparison (does not alter 3B math)."""
    out = {}
    try:
        imported = gstr3b_list_imported_periods() or []
    except Exception:
        imported = []
    for period in imported:
        period = normalize_gst_recon_period(period)
        if not period:
            continue
        # Keep months that belong to selected FY when FY is known.
        try:
            if gstr1_financial_year(period) and gstr1_financial_year(period) != fy:
                continue
        except Exception:
            pass
        try:
            summary = gstr3b_load_summary_for_period(period) or {}
        except Exception:
            summary = {}
        if not summary:
            continue
        tax = (
            gst_payment_num(summary.get("igst"))
            + gst_payment_num(summary.get("cgst"))
            + gst_payment_num(summary.get("sgst"))
            + gst_payment_num(summary.get("cess"))
        )
        out[period] = {
            "tax_payable": round(tax, 2),
            "interest": round(gst_payment_num(summary.get("interest")), 2),
            "late_fee": round(gst_payment_num(summary.get("late_fee")), 2),
            "igst": round(gst_payment_num(summary.get("igst")), 2),
            "cgst": round(gst_payment_num(summary.get("cgst")), 2),
            "sgst": round(gst_payment_num(summary.get("sgst")), 2),
            "cess": round(gst_payment_num(summary.get("cess")), 2),
            "itc_igst": round(gst_payment_num(summary.get("itc_igst") or summary.get("itc_available_igst")), 2),
            "itc_cgst": round(gst_payment_num(summary.get("itc_cgst") or summary.get("itc_available_cgst")), 2),
            "itc_sgst": round(gst_payment_num(summary.get("itc_sgst") or summary.get("itc_available_sgst")), 2),
            "itc_cess": round(gst_payment_num(summary.get("itc_cess") or summary.get("itc_available_cess")), 2),
        }
    return out


def gst_payment_build_status(gstin="", financial_year=""):
    import gst_payment_ledger as gpl

    ctx = gst_portal_resolve_context(gstin, financial_year)
    gstin = gst_text(ctx.get("gstin")).upper()
    fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()

    by_type = {dt: gst_payment_load_rows(dt, gstin=gstin, financial_year=fy) for dt in GST_PAYMENT_DATA_TYPES}
    batches = gst_payment_load_batches(gstin, fy)
    g3_map = gst_payment_gstr3b_by_period(gstin, fy) if gstin else {}

    dashboard = gpl.build_payment_ledger_dashboard(
        by_type["challan_history"],
        by_type["cash_ledger"],
        by_type["credit_ledger"],
        gstr3b_by_period=g3_map,
        financial_year=fy,
    )

    # Persist recon cache (cleared with dataset Clear).
    connection = gst_recon_connection()
    try:
        connection.execute(
            "DELETE FROM gst_payment_recon_results WHERE IFNULL(gstin,'')=? AND IFNULL(financial_year,'')=?",
            (gstin, fy),
        )
        stamp = gst_recon_now()
        for item in (dashboard.get("payment_recon") or {}).get("rows") or []:
            connection.execute(
                """
                INSERT INTO gst_payment_recon_results(gstin, financial_year, recon_type, row_json, status, reconciled_at)
                VALUES (?,?,?,?,?,?)
                """,
                (
                    gstin,
                    fy,
                    "challan_vs_cash",
                    json.dumps(item, ensure_ascii=False),
                    gst_text(item.get("match_status")),
                    stamp,
                ),
            )
        for item in (dashboard.get("tally_adjustments") or {}).get("tally_recon") or []:
            connection.execute(
                """
                INSERT INTO gst_payment_recon_results(gstin, financial_year, recon_type, row_json, status, reconciled_at)
                VALUES (?,?,?,?,?,?)
                """,
                (
                    gstin,
                    fy,
                    "tally_preview",
                    json.dumps(item, ensure_ascii=False),
                    gst_text(item.get("status")),
                    stamp,
                ),
            )
        connection.commit()
    finally:
        connection.close()

    dataset_status = {dt: gst_payment_dataset_status(dt, gstin, fy) for dt in GST_PAYMENT_DATA_TYPES}
    # UI still reads legacy keys on chips
    datasets = {}
    for dt, st in dataset_status.items():
        datasets[dt] = st
        legacy = GST_PAYMENT_LEGACY_KEYS.get(dt)
        if legacy:
            datasets[legacy] = {**st, "data_type": legacy}

    gstr3b_imported = bool(g3_map)
    tally_meta = gst_recon_get_meta("tally_sync", {}) or {}
    tally_connected = bool(tally_meta.get("synced_at") or tally_meta.get("company"))

    cards = dashboard.get("summary_cards") or {}
    return {
        "ok": True,
        "gstin": gstin,
        "financial_year": fy,
        "data_types": list(GST_PAYMENT_DATA_TYPES),
        "datasets": datasets,
        "imports": dataset_status,
        "gstr3b_imported": gstr3b_imported,
        "tally_connected": tally_connected,
        "summary_cards": cards,
        "challan_summary": dashboard.get("challan_summary") or {},
        "cash_ledger": dashboard.get("cash_ledger") or {},
        "itc_ledger": dashboard.get("itc_ledger") or {},
        "payment_recon": dashboard.get("payment_recon") or {},
        "gstr3b_link": dashboard.get("gstr3b_link") or {},
        "tally_adjustments": dashboard.get("tally_adjustments") or {},
        "itc_cross_head_note": (
            (dashboard.get("itc_ledger") or {}).get("matrix_note")
            or "Cross-head utilisation breakup not available in Electronic Credit Ledger source"
        ),
        "note": (
            "Challan History = cash deposits into Electronic Cash Ledger. "
            "Cash Ledger Debits = utilisation. Credit Ledger Debits = ITC utilised. "
            "Tally Adjustment is preview-only."
        ),
    }


def gst_session_get_imports():
    return {
        "GSTR-2B": sorted(GST_SESSION_IMPORTS.get("GSTR-2B") or []),
        "GSTR-1": sorted(GST_SESSION_IMPORTS.get("GSTR-1") or []),
        "GSTR-3B": sorted(GST_SESSION_IMPORTS.get("GSTR-3B") or []),
    }


def gst_session_set_imports(data):
    global GST_SESSION_IMPORTS
    GST_SESSION_IMPORTS = {
        "GSTR-2B": {
            normalize_gst_recon_period(item)
            for item in ((data or {}).get("GSTR-2B") or [])
            if normalize_gst_recon_period(item)
        },
        "GSTR-1": {
            normalize_gst_recon_period(item)
            for item in ((data or {}).get("GSTR-1") or [])
            if normalize_gst_recon_period(item)
        },
        "GSTR-3B": {
            normalize_gst_recon_period(item)
            for item in ((data or {}).get("GSTR-3B") or [])
            if normalize_gst_recon_period(item)
        },
    }


def gst_session_mark_imported(kind, return_period=""):
    """Record a portal import for a GST period in this working session only."""
    kind = normalize_portal_return_type(kind) or gst_text(kind).upper()
    if kind not in {"GSTR-2B", "GSTR-1", "GSTR-3B"}:
        return
    period = normalize_gst_recon_period(return_period)
    if not period:
        return
    GST_SESSION_IMPORTS.setdefault(kind, set()).add(period)


def gst_session_is_imported(kind, return_period=""):
    """Imported badge is true only when this period was imported in the current session."""
    period = normalize_gst_recon_period(return_period)
    if not period:
        return False
    kind = normalize_portal_return_type(kind) or gst_text(kind).upper()
    return period in (GST_SESSION_IMPORTS.get(kind) or set())


def gst_session_reset_portal_imports():
    """
    Start a fresh GST Reconciliation working session:
    clear portal imports + recon results, keep Tally sync rows.
    """
    gst_recon_clear_all()
    return {
        "reset": True,
        "import_status": {
            "gstr2b": "Not Imported",
            "gstr1": "Not Imported",
            "gstr3b": "Not Imported",
            "gstr2b_imported": False,
            "gstr1_imported": False,
            "gstr3b_imported": False,
        },
    }


def format_tally_date(raw_date):
    raw_date = gst_text(raw_date)
    if re.fullmatch(r"\d{8}", raw_date):
        return f"{raw_date[6:8]}-{raw_date[4:6]}-{raw_date[:4]}"
    return raw_date


def tally_purchase_voucher_types():
    return {
        gst_text(item.get("name")).lower()
        for item in TALLY_CACHE.get("voucher_types", [])
        if "purchase" in gst_text(item.get("name")).lower()
        or "purchase" in gst_text(item.get("parent", "")).lower()
    }


def is_tally_purchase_voucher(voucher_type, entries, ledger_parents):
    voucher_type = gst_text(voucher_type).lower()
    if voucher_type in tally_purchase_voucher_types():
        return True
    if "purchase" in voucher_type:
        return True
    for name, _ in entries:
        parent = ledger_parents.get(gst_text(name).lower(), "")
        if parent == "purchase accounts":
            return True
    return False


def normalize_tally_ledger_parent(parent):
    """Normalize Tally parent group for comparisons (XML may use &amp;)."""
    return gst_text(parent).lower().replace("&amp;", "&").strip()


def is_tally_input_tax_ledger(name, parent):
    """
    True only for Input GST under Duties & Taxes (or Input-named tax ledgers).

    Never treat Purchase Accounts such as "Purchase IGST 5%" as tax — those inflate
    Tally Booked above official GSTR-3B Section 4C.
    Source: LEDGERNAME + ledger Parent from Tally masters.
    """
    lname = gst_text(name).lower()
    parent = normalize_tally_ledger_parent(parent)
    if "output" in lname:
        return False
    if parent == "purchase accounts":
        return False
    if parent.endswith("duties & taxes"):
        return any(token in lname for token in ("igst", "cgst", "sgst", "utgst", "cess"))
    # Fallback when parent map is incomplete: Input GST ledger names only.
    if "input" in lname and any(token in lname for token in ("igst", "cgst", "sgst", "utgst", "cess", "gst")):
        return True
    return False


def is_tally_output_igst_ledger(name, parent):
    """
    Official Tally GSTR-1 Output IGST ledger mapping (IGST only).

    XML tags used: ALLLEDGERENTRIES.LIST → LEDGERNAME (+ ledger Parent from masters)
    Accept: Duties & Taxes IGST / Output IGST
    Reject: Purchase Accounts, Purchase IGST*, Input IGST* (purchase-side contamination)

    Does not change CGST / SGST mapping.
    """
    lname = gst_text(name).lower()
    parent = normalize_tally_ledger_parent(parent)
    if "igst" not in lname:
        return False
    if "purchase" in lname or "input" in lname:
        return False
    if parent == "purchase accounts":
        return False
    if parent.endswith("duties & taxes"):
        return True
    if "output" in lname:
        return True
    return False


def extract_tally_output_igst_signed(entries, ledger_parents, has_sales_accounts, signed=1, audit_ctx=""):
    """
    Map Output IGST for Tally Sales / GSTR-1 from voucher ledger lines.

    Official Tally GSTR-1 summary IGST = Output IGST on outward supplies only.
    Do not treat Purchase IGST* or purchase-side Debit Notes as Output IGST.

    Logs: XML tag, raw Tally amount, final mapped value (only when IGST ledgers present).
    """
    igst = 0.0
    xml_tag = "ALLLEDGERENTRIES.LIST/LEDGERNAME+AMOUNT"
    saw_igst_ledger = False
    for name, value in entries or []:
        lname = gst_text(name)
        parent = ledger_parents.get(lname.lower(), "")
        raw = gst_number(value)
        if "igst" not in lname.lower():
            continue
        saw_igst_ledger = True
        accepted = bool(has_sales_accounts and is_tally_output_igst_ledger(name, parent))
        mapped = abs(raw) if accepted else 0.0
        if accepted:
            igst += mapped
        tally_log(
            f"igst-map | {audit_ctx} | xml_tag={xml_tag} | ledger={lname!r} | parent={parent!r} | "
            f"raw_amount={raw:.2f} | has_sales_accounts={bool(has_sales_accounts)} | "
            f"accepted={accepted} | mapped_abs={mapped:.2f}"
        )
    final = round(igst * gst_number(signed or 1), 2)
    if saw_igst_ledger:
        tally_log(
            f"igst-map | {audit_ctx} | xml_tag={xml_tag} | "
            f"raw_component_sum={round(igst, 2):.2f} | signed={signed} | final_ui_igst={final:.2f}"
        )
    return final


def extract_tally_input_tax_signed(entries, ledger_parents):
    """
    Signed Input tax from voucher ledger lines → GSTR-3B Section 4C style ITC.

    XML tags: ALLLEDGERENTRIES.LIST → LEDGERNAME, AMOUNT
    Ledgers: Duties & Taxes / Input (never Purchase Accounts tax-named ledgers)
    Formula: ITC delta = −AMOUNT  (Tally XML Debit is typically negative; Debit to
    Input GST increases ITC; Credit to Input GST decreases ITC).
    """
    igst = cgst = sgst = cess = 0.0
    for name, value in entries:
        parent = ledger_parents.get(gst_text(name).lower(), "")
        if not is_tally_input_tax_ledger(name, parent):
            continue
        delta = -gst_number(value)
        lname = gst_text(name).lower()
        if "igst" in lname:
            igst += delta
        elif "cgst" in lname:
            cgst += delta
        elif "sgst" in lname or "utgst" in lname:
            sgst += delta
        elif "cess" in lname:
            cess += delta
    return round(igst, 2), round(cgst, 2), round(sgst, 2), round(cess, 2)


def is_tally_purchase_itc_voucher(voucher_type, entries, ledger_parents):
    """
    Purchase-side ITC vouchers for GSTR-3B Section 4C style Net ITC:
    Purchase, Credit Note, Debit Note, and ITC reversal journals that touch
    Purchase accounts or Input GST ledgers.
    """
    if is_tally_purchase_voucher(voucher_type, entries, ledger_parents):
        return True
    voucher_type = gst_text(voucher_type).lower()
    touches_purchase_or_input = False
    for name, _ in entries:
        lname = gst_text(name).lower()
        parent = normalize_tally_ledger_parent(ledger_parents.get(lname, ""))
        if parent == "purchase accounts":
            touches_purchase_or_input = True
            break
        if is_tally_input_tax_ledger(name, parent):
            touches_purchase_or_input = True
            break
    if not touches_purchase_or_input:
        return False
    if any(token in voucher_type for token in ("credit note", "debit note", "purchase return", "itc reversal", "reversal")):
        return True
    if "journal" in voucher_type and any(
        is_tally_input_tax_ledger(name, ledger_parents.get(gst_text(name).lower(), ""))
        for name, _ in entries
    ):
        return True
    return False


def tally_purchase_document_type(voucher_type, entries, ledger_parents, signed_tax_total=None):
    """
    Classify for Section 4C formula. ITC ledger sign is authoritative: Tally often
    posts purchase returns / CN as voucher type "Debit Note" while crediting Input GST.
    """
    voucher_type = gst_text(voucher_type).lower()
    has_purchase = any(
        normalize_tally_ledger_parent(ledger_parents.get(gst_text(name).lower(), "")) == "purchase accounts"
        for name, _ in entries
    )
    if signed_tax_total is not None:
        if signed_tax_total < -0.005:
            if "reversal" in voucher_type or ("journal" in voucher_type and not has_purchase):
                return "ITC Reversal"
            return "Credit Note"
        if signed_tax_total > 0.005 and ("debit note" in voucher_type or (
            re.search(r"\bdebit\b", voucher_type) and "credit" not in voucher_type
        )):
            return "Debit Note"
        if "reversal" in voucher_type or ("journal" in voucher_type and not has_purchase):
            return "ITC Reversal"
        return "Invoice"
    if "debit note" in voucher_type:
        return "Debit Note"
    if "credit note" in voucher_type or "purchase return" in voucher_type:
        return "Credit Note"
    if "reversal" in voucher_type or ("journal" in voucher_type and not has_purchase):
        return "ITC Reversal"
    return "Invoice"


def tally_xml_is_yes(value):
    text = gst_text(value).strip().lower()
    return text in {"yes", "y", "true", "1"}


def tally_row_in_gstr3b_section_4c(row):
    """
    Tally GSTR-3B Net ITC Available (4C A−B) excludes uncertain GST vouchers.

    Source XML field on the voucher: VCHGSTSTATUSISUNCERTAIN.
    """
    if not isinstance(row, dict):
        return False
    if tally_xml_is_yes(row.get("gst_uncertain")):
        return False
    return True


def filter_tally_rows_for_gstr3b_4c(rows):
    """Keep only vouchers Tally includes in GSTR-3B Section 4C (A−B)."""
    return [row for row in (rows or []) if tally_row_in_gstr3b_section_4c(row)]


def sync_tally_purchase_vouchers():
    """
    Read Purchase / CN / DN / ITC reversal vouchers with signed Input GST.

    Tax source: Duties & Taxes Input ledgers only (ALLLEDGERENTRIES.AMOUNT signed).
    Inclusion for GSTR-3B Section 4C (A−B): VCHGSTSTATUSISUNCERTAIN must not be Yes
    (same filter Tally applies on the GSTR-3B report).
    """
    cache = sync_tally()
    ledger_lookup = {
        gst_text(item.get("name")).casefold(): gst_text(item.get("gstin"))
        for item in cache.get("ledgers", [])
    }
    ledger_parents = {
        gst_text(item.get("name")).lower(): normalize_tally_ledger_parent(item.get("parent"))
        for item in cache.get("ledgers", [])
    }
    xml = tally_collection_xml(
        "Voucher", "Voucher",
        ["Date", "VoucherTypeName", "VoucherNumber", "Reference", "ReferenceDate",
         "PartyLedgerName", "PartyGSTIN", "VCHGSTSTATUSISUNCERTAIN", "ISELIGIBLEFORITC",
         "AllLedgerEntries.LedgerName", "AllLedgerEntries.Amount"],
        timeout=90,
        purpose="purchase-voucher-sync",
        filter_formula=(
            '$VoucherTypeName = "Purchase" OR '
            '$VoucherTypeName = "Credit Note" OR '
            '$VoucherTypeName = "Debit Note"'
        ),
    )
    rows = []
    for match in re.finditer(r"<VOUCHER\b[^>]*>(.*?)</VOUCHER>", xml, re.I | re.S):
        body = match.group(1)
        voucher_type = tag_value(body, "VOUCHERTYPENAME")
        entries = []
        for entry in re.finditer(r"<ALLLEDGERENTRIES\.LIST>(.*?)</ALLLEDGERENTRIES\.LIST>", body, re.I | re.S):
            entry_body = entry.group(1)
            entries.append((
                tag_value(entry_body, "LEDGERNAME"),
                amount(tag_value(entry_body, "AMOUNT")),
            ))
        if not is_tally_purchase_itc_voucher(voucher_type, entries, ledger_parents):
            continue
        party = tag_value(body, "PARTYLEDGERNAME")
        gstin = gst_text(tag_value(body, "PARTYGSTIN")).upper()
        if not gstin and party:
            gstin = gst_text(ledger_lookup.get(party.casefold(), "")).upper()
        reference = tag_value(body, "REFERENCE") or tag_value(body, "VOUCHERNUMBER")
        invoice_date = format_tally_date(tag_value(body, "REFERENCEDATE")) or format_tally_date(tag_value(body, "DATE"))
        tally_date = format_tally_date(tag_value(body, "DATE"))
        # Signed Input tax (Section 4C). Do not abs() — credit lines reduce ITC.
        igst, cgst, sgst, cess = extract_tally_input_tax_signed(entries, ledger_parents)
        signed_tax = round(igst + cgst + sgst, 2)
        document_type = tally_purchase_document_type(
            voucher_type, entries, ledger_parents, signed_tax_total=signed_tax
        )
        sign = -1 if document_type in {"Credit Note", "ITC Reversal"} else 1
        taxable = 0.0
        for name, value in entries:
            parent = normalize_tally_ledger_parent(ledger_parents.get(gst_text(name).lower(), ""))
            if parent == "purchase accounts":
                # Debit to purchase increases taxable; mirror tax sign convention.
                taxable += -gst_number(value)
        taxable = round(taxable, 2)
        invoice_value = round(taxable + igst + cgst + sgst + cess, 2)
        if abs(invoice_value) <= 0 and abs(taxable) <= 0 and abs(signed_tax + cess) <= 0:
            continue
        gst_rate = gst_rate_for_values(abs(taxable), abs(igst), abs(cgst), abs(sgst), abs(cess))
        gst_uncertain = tally_xml_is_yes(tag_value(body, "VCHGSTSTATUSISUNCERTAIN"))
        rows.append({
            "gstin": gstin,
            "party_name": party,
            "party_ledger": party,
            "invoice_no": reference,
            "invoice_date": invoice_date,
            "tally_entry_date": tally_date,
            "taxable_value": taxable,
            "igst": igst,
            "cgst": cgst,
            "sgst": sgst,
            "cess": cess,
            "gst_rate": gst_rate,
            "invoice_value": invoice_value,
            "document_type": document_type,
            "document_sign": int(sign),
            "_portal_signed": True,
            "source": "Tally",
            "voucher_type": voucher_type,
            "gst_uncertain": gst_uncertain,
            "itc_eligible": tally_xml_is_yes(tag_value(body, "ISELIGIBLEFORITC"))
            if gst_text(tag_value(body, "ISELIGIBLEFORITC")) else True,
            "itc_source": "Tally GSTR-3B 4C A−B (Input Duties & Taxes; excludes VCHGSTSTATUSISUNCERTAIN=Yes)",
            "gstr3b_xml_field": "VCHGSTSTATUSISUNCERTAIN",
        })
    unique = {}
    for row in rows:
        key = (
            row["gstin"],
            re.sub(r"[^A-Z0-9]", "", gst_text(row.get("invoice_no")).upper()).lstrip("0") or "0",
            gst_text(row.get("invoice_date")),
            gst_text(row.get("document_type")),
            gst_text(row.get("voucher_type")),
        )
        unique[key] = row
    included = filter_tally_rows_for_gstr3b_4c(list(unique.values()))
    return {
        "company": cache.get("company", ""),
        "rows": list(unique.values()),
        "count": len(unique),
        "gstr3b_4c_count": len(included),
        "uncertain_excluded": len(unique) - len(included),
        "synced_at": gst_recon_now(),
        "section_4c_source": (
            "Tally GSTR-3B Net Input Tax Credit Available (4C A−B); "
            "XML VCHGSTSTATUSISUNCERTAIN + Input Duties & Taxes amounts"
        ),
    }


def gst_document_bucket(row):
    """Classify a portal/Tally row into a signed-reconciliation document bucket."""
    doc = " ".join(
        gst_text(row.get(key))
        for key in ("document_type", "voucher_type", "note_type", "section", "amendment_direction")
    ).lower()
    if "amend" in doc:
        if "decrease" in doc:
            return "Amendment Decrease"
        if "increase" in doc:
            return "Amendment Increase"
        return "Amendment"
    if "reversal" in doc or "itc reversal" in doc:
        return "ITC Reversal"
    if "debit note" in doc or re.search(r"\bdebit\b", doc):
        if "credit" in doc and "debit note" not in doc:
            pass
        else:
            return "Debit Note"
    if "credit note" in doc or re.search(r"\bcredit\b", doc):
        return "Credit Note"
    if any(token in doc for token in ("cdnr", "cdnur")):
        return "Credit Note"
    return "Invoice"


def gst_document_bucket_sign(bucket):
    if bucket in {"Credit Note", "Amendment Decrease"}:
        return -1.0
    return 1.0


def signed_row_tax_amounts(row):
    """
    Return (bucket, signed tax fields).
    Portal import / GSTR-1 normalize may already store negative Credit Note amounts.
    """
    row = ensure_gst_invoice_fields(row or {})
    bucket = gst_document_bucket(row)
    fields = {
        "taxable_value": gst_number(row.get("taxable_value")),
        "igst": gst_number(row.get("igst")),
        "cgst": gst_number(row.get("cgst")),
        "sgst": gst_number(row.get("sgst")),
        "cess": gst_number(row.get("cess")),
        "invoice_value": gst_number(row.get("invoice_value")),
    }
    already_signed = bool(
        row.get("_portal_signed")
        or gst_number(row.get("document_sign")) == -1
        or gst_text(row.get("return_type")).upper() in {"GSTR1", "GSTR-1"}
    )
    # Credit notes already negative from import — keep as-is.
    if already_signed or fields["taxable_value"] < 0 or fields["igst"] < 0 or fields["cgst"] < 0 or fields["sgst"] < 0:
        signed_total = round(fields["igst"] + fields["cgst"] + fields["sgst"] + fields["cess"], 2)
        return bucket, {**fields, "signed_total": signed_total}
    sign = gst_document_bucket_sign(bucket)
    if sign < 0:
        fields = {key: round(sign * value, 2) for key, value in fields.items()}
    fields["signed_total"] = round(
        fields["igst"] + fields["cgst"] + fields["sgst"] + fields["cess"], 2
    )
    return bucket, fields


def empty_document_type_bucket():
    return {
        "document_type": "",
        "count": 0,
        "taxable_value": 0.0,
        "igst": 0.0,
        "cgst": 0.0,
        "sgst": 0.0,
        "cess": 0.0,
        "signed_total": 0.0,
    }


def build_signed_document_type_summary(rows, source="Portal"):
    """
    Document Type Summary used before reconciliation.
    Net GST = Invoices + Debit Notes - Credit Notes ± Amendments.
    """
    order = [
        "Invoice", "Debit Note", "Credit Note",
        "Amendment Increase", "Amendment Decrease", "Amendment",
    ]
    buckets = {name: empty_document_type_bucket() for name in order}
    for name, bucket in buckets.items():
        bucket["document_type"] = name
    for row in rows or []:
        bucket_name, amounts = signed_row_tax_amounts(row)
        if bucket_name not in buckets:
            buckets[bucket_name] = empty_document_type_bucket()
            buckets[bucket_name]["document_type"] = bucket_name
        bucket = buckets[bucket_name]
        bucket["count"] += 1
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "signed_total"):
            bucket[key] = round(bucket[key] + amounts.get(key, 0), 2)
    rows_out = [buckets[name] for name in order if buckets[name]["count"] or name in {"Invoice", "Debit Note", "Credit Note"}]
    for name, bucket in buckets.items():
        if name not in order and bucket["count"]:
            rows_out.append(bucket)

    inv = buckets["Invoice"]["signed_total"]
    debit = buckets["Debit Note"]["signed_total"]
    credit = buckets["Credit Note"]["signed_total"]  # already negative when signed
    amd_up = buckets["Amendment Increase"]["signed_total"]
    amd_down = buckets["Amendment Decrease"]["signed_total"]  # negative when signed
    amd_other = buckets["Amendment"]["signed_total"]
    # credit/amd_down may already be negative; Net = sum of all signed totals.
    net_gst = round(sum(item["signed_total"] for item in rows_out), 2)
    return {
        "source": source,
        "rows": rows_out,
        "net": {
            "invoices": round(inv, 2),
            "debit_notes": round(debit, 2),
            "credit_notes": round(credit, 2),
            "amendment_increase": round(amd_up, 2),
            "amendment_decrease": round(amd_down, 2),
            "amendments": round(amd_up + amd_down + amd_other, 2),
            "net_gst": net_gst,
            "formula": "Invoices + Debit Notes - Credit Notes ± Amendments",
        },
    }


def compare_signed_document_summaries(portal_summary, tally_summary, tolerance=1.0):
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    portal_map = {row["document_type"]: row for row in (portal_summary or {}).get("rows") or []}
    tally_map = {row["document_type"]: row for row in (tally_summary or {}).get("rows") or []}
    types = []
    for name in (
        "Invoice", "Debit Note", "Credit Note",
        "Amendment Increase", "Amendment Decrease", "Amendment",
    ):
        if name in portal_map or name in tally_map:
            types.append(name)
    for name in sorted(set(portal_map) | set(tally_map)):
        if name not in types:
            types.append(name)
    by_type = []
    for name in types:
        left = portal_map.get(name) or empty_document_type_bucket()
        right = tally_map.get(name) or empty_document_type_bucket()
        diff = round(gst_number(right.get("signed_total")) - gst_number(left.get("signed_total")), 2)
        by_type.append({
            "document_type": name,
            "portal_count": left.get("count", 0),
            "tally_count": right.get("count", 0),
            "portal": left,
            "tally": right,
            "signed_total_difference": diff,
            "matched": abs(diff) <= tolerance,
        })
    portal_net = gst_number(((portal_summary or {}).get("net") or {}).get("net_gst"))
    tally_net = gst_number(((tally_summary or {}).get("net") or {}).get("net_gst"))
    net_diff = round(tally_net - portal_net, 2)
    return {
        "by_type": by_type,
        "net": {
            "portal_net_gst": portal_net,
            "tally_net_gst": tally_net,
            "difference": net_diff,
            "matched": abs(net_diff) <= tolerance,
            "formula": "Invoices + Debit Notes - Credit Notes ± Amendments",
        },
    }


def build_signed_reconciliation_pack(portal_rows, tally_rows, portal_label="Portal", tally_label="Tally", tolerance=1.0):
    portal_summary = build_signed_document_type_summary(portal_rows, portal_label)
    tally_summary = build_signed_document_type_summary(tally_rows, tally_label)
    compare = compare_signed_document_summaries(portal_summary, tally_summary, tolerance)
    return {
        "portal": portal_summary,
        "tally": tally_summary,
        "by_type": compare["by_type"],
        "net": compare["net"],
    }


def gst_invoice_key(row):
    return (
        gst_text(row.get("gstin")).upper(),
        re.sub(r"[^A-Z0-9]", "", gst_text(row.get("invoice_no")).upper()).lstrip("0") or "0",
        gst_document_bucket(row),
    )


def reconcile_gstr2b_tally(rows_2b, rows_tally, tolerance=1.0):
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    left, right = {}, {}
    for row in rows_2b:
        left.setdefault(gst_invoice_key(row), []).append(row)
    for row in rows_tally:
        right.setdefault(gst_invoice_key(row), []).append(row)
    results = []
    for key in sorted(set(left) | set(right)):
        left_rows, right_rows = list(left.get(key, [])), list(right.get(key, []))
        used_right = set()
        for row_2b in left_rows:
            candidates = [(index, row) for index, row in enumerate(right_rows) if index not in used_right]
            if candidates:
                def candidate_score(pair):
                    _, candidate = pair
                    date_penalty = 0 if gst_text(row_2b.get("invoice_date")).replace("/", "-") == gst_text(candidate.get("invoice_date")).replace("/", "-") else 1000000
                    amount_penalty = sum(
                        abs(gst_number(candidate.get(field)) - gst_number(row_2b.get(field)))
                        for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
                    )
                    return date_penalty + amount_penalty
                right_index, row_tally = min(candidates, key=candidate_score)
                used_right.add(right_index)
            else:
                row_tally = None
            if row_tally:
                differences = {
                    field: round(gst_number(row_tally.get(field)) - gst_number(row_2b.get(field)), 2)
                    for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
                }
                date_match = gst_text(row_2b.get("invoice_date")).replace("/", "-") == gst_text(row_tally.get("invoice_date")).replace("/", "-")
                value_match = all(abs(value) <= tolerance for value in differences.values())
                status = "Matched" if date_match and value_match else ("Date Mismatch" if value_match else "Amount/Tax Mismatch")
                base = row_2b
            else:
                differences, status, base = {}, "Only in GSTR-2B", row_2b
            results.append({
                **base,
                "status": status,
                "gstr2b": row_2b,
                "tally": row_tally,
                "differences": differences,
                "document_bucket": gst_document_bucket(base),
                "party_ledger": gst_text((row_tally or {}).get("party_ledger") or base.get("party_name")),
            })
        for index, row_tally in enumerate(right_rows):
            if index in used_right:
                continue
            results.append({
                **row_tally,
                "status": "Only in Tally",
                "gstr2b": None,
                "tally": row_tally,
                "differences": {},
                "document_bucket": gst_document_bucket(row_tally),
                "party_ledger": gst_text(row_tally.get("party_ledger") or row_tally.get("party_name")),
            })
    counts = {}
    for row in results:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    document_summary = build_signed_reconciliation_pack(
        rows_2b, rows_tally, "GSTR-2B", "Tally Purchase", tolerance
    )
    return results, counts, document_summary


def is_gst_all_periods_selection(value):
    """True for ALL / FY aggregate selectors (not a single MMYYYY month)."""
    text = gst_text(value).strip().upper()
    if not text:
        return False
    compact = re.sub(r"[\s_\-/]", "", text)
    return compact in {
        "ALL", "ALLPERIODS", "FY202526", "FY2526", "202526",
        "ALLFY202526", "ALLFY2526", "ALLPERIODSFY202526",
    } or compact.startswith("ALLFY") or text.startswith("ALL /")


def gst_fy_2025_26_periods():
    """Apr-25 … Mar-26 return periods for FY 2025-26."""
    return [
        ("042025", "Apr-25"),
        ("052025", "May-25"),
        ("062025", "Jun-25"),
        ("072025", "Jul-25"),
        ("082025", "Aug-25"),
        ("092025", "Sep-25"),
        ("102025", "Oct-25"),
        ("112025", "Nov-25"),
        ("122025", "Dec-25"),
        ("012026", "Jan-26"),
        ("022026", "Feb-26"),
        ("032026", "Mar-26"),
    ]


def normalize_gst_recon_period(value):
    # ALL / FY selectors mean "no month filter" — same as blank.
    if is_gst_all_periods_selection(value):
        return ""
    text = gst_text(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) == 6:
        return digits
    if len(digits) == 4:
        return digits
    if len(digits) == 8:
        # Invoice dates often arrive as DD-MM-YYYY (Portal/Tally) without source_period.
        dd, mm, yyyy = digits[0:2], digits[2:4], digits[4:8]
        try:
            if 1 <= int(mm) <= 12 and 1 <= int(dd) <= 31 and int(yyyy) >= 2017:
                return f"{mm}{yyyy}"
        except ValueError:
            pass
        yyyy, mm, dd = digits[0:4], digits[4:6], digits[6:8]
        try:
            if 1 <= int(mm) <= 12 and 1 <= int(dd) <= 31 and int(yyyy) >= 2017:
                return f"{mm}{yyyy}"
        except ValueError:
            pass
    match = re.fullmatch(r"(\d{2})[-/](\d{4})", text)
    if match:
        return f"{match.group(1)}{match.group(2)}"
    return digits


def build_gstr1_period_breakdown(rows, financial_year="2025-26"):
    """
    Month-wise GSTR-1 totals using the same per-row tax fields as monthly view.
    Does not change GSTR-1 parsing or tax math — only groups existing rows.
    """
    fy = gst_text(financial_year) or "2025-26"
    period_defs = gst_fy_2025_26_periods() if fy in {"2025-26", "2526", "FY2025-26"} else gst_fy_2025_26_periods()
    by_period = {}
    for row in rows or []:
        period = normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date"))
        if not period:
            continue
        bucket = by_period.setdefault(
            period,
            {"period": period, "count": 0, "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0, "output_gst": 0.0},
        )
        item = ensure_gst_invoice_fields(row)
        bucket["count"] += 1
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess"):
            bucket[key] = round(bucket[key] + gst_number(item.get(key)), 2)
        bucket["output_gst"] = round(
            bucket["igst"] + bucket["cgst"] + bucket["sgst"] + bucket["cess"], 2
        )
    months = []
    totals = {
        "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
        "output_gst": 0.0, "count": 0, "periods_present": 0,
    }
    for period, label in period_defs:
        bucket = by_period.get(period) or {
            "period": period, "count": 0, "taxable_value": 0.0, "igst": 0.0,
            "cgst": 0.0, "sgst": 0.0, "cess": 0.0, "output_gst": 0.0,
        }
        row = {
            **bucket,
            "period_label": label,
            "present": bucket["count"] > 0,
        }
        months.append(row)
        if row["present"]:
            totals["periods_present"] += 1
        totals["count"] += row["count"]
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "output_gst"):
            totals[key] = round(totals[key] + row[key], 2)
    return {
        "financial_year": "2025-26",
        "label": "ALL / FY 2025-26",
        "months": months,
        "totals": totals,
        "formula": "Sum of monthly GSTR-1 Output GST (IGST+CGST+SGST+CESS) for Apr-25…Mar-26",
    }


def gstr2b_document_sign(row):
    """
    Portal GSTR-2B net ITC sign before import-time negation.
    Credit Notes / CDN decreases reduce ITC. Debit Notes add.
    If amounts are already portal-signed (negative Credit Notes), return +1
    so callers never double-flip.
    """
    if row.get("_portal_signed") or gst_number(row.get("document_sign")) == -1:
        return 1.0
    # Already-negated tax lines (import / prior signing) — do not flip again.
    if any(gst_number(row.get(key)) < 0 for key in ("igst", "cgst", "sgst", "cess", "taxable_value")):
        return 1.0
    doc = gst_text(row.get("document_type")).lower()
    note = gst_text(row.get("note_type")).lower()
    section = gst_text(row.get("section")).lower()
    direction = gst_text(row.get("amendment_direction")).lower()
    if "debit note" in doc or note.startswith("debit") or note == "debit note":
        return 1.0
    if "credit note" in doc or note.startswith("credit") or note == "credit note":
        return -1.0
    # Ambiguous "Credit/Debit Note" on CDNR/CDNUR sheets → Credit unless Debit labeled.
    if any(token in section for token in ("cdnr", "cdnur")):
        if "debit" in note or "debit" in doc:
            return 1.0
        return -1.0
    if direction == "decrease" or ("amendment" in doc and "decrease" in doc):
        return -1.0
    return 1.0


def apply_gstr2b_portal_signs(rows):
    """
    Import-time netting: store Credit Note / decrease amounts as negative so
    every downstream sum matches the GST Portal Comparison Report.
    Also drops exact duplicate invoice keys within the same import batch.
    """
    signed_rows = []
    seen = set()
    for row in rows or []:
        item = ensure_gst_invoice_fields(dict(row))
        if not item.get("_portal_signed"):
            sign = gstr2b_document_sign(item)
            if sign < 0:
                for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "invoice_value"):
                    item[key] = round(sign * gst_number(item.get(key)), 2)
                for sub in item.get("items") or []:
                    for key in ("taxable_value", "igst", "cgst", "sgst", "cess"):
                        if key in sub:
                            sub[key] = round(sign * gst_number(sub.get(key)), 2)
                doc = gst_text(item.get("document_type"))
                if "credit/debit" in doc.lower():
                    item["document_type"] = (
                        "Credit Note Amendment" if "amend" in doc.lower() else "Credit Note"
                    )
                item["document_sign"] = -1
            else:
                item["document_sign"] = 1
            item["_portal_signed"] = True
        key = (
            gst_text(item.get("gstin")).upper(),
            normalize_invoice_number(item.get("invoice_no")),
            gst_text(item.get("invoice_date")),
            gst_text(item.get("document_type")),
            normalize_gst_recon_period(item.get("source_period") or item.get("gstr2b_period")),
        )
        if key in seen:
            continue
        seen.add(key)
        signed_rows.append(item)
    return signed_rows


def gst_recon_period_totals(rows, signed=False):
    totals = {}
    for row in rows or []:
        period = normalize_gst_recon_period(
            row.get("source_period") or row.get("gstr2b_period") or row.get("invoice_date")
        )
        if not period:
            continue
        sign = gstr2b_document_sign(row) if signed else 1.0
        bucket = totals.setdefault(
            period,
            {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0, "invoices": 0},
        )
        bucket["invoices"] += 1
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess"):
            bucket[key] = round(bucket[key] + sign * gst_number(row.get(key)), 2)
    return totals


def gstr2b_tax_totals(rows):
    """Net GSTR-2B ITC totals with Credit Notes subtracted (portal-aligned)."""
    totals = {
        "invoices": 0,
        "invoice_count": 0,
        "credit_note_count": 0,
        "debit_note_count": 0,
        "amendment_count": 0,
        "taxable_value": 0.0,
        "igst": 0.0,
        "cgst": 0.0,
        "sgst": 0.0,
        "cess": 0.0,
        "invoice_value": 0.0,
        "output_gst": 0.0,
        "gross_invoice_itc": 0.0,
        "credit_note_itc": 0.0,
        "debit_note_itc": 0.0,
        "amendment_itc": 0.0,
        "net_itc": 0.0,
    }
    for row in rows or []:
        item = ensure_gst_invoice_fields(row)
        bucket, amounts = signed_row_tax_amounts(item)
        itc = round(amounts["signed_total"], 2)
        totals["invoices"] += 1
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "invoice_value"):
            totals[key] = round(totals[key] + amounts.get(key, 0), 2)
        if bucket == "Invoice":
            totals["invoice_count"] += 1
            totals["gross_invoice_itc"] = round(totals["gross_invoice_itc"] + abs(itc), 2)
        elif bucket == "Credit Note":
            totals["credit_note_count"] += 1
            totals["credit_note_itc"] = round(totals["credit_note_itc"] + abs(itc), 2)
        elif bucket == "Debit Note":
            totals["debit_note_count"] += 1
            totals["debit_note_itc"] = round(totals["debit_note_itc"] + abs(itc), 2)
        else:
            totals["amendment_count"] += 1
            totals["amendment_itc"] = round(totals["amendment_itc"] + itc, 2)
    totals["output_gst"] = round(
        totals["igst"] + totals["cgst"] + totals["sgst"] + totals["cess"], 2
    )
    totals["net_itc"] = totals["output_gst"]
    return totals


def build_gstr2b_gross_net_summary(rows):
    """Gross components + Net GSTR-2B ITC for dashboards and exports."""
    totals = gstr2b_tax_totals(rows)
    return {
        "invoice_count": totals["invoice_count"],
        "credit_note_count": totals["credit_note_count"],
        "debit_note_count": totals["debit_note_count"],
        "amendment_count": totals["amendment_count"],
        "document_count": totals["invoices"],
        "gross_invoice_itc": totals["gross_invoice_itc"],
        "credit_note_itc": totals["credit_note_itc"],
        "debit_note_itc": totals["debit_note_itc"],
        "amendment_itc": totals["amendment_itc"],
        "net_itc": totals["net_itc"],
        "taxable_value": totals["taxable_value"],
        "igst": totals["igst"],
        "cgst": totals["cgst"],
        "sgst": totals["sgst"],
        "cess": totals["cess"],
        "formula": "Gross Invoice ITC − Credit Note ITC + Debit Note ITC ± Amendments",
    }


def normalize_tally_purchase_itc_rows(rows):
    """
    Ensure TALLY_PURCHASE rows use Section 4C signs.

    Legacy sync stored abs() tax and left purchase returns as voucher type
    "Debit Note" with positive amounts (inflating Tally Booked by 2× CN tax).
    New sync sets itc_source and signed Input GST — left unchanged here.
    """
    normalized = []
    for row in rows or []:
        item = dict(ensure_gst_invoice_fields(row))
        voucher = gst_text(item.get("voucher_type")).lower()
        doc = gst_text(item.get("document_type")).lower()
        tax = round(
            gst_number(item.get("igst")) + gst_number(item.get("cgst")) + gst_number(item.get("sgst")),
            2,
        )
        legacy_misnamed_cn = (
            not item.get("itc_source")
            and tax > 0.005
            and ("debit note" in voucher or "debit note" in doc)
        )
        if legacy_misnamed_cn:
            for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "invoice_value"):
                item[key] = round(-abs(gst_number(item.get(key))), 2)
            item["document_type"] = "Credit Note"
            item["document_sign"] = -1
            item["_portal_signed"] = True
        normalized.append(item)
    return normalized


def ensure_tally_purchase_rows_for_itc(return_period=""):
    """
    Keep TALLY_PURCHASE aligned with Tally GSTR-3B Section 4C (A−B).

    Re-syncs when empty/stale vs tally_sync count, or when gst_uncertain flags
    are missing (required to match Tally GSTR-3B uncertain exclusions).
    """
    period = normalize_gst_recon_period(return_period)
    rows = gst_recon_load_rows("TALLY_PURCHASE") or []
    if period:
        scoped = [
            row for row in rows
            if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period
        ]
        if scoped:
            return rows, False
    sync_meta = gst_recon_get_meta("tally_sync", {}) or {}
    meta_count = int(gst_number(sync_meta.get("count")))
    current = len(rows)
    flags_missing = bool(rows) and not any("gst_uncertain" in (row or {}) for row in rows)
    needs_refresh = current == 0 or (meta_count > 0 and current < meta_count) or flags_missing
    if not needs_refresh:
        return rows, False
    try:
        synced = sync_tally_purchase_vouchers()
        new_rows = synced.get("rows") or []
        if not new_rows:
            return rows, False
        gst_recon_save_rows("TALLY_PURCHASE", new_rows)
        booked = build_tally_booked_itc_summary(normalize_tally_purchase_itc_rows(new_rows))
        gst_recon_set_meta("tally_sync", {
            "company": synced.get("company", ""),
            "count": synced.get("count", len(new_rows)),
            "synced_at": synced.get("synced_at") or gst_recon_now(),
            "ok": True,
            "net_itc": booked.get("net_itc"),
            "igst": booked.get("igst"),
            "cgst": booked.get("cgst"),
            "sgst": booked.get("sgst"),
            "cess": booked.get("cess"),
            "section_4c_source": synced.get("section_4c_source"),
            "uncertain_excluded": synced.get("uncertain_excluded"),
            "formula": booked.get("formula"),
            "function": booked.get("function"),
        })
        return new_rows, True
    except Exception:
        return rows, False


def build_tally_booked_itc_summary(rows, gstr3b_eligible_only=True):
    """
    Tally Booked = Net Input Tax Credit Available on Tally GSTR-3B (Section 4C A−B).

    By default only vouchers Tally includes in GSTR-3B are used
    (VCHGSTSTATUSISUNCERTAIN ≠ Yes). Tax amounts still come from signed Input
    Duties & Taxes ledger lines on those vouchers — the same pool the report uses.

    Purchase ITC
    − Credit Note adjustment
    − ITC Reversal
    + Debit Note
    = Final Tally Booked (IGST + CGST + SGST; CESS shown separately)
    """
    source_rows = list(rows or [])
    excluded_uncertain = 0
    if gstr3b_eligible_only:
        filtered = filter_tally_rows_for_gstr3b_4c(source_rows)
        excluded_uncertain = len(source_rows) - len(filtered)
        source_rows = filtered
    buckets = {
        "Purchase": {"count": 0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0},
        "Credit Note": {"count": 0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0},
        "Debit Note": {"count": 0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0},
        "ITC Reversal": {"count": 0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0},
    }
    signed_igst = signed_cgst = signed_sgst = signed_cess = 0.0
    for row in source_rows:
        item = ensure_gst_invoice_fields(row)
        igst = gst_number(item.get("igst"))
        cgst = gst_number(item.get("cgst"))
        sgst = gst_number(item.get("sgst"))
        cess = gst_number(item.get("cess"))
        tax_total = round(igst + cgst + sgst, 2)
        doc = gst_text(item.get("document_type")).lower()
        voucher = gst_text(item.get("voucher_type")).lower()
        if "reversal" in doc or "reversal" in voucher:
            key = "ITC Reversal"
        elif tax_total < -0.005 or "credit note" in doc or "purchase return" in voucher or "credit note" in voucher:
            # Negative signed tax always reduces Section 4C (Credit Note / return).
            key = "Credit Note"
        elif "debit note" in doc or "debit note" in voucher:
            # New sync stores genuine ITC-increasing DN as positive with itc_source.
            # Legacy abs()-synced "Debit Note" rows were purchase returns → Credit Note.
            if tax_total > 0.005 and item.get("itc_source"):
                key = "Debit Note"
            else:
                key = "Credit Note"
        else:
            key = "Purchase"
        bucket = buckets[key]
        bucket["count"] += 1
        for tax_name, tax_val in (("igst", igst), ("cgst", cgst), ("sgst", sgst), ("cess", cess)):
            bucket[tax_name] = round(bucket[tax_name] + abs(tax_val), 2)
        # Authoritative Section 4C components = sum of signed stored tax.
        if key in {"Credit Note", "ITC Reversal"} and tax_total > 0.005:
            # Legacy abs-stored decrease lines: subtract.
            signed_igst = round(signed_igst - abs(igst), 2)
            signed_cgst = round(signed_cgst - abs(cgst), 2)
            signed_sgst = round(signed_sgst - abs(sgst), 2)
            signed_cess = round(signed_cess - abs(cess), 2)
        else:
            signed_igst = round(signed_igst + igst, 2)
            signed_cgst = round(signed_cgst + cgst, 2)
            signed_sgst = round(signed_sgst + sgst, 2)
            signed_cess = round(signed_cess + cess, 2)

    purchase_itc = round(sum(buckets["Purchase"][k] for k in ("igst", "cgst", "sgst")), 2)
    credit_itc = round(sum(buckets["Credit Note"][k] for k in ("igst", "cgst", "sgst")), 2)
    debit_itc = round(sum(buckets["Debit Note"][k] for k in ("igst", "cgst", "sgst")), 2)
    reversal_itc = round(sum(buckets["ITC Reversal"][k] for k in ("igst", "cgst", "sgst")), 2)
    # Prefer signed ledger sum; fall back to formula on abs buckets (must match).
    formula_igst = round(
        buckets["Purchase"]["igst"] + buckets["Debit Note"]["igst"]
        - buckets["Credit Note"]["igst"] - buckets["ITC Reversal"]["igst"], 2
    )
    formula_cgst = round(
        buckets["Purchase"]["cgst"] + buckets["Debit Note"]["cgst"]
        - buckets["Credit Note"]["cgst"] - buckets["ITC Reversal"]["cgst"], 2
    )
    formula_sgst = round(
        buckets["Purchase"]["sgst"] + buckets["Debit Note"]["sgst"]
        - buckets["Credit Note"]["sgst"] - buckets["ITC Reversal"]["sgst"], 2
    )
    formula_cess = round(
        buckets["Purchase"]["cess"] + buckets["Debit Note"]["cess"]
        - buckets["Credit Note"]["cess"] - buckets["ITC Reversal"]["cess"], 2
    )
    net_igst = signed_igst if abs(signed_igst - formula_igst) < 0.02 else formula_igst
    net_cgst = signed_cgst if abs(signed_cgst - formula_cgst) < 0.02 else formula_cgst
    net_sgst = signed_sgst if abs(signed_sgst - formula_sgst) < 0.02 else formula_sgst
    net_cess = signed_cess if abs(signed_cess - formula_cess) < 0.02 else formula_cess
    # When signed and formula disagree slightly after mixed legacy rows, use formula
    # built from corrected buckets (legacy Debit→Credit remap).
    if abs((signed_igst + signed_cgst + signed_sgst) - (formula_igst + formula_cgst + formula_sgst)) >= 0.02:
        net_igst, net_cgst, net_sgst, net_cess = formula_igst, formula_cgst, formula_sgst, formula_cess
    # Section 4C on Tally GSTR-3B screen = IGST+CGST+SGST (cess separate).
    net_itc = round(net_igst + net_cgst + net_sgst, 2)
    return {
        "source": "Tally GSTR-3B Net Input Tax Credit Available (4C A−B)",
        "xml_tags": "VCHGSTSTATUSISUNCERTAIN + ALLLEDGERENTRIES.LIST / LEDGERNAME / AMOUNT",
        "ledgers": "Duties & Taxes Input (IGST/CGST/SGST); uncertain vouchers excluded",
        "function": "build_tally_booked_itc_summary ← sync_tally_purchase_vouchers (VCHGSTSTATUSISUNCERTAIN)",
        "formula": "Purchase ITC − Credit Note − ITC Reversal + Debit Note (GSTR-3B eligible only)",
        "gstr3b_eligible_only": bool(gstr3b_eligible_only),
        "uncertain_excluded": excluded_uncertain,
        "purchase_itc": purchase_itc,
        "credit_note_itc": credit_itc,
        "debit_note_itc": debit_itc,
        "reversal_itc": reversal_itc,
        "net_itc": net_itc,
        "igst": net_igst,
        "cgst": net_cgst,
        "sgst": net_sgst,
        "cess": net_cess,
        "net_itc_with_cess": round(net_itc + net_cess, 2),
        "buckets": buckets,
        "lines": [
            {"particulars": "Purchase ITC", "sign": "+", "amount": purchase_itc, "count": buckets["Purchase"]["count"]},
            {"particulars": "Less Credit Note adjustment", "sign": "−", "amount": credit_itc, "count": buckets["Credit Note"]["count"]},
            {"particulars": "Less ITC Reversal", "sign": "−", "amount": reversal_itc, "count": buckets["ITC Reversal"]["count"]},
            {"particulars": "Plus Debit Note", "sign": "+", "amount": debit_itc, "count": buckets["Debit Note"]["count"]},
            {"particulars": "Final Tally Booked (Section 4C A−B)", "sign": "=", "amount": net_itc, "count": 0},
        ],
    }


def normalize_invoice_date_key(value):
    text = gst_text(value).replace("/", "-").strip()
    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        # DDMMYYYY or YYYYMMDD
        if int(digits[4:8]) >= 2017:
            return f"{digits[0:2]}-{digits[2:4]}-{digits[4:8]}"
        return f"{digits[6:8]}-{digits[4:6]}-{digits[0:4]}"
    return text


def normalize_party_key(value):
    return re.sub(r"[^a-z0-9]", "", gst_text(value).lower())


def portal_row_itc(row):
    """Net GSTR-2B ITC contribution (IGST+CGST+SGST+CESS, signed)."""
    _, amounts = signed_row_tax_amounts(row)
    return round(
        amounts["igst"] + amounts["cgst"] + amounts["sgst"] + amounts["cess"],
        2,
    )


def tally_row_section_4c_itc(row):
    """Tally Booked (Section 4C) contribution (IGST+CGST+SGST only, signed)."""
    item = ensure_gst_invoice_fields(row or {})
    return round(
        gst_number(item.get("igst")) + gst_number(item.get("cgst")) + gst_number(item.get("sgst")),
        2,
    )


def _itc_diff_field_amounts(row):
    _, amounts = signed_row_tax_amounts(row)
    return {
        "taxable_value": round(amounts.get("taxable_value", 0), 2),
        "igst": round(amounts.get("igst", 0), 2),
        "cgst": round(amounts.get("cgst", 0), 2),
        "sgst": round(amounts.get("sgst", 0), 2),
        "cess": round(amounts.get("cess", 0), 2),
    }


def _itc_diff_match_score(portal_row, tally_row, tolerance):
    p = _itc_diff_field_amounts(portal_row)
    # Tally rows are already Section-4C signed; compare tax components including stored cess.
    t = {
        "taxable_value": gst_number(tally_row.get("taxable_value")),
        "igst": gst_number(tally_row.get("igst")),
        "cgst": gst_number(tally_row.get("cgst")),
        "sgst": gst_number(tally_row.get("sgst")),
        "cess": gst_number(tally_row.get("cess")),
    }
    date_penalty = 0 if normalize_invoice_date_key(portal_row.get("invoice_date")) == normalize_invoice_date_key(
        tally_row.get("invoice_date")
    ) else 1_000_000
    amount_penalty = sum(abs(t[field] - p[field]) for field in ("taxable_value", "igst", "cgst", "sgst", "cess"))
    inv_p = re.sub(r"[^A-Z0-9]", "", gst_text(portal_row.get("invoice_no")).upper()).lstrip("0") or "0"
    inv_t = re.sub(r"[^A-Z0-9]", "", gst_text(tally_row.get("invoice_no")).upper()).lstrip("0") or "0"
    inv_penalty = 0 if inv_p == inv_t else 500_000
    gstin_penalty = 0 if gst_text(portal_row.get("gstin")).upper() == gst_text(tally_row.get("gstin")).upper() else 250_000
    return date_penalty + inv_penalty + gstin_penalty + amount_penalty


def _classify_itc_difference_pair(portal_row, tally_row, tolerance):
    """Return (status, reason) for a paired portal/tally voucher."""
    p_gstin = gst_text((portal_row or {}).get("gstin")).upper()
    t_gstin = gst_text((tally_row or {}).get("gstin")).upper()
    p_inv = re.sub(r"[^A-Z0-9]", "", gst_text((portal_row or {}).get("invoice_no")).upper()).lstrip("0") or "0"
    t_inv = re.sub(r"[^A-Z0-9]", "", gst_text((tally_row or {}).get("invoice_no")).upper()).lstrip("0") or "0"
    p_date = normalize_invoice_date_key((portal_row or {}).get("invoice_date"))
    t_date = normalize_invoice_date_key((tally_row or {}).get("invoice_date"))
    p_party = normalize_party_key((portal_row or {}).get("party_name") or (portal_row or {}).get("party_ledger"))
    t_party = normalize_party_key((tally_row or {}).get("party_name") or (tally_row or {}).get("party_ledger"))
    p_amt = _itc_diff_field_amounts(portal_row) if portal_row else None
    t_amt = {
        "taxable_value": gst_number(tally_row.get("taxable_value")),
        "igst": gst_number(tally_row.get("igst")),
        "cgst": gst_number(tally_row.get("cgst")),
        "sgst": gst_number(tally_row.get("sgst")),
        "cess": gst_number(tally_row.get("cess")),
    } if tally_row else None

    if portal_row and not tally_row:
        return "Missing in Tally", "Present in GSTR-2B but no matching Tally purchase voucher"
    if tally_row and not portal_row:
        return "Missing in GSTR-2B", "Present in Tally but no matching GSTR-2B document"

    tax_fields = ("igst", "cgst", "sgst", "cess")
    tax_diff = any(abs(t_amt[f] - p_amt[f]) > tolerance for f in tax_fields)
    taxable_diff = abs(t_amt["taxable_value"] - p_amt["taxable_value"]) > tolerance
    date_diff = p_date != t_date
    inv_diff = p_inv != t_inv
    gstin_diff = p_gstin != t_gstin
    party_diff = bool(p_party and t_party and p_party != t_party)

    portal_itc = portal_row_itc(portal_row)
    tally_itc = tally_row_section_4c_itc(tally_row)
    contribution = round(portal_itc - tally_itc, 2)

    # Priority: identity mismatches that explain pairing quality first.
    if gstin_diff:
        return "GSTIN Mismatch", f"GSTIN differs (Portal {p_gstin or '—'} vs Tally {t_gstin or '—'})"
    if inv_diff:
        return "Invoice Number Difference", (
            f"Invoice number differs (Portal {gst_text(portal_row.get('invoice_no')) or '—'} "
            f"vs Tally {gst_text(tally_row.get('invoice_no')) or '—'})"
        )
    if date_diff:
        return "Invoice Date Difference", (
            f"Invoice date differs (Portal {gst_text(portal_row.get('invoice_date')) or '—'} "
            f"vs Tally {gst_text(tally_row.get('invoice_date')) or '—'})"
        )
    if taxable_diff and not tax_diff:
        return "Taxable Difference", (
            f"Taxable value differs by ₹{abs(t_amt['taxable_value'] - p_amt['taxable_value']):.2f}"
        )
    if tax_diff:
        parts = []
        for f in tax_fields:
            delta = round(p_amt[f] - t_amt[f], 2)
            if abs(delta) > tolerance:
                parts.append(f"{f.upper()} ₹{delta:.2f}")
        # Section 4C excludes CESS from Tally Booked even when cess lines match.
        cess_only_4c = (
            abs(contribution) > tolerance
            and abs(p_amt["igst"] - t_amt["igst"]) <= tolerance
            and abs(p_amt["cgst"] - t_amt["cgst"]) <= tolerance
            and abs(p_amt["sgst"] - t_amt["sgst"]) <= tolerance
            and abs(p_amt["cess"]) > tolerance
        )
        if cess_only_4c and abs(p_amt["cess"] - t_amt["cess"]) <= tolerance:
            return (
                "GST Difference",
                f"CESS ₹{abs(p_amt['cess']):.2f} is included in Net GSTR-2B ITC but excluded from Tally Booked (Section 4C)",
            )
        return "GST Difference", "Tax differs: " + (", ".join(parts) if parts else "IGST/CGST/SGST/CESS")
    if party_diff:
        return "Party Mismatch", "Party name differs between GSTR-2B and Tally"
    # Perfect field match — still explain Section 4C CESS exclusion if contribution remains.
    if abs(contribution) > 0.02 and abs(p_amt["cess"]) > 0.02:
        return (
            "GST Difference",
            f"CESS ₹{abs(p_amt['cess']):.2f} is included in Net GSTR-2B ITC but excluded from Tally Booked (Section 4C)",
        )
    if abs(contribution) > 0.02:
        return "GST Difference", f"Portal ITC − Tally Booked (4C) = ₹{contribution:.2f}"
    return "Matched", "All key fields match within tolerance"


def build_itc_available_difference_recon(rows_2b, rows_tally, tolerance=1.0):
    """
    Voucher-level explanation of Available ITC Difference:
    Net GSTR-2B ITC − Tally Booked (Section 4C).

    Does not alter dashboard calculations — read-only explainer.
    Sum of row Difference must equal dashboard Available ITC (when 3B claimed ≤ Tally).
    """
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    portal_rows = [ensure_gst_invoice_fields(row) for row in (rows_2b or [])]
    tally_rows = normalize_tally_purchase_itc_rows(rows_tally)

    # Primary index: GSTIN + invoice no + document bucket
    left, right = {}, {}
    for row in portal_rows:
        left.setdefault(gst_invoice_key(row), []).append(row)
    for row in tally_rows:
        right.setdefault(gst_invoice_key(row), []).append(row)

    pairs = []  # (portal, tally)
    used_portal = set()
    used_tally = set()

    def portal_id(row):
        return id(row)

    def tally_id(row):
        return id(row)

    # Pass 1 — same GSTIN + invoice + bucket
    for key in sorted(set(left) | set(right)):
        left_rows = list(left.get(key, []))
        right_rows = list(right.get(key, []))
        for portal in left_rows:
            candidates = [
                row for row in right_rows
                if tally_id(row) not in used_tally
            ]
            if not candidates:
                continue
            best = min(candidates, key=lambda row: _itc_diff_match_score(portal, row, tolerance))
            pairs.append((portal, best))
            used_portal.add(portal_id(portal))
            used_tally.add(tally_id(best))

    unmatched_portal = [row for row in portal_rows if portal_id(row) not in used_portal]
    unmatched_tally = [row for row in tally_rows if tally_id(row) not in used_tally]

    # Pass 2 — GSTIN + date + tax proximity (invoice number difference)
    still_portal = []
    for portal in unmatched_portal:
        candidates = []
        p_gstin = gst_text(portal.get("gstin")).upper()
        p_date = normalize_invoice_date_key(portal.get("invoice_date"))
        for tally in unmatched_tally:
            if tally_id(tally) in used_tally:
                continue
            if gst_text(tally.get("gstin")).upper() != p_gstin:
                continue
            if normalize_invoice_date_key(tally.get("invoice_date")) != p_date:
                continue
            candidates.append(tally)
        if candidates:
            best = min(candidates, key=lambda row: _itc_diff_match_score(portal, row, tolerance))
            pairs.append((portal, best))
            used_portal.add(portal_id(portal))
            used_tally.add(tally_id(best))
        else:
            still_portal.append(portal)
    unmatched_portal = still_portal
    unmatched_tally = [row for row in unmatched_tally if tally_id(row) not in used_tally]

    # Pass 3 — invoice no + date (GSTIN mismatch)
    still_portal = []
    for portal in unmatched_portal:
        p_inv = re.sub(r"[^A-Z0-9]", "", gst_text(portal.get("invoice_no")).upper()).lstrip("0") or "0"
        p_date = normalize_invoice_date_key(portal.get("invoice_date"))
        candidates = []
        for tally in unmatched_tally:
            if tally_id(tally) in used_tally:
                continue
            t_inv = re.sub(r"[^A-Z0-9]", "", gst_text(tally.get("invoice_no")).upper()).lstrip("0") or "0"
            if t_inv != p_inv:
                continue
            if normalize_invoice_date_key(tally.get("invoice_date")) != p_date:
                continue
            candidates.append(tally)
        if candidates:
            best = min(candidates, key=lambda row: _itc_diff_match_score(portal, row, tolerance))
            pairs.append((portal, best))
            used_portal.add(portal_id(portal))
            used_tally.add(tally_id(best))
        else:
            still_portal.append(portal)
    unmatched_portal = still_portal
    unmatched_tally = [row for row in unmatched_tally if tally_id(row) not in used_tally]

    for portal in unmatched_portal:
        pairs.append((portal, None))
    for tally in unmatched_tally:
        pairs.append((None, tally))

    rows = []
    counts = {}
    voucher_diff_total = 0.0
    for portal, tally in pairs:
        status, reason = _classify_itc_difference_pair(portal, tally, tolerance)
        portal_taxable = gst_number((_itc_diff_field_amounts(portal) if portal else {}).get("taxable_value"))
        tally_taxable = gst_number((tally or {}).get("taxable_value"))
        portal_itc = portal_row_itc(portal) if portal else 0.0
        tally_itc = tally_row_section_4c_itc(tally) if tally else 0.0
        difference = round(portal_itc - tally_itc, 2)
        voucher_diff_total = round(voucher_diff_total + difference, 2)
        base = portal or tally or {}
        gstin = gst_text((portal or {}).get("gstin") or (tally or {}).get("gstin")).upper()
        party = gst_text(
            (portal or {}).get("party_name")
            or (portal or {}).get("party_ledger")
            or (tally or {}).get("party_name")
            or (tally or {}).get("party_ledger")
        )
        invoice_no = gst_text((portal or {}).get("invoice_no") or (tally or {}).get("invoice_no"))
        invoice_date = gst_text((portal or {}).get("invoice_date") or (tally or {}).get("invoice_date"))
        row = {
            "status": status,
            "reason": reason,
            "gstin": gstin,
            "party_name": party,
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "portal_taxable": portal_taxable if portal else None,
            "tally_taxable": tally_taxable if tally else None,
            "portal_itc": portal_itc if portal else None,
            "tally_itc": tally_itc if tally else None,
            "difference": difference,
            "document_bucket": gst_document_bucket(base),
        }
        rows.append(row)
        counts[status] = counts.get(status, 0) + 1

    # Stable sort: non-zero differences first, then status, then invoice.
    status_order = {
        "Missing in Tally": 0,
        "Missing in GSTR-2B": 1,
        "GST Difference": 2,
        "Taxable Difference": 3,
        "Invoice Date Difference": 4,
        "Invoice Number Difference": 5,
        "Party Mismatch": 6,
        "GSTIN Mismatch": 7,
        "Matched": 8,
    }
    rows.sort(
        key=lambda r: (
            0 if abs(gst_number(r.get("difference"))) > 0.005 else 1,
            status_order.get(r.get("status"), 99),
            gst_text(r.get("invoice_date")),
            gst_text(r.get("invoice_no")),
        )
    )

    net_pack = build_gstr2b_gross_net_summary(portal_rows)
    tally_booked = build_tally_booked_itc_summary(tally_rows)
    portal_net = round(gst_number(net_pack.get("net_itc")), 2)
    tally_net = round(gst_number(tally_booked.get("net_itc")), 2)
    available_diff = round(portal_net - tally_net, 2)

    return {
        "rows": rows,
        "counts": counts,
        "summary": {
            "portal_net_itc": portal_net,
            "tally_booked_itc": tally_net,
            "available_itc_difference": available_diff,
            "voucher_difference_total": voucher_diff_total,
            "balanced": abs(voucher_diff_total - available_diff) <= max(tolerance, 0.02),
            "formula": "Net GSTR-2B ITC − Tally Booked (Section 4C)",
            "row_formula": "Difference = Portal ITC (IGST+CGST+SGST+CESS) − Tally Booked ITC (IGST+CGST+SGST)",
            "document_count": len(rows),
            "nonzero_count": sum(1 for r in rows if abs(gst_number(r.get("difference"))) > 0.005),
        },
        "tolerance": tolerance,
    }


def make_itc_difference_export(recon, title="Available ITC Difference"):
    rows = (recon or {}).get("rows") or []
    summary = (recon or {}).get("summary") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Voucher Differences"
    headers = [
        "Status", "Reason", "GSTIN", "Party Name", "Invoice Number", "Invoice Date",
        "Portal Taxable", "Tally Taxable", "Portal ITC", "Tally ITC", "Difference",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.get("status"),
            row.get("reason"),
            row.get("gstin"),
            row.get("party_name"),
            row.get("invoice_no"),
            row.get("invoice_date"),
            row.get("portal_taxable"),
            row.get("tally_taxable"),
            row.get("portal_itc"),
            row.get("tally_itc"),
            row.get("difference"),
        ])
    fill = PatternFill("solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    for col in ("G", "H", "I", "J", "K"):
        for cell in ws[col][1:]:
            cell.number_format = "#,##0.00"
    widths = [22, 56, 18, 32, 18, 14, 14, 14, 14, 14, 14]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for label, key in (
        ("Net GSTR-2B ITC", "portal_net_itc"),
        ("Tally Booked (Section 4C)", "tally_booked_itc"),
        ("Available ITC Difference", "available_itc_difference"),
        ("Sum of voucher differences", "voucher_difference_total"),
        ("Balanced", "balanced"),
        ("Formula", "formula"),
        ("Row formula", "row_formula"),
        ("Documents", "document_count"),
        ("Non-zero differences", "nonzero_count"),
    ):
        ws2.append([label, summary.get(key)])
    counts = (recon or {}).get("counts") or {}
    ws2.append([])
    ws2.append(["Status", "Count"])
    for status, count in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
        ws2.append([status, count])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _gstr1_output_gst_amount(row):
    if not row:
        return 0.0
    return round(
        gst_number(row.get("igst"))
        + gst_number(row.get("cgst"))
        + gst_number(row.get("sgst"))
        + gst_number(row.get("cess")),
        2,
    )


def _classify_gstr1_tally_difference_pair(g1_row, tally_row, tolerance):
    """
    Same badge statuses as GSTR-2B vs Tally:
    Matched | Only in GSTR-1 | Only in Tally | Date Mismatch | Amount/Tax Mismatch
    """
    if g1_row and not tally_row:
        return "Only in GSTR-1", "Present in GSTR-1 but no matching Tally Sales voucher"
    if tally_row and not g1_row:
        return "Only in Tally", "Present in Tally Sales but no matching GSTR-1 document"

    fields = ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
    differences = {
        field: round(gst_number(tally_row.get(field)) - gst_number(g1_row.get(field)), 2)
        for field in fields
    }
    date_match = (
        gst_text(g1_row.get("invoice_date")).replace("/", "-")
        == gst_text(tally_row.get("invoice_date")).replace("/", "-")
    )
    value_match = all(abs(value) <= tolerance for value in differences.values())
    if date_match and value_match:
        return "Matched", "All key fields match within tolerance"
    if value_match and not date_match:
        return (
            "Date Mismatch",
            f"Invoice date differs (GSTR-1 {gst_text(g1_row.get('invoice_date')) or '—'} "
            f"vs Tally {gst_text(tally_row.get('invoice_date')) or '—'})",
        )
    parts = []
    for field in ("taxable_value", "igst", "cgst", "sgst", "cess"):
        delta = differences[field]
        if abs(delta) > tolerance:
            label = "Taxable" if field == "taxable_value" else field.upper()
            parts.append(f"{label} ₹{delta:.2f}")
    return "Amount/Tax Mismatch", "Amount/tax differs: " + (", ".join(parts) if parts else "value/tax fields")


def build_gstr1_output_difference_recon(rows_g1, rows_tally, tolerance=1.0, return_period=""):
    """
    Voucher-level Output GST Difference explainer (GSTR-1 vs Tally Sales).

    Mirrors Available ITC Difference / GSTR-2B vs Tally pattern.
    Difference per row = Tally Output GST − GSTR-1 Output GST so the sum equals
    the dashboard Output GST Difference.

    ALL / FY mode: restrict the voucher grid to month(s) whose monthly Output GST
    Diff is non-zero. Months with ₹0.00 monthly difference are excluded from the
    Difference View filter only (import + reconcile_gstr1_tally unchanged).
    """
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    raw_period = gst_text(return_period)
    all_mode = is_gst_all_periods_selection(raw_period) or not normalize_gst_recon_period(raw_period)
    period = "" if all_mode else normalize_gst_recon_period(return_period)
    portal_rows = [ensure_gst_invoice_fields(row) for row in (rows_g1 or [])]
    tally_rows = [ensure_gst_invoice_fields(row) for row in (rows_tally or [])]
    contributing_periods = []
    monthly_diff_map = {}

    def row_period(row):
        return normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date"))

    if period:
        portal_rows = [row for row in portal_rows if row_period(row) == period]
        tally_rows = [row for row in tally_rows if row_period(row) == period]
        contributing_periods = [period]
    elif all_mode:
        # Difference View filter only: keep months that actually move Output GST Diff.
        fy_codes = [code for code, _label in gst_fy_2025_26_periods()]
        present = {
            row_period(row)
            for row in (portal_rows + tally_rows)
            if row_period(row)
        }
        scan_periods = [code for code in fy_codes if code in present] or sorted(present)
        for code in scan_periods:
            g_month = [row for row in portal_rows if row_period(row) == code]
            t_month = [row for row in tally_rows if row_period(row) == code]
            g_tot = gstr1_tax_totals(g_month)
            t_tot = gstr1_tax_totals(t_month)
            month_diff = round(
                gst_number(t_tot.get("output_gst")) - gst_number(g_tot.get("output_gst")),
                2,
            )
            monthly_diff_map[code] = month_diff
            # Zero at money precision → exclude from Difference View.
            if abs(month_diff) >= 0.01:
                contributing_periods.append(code)
        if contributing_periods:
            keep = set(contributing_periods)
            portal_rows = [row for row in portal_rows if row_period(row) in keep]
            tally_rows = [row for row in tally_rows if row_period(row) in keep]
        else:
            # No month contributes — show an empty difference grid (not the full FY).
            portal_rows = []
            tally_rows = []

    left, right = {}, {}
    for row in portal_rows:
        left.setdefault(gst_invoice_key(row), []).append(row)
    for row in tally_rows:
        right.setdefault(gst_invoice_key(row), []).append(row)

    pairs = []
    for key in sorted(set(left) | set(right)):
        left_rows = list(left.get(key, []))
        right_rows = list(right.get(key, []))
        used_right = set()
        for row_g1 in left_rows:
            candidates = [(index, row) for index, row in enumerate(right_rows) if index not in used_right]
            if candidates:
                def candidate_score(pair, base=row_g1):
                    _, candidate = pair
                    date_penalty = 0 if (
                        gst_text(base.get("invoice_date")).replace("/", "-")
                        == gst_text(candidate.get("invoice_date")).replace("/", "-")
                    ) else 1000000
                    amount_penalty = sum(
                        abs(gst_number(candidate.get(field)) - gst_number(base.get(field)))
                        for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
                    )
                    return date_penalty + amount_penalty
                right_index, row_tally = min(candidates, key=candidate_score)
                used_right.add(right_index)
                pairs.append((row_g1, row_tally))
            else:
                pairs.append((row_g1, None))
        for index, row_tally in enumerate(right_rows):
            if index not in used_right:
                pairs.append((None, row_tally))

    rows = []
    counts = {
        "Matched": 0,
        "Only in GSTR-1": 0,
        "Only in Tally": 0,
        "Date Mismatch": 0,
        "Amount/Tax Mismatch": 0,
    }
    voucher_diff_total = 0.0
    for g1_row, tally_row in pairs:
        status, reason = _classify_gstr1_tally_difference_pair(g1_row, tally_row, tolerance)
        g1_taxable = gst_number((g1_row or {}).get("taxable_value")) if g1_row else None
        t_taxable = gst_number((tally_row or {}).get("taxable_value")) if tally_row else None
        g1_igst = gst_number((g1_row or {}).get("igst")) if g1_row else None
        t_igst = gst_number((tally_row or {}).get("igst")) if tally_row else None
        g1_cgst = gst_number((g1_row or {}).get("cgst")) if g1_row else None
        t_cgst = gst_number((tally_row or {}).get("cgst")) if tally_row else None
        g1_sgst = gst_number((g1_row or {}).get("sgst")) if g1_row else None
        t_sgst = gst_number((tally_row or {}).get("sgst")) if tally_row else None
        g1_cess = gst_number((g1_row or {}).get("cess")) if g1_row else None
        t_cess = gst_number((tally_row or {}).get("cess")) if tally_row else None
        g1_out = _gstr1_output_gst_amount(g1_row)
        t_out = _gstr1_output_gst_amount(tally_row)
        # Same sign as dashboard Output GST Diff: Tally − GSTR-1
        difference = round(t_out - g1_out, 2)
        taxable_difference = round(
            (t_taxable if t_taxable is not None else 0.0) - (g1_taxable if g1_taxable is not None else 0.0),
            2,
        )
        voucher_diff_total = round(voucher_diff_total + difference, 2)
        base = g1_row or tally_row or {}
        voucher_type = gst_text(
            (tally_row or {}).get("voucher_type")
            or (tally_row or {}).get("document_type")
            or (g1_row or {}).get("document_type")
            or (g1_row or {}).get("voucher_type")
        )
        g1_gstin = gst_text((g1_row or {}).get("gstin")).upper() if g1_row else ""
        t_gstin = gst_text((tally_row or {}).get("gstin")).upper() if tally_row else ""
        g1_inv = gst_text((g1_row or {}).get("invoice_no")) if g1_row else ""
        t_inv = gst_text((tally_row or {}).get("invoice_no")) if tally_row else ""
        g1_date = gst_text((g1_row or {}).get("invoice_date")).replace("/", "-") if g1_row else ""
        t_date = gst_text((tally_row or {}).get("invoice_date")).replace("/", "-") if tally_row else ""
        field_mismatches = []
        if g1_row and tally_row:
            if g1_gstin != t_gstin and (g1_gstin or t_gstin):
                field_mismatches.append("GSTIN Mismatch")
            g1_inv_norm = normalize_invoice_number(g1_inv)
            t_inv_norm = normalize_invoice_number(t_inv)
            if g1_inv_norm != t_inv_norm and (g1_inv_norm or t_inv_norm):
                field_mismatches.append("Invoice Number Mismatch")
            if g1_date != t_date and (g1_date or t_date):
                field_mismatches.append("Invoice Date Mismatch")
            if abs((t_taxable or 0.0) - (g1_taxable or 0.0)) > tolerance:
                field_mismatches.append("Taxable Mismatch")
            if abs((t_igst or 0.0) - (g1_igst or 0.0)) > tolerance:
                field_mismatches.append("IGST Mismatch")
            if abs((t_cgst or 0.0) - (g1_cgst or 0.0)) > tolerance:
                field_mismatches.append("CGST Mismatch")
            if abs((t_sgst or 0.0) - (g1_sgst or 0.0)) > tolerance:
                field_mismatches.append("SGST Mismatch")
            if abs((t_cess or 0.0) - (g1_cess or 0.0)) > tolerance:
                field_mismatches.append("CESS Mismatch")
        elif g1_row and not tally_row:
            field_mismatches.append("Only in GSTR-1")
        elif tally_row and not g1_row:
            field_mismatches.append("Only in Tally")
        if len(field_mismatches) > 1:
            difference_reason = "Multiple Field Mismatch"
        elif field_mismatches:
            difference_reason = field_mismatches[0]
        else:
            difference_reason = reason or status
        row = {
            "status": status,
            "reason": reason,
            "difference_reason": difference_reason,
            "field_mismatches": field_mismatches,
            "gstin": gst_text(g1_gstin or t_gstin).upper(),
            "gstr1_gstin": g1_gstin or None,
            "tally_gstin": t_gstin or None,
            "party_name": gst_text(
                (g1_row or {}).get("party_name")
                or (g1_row or {}).get("party_ledger")
                or (tally_row or {}).get("party_name")
                or (tally_row or {}).get("party_ledger")
            ),
            "invoice_no": gst_text(g1_inv or t_inv),
            "gstr1_invoice_no": g1_inv or None,
            "tally_invoice_no": t_inv or None,
            "invoice_date": gst_text(g1_date or t_date),
            "gstr1_invoice_date": g1_date or None,
            "tally_invoice_date": t_date or None,
            "voucher_type": voucher_type,
            "gstr1_taxable": g1_taxable,
            "tally_taxable": t_taxable,
            "taxable_difference": taxable_difference,
            "difference": difference,
            "gstr1_igst": g1_igst,
            "tally_igst": t_igst,
            "gstr1_cgst": g1_cgst,
            "tally_cgst": t_cgst,
            "gstr1_sgst": g1_sgst,
            "tally_sgst": t_sgst,
            "gstr1_cess": g1_cess,
            "tally_cess": t_cess,
            "cess": round((t_cess or 0.0) - (g1_cess or 0.0), 2),
            "gstr1_output_gst": g1_out if g1_row else None,
            "tally_output_gst": t_out if tally_row else None,
            "document_bucket": gst_document_bucket(base),
            "source_period": normalize_gst_recon_period(
                base.get("source_period") or base.get("invoice_date")
            ),
        }
        rows.append(row)
        counts[status] = counts.get(status, 0) + 1

    status_order = {
        "Only in GSTR-1": 0,
        "Only in Tally": 1,
        "Amount/Tax Mismatch": 2,
        "Date Mismatch": 3,
        "Matched": 4,
    }
    rows.sort(
        key=lambda r: (
            0 if abs(gst_number(r.get("difference"))) > 0.005 else 1,
            status_order.get(r.get("status"), 99),
            gst_text(r.get("invoice_date")),
            gst_text(r.get("invoice_no")),
        )
    )

    g1_totals = gstr1_tax_totals(portal_rows)
    tally_totals = gstr1_tax_totals(tally_rows)
    output_gst_difference = round(
        gst_number(tally_totals.get("output_gst")) - gst_number(g1_totals.get("output_gst")),
        2,
    )
    period_labels = {code: label for code, label in gst_fy_2025_26_periods()}
    return {
        "rows": rows,
        "counts": counts,
        "summary": {
            "gstr1_output_gst": gst_number(g1_totals.get("output_gst")),
            "tally_output_gst": gst_number(tally_totals.get("output_gst")),
            "output_gst_difference": output_gst_difference,
            "voucher_difference_total": voucher_diff_total,
            "balanced": abs(voucher_diff_total - output_gst_difference) <= max(tolerance, 0.02),
            "formula": "Tally Output GST − GSTR-1 Output GST",
            "row_formula": "Difference = (Tally IGST+CGST+SGST+CESS) − (GSTR-1 IGST+CGST+SGST+CESS)",
            "document_count": len(rows),
            "nonzero_count": sum(1 for r in rows if abs(gst_number(r.get("difference"))) > 0.005),
            "return_period": period or "ALL",
            "period_mode": "fy_all" if all_mode else "month",
            "contributing_periods": contributing_periods,
            "contributing_period_labels": [
                period_labels.get(code, code) for code in contributing_periods
            ],
            "monthly_output_gst_differences": monthly_diff_map,
            "gstr1_taxable": gst_number(g1_totals.get("taxable_value")),
            "tally_taxable": gst_number(tally_totals.get("taxable_value")),
        },
        "tolerance": tolerance,
    }


def make_gstr1_difference_export(recon, title="GSTR-1 vs Tally Difference"):
    rows = (recon or {}).get("rows") or []
    summary = (recon or {}).get("summary") or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Voucher Differences"
    headers = [
        "Status", "GSTIN", "Party Name", "Invoice Number", "Invoice Date", "Voucher Type",
        "GSTR-1 Taxable", "Tally Taxable", "Difference",
        "GSTR-1 IGST", "Tally IGST", "GSTR-1 CGST", "Tally CGST",
        "GSTR-1 SGST", "Tally SGST", "GSTR-1 CESS", "Tally CESS", "Difference Reason",
        "Field Mismatches",
    ]
    ws.append(headers)
    for row in rows:
        mismatches = row.get("field_mismatches") or []
        if isinstance(mismatches, list):
            mismatch_text = "; ".join(str(x) for x in mismatches if x)
        else:
            mismatch_text = gst_text(mismatches)
        ws.append([
            row.get("status"),
            row.get("gstin"),
            row.get("party_name"),
            row.get("invoice_no"),
            row.get("invoice_date"),
            row.get("voucher_type"),
            row.get("gstr1_taxable"),
            row.get("tally_taxable"),
            row.get("difference"),
            row.get("gstr1_igst"),
            row.get("tally_igst"),
            row.get("gstr1_cgst"),
            row.get("tally_cgst"),
            row.get("gstr1_sgst"),
            row.get("tally_sgst"),
            row.get("gstr1_cess"),
            row.get("tally_cess"),
            row.get("difference_reason") or row.get("reason"),
            mismatch_text,
        ])
    fill = PatternFill("solid", fgColor="17365D")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
    for col_idx in range(7, 18):
        letter = get_column_letter(col_idx)
        for cell in ws[letter][1:]:
            cell.number_format = "#,##0.00"
    widths = [20, 18, 28, 16, 12, 14, 14, 14, 12, 12, 12, 12, 12, 12, 12, 12, 12, 28, 40]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for label, key in (
        ("GSTR-1 Output GST", "gstr1_output_gst"),
        ("Tally Output GST", "tally_output_gst"),
        ("Output GST Difference", "output_gst_difference"),
        ("Sum of voucher differences", "voucher_difference_total"),
        ("Balanced", "balanced"),
        ("Formula", "formula"),
        ("Row formula", "row_formula"),
        ("Documents", "document_count"),
        ("Non-zero differences", "nonzero_count"),
        ("Return period", "return_period"),
    ):
        ws2.append([label, summary.get(key)])
    counts = (recon or {}).get("counts") or {}
    ws2.append([])
    ws2.append(["Status", "Count"])
    for status in ("Matched", "Only in GSTR-1", "Only in Tally", "Date Mismatch", "Amount/Tax Mismatch"):
        ws2.append([status, counts.get(status, 0)])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_itc_dashboard(rows_2b, gstr3b_data, rows_tally, tolerance=1.0):
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    # GSTR-2B period ITC must net Credit Notes (portal Comparison / ITC Available sheet).
    gstr2b_periods = gst_recon_period_totals(rows_2b, signed=True)
    # Tally Booked uses signed CN/DN/reversal rows — never unsigned purchase gross.
    rows_tally = normalize_tally_purchase_itc_rows(rows_tally)
    tally_periods = gst_recon_period_totals(rows_tally, signed=True)
    tally_booked = build_tally_booked_itc_summary(rows_tally)
    gstr3b_periods = {}
    raw_periods = (gstr3b_data or {}).get("net_periods") or (gstr3b_data or {}).get("periods") or {}
    for period, values in raw_periods.items():
        key = normalize_gst_recon_period(period)
        if key:
            gstr3b_periods[key] = {
                "igst": round(gst_number(values.get("igst")), 2),
                "cgst": round(gst_number(values.get("cgst")), 2),
                "sgst": round(gst_number(values.get("sgst")), 2),
                "cess": round(gst_number(values.get("cess")), 2),
            }
    all_claim = None
    if not gstr3b_periods and gstr3b_data:
        all_claim = {
            "igst": round(gst_number((gstr3b_data.get("net_itc") or gstr3b_data).get("igst")), 2),
            "cgst": round(gst_number((gstr3b_data.get("net_itc") or gstr3b_data).get("cgst")), 2),
            "sgst": round(gst_number((gstr3b_data.get("net_itc") or gstr3b_data).get("sgst")), 2),
            "cess": round(gst_number((gstr3b_data.get("net_itc") or gstr3b_data).get("cess")), 2),
        }
    # Never broadcast a single/ALL 3B claim onto every 2B month — that inflated
    # "GSTR-3B Claimed" to crores (claim × month count).
    period_keys = set(gstr2b_periods) | set(tally_periods) | set(gstr3b_periods)
    periods = sorted(period_keys)
    rows = []
    summary = {"gstr2b_itc": 0.0, "gstr3b_itc": 0.0, "tally_itc": 0.0, "available_itc": 0.0}
    for period in periods:
        g2b = gstr2b_periods.get(period, {})
        tally = tally_periods.get(period, {})
        g3b = gstr3b_periods.get(period, {})
        g2b_itc = round(
            gst_number(g2b.get("igst")) + gst_number(g2b.get("cgst"))
            + gst_number(g2b.get("sgst")) + gst_number(g2b.get("cess")),
            2,
        )
        g3b_itc = round(
            gst_number(g3b.get("igst")) + gst_number(g3b.get("cgst"))
            + gst_number(g3b.get("sgst")) + gst_number(g3b.get("cess")),
            2,
        )
        # Period Tally Booked = IGST+CGST+SGST (Section 4C style; cess tracked separately).
        tally_itc = round(
            gst_number(tally.get("igst")) + gst_number(tally.get("cgst"))
            + gst_number(tally.get("sgst")),
            2,
        )
        booked = max(g3b_itc, tally_itc)
        available = round(max(0.0, g2b_itc - booked), 2)
        action = "Matched"
        if g3b_itc <= 0 and tally_itc <= 0 and g2b_itc > tolerance:
            action = "ITC available but not fully booked/claimed"
        elif g2b_itc - g3b_itc > tolerance and g2b_itc - tally_itc > tolerance:
            action = "ITC available but not fully booked/claimed"
        elif g3b_itc - g2b_itc > tolerance:
            action = "Possible excess ITC claim in GSTR-3B"
        elif tally_itc - g2b_itc > tolerance:
            action = "Tally books more ITC than GSTR-2B"
        if len(period) == 6:
            period_label = f"{period[0:2]}/{period[2:6]}"
        else:
            period_label = period
        rows.append({
            "period": period,
            "period_label": period_label,
            "gstr2b_invoices": int(g2b.get("invoices", 0)),
            "gstr2b_itc": g2b_itc,
            "gstr3b_itc": g3b_itc,
            "tally_itc": tally_itc,
            "available_itc": available,
            "taxable_value": round(gst_number(g2b.get("taxable_value")), 2),
            "igst": round(gst_number(g2b.get("igst")), 2),
            "cgst": round(gst_number(g2b.get("cgst")), 2),
            "sgst": round(gst_number(g2b.get("sgst")), 2),
            "action": action,
        })
        summary["gstr2b_itc"] = round(summary["gstr2b_itc"] + g2b_itc, 2)
        summary["gstr3b_itc"] = round(summary["gstr3b_itc"] + g3b_itc, 2)
        summary["tally_itc"] = round(summary["tally_itc"] + tally_itc, 2)
        summary["available_itc"] = round(summary["available_itc"] + available, 2)
    if summary["gstr3b_itc"] == 0 and all_claim:
        # Single unscoped 3B import: show once in summary only, not per month.
        summary["gstr3b_itc"] = round(
            gst_number(all_claim.get("igst"))
            + gst_number(all_claim.get("cgst"))
            + gst_number(all_claim.get("sgst"))
            + gst_number(all_claim.get("cess")),
            2,
        )
    net_pack = build_gstr2b_gross_net_summary(rows_2b)
    # Always expose portal-aligned Net GSTR-2B ITC (never gross) as the headline figure.
    summary["gstr2b_itc"] = net_pack["net_itc"]
    summary["gstr2b_net_itc"] = net_pack["net_itc"]
    summary["gstr2b_gross_invoice_itc"] = net_pack["gross_invoice_itc"]
    summary["gstr2b_credit_note_itc"] = net_pack["credit_note_itc"]
    summary["gstr2b_debit_note_itc"] = net_pack["debit_note_itc"]
    summary["gstr2b_amendment_itc"] = net_pack["amendment_itc"]
    summary["invoice_count"] = net_pack["invoice_count"]
    summary["credit_note_count"] = net_pack["credit_note_count"]
    summary["debit_note_count"] = net_pack["debit_note_count"]
    summary["amendment_count"] = net_pack["amendment_count"]
    # Headline Tally Booked = Section 4C Net ITC (not purchase voucher gross).
    summary["tally_itc"] = tally_booked["net_itc"]
    summary["tally_booked_purchase_itc"] = tally_booked["purchase_itc"]
    summary["tally_booked_credit_note_itc"] = tally_booked["credit_note_itc"]
    summary["tally_booked_debit_note_itc"] = tally_booked["debit_note_itc"]
    summary["tally_booked_reversal_itc"] = tally_booked["reversal_itc"]
    summary["tally_booked_igst"] = tally_booked["igst"]
    summary["tally_booked_cgst"] = tally_booked["cgst"]
    summary["tally_booked_sgst"] = tally_booked["sgst"]
    summary["tally_booked_cess"] = tally_booked["cess"]
    summary["tally_booked_formula"] = tally_booked["formula"]
    # Available ITC uses the same Section 4C Tally Booked figure as the dashboard card.
    booked_headline = max(summary["gstr3b_itc"], summary["tally_itc"])
    summary["available_itc"] = round(max(0.0, summary["gstr2b_itc"] - booked_headline), 2)
    return {
        "rows": rows,
        "summary": summary,
        "periods": len(rows),
        "gross_net": net_pack,
        "tally_booked": tally_booked,
    }


# ---------------------------------------------------------------------------
# Phase 2 — GSTR-1 vs Tally Sales (GST Reconciliation module only)
# Reuses parse_gst_file / tally_collection_xml / gst_recon_* / invoice keys.
# ---------------------------------------------------------------------------

GSTR1_SECTION_MAP = {
    "b2b": ("B2B", "Invoice"),
    "b2ba": ("B2B Amendment", "Invoice Amendment"),
    "b2cl": ("B2CL", "Invoice"),
    "b2cla": ("B2CL Amendment", "Invoice Amendment"),
    "b2cs": ("B2CS", "B2C Invoice"),
    "b2csa": ("B2CS Amendment", "B2C Invoice Amendment"),
    "cdnr": ("CDNR", "Credit/Debit Note"),
    "cdnra": ("CDNRA", "Credit/Debit Note Amendment"),
    "cdnur": ("CDNUR", "Credit/Debit Note"),
    "cdnura": ("CDNURA", "Credit/Debit Note Amendment"),
    "exp": ("Export", "Export Invoice"),
    "expa": ("Export Amendment", "Export Invoice Amendment"),
    "sez": ("SEZ", "SEZ Invoice"),
    "sezwp": ("SEZ", "SEZ Invoice"),
    "sezwop": ("SEZ", "SEZ Invoice"),
}

# Portal meta / non-supply blocks — never treat as outward invoices.
GSTR1_SKIP_KEYS = {
    "gstin", "fp", "filing_typ", "gt", "cur_gt", "fil_dt", "chksum", "version",
    "hsn", "nil", "doc_issue", "docdata", "doc_det", "txp", "txpd", "at", "ata",
    "supeco", "eco", "ecom", "einv",
}


def normalize_invoice_number(value):
    text = gst_text(value)
    compact = re.sub(r"[\s\-_/\\.]+", "", text).upper()
    return compact.lstrip("0") or "0"


def gstr1_financial_year(period):
    digits = re.sub(r"\D", "", gst_text(period))
    if len(digits) == 6:
        month, year = int(digits[:2]), int(digits[2:])
        start = year if month >= 4 else year - 1
        return f"{start}-{str(start + 1)[-2:]}"
    return ""


def gstr1_document_sign(document_type, section=""):
    """Credit notes reduce outward liability; debit notes / invoices add."""
    doc = gst_text(document_type).lower()
    section = gst_text(section).lower()
    if "debit note" in doc:
        return 1
    if "credit note" in doc:
        return -1
    # Ambiguous CDN section label without ntty — default Credit Note (portal majority).
    if "cdn" in section and "debit" not in doc:
        if "credit" in doc or doc in {"credit/debit note", "credit/debit note amendment", ""}:
            return -1
    return 1


def classify_gstr1_note_type(record, section=""):
    raw = gst_text(
        gst_pick(record, "ntty", "note type", "document type", "type", "doctyp")
    ).upper()
    if raw in {"C", "CR", "CREDIT", "CREDIT NOTE"}:
        return "Credit Note"
    if raw in {"D", "DR", "DEBIT", "DEBIT NOTE"}:
        return "Debit Note"
    section = gst_text(section).upper()
    if "CDN" in section:
        return "Credit/Debit Note"
    return ""


def normalize_gstr1_row(row, section="", document_type="", source_period="", gstin="", taxpayer_gstin=""):
    if not row:
        return None
    row = ensure_gst_invoice_fields(row)
    section_name = section or gst_text(row.get("section")) or "GSTR-1"
    section_upper = section_name.upper()
    is_b2cs = section_upper.startswith("B2CS")
    invoice_no = gst_text(row.get("invoice_no"))
    if not invoice_no and is_b2cs:
        pos = gst_text(row.get("place_of_supply") or row.get("pos") or "NA")
        rate = gst_text(row.get("gst_rate") or row.get("rt") or "0")
        period = normalize_gst_recon_period(source_period or row.get("source_period")) or "NOPERIOD"
        invoice_no = f"B2CS-{period}-{pos}-{rate}"
        row["invoice_no"] = invoice_no
    if not invoice_no:
        return None
    note_type = classify_gstr1_note_type(row, section) or gst_text(row.get("document_type"))
    doc_type = document_type or note_type or gst_text(row.get("document_type")) or "Invoice"
    if note_type in {"Credit Note", "Debit Note"}:
        doc_type = note_type
    # Party GSTIN: never stamp filing taxpayer GSTIN onto B2CS / B2CL summary rows.
    party_gstin = gst_text(row.get("gstin")).upper()
    if not party_gstin and not is_b2cs and not section_upper.startswith("B2CL"):
        candidate = gst_text(gstin).upper()
        taxpayer = gst_text(taxpayer_gstin).upper()
        if candidate and candidate != taxpayer:
            party_gstin = candidate
    is_b2c = (
        section_upper.startswith("B2C")
        or doc_type.upper().startswith("B2C")
        or (not party_gstin and section_upper not in {"B2B", "B2B AMENDMENT", "CDNR", "CDNRA"})
    )
    signed = gstr1_document_sign(doc_type, section_name)
    taxable = round(gst_number(row.get("taxable_value")) * signed, 2)
    igst = round(gst_number(row.get("igst")) * signed, 2)
    cgst = round(gst_number(row.get("cgst")) * signed, 2)
    sgst = round(gst_number(row.get("sgst")) * signed, 2)
    cess = round(gst_number(row.get("cess")) * signed, 2)
    invoice_value = round(gst_number(row.get("invoice_value")) * signed, 2)
    if not invoice_value:
        invoice_value = round(taxable + igst + cgst + sgst + cess, 2)
    # Preserve portal tax figures exactly — do not rewrite CGST/SGST to match val.
    period = normalize_gst_recon_period(source_period or row.get("source_period") or row.get("invoice_date"))
    return ensure_gst_invoice_fields({
        **row,
        "gstin": party_gstin,
        "party_name": gst_text(row.get("party_name")) or ("B2C" if is_b2c else ""),
        "invoice_no": invoice_no,
        "invoice_no_norm": normalize_invoice_number(invoice_no),
        "invoice_date": gst_text(row.get("invoice_date")),
        "taxable_value": taxable,
        "igst": igst,
        "cgst": cgst,
        "sgst": sgst,
        "cess": cess,
        "invoice_value": invoice_value,
        "gst_rate": gst_rate_for_values(abs(taxable), abs(igst), abs(cgst), abs(sgst), abs(cess)),
        "section": section_name,
        "document_type": doc_type,
        "place_of_supply": gst_text(row.get("place_of_supply") or row.get("pos") or row.get("Place of Supply")),
        "hsn_code": gst_text(row.get("hsn_code") or row.get("hsn") or row.get("HSN")),
        "original_invoice_no": gst_text(row.get("original_invoice_no") or row.get("oinum") or row.get("ont_num")),
        "source_period": period,
        "financial_year": gstr1_financial_year(period),
        "is_b2c": bool(is_b2c),
        "return_type": "GSTR1",
        "taxpayer_gstin": gst_text(taxpayer_gstin or row.get("taxpayer_gstin")).upper(),
        "source": gst_text(row.get("source")) or "GSTR-1",
        "document_sign": signed,
        "_portal_signed": True,
    })


def gstr1_section_bucket(section):
    text = gst_text(section).upper()
    if text.startswith("B2B"):
        return "B2B" if "AMEND" not in text else "B2B Amendment"
    if text.startswith("B2CS"):
        return "B2CS" if "AMEND" not in text else "B2CS Amendment"
    if text.startswith("B2CL"):
        return "B2CL" if "AMEND" not in text else "B2CL Amendment"
    if text.startswith("CDN"):
        return "CDNR/CDNUR"
    if text.startswith("EXP") or "EXPORT" in text:
        return "Export"
    if "AMEND" in text:
        return "Amendment"
    if text.startswith("SEZ"):
        return "SEZ"
    return text or "Other"


def build_gstr1_section_totals(rows):
    """
    Section-wise GSTR-1 outward totals for verification.
    Net Output GST = IGST + CGST + SGST + CESS (credit notes already signed negative).
    """
    buckets = {}
    for row in rows or []:
        item = ensure_gst_invoice_fields(row)
        name = gstr1_section_bucket(item.get("section"))
        bucket = buckets.setdefault(name, {
            "section": name, "count": 0,
            "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
            "output_gst": 0.0,
        })
        bucket["count"] += 1
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess"):
            bucket[key] = round(bucket[key] + gst_number(item.get(key)), 2)
        bucket["output_gst"] = round(
            bucket["igst"] + bucket["cgst"] + bucket["sgst"] + bucket["cess"], 2
        )
    gross = {
        "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
        "output_gst": 0.0, "count": 0,
    }
    for bucket in buckets.values():
        gross["count"] += bucket["count"]
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "output_gst"):
            gross[key] = round(gross[key] + bucket[key], 2)
    amendment = round(
        sum(
            bucket["output_gst"]
            for name, bucket in buckets.items()
            if "AMEND" in name.upper()
        ),
        2,
    )
    return {
        "sections": buckets,
        "gross_taxable": gross["taxable_value"],
        "igst": gross["igst"],
        "cgst": gross["cgst"],
        "sgst": gross["sgst"],
        "cess": gross["cess"],
        "net_output_gst": gross["output_gst"],
        "document_count": gross["count"],
        "b2b_total": gst_number((buckets.get("B2B") or {}).get("output_gst")),
        "b2cs_total": gst_number((buckets.get("B2CS") or {}).get("output_gst")),
        "b2cl_total": gst_number((buckets.get("B2CL") or {}).get("output_gst")),
        "cdn_total": gst_number((buckets.get("CDNR/CDNUR") or {}).get("output_gst")),
        "export_total": gst_number((buckets.get("Export") or {}).get("output_gst")),
        "amendment_adjustment": amendment,
        "final_net_gstr1_total": gross["output_gst"],
        "formula": "Gross outward tax (B2B+B2CS+B2CL+CDN+Export+…) with Credit Notes signed negative",
    }


def log_gstr1_section_totals(summary, source=""):
    summary = summary or {}
    print(
        f"[GSTR-1] {gst_text(source) or 'import'} | "
        f"Gross Taxable={summary.get('gross_taxable', 0):.2f} | "
        f"IGST={summary.get('igst', 0):.2f} CGST={summary.get('cgst', 0):.2f} "
        f"SGST={summary.get('sgst', 0):.2f} CESS={summary.get('cess', 0):.2f} | "
        f"Net Output GST={summary.get('net_output_gst', 0):.2f}"
    )
    print(
        f"[GSTR-1] sections | B2B={summary.get('b2b_total', 0):.2f} "
        f"B2CS={summary.get('b2cs_total', 0):.2f} B2CL={summary.get('b2cl_total', 0):.2f} "
        f"CDN={summary.get('cdn_total', 0):.2f} Export={summary.get('export_total', 0):.2f} "
        f"Amendment={summary.get('amendment_adjustment', 0):.2f} | "
        f"Final Net={summary.get('final_net_gstr1_total', 0):.2f}"
    )
    for name, bucket in sorted((summary.get("sections") or {}).items()):
        print(
            f"[GSTR-1]   {name}: count={bucket.get('count', 0)} "
            f"taxable={bucket.get('taxable_value', 0):.2f} "
            f"output_gst={bucket.get('output_gst', 0):.2f}"
        )


def gstr1_rows_from_portal_json(data, source=""):
    """
    Walk GST Portal GSTR-1 JSON with all outward-supply sections:
    B2B, B2CL, B2CS, CDNR/CDNUR, EXP/SEZ and their amendments.
    """
    rows = []
    root_gstin = gst_text(gst_pick(data, "gstin", "GSTIN")).upper() if isinstance(data, dict) else ""
    root_period = ""
    if isinstance(data, dict):
        root_period = normalize_gst_recon_period(
            gst_pick(data, "fp", "ret_period", "return period") or infer_gst_period(source)
        )

    def add_invoice(record, context):
        if not isinstance(record, dict):
            return
        section = gst_text(context.get("section"))
        section_upper = section.upper()
        candidate = normalize_gst_invoice(record, context)
        # B2CS / rate-wise summary rows have no inum — build from txval/tax keys.
        if not candidate and (
            section_upper.startswith("B2CS")
            or (
                gst_pick(record, "txval", "taxable value") not in (None, "")
                and not gst_pick(record, "inum", "nt_num", "invoice number")
            )
        ):
            txval = gst_number(gst_pick(record, "txval", "taxable value"))
            igst = gst_number(gst_pick(record, "iamt", "igst"))
            cgst = gst_number(gst_pick(record, "camt", "cgst"))
            sgst = gst_number(gst_pick(record, "samt", "sgst"))
            cess = gst_number(gst_pick(record, "csamt", "cess"))
            if abs(txval) <= 0 and abs(igst + cgst + sgst + cess) <= 0:
                return
            pos = gst_text(gst_pick(record, "pos", "place of supply"))
            rate = gst_pick(record, "rt", "rate", "gst_rate") or 0
            period = context.get("source_period") or root_period or "NOPERIOD"
            sply = gst_pick(record, "sply_ty", "supply type")
            candidate = {
                "gstin": "",
                "party_name": gst_text(sply) or ("B2C" if section_upper.startswith("B2C") else "GSTR-1"),
                "invoice_no": (
                    gst_text(gst_pick(record, "inum", "invoice number", "nt_num", "note number"))
                    or f"{section_upper or 'B2CS'}-{period}-{pos or 'NA'}-{rate}"
                ),
                "invoice_date": gst_text(gst_pick(record, "idt", "invoice date", "nt_dt", "note date")),
                "invoice_value": round(txval + igst + cgst + sgst + cess, 2),
                "taxable_value": txval,
                "igst": igst,
                "cgst": cgst,
                "sgst": sgst,
                "cess": cess,
                "place_of_supply": pos,
                "rt": rate,
                "gst_rate": rate,
                "source": source,
                "section": section or "B2CS",
                "document_type": context.get("document_type", "B2C Invoice"),
            }
        if not candidate:
            return
        note = classify_gstr1_note_type(record, section)
        if note:
            candidate["document_type"] = note
        candidate["place_of_supply"] = gst_text(
            candidate.get("place_of_supply") or gst_pick(record, "pos", "place of supply")
        )
        candidate["original_invoice_no"] = gst_text(
            gst_pick(record, "oinum", "ont_num", "original invoice number", "against invoice")
        )
        # Counterparty GSTIN only — never fall back to filing GSTIN for party field.
        party_gstin = gst_text(context.get("gstin") or candidate.get("gstin")).upper()
        if party_gstin == root_gstin:
            party_gstin = ""
        normalized = normalize_gstr1_row(
            candidate,
            section=section,
            document_type=candidate.get("document_type") or context.get("document_type", "Invoice"),
            source_period=context.get("source_period") or root_period,
            gstin=party_gstin,
            taxpayer_gstin=root_gstin,
        )
        if normalized:
            rows.append(normalized)

    def walk_section(node, context):
        if isinstance(node, dict):
            next_context = dict(context)
            gstin = gst_pick(node, "ctin", "gstin")
            if gstin:
                party = gst_text(gstin).upper()
                if party != root_gstin:
                    next_context["gstin"] = party
                    next_context["party_name"] = gst_text(
                        gst_pick(node, "trade_name", "lgl_nm", "party", "cname")
                        or next_context.get("party_name", "")
                    )
            for key in ("inv", "nt", "inv_data", "invoices"):
                child = node.get(key)
                if isinstance(child, list):
                    for item in child:
                        if isinstance(item, dict):
                            add_invoice(item, next_context)
            if any(gst_pick(node, "nt_num", "inum", "invoice number", "note number")):
                add_invoice(node, next_context)
            for key, value in node.items():
                lower = str(key).lower()
                if lower in GSTR1_SKIP_KEYS or lower in {"inv", "nt", "itms", "item", "items", "inv_data", "invoices"}:
                    continue
                if lower in GSTR1_SECTION_MAP:
                    section, doc = GSTR1_SECTION_MAP[lower]
                    walk_section(value, {**next_context, "section": section, "document_type": doc})
                else:
                    walk_section(value, next_context)
        elif isinstance(node, list):
            section_upper = gst_text(context.get("section")).upper()
            for value in node:
                if isinstance(value, dict) and (
                    section_upper.startswith("B2CS")
                    or section_upper.startswith("B2CL")
                    or (
                        gst_pick(value, "txval") not in (None, "")
                        and not gst_pick(value, "inum", "nt_num", "ctin")
                    )
                ):
                    add_invoice(value, context)
                else:
                    walk_section(value, context)

    if isinstance(data, dict):
        matched_section = False
        for key, value in data.items():
            lower = str(key).lower()
            if lower in GSTR1_SKIP_KEYS:
                continue
            if lower in GSTR1_SECTION_MAP:
                matched_section = True
                section, doc = GSTR1_SECTION_MAP[lower]
                walk_section(value, {
                    "source": source,
                    "source_period": root_period,
                    "section": section,
                    "document_type": doc,
                })
        if not matched_section:
            for row in gst_rows_from_json(data, source):
                normalized = normalize_gstr1_row(
                    row, source_period=root_period, taxpayer_gstin=root_gstin
                )
                if normalized:
                    rows.append(normalized)
    else:
        for row in gst_rows_from_json(data, source):
            normalized = normalize_gstr1_row(
                row, source_period=root_period, taxpayer_gstin=root_gstin
            )
            if normalized:
                rows.append(normalized)

    unique = {}
    for row in rows:
        key = (
            row.get("source_period"),
            row.get("section"),
            row.get("gstin"),
            row.get("invoice_no_norm"),
            row.get("invoice_date"),
            row.get("document_type"),
            round(gst_number(row.get("taxable_value")), 2),
            round(gst_number(row.get("cgst")), 2),
            round(gst_number(row.get("sgst")), 2),
            round(gst_number(row.get("igst")), 2),
        )
        unique[key] = row
    return list(unique.values())


def gstr1_rows_from_csv(raw, source=""):
    text = raw.decode("utf-8-sig", errors="replace")
    grid = list(csv.reader(io.StringIO(text)))
    if not grid:
        return []
    header_index = next(
        (
            i for i, line in enumerate(grid[:30])
            if any("invoice" in gst_text(v).lower() and "no" in gst_text(v).lower() for v in line)
            or any(gst_text(v).lower() in {"inum", "invoice_no", "inv no"} for v in line)
        ),
        0,
    )
    headers = [gst_text(value) for value in grid[header_index]]
    rows = []
    period = infer_gst_period(source)
    for values in grid[header_index + 1:]:
        if not any(gst_text(v) for v in values):
            continue
        record = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        candidate = normalize_gst_invoice(record, {"source": source, "section": "CSV", "document_type": "Invoice"})
        if not candidate:
            continue
        section = gst_text(gst_pick(record, "section", "return section", "type")) or "CSV"
        normalized = normalize_gstr1_row(candidate, section=section, source_period=period)
        if normalized:
            rows.append(normalized)
    return rows


def parse_gstr1_file(name, raw):
    """Import GSTR-1 JSON/Excel/CSV using existing parsers where possible."""
    suffix = Path(name).suffix.lower()
    period = infer_gst_period(name)
    if suffix == ".json":
        data = json.loads(raw.decode("utf-8-sig"))
        rows = gstr1_rows_from_portal_json(data, name)
        summary = build_gstr1_section_totals(rows)
        log_gstr1_section_totals(summary, name)
        return rows
    if suffix == ".csv":
        rows = gstr1_rows_from_csv(raw, name)
        log_gstr1_section_totals(build_gstr1_section_totals(rows), name)
        return rows
    if suffix == ".zip":
        rows = []
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            for member in archive.infolist():
                if member.is_dir():
                    continue
                if Path(member.filename).suffix.lower() in {".json", ".xlsx", ".xlsm", ".xls", ".csv"}:
                    rows.extend(parse_gstr1_file(member.filename, archive.read(member)))
        if rows:
            log_gstr1_section_totals(build_gstr1_section_totals(rows), name)
        return rows
    # Excel / XLS — reuse existing GST excel parser then normalize as GSTR-1
    rows = parse_gst_file(name, raw)
    normalized = []
    for row in rows:
        section = gst_text(row.get("section")) or Path(name).stem
        doc = gst_text(row.get("document_type")) or "Invoice"
        item = normalize_gstr1_row(row, section=section, document_type=doc, source_period=period or row.get("source_period"))
        if item:
            normalized.append(item)
    log_gstr1_section_totals(build_gstr1_section_totals(normalized), name)
    return normalized


def gstr1_file_digest(raw):
    return hashlib.sha256(raw).hexdigest()


def gstr1_save_import_batch(file_name, raw, rows, gstin="", return_period=""):
    """Persist GSTR-1 rows with duplicate-file protection."""
    rows = rows or []
    if not rows:
        raise ValueError("No GSTR-1 invoice rows were found in the selected file.")
    digest = gstr1_file_digest(raw)
    period = normalize_gst_recon_period(return_period or rows[0].get("source_period") or infer_gst_period(file_name))
    # Filing taxpayer GSTIN only — never fall back to counterparty (ctin) on invoice rows.
    batch_gstin = gst_text(
        gstin
        or next((row.get("taxpayer_gstin") or row.get("filing_gstin") for row in rows if row.get("taxpayer_gstin") or row.get("filing_gstin")), "")
        or infer_taxpayer_gstin_from_text(file_name)
        or (gst_portal_get_context().get("gstin") if gst_portal_get_context() else "")
    ).upper()
    fy = gstr1_financial_year(period) or gst_portal_default_fy()
    if batch_gstin:
        gst_portal_set_context(batch_gstin, fy)
    stamped_import_rows = [
        gst_stamp_portal_row(row, taxpayer_gstin=batch_gstin, financial_year=fy, return_type="GSTR-1")
        for row in rows
    ]
    connection = gst_recon_connection()
    try:
        existing = connection.execute(
            "SELECT id, record_count FROM gst_1_import_batches WHERE file_digest=? AND return_period=? AND gstin=?",
            (digest, period, batch_gstin),
        ).fetchone()
        if existing:
            gst_session_mark_imported("GSTR-1", period)
            return {
                "duplicate": True,
                "batch_id": existing["id"],
                "record_count": existing["record_count"],
                "message": "This GSTR-1 file was already imported for the same period.",
                "rows": gst_recon_load_rows("GSTR-1", gstin=batch_gstin, financial_year=fy),
            }
        stamp = gst_recon_now()
        cursor = connection.execute(
            """INSERT INTO gst_1_import_batches
            (gstin, financial_year, return_period, import_date, file_name, record_count, return_type, file_digest)
            VALUES (?,?,?,?,?,?, 'GSTR1', ?)""",
            (batch_gstin, fy, period, stamp, file_name, len(stamped_import_rows), digest),
        )
        batch_id = cursor.lastrowid
        for row in stamped_import_rows:
            connection.execute(
                """INSERT INTO gst_1_invoices
                (batch_id, row_json, gstin, invoice_no, invoice_no_norm, invoice_date, section,
                 document_type, source_period, taxable_value, igst, cgst, sgst, cess, invoice_value, imported_at,
                 taxpayer_gstin, financial_year)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    batch_id,
                    json.dumps(row, ensure_ascii=False),
                    gst_text(row.get("gstin")),
                    gst_text(row.get("invoice_no")),
                    gst_text(row.get("invoice_no_norm") or normalize_invoice_number(row.get("invoice_no"))),
                    gst_text(row.get("invoice_date")),
                    gst_text(row.get("section")),
                    gst_text(row.get("document_type")),
                    gst_text(row.get("source_period") or period),
                    gst_number(row.get("taxable_value")),
                    gst_number(row.get("igst")),
                    gst_number(row.get("cgst")),
                    gst_number(row.get("sgst")),
                    gst_number(row.get("cess")),
                    gst_number(row.get("invoice_value")),
                    stamp,
                    batch_gstin,
                    fy,
                ),
            )
        connection.commit()
    finally:
        connection.close()
    # Also mirror into shared recon dataset for one-click / load APIs.
    # Include source_period so B2CS rate-lines from different months never collide.
    existing_rows = gst_recon_load_rows("GSTR-1", gstin=batch_gstin, financial_year=fy)
    merged = [ensure_gst_invoice_fields(row) for row in (existing_rows + stamped_import_rows)]
    unique = {}
    for row in merged:
        key = (
            normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")),
            gst_text(row.get("section")),
            gst_text(row.get("gstin")),
            row.get("invoice_no_norm") or normalize_invoice_number(row.get("invoice_no")),
            gst_text(row.get("invoice_date")),
            gst_text(row.get("document_type")),
            round(gst_number(row.get("taxable_value")), 2),
            round(gst_number(row.get("igst")), 2),
            round(gst_number(row.get("cgst")), 2),
            round(gst_number(row.get("sgst")), 2),
        )
        unique[key] = row
    saved = list(unique.values())
    gst_recon_save_rows("GSTR-1", saved, gstin=batch_gstin, financial_year=fy)
    section_summary = build_gstr1_section_totals(stamped_import_rows)
    log_gstr1_section_totals(section_summary, file_name)
    gst_session_mark_imported("GSTR-1", period)
    gst_recon_set_meta("gstr1_last_import", {
        "batch_id": batch_id,
        "file_name": file_name,
        "record_count": len(stamped_import_rows),
        "section_summary": section_summary,
        "return_period": period,
        "gstin": batch_gstin,
        "financial_year": fy,
        "imported_at": stamp,
    })
    return {
        "duplicate": False,
        "batch_id": batch_id,
        "record_count": len(stamped_import_rows),
        "total_rows": len(saved),
        "return_period": period,
        "gstin": batch_gstin,
        "rows": saved,
        "summary": gst_summary(stamped_import_rows),
        "section_summary": section_summary,
    }


def gstr1_rows_from_invoice_table(return_period="", gstin="", financial_year=""):
    """Authoritative GSTR-1 rows from gst_1_invoices (parser persistence)."""
    period = normalize_gst_recon_period(return_period)
    connection = gst_recon_connection()
    try:
        gst_portal_backfill_scope_columns(connection)
        if period:
            records = connection.execute(
                "SELECT row_json, taxpayer_gstin, financial_year, source_period FROM gst_1_invoices "
                "WHERE source_period=? ORDER BY id",
                (period,),
            ).fetchall()
        else:
            records = connection.execute(
                "SELECT row_json, taxpayer_gstin, financial_year, source_period FROM gst_1_invoices ORDER BY id"
            ).fetchall()
    finally:
        connection.close()
    rows = []
    for record in records:
        try:
            item = json.loads(record["row_json"] or "{}")
        except (TypeError, json.JSONDecodeError, ValueError):
            continue
        if not item.get("taxpayer_gstin"):
            item["taxpayer_gstin"] = gst_text(record["taxpayer_gstin"]).upper()
        if not item.get("financial_year"):
            item["financial_year"] = gst_text(record["financial_year"]) or gstr1_financial_year(
                record["source_period"]
            )
        if not gst_portal_row_in_scope(item, gstin=gstin, financial_year=financial_year):
            continue
        normalized = ensure_gst_invoice_fields(item)
        if normalized.get("invoice_no") or normalized.get("taxable_value") or normalized.get("igst"):
            rows.append(normalized)
    return rows


def sync_gstr1_recon_rows_from_invoices(gstin="", financial_year=""):
    """Rebuild gst_recon_rows GSTR-1 from invoice table (fixes cross-period B2CS collapse)."""
    ctx = gst_portal_resolve_context(gstin, financial_year)
    scope_gstin = ctx.get("gstin")
    scope_fy = ctx.get("financial_year") or gst_portal_default_fy()
    rows = gstr1_rows_from_invoice_table(gstin=scope_gstin, financial_year=scope_fy)
    unique = {}
    for row in rows:
        key = (
            normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")),
            gst_text(row.get("section")),
            gst_text(row.get("gstin")),
            row.get("invoice_no_norm") or normalize_invoice_number(row.get("invoice_no")),
            gst_text(row.get("invoice_date")),
            gst_text(row.get("document_type")),
            round(gst_number(row.get("taxable_value")), 2),
            round(gst_number(row.get("igst")), 2),
            round(gst_number(row.get("cgst")), 2),
            round(gst_number(row.get("sgst")), 2),
        )
        unique[key] = ensure_gst_invoice_fields(row)
    saved = list(unique.values())
    gst_recon_save_rows("GSTR-1", saved, gstin=scope_gstin, financial_year=scope_fy)
    return saved


def gstr1_load_invoices(return_period="", gstin="", financial_year=""):
    period = normalize_gst_recon_period(return_period)
    # Prefer invoice table — it retains every parsed section (B2B/B2CS/CDN/…).
    table_rows = gstr1_rows_from_invoice_table(period, gstin=gstin, financial_year=financial_year)
    if table_rows:
        # Keep recon mirror aligned whenever invoice table has more complete data.
        recon_rows = [
            row for row in gst_recon_load_rows("GSTR-1", gstin=gstin, financial_year=financial_year)
            if not period or normalize_gst_recon_period(row.get("source_period")) == period
        ]
        if len(table_rows) != len(recon_rows):
            sync_gstr1_recon_rows_from_invoices()
        return [ensure_gst_invoice_fields(row) for row in table_rows]
    rows = [
        ensure_gst_invoice_fields(row)
        for row in gst_recon_load_rows("GSTR-1", gstin=gstin, financial_year=financial_year)
    ]
    if period:
        rows = [row for row in rows if normalize_gst_recon_period(row.get("source_period")) == period]
    return rows


def gstr1_save_reconciliation(results, return_period=""):
    stamp = gst_recon_now()
    period = normalize_gst_recon_period(return_period)
    connection = gst_recon_connection()
    try:
        if period:
            connection.execute("DELETE FROM gst_1_reconciliation WHERE return_period=?", (period,))
        else:
            connection.execute("DELETE FROM gst_1_reconciliation")
        for row in results or []:
            connection.execute(
                "INSERT INTO gst_1_reconciliation(return_period,row_json,status,review_action,reconciled_at) VALUES (?,?,?,?,?)",
                (
                    period or gst_text(row.get("source_period")),
                    json.dumps(row, ensure_ascii=False),
                    gst_text(row.get("status")),
                    gst_text(row.get("review_action")),
                    stamp,
                ),
            )
        connection.commit()
    finally:
        connection.close()
    gst_recon_save_results(results, "gstr1_tally")


def gstr1_load_reconciliation(return_period=""):
    period = normalize_gst_recon_period(return_period)
    connection = gst_recon_connection()
    try:
        if period:
            query = "SELECT row_json FROM gst_1_reconciliation WHERE return_period=? ORDER BY id"
            params = (period,)
        else:
            query = "SELECT row_json FROM gst_1_reconciliation ORDER BY id"
            params = ()
        rows = []
        for item in connection.execute(query, params):
            try:
                rows.append(json.loads(item["row_json"]))
            except json.JSONDecodeError:
                continue
        return rows
    finally:
        connection.close()


def tally_sales_voucher_types():
    return {
        gst_text(item.get("name")).lower()
        for item in TALLY_CACHE.get("voucher_types", [])
        if any(token in gst_text(item.get("name")).lower() for token in ("sales", "credit note", "debit note"))
        or any(token in gst_text(item.get("parent", "")).lower() for token in ("sales", "credit note", "debit note"))
    }


def is_tally_sales_voucher(voucher_type, entries, ledger_parents):
    voucher_type = gst_text(voucher_type).lower()
    if voucher_type in tally_sales_voucher_types():
        return True
    if any(token in voucher_type for token in ("sales", "credit note", "debit note")):
        return True
    for name, _ in entries:
        parent = ledger_parents.get(gst_text(name).lower(), "")
        if parent == "sales accounts":
            return True
    return False


def sync_tally_sales_vouchers():
    """
    Read Sales / Credit Note / Debit Note vouchers from the open Tally company.
    Tests connectivity first. Raises ValueError on connection/export failure so
    callers do not overwrite existing TALLY_SALES rows with zeros.
    """
    probe = tally_test_connection(timeout=15)
    if not probe.get("ok"):
        raise ValueError(
            probe.get("error")
            or "TallyPrime connection test failed before sales sync."
        )
    cache = sync_tally()
    ledger_lookup = {
        gst_text(item.get("name")).casefold(): gst_text(item.get("gstin"))
        for item in cache.get("ledgers", [])
    }
    ledger_parents = {
        gst_text(item.get("name")).lower(): normalize_tally_ledger_parent(item.get("parent"))
        for item in cache.get("ledgers", [])
    }
    # Full voucher export is heavy — use the same long timeout as purchase sync.
    xml = tally_collection_xml(
        "Voucher", "Voucher",
        ["Date", "VoucherTypeName", "VoucherNumber", "Reference", "ReferenceDate",
         "PartyLedgerName", "PartyGSTIN", "PlaceOfSupply", "AllLedgerEntries.LedgerName",
         "AllLedgerEntries.Amount", "AllInventoryEntries.StockItemName",
         "AllInventoryEntries.HSNCode", "AllInventoryEntries.Amount"],
        timeout=120,
        purpose="sales-voucher-export",
    )
    if not gst_text(xml):
        raise ValueError("TallyPrime returned empty XML for sales voucher export.")
    rows = []
    for match in re.finditer(r"<VOUCHER\b[^>]*>(.*?)</VOUCHER>", xml, re.I | re.S):
        body = match.group(1)
        voucher_type = tag_value(body, "VOUCHERTYPENAME")
        entries = []
        for entry in re.finditer(r"<ALLLEDGERENTRIES\.LIST>(.*?)</ALLLEDGERENTRIES\.LIST>", body, re.I | re.S):
            entry_body = entry.group(1)
            entries.append((
                tag_value(entry_body, "LEDGERNAME"),
                amount(tag_value(entry_body, "AMOUNT")),
            ))
        if not is_tally_sales_voucher(voucher_type, entries, ledger_parents):
            continue
        party = tag_value(body, "PARTYLEDGERNAME")
        gstin = gst_text(tag_value(body, "PARTYGSTIN")).upper()
        if not gstin and party:
            gstin = gst_text(ledger_lookup.get(party.casefold(), "")).upper()
        reference = tag_value(body, "REFERENCE") or tag_value(body, "VOUCHERNUMBER")
        voucher_number = tag_value(body, "VOUCHERNUMBER")
        invoice_date = format_tally_date(tag_value(body, "REFERENCEDATE")) or format_tally_date(tag_value(body, "DATE"))
        tally_date = format_tally_date(tag_value(body, "DATE"))
        place = tag_value(body, "PLACEOFSUPPLY")
        hsn_codes = []
        for inv in re.finditer(r"<ALLINVENTORYENTRIES\.LIST>(.*?)</ALLINVENTORYENTRIES\.LIST>", body, re.I | re.S):
            hsn = tag_value(inv.group(1), "HSNCODE")
            if hsn:
                hsn_codes.append(hsn)
        taxable = igst = cgst = sgst = cess = 0.0
        has_sales_accounts = any(
            ledger_parents.get(gst_text(name).lower(), "") == "sales accounts"
            for name, _ in entries
        )
        vtype_lower = gst_text(voucher_type).lower()
        if "credit note" in vtype_lower:
            document_type = "Credit Note"
            signed = -1
        elif "debit note" in vtype_lower:
            document_type = "Debit Note"
            signed = 1
        else:
            document_type = "Sales Invoice"
            signed = 1
        # IGST only: official Tally GSTR-1 Output IGST mapping (not Purchase/Input IGST).
        igst = extract_tally_output_igst_signed(
            entries,
            ledger_parents,
            has_sales_accounts,
            signed=1,  # apply voucher sign below with CGST/SGST
            audit_ctx=f"voucher={gst_text(voucher_type)}/{gst_text(tag_value(body, 'VOUCHERNUMBER'))}",
        )
        for name, value in entries:
            lname = gst_text(name).lower()
            parent = ledger_parents.get(lname, "")
            abs_val = abs(value)
            if parent == "sales accounts":
                taxable += abs_val
            elif parent.endswith("duties & taxes") or "output" in lname or "cgst" in lname or "sgst" in lname or "cess" in lname:
                # CGST / SGST / CESS unchanged — do not alter their mapping.
                if "input" in lname:
                    continue
                if "igst" in lname:
                    continue  # IGST handled by extract_tally_output_igst_signed
                if "cgst" in lname:
                    cgst += abs_val
                elif "sgst" in lname or "utgst" in lname:
                    sgst += abs_val
                elif "cess" in lname:
                    cess += abs_val
        taxable = round(taxable * signed, 2)
        igst = round(gst_number(igst) * signed, 2)
        cgst, sgst, cess = round(cgst * signed, 2), round(sgst * signed, 2), round(cess * signed, 2)
        invoice_value = round(taxable + igst + cgst + sgst + cess, 2)
        if invoice_value == 0 and taxable == 0:
            continue
        gst_rate = gst_rate_for_values(abs(taxable), abs(igst), abs(cgst), abs(sgst), abs(cess))
        rows.append({
            "gstin": gstin,
            "party_name": party,
            "party_ledger": party,
            "invoice_no": reference,
            "invoice_no_norm": normalize_invoice_number(reference),
            "invoice_date": invoice_date,
            "tally_entry_date": tally_date,
            "voucher_number": voucher_number,
            "voucher_type": voucher_type,
            "taxable_value": taxable,
            "igst": igst,
            "cgst": cgst,
            "sgst": sgst,
            "cess": cess,
            "gst_rate": gst_rate,
            "invoice_value": invoice_value,
            "document_type": document_type,
            "place_of_supply": place,
            "hsn_code": ", ".join(dict.fromkeys(hsn_codes)),
            "is_b2c": not bool(gstin),
            "source": "Tally",
            "source_period": normalize_gst_recon_period(invoice_date),
            "return_type": "TALLY_SALES",
            "igst_xml_tag": "ALLLEDGERENTRIES.LIST/LEDGERNAME+AMOUNT",
            "igst_mapping": "tally_gstr1_output_igst",
        })
    unique = {}
    for row in rows:
        safe = ensure_gst_invoice_fields(row)
        key = (
            safe.get("gstin"),
            safe.get("invoice_no_norm") or normalize_invoice_number(safe.get("invoice_no")),
            safe.get("invoice_date"),
            safe.get("document_type"),
            safe.get("voucher_number"),
        )
        unique[key] = safe
    saved = list(unique.values())
    totals = gstr1_tax_totals(saved)
    tally_log(
        f"sales-sync | company={cache.get('company', '')} | vouchers={len(saved)} | "
        f"taxable={totals.get('taxable_value', 0):.2f} | output_gst={totals.get('output_gst', 0):.2f}"
    )
    tally_log(
        f"igst-map | FINAL UI | xml_tag=ALLLEDGERENTRIES.LIST/LEDGERNAME+AMOUNT | "
        f"raw_mapped_output_igst={totals.get('igst', 0):.2f} | "
        f"final_value_displayed_in_ui={totals.get('igst', 0):.2f} | "
        f"cgst_unchanged={totals.get('cgst', 0):.2f} | sgst_unchanged={totals.get('sgst', 0):.2f}"
    )
    return {
        "ok": True,
        "company": cache.get("company", "") or probe.get("company", ""),
        "rows": saved,
        "count": len(saved),
        "synced_at": gst_recon_now(),
        "connection": probe,
        "taxable_value": totals.get("taxable_value", 0),
        "igst": totals.get("igst", 0),
        "cgst": totals.get("cgst", 0),
        "sgst": totals.get("sgst", 0),
        "output_gst": totals.get("output_gst", 0),
        "igst_xml_tag": "ALLLEDGERENTRIES.LIST/LEDGERNAME+AMOUNT",
        "igst_final_ui": totals.get("igst", 0),
        "url": TALLY_HTTP_URL,
    }


def repair_stored_tally_sales_igst(rows=None, persist=True):
    """
    One-shot IGST mapping repair for already-synced TALLY_SALES rows.

    Purchase-side Debit Notes were stored with Output IGST = Purchase IGST + Input IGST
    (e.g. 529200) while official Tally GSTR-1 Output IGST = 0. Outward Output IGST
    always rides with Sales Accounts taxable; taxable==0 ⇒ not GSTR-1 Output IGST.
    CGST / SGST fields are left unchanged.
    """
    source = list(rows) if rows is not None else gst_recon_load_rows("TALLY_SALES")
    xml_tag = "ALLLEDGERENTRIES.LIST/LEDGERNAME+AMOUNT"
    fixed = []
    changed = 0
    for row in source:
        item = dict(row)
        raw_igst = gst_number(item.get("igst"))
        taxable = gst_number(item.get("taxable_value"))
        if abs(raw_igst) > 0 and abs(taxable) == 0:
            tally_log(
                f"igst-map | repair-stored | xml_tag={xml_tag} | "
                f"voucher={gst_text(item.get('voucher_type'))}/{gst_text(item.get('voucher_number') or item.get('invoice_no'))} | "
                f"raw_value_returned={raw_igst:.2f} | reason=no_sales_accounts_taxable | "
                f"final_value_displayed_in_ui=0.00"
            )
            item["igst"] = 0.0
            item["invoice_value"] = round(
                taxable
                + gst_number(item.get("cgst"))
                + gst_number(item.get("sgst"))
                + gst_number(item.get("cess")),
                2,
            )
            item["igst_xml_tag"] = xml_tag
            item["igst_mapping"] = "tally_gstr1_output_igst"
            changed += 1
        fixed.append(ensure_gst_invoice_fields(item))
    totals = gstr1_tax_totals(fixed)
    tally_log(
        f"igst-map | FINAL UI | xml_tag={xml_tag} | "
        f"raw_mapped_output_igst={totals.get('igst', 0):.2f} | "
        f"final_value_displayed_in_ui={totals.get('igst', 0):.2f} | "
        f"rows_repaired={changed} | cgst_unchanged={totals.get('cgst', 0):.2f} | "
        f"sgst_unchanged={totals.get('sgst', 0):.2f}"
    )
    if persist and fixed:
        gst_recon_save_rows("TALLY_SALES", fixed)
        meta = dict(gst_recon_get_meta("tally_sales_sync", {}) or {})
        meta.update({
            "igst": totals.get("igst", 0),
            "cgst": totals.get("cgst", 0),
            "sgst": totals.get("sgst", 0),
            "output_gst": totals.get("output_gst", 0),
            "igst_xml_tag": xml_tag,
            "igst_final_ui": totals.get("igst", 0),
            "igst_mapping_repaired": True,
        })
        gst_recon_set_meta("tally_sales_sync", meta)
    return {"rows": fixed, "changed": changed, "totals": totals, "igst_xml_tag": xml_tag}


def gstr1_tax_totals(rows):
    totals = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0, "invoice_value": 0.0, "invoices": 0}
    for row in rows or []:
        safe = ensure_gst_invoice_fields(row)
        totals["invoices"] += 1
        for key in ("taxable_value", "igst", "cgst", "sgst", "cess", "invoice_value"):
            totals[key] = round(totals[key] + gst_number(safe.get(key)), 2)
    totals["output_gst"] = round(totals["igst"] + totals["cgst"] + totals["sgst"] + totals["cess"], 2)
    return totals


def _gstr1_amount_diffs(left, right):
    return {
        field: round(gst_number(right.get(field)) - gst_number(left.get(field)), 2)
        for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
    }


def _gstr1_match_status(row_g1, row_tally, tolerance):
    """Return detailed Phase-2 status; never force Exact Match on uncertain rows."""
    if not row_g1 or not row_tally:
        return "Possible Match", {}
    diffs = _gstr1_amount_diffs(row_g1, row_tally)
    g1_gstin = gst_text(row_g1.get("gstin")).upper()
    t_gstin = gst_text(row_tally.get("gstin")).upper()
    g1_no = gst_text(row_g1.get("invoice_no_norm") or normalize_invoice_number(row_g1.get("invoice_no")))
    t_no = gst_text(row_tally.get("invoice_no_norm") or normalize_invoice_number(row_tally.get("invoice_no")))
    g1_date = gst_text(row_g1.get("invoice_date")).replace("/", "-")
    t_date = gst_text(row_tally.get("invoice_date")).replace("/", "-")
    tax_fields = ("igst", "cgst", "sgst", "cess")
    value_fields = ("invoice_value", "taxable_value")
    tax_ok = all(abs(diffs[field]) <= tolerance for field in tax_fields)
    value_ok = all(abs(diffs[field]) <= tolerance for field in value_fields)
    # B2C: GSTIN may be blank on both sides
    both_b2c = bool(row_g1.get("is_b2c") or row_tally.get("is_b2c") or (not g1_gstin and not t_gstin))
    if not both_b2c and g1_gstin and t_gstin and g1_gstin != t_gstin:
        return "GSTIN Difference", diffs
    if g1_no != t_no:
        return "Invoice Number Difference", diffs
    if g1_date and t_date and g1_date != t_date:
        return "Date Difference", diffs
    if not tax_ok and value_ok:
        return "Tax Difference", diffs
    if not value_ok and tax_ok:
        return "Value Difference", diffs
    if not value_ok or not tax_ok:
        return "Value Difference" if abs(diffs["taxable_value"]) >= abs(diffs["igst"]) else "Tax Difference", diffs
    if (not both_b2c) and ((g1_gstin and not t_gstin) or (t_gstin and not g1_gstin)):
        return "Possible Match", diffs
    return "Exact Match", diffs


def gstr1_match_doc_group(row):
    return gst_document_bucket(row)


def reconcile_gstr1_tally(rows_gstr1, rows_tally, tolerance=1.0, return_period=""):
    """B2B and B2C matching kept separate; signed document-type net GST vs Tally."""
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    period = normalize_gst_recon_period(return_period)
    rows_gstr1 = [ensure_gst_invoice_fields(row) for row in rows_gstr1 or []]
    rows_tally = [ensure_gst_invoice_fields(row) for row in rows_tally or []]

    def in_period(row):
        if not period:
            return True
        return normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period

    left_rows = [row for row in (rows_gstr1 or []) if in_period(row)]
    right_rows = [row for row in (rows_tally or []) if in_period(row)]

    def bucket_key(row, b2c=False):
        inv = gst_text(row.get("invoice_no_norm") or normalize_invoice_number(row.get("invoice_no")))
        date = gst_text(row.get("invoice_date")).replace("/", "-")
        doc = gstr1_match_doc_group(row)
        if b2c or row.get("is_b2c") or not gst_text(row.get("gstin")):
            return ("B2C", inv, date, doc, round(gst_number(row.get("taxable_value")), 2))
        return ("B2B", gst_text(row.get("gstin")).upper(), inv, date, doc)

    left_b2b, left_b2c, right_b2b, right_b2c = {}, {}, {}, {}
    for row in left_rows:
        target = left_b2c if (row.get("is_b2c") or not row.get("gstin")) else left_b2b
        target.setdefault(bucket_key(row, b2c=row.get("is_b2c") or not row.get("gstin")), []).append(row)
    for row in right_rows:
        target = right_b2c if (row.get("is_b2c") or not row.get("gstin")) else right_b2b
        target.setdefault(bucket_key(row, b2c=row.get("is_b2c") or not row.get("gstin")), []).append(row)

    results = []

    def consume(left_map, right_map):
        for key in sorted(set(left_map) | set(right_map), key=lambda item: str(item)):
            left_list = list(left_map.get(key, []))
            right_list = list(right_map.get(key, []))
            used = set()
            for row_g1 in left_list:
                candidates = [(index, row) for index, row in enumerate(right_list) if index not in used]
                if candidates:
                    def score(pair, base=row_g1):
                        _, candidate = pair
                        date_penalty = 0 if gst_text(base.get("invoice_date")).replace("/", "-") == gst_text(candidate.get("invoice_date")).replace("/", "-") else 1000000
                        amount_penalty = sum(
                            abs(gst_number(candidate.get(field)) - gst_number(base.get(field)))
                            for field in ("invoice_value", "taxable_value", "igst", "cgst", "sgst", "cess")
                        )
                        return date_penalty + amount_penalty
                    right_index, row_t = min(candidates, key=score)
                    used.add(right_index)
                    status, diffs = _gstr1_match_status(row_g1, row_t, tolerance)
                    if len(left_list) > 1:
                        same_value = [
                            other for other in left_list
                            if other is not row_g1 and abs(gst_number(other.get("taxable_value")) - gst_number(row_g1.get("taxable_value"))) <= tolerance
                        ]
                        if same_value:
                            status = "Duplicate"
                    results.append({
                        **row_g1,
                        "status": status,
                        "gstr1": row_g1,
                        "tally": row_t,
                        "differences": diffs,
                        "party_name": gst_text(row_g1.get("party_name") or row_t.get("party_name")),
                        "voucher_type": gst_text(row_t.get("voucher_type") or row_g1.get("document_type")),
                        "gstr1_taxable": gst_number(row_g1.get("taxable_value")),
                        "tally_taxable": gst_number(row_t.get("taxable_value")),
                        "gstr1_igst": gst_number(row_g1.get("igst")),
                        "tally_igst": gst_number(row_t.get("igst")),
                        "gstr1_cgst": gst_number(row_g1.get("cgst")),
                        "tally_cgst": gst_number(row_t.get("cgst")),
                        "gstr1_sgst": gst_number(row_g1.get("sgst")),
                        "tally_sgst": gst_number(row_t.get("sgst")),
                        "gstr1_cess": gst_number(row_g1.get("cess")),
                        "tally_cess": gst_number(row_t.get("cess")),
                        "total_difference": round(sum(abs(v) for v in diffs.values()), 2),
                        "review_action": "",
                        "match_side": "B2C" if (row_g1.get("is_b2c") or not row_g1.get("gstin")) else "B2B",
                    })
                else:
                    results.append({
                        **row_g1,
                        "status": "Missing in Tally",
                        "gstr1": row_g1,
                        "tally": None,
                        "differences": {},
                        "party_name": gst_text(row_g1.get("party_name")),
                        "voucher_type": gst_text(row_g1.get("document_type")),
                        "gstr1_taxable": gst_number(row_g1.get("taxable_value")),
                        "tally_taxable": 0,
                        "gstr1_igst": gst_number(row_g1.get("igst")),
                        "tally_igst": 0,
                        "gstr1_cgst": gst_number(row_g1.get("cgst")),
                        "tally_cgst": 0,
                        "gstr1_sgst": gst_number(row_g1.get("sgst")),
                        "tally_sgst": 0,
                        "gstr1_cess": gst_number(row_g1.get("cess")),
                        "tally_cess": 0,
                        "total_difference": round(abs(gst_number(row_g1.get("invoice_value"))), 2),
                        "review_action": "",
                        "match_side": "B2C" if (row_g1.get("is_b2c") or not row_g1.get("gstin")) else "B2B",
                    })
            for index, row_t in enumerate(right_list):
                if index in used:
                    continue
                results.append({
                    **row_t,
                    "status": "Missing in GSTR-1",
                    "gstr1": None,
                    "tally": row_t,
                    "differences": {},
                    "party_name": gst_text(row_t.get("party_name") or row_t.get("party_ledger")),
                    "voucher_type": gst_text(row_t.get("voucher_type") or row_t.get("document_type")),
                    "gstr1_taxable": 0,
                    "tally_taxable": gst_number(row_t.get("taxable_value")),
                    "gstr1_igst": 0,
                    "tally_igst": gst_number(row_t.get("igst")),
                    "gstr1_cgst": 0,
                    "tally_cgst": gst_number(row_t.get("cgst")),
                    "gstr1_sgst": 0,
                    "tally_sgst": gst_number(row_t.get("sgst")),
                    "gstr1_cess": 0,
                    "tally_cess": gst_number(row_t.get("cess")),
                    "total_difference": round(abs(gst_number(row_t.get("invoice_value"))), 2),
                    "review_action": "",
                    "match_side": "B2C" if (row_t.get("is_b2c") or not row_t.get("gstin")) else "B2B",
                })

    consume(left_b2b, right_b2b)
    consume(left_b2c, right_b2c)

    # Soft possible matches for leftover Missing pairs by invoice number only
    missing_tally = [row for row in results if row.get("status") == "Missing in Tally"]
    missing_gstr1 = [row for row in results if row.get("status") == "Missing in GSTR-1"]
    used_missing = set()
    soft_rows = []
    remove_ids = set()
    for left in missing_tally:
        for right in missing_gstr1:
            rid = id(right)
            if rid in used_missing:
                continue
            if gst_text(left.get("invoice_no_norm") or normalize_invoice_number(left.get("invoice_no"))) != gst_text(
                right.get("invoice_no_norm") or normalize_invoice_number(right.get("invoice_no"))
            ):
                continue
            if gstr1_match_doc_group(left) != gstr1_match_doc_group(right):
                continue
            if left.get("match_side") != right.get("match_side"):
                continue
            status, diffs = _gstr1_match_status(left.get("gstr1") or left, right.get("tally") or right, tolerance)
            # Cross-bucket recovery: keep Exact Match when values truly match; otherwise keep detailed status.
            # Only use Possible Match when identity is weak (blank GSTIN asymmetry already handled).
            base = left.get("gstr1") or left
            tally = right.get("tally") or right
            soft_rows.append({
                **base,
                "status": status,
                "gstr1": base,
                "tally": tally,
                "differences": diffs,
                "party_name": gst_text(base.get("party_name") or tally.get("party_name")),
                "voucher_type": gst_text(tally.get("voucher_type") or base.get("document_type")),
                "gstr1_taxable": gst_number(base.get("taxable_value")),
                "tally_taxable": gst_number(tally.get("taxable_value")),
                "gstr1_igst": gst_number(base.get("igst")),
                "tally_igst": gst_number(tally.get("igst")),
                "gstr1_cgst": gst_number(base.get("cgst")),
                "tally_cgst": gst_number(tally.get("cgst")),
                "gstr1_sgst": gst_number(base.get("sgst")),
                "tally_sgst": gst_number(tally.get("sgst")),
                "gstr1_cess": gst_number(base.get("cess")),
                "tally_cess": gst_number(tally.get("cess")),
                "total_difference": round(sum(abs(v) for v in diffs.values()), 2),
                "review_action": "",
                "match_side": left.get("match_side") or "B2B",
            })
            remove_ids.add(id(left))
            remove_ids.add(rid)
            used_missing.add(rid)
            break
    if soft_rows:
        results = [row for row in results if id(row) not in remove_ids] + soft_rows

    counts = {}
    for row in results:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
        row["document_bucket"] = gst_document_bucket(row.get("gstr1") or row.get("tally") or row)
    document_summary = build_signed_reconciliation_pack(
        left_rows, right_rows, "GSTR-1", "Tally Sales", tolerance
    )
    return results, counts, document_summary


def build_sales_recon_dashboard(rows_gstr1, rows_tally, results=None, tolerance=1.0, return_period=""):
    raw_period = gst_text(return_period)
    all_periods = is_gst_all_periods_selection(raw_period) or not normalize_gst_recon_period(raw_period)
    period = "" if all_periods else normalize_gst_recon_period(return_period)
    document_summary = None
    # ALL / FY view: aggregate Apr-25…Mar-26 only (keeps monthly math unchanged).
    fy_keys = {key for key, _ in gst_fy_2025_26_periods()}
    scoped_g1 = list(rows_gstr1 or [])
    scoped_tally = list(rows_tally or [])
    if all_periods:
        fy_g1 = [
            row for row in scoped_g1
            if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) in fy_keys
        ]
        fy_tally = [
            row for row in scoped_tally
            if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) in fy_keys
        ]
        if fy_g1:
            scoped_g1 = fy_g1
        if fy_tally:
            scoped_tally = fy_tally
    if results is None:
        results, counts, document_summary = reconcile_gstr1_tally(scoped_g1, scoped_tally, tolerance, period)
    else:
        counts = {}
        for row in results:
            counts[row.get("status", "")] = counts.get(row.get("status", ""), 0) + 1
        period_g1 = [
            row for row in scoped_g1
            if not period or normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period
        ]
        period_tally = [
            row for row in scoped_tally
            if not period or normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period
        ]
        document_summary = build_signed_reconciliation_pack(
            period_g1, period_tally, "GSTR-1", "Tally Sales", tolerance
        )
    g1_source = scoped_g1 if not period else [
        row for row in scoped_g1
        if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period
    ]
    tally_source = scoped_tally if not period else [
        row for row in scoped_tally
        if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period
    ]
    g1_totals = gstr1_tax_totals(g1_source)
    tally_totals = gstr1_tax_totals(tally_source)
    # Prefer signed Net GST for portal vs Tally comparison cards.
    portal_net = gst_number(((document_summary or {}).get("net") or {}).get("portal_net_gst"))
    tally_net = gst_number(((document_summary or {}).get("net") or {}).get("tally_net_gst"))
    if document_summary:
        g1_totals = dict(g1_totals)
        tally_totals = dict(tally_totals)
        g1_totals["output_gst"] = portal_net
        tally_totals["output_gst"] = tally_net
    exact = counts.get("Exact Match", 0)
    mismatch_keys = (
        "Value Difference", "Tax Difference", "Date Difference", "GSTIN Difference",
        "Invoice Number Difference", "Possible Match", "Duplicate",
    )
    mismatch = sum(counts.get(key, 0) for key in mismatch_keys)
    missing_tally = counts.get("Missing in Tally", 0)
    missing_g1 = counts.get("Missing in GSTR-1", 0)
    output_diff = {
        "taxable_value": round(tally_totals["taxable_value"] - g1_totals["taxable_value"], 2),
        "igst": round(tally_totals["igst"] - g1_totals["igst"], 2),
        "cgst": round(tally_totals["cgst"] - g1_totals["cgst"], 2),
        "sgst": round(tally_totals["sgst"] - g1_totals["sgst"], 2),
        "cess": round(tally_totals["cess"] - g1_totals["cess"], 2),
        "output_gst": round(tally_totals["output_gst"] - g1_totals["output_gst"], 2),
    }
    cards = {
        "gstr1_invoice_count": g1_totals["invoices"],
        "tally_sales_count": tally_totals["invoices"],
        "exact_match": exact,
        "mismatch": mismatch,
        "missing_in_tally": missing_tally,
        "missing_in_gstr1": missing_g1,
        "gstr1_taxable": g1_totals["taxable_value"],
        "tally_taxable": tally_totals["taxable_value"],
        "output_gst_difference": output_diff["output_gst"],
        "portal_net_gst": portal_net,
        "tally_net_gst": tally_net,
        "net_gst_matched": bool(((document_summary or {}).get("net") or {}).get("matched")),
    }
    output_summary = {
        "gstr1": {
            "igst": g1_totals["igst"], "cgst": g1_totals["cgst"], "sgst": g1_totals["sgst"],
            "cess": g1_totals["cess"], "output_gst": g1_totals["output_gst"],
            "taxable_value": g1_totals["taxable_value"],
        },
        "tally": {
            "igst": tally_totals["igst"], "cgst": tally_totals["cgst"], "sgst": tally_totals["sgst"],
            "cess": tally_totals["cess"], "output_gst": tally_totals["output_gst"],
            "taxable_value": tally_totals["taxable_value"],
        },
        "difference": output_diff,
    }
    period_breakdown = build_gstr1_period_breakdown(scoped_g1) if all_periods else None
    return {
        "cards": cards,
        "counts": counts,
        "output_summary": output_summary,
        "document_summary": document_summary,
        "rows": results,
        "return_period": "ALL" if all_periods else period,
        "period_mode": "fy_all" if all_periods else "month",
        "period_breakdown": period_breakdown,
    }


def filter_gst_rows_by_period(rows, return_period=""):
    period = normalize_gst_recon_period(return_period)
    if not period:
        return list(rows or [])
    filtered = []
    for row in rows or []:
        row_period = normalize_gst_recon_period(
            row.get("source_period")
            or row.get("gstr2b_period")
            or row.get("return_period")
            or row.get("invoice_date")
        )
        if row_period == period:
            filtered.append(row)
    return filtered


def gstr3b_imported_for_period(return_period="", gstin="", financial_year=""):
    period = normalize_gst_recon_period(return_period)
    ctx = gst_portal_resolve_context(gstin, financial_year)
    scope_gstin = ctx.get("gstin")
    scope_fy = ctx.get("financial_year") or gst_portal_default_fy()
    imported = set(gstr3b_list_imported_periods(gstin=scope_gstin, financial_year=scope_fy))
    if period and period not in imported:
        return False, {}
    # Authoritative source is gst_3b_summary (via gstr3b_load_summary), never stale meta.
    summary = gstr3b_load_summary(period) if period else gstr3b_load_summary()
    if not summary:
        return False, {}
    if not period:
        return True, summary
    summary_period = normalize_gst_recon_period(summary.get("return_period"))
    periods = summary.get("periods") or summary.get("net_periods") or {}
    period_keys = {normalize_gst_recon_period(key) for key in periods.keys()}
    if summary_period == period or period in period_keys:
        return True, summary
    # Monthly loader may return a single-month summary without periods map.
    if period and period in imported and gstr3b_load_summary_for_period(period):
        return True, summary
    return False, {}


def build_gst_recon_overview(return_period="", tolerance=1.0, gstin="", financial_year=""):
    """
    Period-scoped overview. Import badges and GST totals come from the same
    SQLite sources as the GSTR-2B / GSTR-1 / GSTR-3B tabs (not session flags).
    Blank / ALL period = full FY aggregate of whatever is stored for the active GSTIN+FY.
    """
    period = normalize_gst_recon_period(return_period)
    tolerance = max(0.0, min(gst_number(tolerance), 1000.0))
    ctx = gst_portal_resolve_context(gstin, financial_year)
    scope_gstin = ctx.get("gstin")
    scope_fy = ctx.get("financial_year") or gst_portal_default_fy()

    rows_2b_all = gst_recon_load_rows("GSTR-2B", gstin=scope_gstin, financial_year=scope_fy)
    rows_g1_all = gstr1_load_invoices(gstin=scope_gstin, financial_year=scope_fy)
    rows_2b = filter_gst_rows_by_period(rows_2b_all, period) if period else list(rows_2b_all or [])
    rows_g1 = filter_gst_rows_by_period(rows_g1_all, period) if period else list(rows_g1_all or [])
    rows_purchase = (
        filter_gst_rows_by_period(gst_recon_load_rows("TALLY_PURCHASE"), period)
        if period
        else gst_recon_load_rows("TALLY_PURCHASE")
    )
    rows_sales = (
        filter_gst_rows_by_period(gst_recon_load_rows("TALLY_SALES"), period)
        if period
        else gst_recon_load_rows("TALLY_SALES")
    )

    # Single source of truth: actual imported portal records in the database.
    gstr2b_imported = len(rows_2b) > 0
    gstr1_imported = len(rows_g1) > 0
    if period:
        gstr3b_data_ok, gstr3b = gstr3b_imported_for_period(
            period, gstin=scope_gstin, financial_year=scope_fy
        )
    else:
        gstr3b = gstr3b_load_summary() or {}
        imported_periods = gstr3b.get("imported_periods") or gstr3b_list_imported_periods(
            gstin=scope_gstin, financial_year=scope_fy
        )
        # Ignore unscoped aggregate leftovers when this GSTIN+FY has no imported months.
        if not imported_periods:
            gstr3b = {}
            gstr3b_data_ok = False
        else:
            gstr3b_data_ok = bool(
                gstr3b.get("periods") or gstr3b.get("net_periods") or gstr3b.get("return_period") or imported_periods
            )
            if gstr3b_data_ok and not gstr3b:
                gstr3b = gstr3b_load_summary() or {}
    gstr3b_imported = bool(gstr3b_data_ok)
    tally_connected = bool(TALLY_CACHE.get("connected")) or bool(rows_purchase or rows_sales) or bool(
        gst_recon_get_meta("tally_sync") or gst_recon_get_meta("tally_sales_sync")
    )

    # Purchase side — only when GSTR-2B rows exist for the selected scope.
    purchase = {
        "imported": gstr2b_imported,
        "status": "Imported" if gstr2b_imported else "Not Imported",
        "gstr2b_available_itc": None,
        "gstr2b_itc": None,
        "purchase_mismatch": None,
        "missing_purchase": None,
        "matched": None,
        "gstr2b_count": len(rows_2b) if gstr2b_imported else 0,
        "tally_purchase_count": len(rows_purchase),
        "ready_to_reconcile": bool(gstr2b_imported and rows_purchase),
    }
    itc = {}
    if gstr2b_imported:
        purchase_results = gst_recon_load_results("2b_tally") if rows_purchase else []
        purchase_counts = {}
        for row in purchase_results:
            purchase_counts[row.get("status", "")] = purchase_counts.get(row.get("status", ""), 0) + 1
        itc = build_itc_dashboard(rows_2b, gstr3b if gstr3b_imported else {}, rows_purchase, tolerance)
        itc_summary = (itc or {}).get("summary") or {}
        purchase.update({
            "gstr2b_available_itc": itc_summary.get("available_itc", itc_summary.get("gstr2b_itc", 0)),
            "gstr2b_itc": itc_summary.get("gstr2b_itc", 0),
            "purchase_mismatch": sum(
                purchase_counts.get(key, 0)
                for key in ("Amount/Tax Mismatch", "Date Mismatch")
            ) if rows_purchase else None,
            "missing_purchase": (
                purchase_counts.get("Only in GSTR-2B", 0) + purchase_counts.get("Only in Tally", 0)
            ) if rows_purchase else None,
            "matched": purchase_counts.get("Matched", 0) if rows_purchase else None,
        })

    # Sales side — same GSTR-1 invoice store used by the GSTR-1 vs Tally tab.
    sales = {
        "imported": gstr1_imported,
        "status": "Imported" if gstr1_imported else "Not Imported",
        "output_gst": None,
        "gstr1_match": None,
        "gstr1_mismatch": None,
        "missing_sales": None,
        "gstr1_count": len(rows_g1) if gstr1_imported else 0,
        "tally_sales_count": len(rows_sales),
        "output_gst_difference": None,
        "ready_to_reconcile": bool(gstr1_imported and rows_sales),
    }
    sales_dash = {}
    if gstr1_imported and rows_sales:
        sales_results = gstr1_load_reconciliation(period) or None
        sales_dash = build_sales_recon_dashboard(rows_g1, rows_sales, sales_results, tolerance, period or "ALL")
        sales.update({
            "output_gst": sales_dash["output_summary"]["tally"]["output_gst"],
            "gstr1_match": sales_dash["cards"]["exact_match"],
            "gstr1_mismatch": sales_dash["cards"]["mismatch"],
            "missing_sales": sales_dash["cards"]["missing_in_tally"] + sales_dash["cards"]["missing_in_gstr1"],
            "gstr1_count": sales_dash["cards"]["gstr1_invoice_count"],
            "tally_sales_count": sales_dash["cards"]["tally_sales_count"],
            "output_gst_difference": sales_dash["cards"]["output_gst_difference"],
        })
    elif gstr1_imported:
        # Portal imported but Tally sales not synced for scope — show portal totals only.
        g1_totals = gstr1_tax_totals(rows_g1)
        sales["output_gst"] = g1_totals.get("output_gst", 0)
        sales["gstr1_count"] = g1_totals.get("invoices", len(rows_g1))
        sales_results = gstr1_load_reconciliation(period) or []
        if sales_results:
            counts = {}
            for row in sales_results:
                counts[row.get("status", "")] = counts.get(row.get("status", ""), 0) + 1
            sales["gstr1_match"] = counts.get("Exact Match", 0)
            sales["gstr1_mismatch"] = sum(
                counts.get(key, 0)
                for key in counts
                if key and key != "Exact Match"
            )
            sales["missing_sales"] = (
                counts.get("Missing in Tally", 0) + counts.get("Missing in GSTR-1", 0)
            )

    # GSTR-3B side — gst_3b_summary store only.
    gstr3b_block = {
        "imported": gstr3b_imported,
        "status": "Imported" if gstr3b_imported else "Not Imported",
        "output_gst": None,
        "available_itc": None,
        "claimed_itc": None,
        "net_gst_payable": None,
        "net_gst_interest_payable": None,
        "net_gst_late_fee_payable": None,
        "total_cash_payable": None,
        "cash_required": None,
        "books_vs_3b_difference": None,
        "interest": None,
        "late_fee": None,
        "cash_payable_breakdown": None,
        "ready_to_reconcile": bool(gstr3b_imported and (rows_sales or rows_g1 or gstr2b_imported)),
    }
    gstr3b_dash = {}
    if gstr3b_imported and (rows_sales or rows_g1 or gstr2b_imported):
        try:
            gstr3b_dash = reconcile_gstr3b_vs_books(period or "ALL", tolerance)
            cards3 = (gstr3b_dash or {}).get("cards") or {}
            gstr3b_block.update({
                "output_gst": cards3.get("output_gst", 0),
                "available_itc": cards3.get("available_itc", 0),
                "claimed_itc": cards3.get("claimed_itc", 0),
                "net_gst_payable": cards3.get("net_gst_payable", 0),
                "net_gst_interest_payable": cards3.get("net_gst_interest_payable", cards3.get("interest", 0)),
                "net_gst_late_fee_payable": cards3.get("net_gst_late_fee_payable", cards3.get("late_fee", 0)),
                "total_cash_payable": cards3.get("total_cash_payable", 0),
                "cash_required": cards3.get("cash_required", 0),
                "books_vs_3b_difference": cards3.get("books_vs_3b_difference", 0),
                "interest": cards3.get("interest", 0),
                "late_fee": cards3.get("late_fee", 0),
                "cash_payable_breakdown": (gstr3b_dash or {}).get("cash_payable_breakdown"),
            })
        except Exception:
            gstr3b_dash = {}
    elif gstr3b_imported:
        # Portal 3B present; books sides empty — still expose portal headline figures.
        claimed = tax_bucket_from((gstr3b or {}).get("itc_claimed") or (gstr3b or {}).get("net_itc") or gstr3b)
        outward = tax_bucket_from((gstr3b or {}).get("outward_supplies") or gstr3b)
        interest_amt = gst_number(((gstr3b or {}).get("interest") or {}).get("output_gst"))
        late_amt = gst_number(((gstr3b or {}).get("late_fee") or {}).get("output_gst"))
        tax_only = round(max(0.0, outward.get("output_gst", 0) - claimed.get("output_gst", 0)), 2)
        breakdown = build_gstr3b_cash_payable_breakdown(period or "ALL")
        gstr3b_block.update({
            "output_gst": outward.get("output_gst", 0),
            "available_itc": claimed.get("output_gst", 0),
            "claimed_itc": claimed.get("output_gst", 0),
            "net_gst_payable": tax_only,
            "net_gst_interest_payable": interest_amt,
            "net_gst_late_fee_payable": late_amt,
            "total_cash_payable": round(tax_only + interest_amt + late_amt, 2),
            "interest": interest_amt,
            "late_fee": late_amt,
            "cash_payable_breakdown": breakdown,
        })

    import_status = {
        "return_period": period or "ALL",
        "tally_connected": tally_connected,
        "gstr2b": "Imported" if gstr2b_imported else "Not Imported",
        "gstr1": "Imported" if gstr1_imported else "Not Imported",
        "gstr3b": "Imported" if gstr3b_imported else "Not Imported",
        "gstr2b_imported": gstr2b_imported,
        "gstr1_imported": gstr1_imported,
        "gstr3b_imported": gstr3b_imported,
        "gstr2b_count": len(rows_2b),
        "gstr1_count": len(rows_g1),
        "gstr3b_periods": (gstr3b or {}).get("imported_periods") or gstr3b_list_imported_periods(),
        "session_imports": gst_session_get_imports(),
    }
    return {
        "return_period": period or "ALL",
        "import_status": import_status,
        "purchase": purchase,
        "sales": sales,
        "gstr3b": gstr3b_block,
        "sales_dashboard": sales_dash if gstr1_imported and rows_sales else {},
        "itc_dashboard": itc if gstr2b_imported else {},
        "gstr3b_dashboard": gstr3b_dash if gstr3b_imported else {},
    }



def apply_gstr1_review_action(results, invoice_no, gstin, action, note=""):
    action = gst_text(action)
    allowed = {"Review", "Accept Match", "Mark Corrected", "Ignore", "View Tally Voucher", "View Portal Record"}
    if action not in allowed:
        raise ValueError("Unsupported review action.")
    invoice_norm = normalize_invoice_number(invoice_no)
    gstin = gst_text(gstin).upper()
    updated = []
    changed = 0
    for row in results or []:
        item = dict(row)
        same = normalize_invoice_number(item.get("invoice_no")) == invoice_norm
        if gstin:
            same = same and gst_text(item.get("gstin")).upper() == gstin
        if same:
            item["review_action"] = action
            item["review_note"] = gst_text(note)
            if action == "Accept Match" and item.get("status") != "Exact Match":
                item["status"] = "Exact Match"
                item["manual_accept"] = True
            changed += 1
        updated.append(item)
    if not changed:
        raise ValueError("Matching reconciliation row was not found.")
    return updated


def make_gstr1_recon_export(rows, title="GSTR-1 vs Tally"):
    """Sales reconciliation export with GSTR-1 / Tally side-by-side columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\\/*?:\[\]]", " ", gst_text(title))[:31] or "GSTR1_Tally"
    headers = [
        "Status", "GSTIN", "Party Name", "Invoice No", "Invoice Date", "Voucher Type", "Document Type",
        "GSTR-1 Taxable", "Tally Taxable", "Taxable Diff",
        "GSTR-1 IGST", "Tally IGST", "GSTR-1 CGST", "Tally CGST",
        "GSTR-1 SGST", "Tally SGST", "GSTR-1 CESS", "Tally CESS", "Total Difference",
        "Review Action", "Match Side",
    ]
    ws.append(headers)
    for row in rows or []:
        g1_tax = gst_number(row.get("gstr1_taxable", row.get("taxable_value")))
        t_tax = gst_number(row.get("tally_taxable"))
        ws.append([
            row.get("status"), row.get("gstin"), row.get("party_name"), row.get("invoice_no"),
            row.get("invoice_date"), row.get("voucher_type"), row.get("document_type"),
            g1_tax, t_tax, round(t_tax - g1_tax, 2),
            gst_number(row.get("gstr1_igst")), gst_number(row.get("tally_igst")),
            gst_number(row.get("gstr1_cgst")), gst_number(row.get("tally_cgst")),
            gst_number(row.get("gstr1_sgst")), gst_number(row.get("tally_sgst")),
            gst_number(row.get("gstr1_cess")), gst_number(row.get("tally_cess")),
            gst_number(row.get("total_difference")),
            row.get("review_action"), row.get("match_side"),
        ])
    fill = PatternFill("solid", fgColor="0B5CAB")
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = fill, Font(bold=True, color="FFFFFF"), Alignment(wrap_text=True)
    for column in range(8, 20):
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            cell[0].number_format = "#,##0.00"
    for index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 14
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# GSTR-3B vs Books / Liability / ITC / Utilisation / Net Payable

TAX_KEYS = ("igst", "cgst", "sgst", "cess")


def empty_tax_bucket(taxable=0.0):
    return {
        "taxable_value": round(gst_number(taxable), 2),
        "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
        "output_gst": 0.0,
    }


def tax_bucket_from(values):
    values = values or {}
    bucket = empty_tax_bucket(values.get("taxable_value"))
    for key in TAX_KEYS:
        bucket[key] = round(gst_number(values.get(key)), 2)
    bucket["output_gst"] = round(sum(bucket[key] for key in TAX_KEYS), 2)
    return bucket


def tax_bucket_diff(left, right):
    left, right = tax_bucket_from(left), tax_bucket_from(right)
    return {
        key: round(gst_number(right.get(key)) - gst_number(left.get(key)), 2)
        for key in ("taxable_value", *TAX_KEYS, "output_gst")
    }


def enrich_gstr3b_from_json(data, base=None):
    """
    Map GST Portal GSTR-3B JSON into normalized buckets.

    Official portal sections (Table 4):
      itc_avl  → ITC available (gross)
      itc_rev  → ITC reversed
      itc_net  → Net ITC Available (= claimed for utilisation)
      itc_inelg → Ineligible ITC

    Claimed ITC must come from itc_net only — never itc_avl + itc_net.
    """
    base = dict(base or {})
    if not isinstance(data, dict):
        return base
    gstin = gst_text(gst_pick(data, "gstin", "GSTIN")).upper()
    period = normalize_gst_recon_period(
        gst_pick(data, "ret_period", "fp", "return period") or base.get("return_period")
    )
    outward = empty_tax_bucket()
    outward_nil = empty_tax_bucket()
    available_gross = empty_tax_bucket()
    claimed = empty_tax_bucket()
    reversed_itc = empty_tax_bucket()
    ineligible = empty_tax_bucket()
    interest = empty_tax_bucket()
    late_fee = empty_tax_bucket()

    def absorb(target, node):
        if not isinstance(node, dict):
            return
        target["taxable_value"] = round(target["taxable_value"] + gst_number(gst_pick(node, "txval", "taxable_value")), 2)
        target["igst"] = round(target["igst"] + gst_number(gst_pick(node, "iamt", "igst")), 2)
        target["cgst"] = round(target["cgst"] + gst_number(gst_pick(node, "camt", "cgst")), 2)
        target["sgst"] = round(target["sgst"] + gst_number(gst_pick(node, "samt", "sgst")), 2)
        target["cess"] = round(target["cess"] + gst_number(gst_pick(node, "csamt", "cess")), 2)

    def walk(node, path=""):
        if isinstance(node, dict):
            lower_path = path.lower()
            for key, value in node.items():
                key_l = str(key).lower()
                next_path = f"{lower_path}.{key_l}" if lower_path else key_l
                # Table 3.1(a)/(b) taxable / zero-rated — keep nil/exempt separate.
                if key_l in {"osup_det", "osup_zero"} or (key_l == "outward" and isinstance(value, dict) and "nil" not in next_path):
                    if isinstance(value, dict):
                        absorb(outward, value)
                    elif isinstance(value, list):
                        for item in value:
                            absorb(outward, item)
                if key_l in {"osup_nil_exmp", "osup_nongst"} or "nil_exmp" in key_l or "nil_exempt" in key_l:
                    if isinstance(value, dict):
                        absorb(outward_nil, value)
                    elif isinstance(value, list):
                        for item in value:
                            absorb(outward_nil, item)
                # Gross available only — do not treat as claimed.
                if key_l in {"itc_avl", "itc_available"} and isinstance(value, list):
                    for item in value:
                        absorb(available_gross, item)
                # Net ITC Available = Claimed ITC for GSTR-3B utilisation.
                if key_l == "itc_net" and isinstance(value, list):
                    for item in value:
                        absorb(claimed, item)
                if key_l in {"itc_rev", "itc_reversed"} and isinstance(value, list):
                    for item in value:
                        absorb(reversed_itc, item)
                if key_l in {"itc_inelg", "itc_ineligible"} and isinstance(value, list):
                    for item in value:
                        absorb(ineligible, item)
                if key_l in {"intr_details", "interest"} and isinstance(value, dict):
                    absorb(interest, value)
                if key_l in {"ltfee_details", "late_fee", "latefee"} and isinstance(value, dict):
                    absorb(late_fee, value)
                walk(value, next_path)
        elif isinstance(node, list):
            for item in node:
                walk(item, path)

    walk(data)
    for bucket in (outward, outward_nil, available_gross, claimed, reversed_itc, ineligible, interest, late_fee):
        bucket["output_gst"] = round(sum(bucket[key] for key in TAX_KEYS), 2)

    # Fallback when JSON omitted itc_net: Net = Available − Reversed (portal formula).
    if not any(claimed[key] for key in TAX_KEYS):
        pdf_net = tax_bucket_from(base.get("net_itc") or {})
        if any(pdf_net[key] for key in TAX_KEYS):
            claimed = pdf_net
        elif any(available_gross[key] for key in TAX_KEYS):
            claimed = empty_tax_bucket(available_gross.get("taxable_value"))
            for key in TAX_KEYS:
                claimed[key] = round(max(0.0, available_gross[key] - reversed_itc[key]), 2)
            claimed["output_gst"] = round(sum(claimed[key] for key in TAX_KEYS), 2)
        else:
            claimed = tax_bucket_from(base)

    if not any(available_gross[key] for key in TAX_KEYS):
        available_gross = tax_bucket_from(base)

    if not any(outward_nil[key] for key in ("taxable_value", *TAX_KEYS)):
        outward_nil = tax_bucket_from(base.get("outward_nil_exempt"))

    return {
        **base,
        "gstin": gstin or base.get("gstin", ""),
        "return_period": period or base.get("return_period", ""),
        "financial_year": gstr1_financial_year(period or base.get("return_period", "")),
        "outward_supplies": outward,
        "outward_nil_exempt": outward_nil,
        "itc_claimed": claimed,
        "itc_available_gross": available_gross,
        "itc_reversed": reversed_itc,
        "itc_ineligible": ineligible,
        "interest": interest,
        "late_fee": late_fee,
        "reverse_charge": tax_bucket_from(base.get("reverse_charge")),
        # Keep net_itc aligned with portal "Net ITC Available" / claimed.
        "net_itc": claimed,
    }


def parse_gstr3b_full(name, raw):
    period = infer_gst_period(name)
    base = parse_gstr3b_totals(name, raw)
    # Prefer period embedded in PDF/JSON over filename when present.
    base["return_period"] = normalize_gst_recon_period(
        base.get("return_period") or period
    )
    base["source_file"] = name
    suffix = Path(name).suffix.lower()
    if suffix == ".json":
        data = json.loads(raw.decode("utf-8-sig"))
        return enrich_gstr3b_from_json(data, base)
    if suffix == ".zip":
        enriched = dict(base)
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            for member in archive.infolist():
                if member.is_dir():
                    continue
                if Path(member.filename).suffix.lower() == ".json":
                    child = parse_gstr3b_full(member.filename, archive.read(member))
                    for key in ("outward_supplies", "outward_nil_exempt", "itc_claimed", "interest", "late_fee", "itc_reversed", "itc_ineligible"):
                        if child.get(key) and not enriched.get(key):
                            enriched[key] = child.get(key)
                    if child.get("gstin"):
                        enriched["gstin"] = child["gstin"]
                    if child.get("return_period"):
                        enriched["return_period"] = child["return_period"]
        enriched.setdefault("outward_supplies", empty_tax_bucket())
        enriched.setdefault("outward_nil_exempt", empty_tax_bucket())
        enriched.setdefault("itc_claimed", tax_bucket_from(enriched.get("net_itc") or enriched))
        enriched.setdefault("interest", empty_tax_bucket())
        enriched.setdefault("late_fee", empty_tax_bucket())
        enriched.setdefault("itc_reversed", empty_tax_bucket())
        enriched.setdefault("itc_ineligible", empty_tax_bucket())
        enriched.setdefault("itc_available_gross", tax_bucket_from(enriched))
        enriched["financial_year"] = gstr1_financial_year(enriched.get("return_period", ""))
        return enriched
    base.setdefault("outward_supplies", empty_tax_bucket())
    base.setdefault("outward_nil_exempt", empty_tax_bucket())
    base.setdefault("itc_claimed", tax_bucket_from(base.get("itc_claimed") or base.get("net_itc") or base))
    base.setdefault(
        "itc_available_gross",
        tax_bucket_from(base.get("itc_available_gross") or base.get("net_itc") or base),
    )
    base.setdefault("itc_reversed", tax_bucket_from(base.get("itc_reversed")))
    base.setdefault("itc_ineligible", tax_bucket_from(base.get("itc_ineligible")))
    base.setdefault("interest", tax_bucket_from(base.get("interest")))
    base.setdefault("late_fee", tax_bucket_from(base.get("late_fee")))
    # Keep net_itc aligned with claimed / Table 4C.
    base["net_itc"] = tax_bucket_from(base.get("itc_claimed") or base.get("net_itc") or base)
    base["financial_year"] = gstr1_financial_year(base.get("return_period", ""))
    return base


def gstr3b_period_sort_key(period):
    digits = normalize_gst_recon_period(period)
    if len(digits) != 6:
        return (99, 9999, digits)
    month, year = int(digits[:2]), int(digits[2:])
    # FY 2025-26 order: Apr..Mar
    fy_order = month - 3 if month >= 4 else month + 9
    return (year if month >= 4 else year - 1, fy_order, digits)


def gstr3b_list_imported_periods(gstin="", financial_year=""):
    """Return distinct imported GSTR-3B periods (MMYYYY), FY-sorted, scoped to GSTIN+FY."""
    ctx = gst_portal_resolve_context(gstin, financial_year)
    want_gstin = gst_text(ctx.get("gstin")).upper()
    want_fy = gst_text(ctx.get("financial_year")) or gst_portal_default_fy()
    period_set = set(gst_fy_period_values(want_fy))
    connection = gst_recon_connection()
    try:
        rows = connection.execute(
            """SELECT s.return_period AS return_period, b.gstin AS gstin, b.financial_year AS financial_year,
                      b.file_name AS file_name
               FROM gst_3b_summary s
               LEFT JOIN gst_3b_import_batches b ON b.id = s.batch_id
               WHERE s.return_period IS NOT NULL AND TRIM(s.return_period) != ''"""
        ).fetchall()
    finally:
        connection.close()
    periods = []
    seen = set()
    for row in rows or []:
        period = normalize_gst_recon_period(row["return_period"] if hasattr(row, "keys") else row[0])
        if not period or period in seen:
            continue
        if period not in period_set:
            continue
        batch_gstin = gst_text(row["gstin"] if hasattr(row, "keys") else "").upper()
        if not batch_gstin and hasattr(row, "keys"):
            batch_gstin = infer_taxpayer_gstin_from_text(row["file_name"])
        if want_gstin and batch_gstin and batch_gstin != want_gstin:
            continue
        seen.add(period)
        periods.append(period)
    periods.sort(key=gstr3b_period_sort_key)
    return periods


def gstr3b_add_tax_buckets(left, right):
    out = empty_tax_bucket()
    left = tax_bucket_from(left)
    right = tax_bucket_from(right)
    for key in ("taxable_value", *TAX_KEYS, "output_gst"):
        out[key] = round(gst_number(left.get(key)) + gst_number(right.get(key)), 2)
    return out


def gstr3b_aggregate_summaries(summaries):
    """Merge month summaries into one FY dataset (Table 3.1 + ITC buckets)."""
    aggregated = {
        "return_period": "ALL",
        "financial_year": "2025-26",
        "outward_supplies": empty_tax_bucket(),
        "outward_nil_exempt": empty_tax_bucket(),
        "itc_claimed": empty_tax_bucket(),
        "itc_available_gross": empty_tax_bucket(),
        "itc_reversed": empty_tax_bucket(),
        "itc_ineligible": empty_tax_bucket(),
        "interest": empty_tax_bucket(),
        "late_fee": empty_tax_bucket(),
        "reverse_charge": empty_tax_bucket(),
        "net_itc": empty_tax_bucket(),
        "gstin": "",
        "source_files": [],
    }
    for summary in summaries or []:
        if not isinstance(summary, dict):
            continue
        if not aggregated.get("gstin"):
            aggregated["gstin"] = gst_text(summary.get("gstin")).upper()
        fy = gst_text(summary.get("financial_year"))
        if fy:
            aggregated["financial_year"] = fy
        source = gst_text(summary.get("source_file"))
        if source:
            aggregated["source_files"].append(source)
        aggregated["outward_supplies"] = gstr3b_add_tax_buckets(
            aggregated["outward_supplies"], summary.get("outward_supplies")
        )
        aggregated["outward_nil_exempt"] = gstr3b_add_tax_buckets(
            aggregated["outward_nil_exempt"], summary.get("outward_nil_exempt")
        )
        claimed = summary.get("itc_claimed") or summary.get("net_itc") or {}
        aggregated["itc_claimed"] = gstr3b_add_tax_buckets(aggregated["itc_claimed"], claimed)
        aggregated["net_itc"] = gstr3b_add_tax_buckets(aggregated["net_itc"], claimed)
        aggregated["itc_available_gross"] = gstr3b_add_tax_buckets(
            aggregated["itc_available_gross"], summary.get("itc_available_gross") or summary
        )
        aggregated["itc_reversed"] = gstr3b_add_tax_buckets(
            aggregated["itc_reversed"], summary.get("itc_reversed")
        )
        aggregated["itc_ineligible"] = gstr3b_add_tax_buckets(
            aggregated["itc_ineligible"], summary.get("itc_ineligible")
        )
        aggregated["interest"] = gstr3b_add_tax_buckets(aggregated["interest"], summary.get("interest"))
        aggregated["late_fee"] = gstr3b_add_tax_buckets(aggregated["late_fee"], summary.get("late_fee"))
        aggregated["reverse_charge"] = gstr3b_add_tax_buckets(
            aggregated["reverse_charge"], summary.get("reverse_charge")
        )
    claimed = aggregated["itc_claimed"]
    aggregated["igst"] = gst_number(claimed.get("igst"))
    aggregated["cgst"] = gst_number(claimed.get("cgst"))
    aggregated["sgst"] = gst_number(claimed.get("sgst"))
    aggregated["cess"] = gst_number(claimed.get("cess"))
    return aggregated


def gstr3b_load_summary_for_period(period):
    """Load one month summary from gst_3b_summary (no meta fallback)."""
    period = normalize_gst_recon_period(period)
    if not period:
        return None
    connection = gst_recon_connection()
    try:
        row = connection.execute(
            "SELECT summary_json FROM gst_3b_summary WHERE return_period=? ORDER BY id DESC LIMIT 1",
            (period,),
        ).fetchone()
        if not row:
            return None
        try:
            return json.loads(row["summary_json"])
        except json.JSONDecodeError:
            return None
    finally:
        connection.close()


def gstr3b_rebuild_period_index():
    """
    Rebuild FY GSTR-3B meta from all stored monthly summaries.
    Never drops previously imported months that still exist in gst_3b_summary.
    When no months remain, delete cached meta so dashboards cannot show stale totals.
    """
    imported = gstr3b_list_imported_periods()
    if not imported:
        gst_recon_delete_meta_keys(
            ("GSTR-3B", "gstr3b_last_import", "gstr3b_dashboard")
        )
        return {
            "return_period": "",
            "periods": {},
            "net_periods": {},
            "outward_periods": {},
            "nil_periods": {},
            "imported_periods": [],
        }
    periods = {}
    net_periods = {}
    outward_periods = {}
    nil_periods = {}
    summaries = []
    for period in imported:
        summary = gstr3b_load_summary_for_period(period)
        if not summary:
            continue
        summaries.append(summary)
        claimed = tax_bucket_from(summary.get("itc_claimed") or summary.get("net_itc") or summary)
        periods[period] = {
            "taxable_value": gst_number(summary.get("taxable_value")),
            "igst": gst_number(claimed.get("igst")),
            "cgst": gst_number(claimed.get("cgst")),
            "sgst": gst_number(claimed.get("sgst")),
            "cess": gst_number(claimed.get("cess")),
        }
        net_periods[period] = {
            "taxable_value": gst_number(claimed.get("taxable_value")),
            "igst": gst_number(claimed.get("igst")),
            "cgst": gst_number(claimed.get("cgst")),
            "sgst": gst_number(claimed.get("sgst")),
            "cess": gst_number(claimed.get("cess")),
            "output_gst": gst_number(claimed.get("output_gst")),
        }
        outward_periods[period] = tax_bucket_from(summary.get("outward_supplies"))
        nil_periods[period] = tax_bucket_from(summary.get("outward_nil_exempt"))
    aggregated = gstr3b_aggregate_summaries(summaries) if summaries else {
        "return_period": "ALL",
        "periods": {},
        "net_periods": {},
    }
    meta = {
        **aggregated,
        "periods": periods,
        "net_periods": net_periods,
        "outward_periods": outward_periods,
        "nil_periods": nil_periods,
        "imported_periods": imported,
        "return_period": "ALL" if len(imported) > 1 else (imported[0] if imported else ""),
    }
    gst_recon_set_meta("GSTR-3B", meta)
    return meta


def gstr3b_save_import_batch(file_name, raw, summary, gstin="", return_period="", rebuild_index=True):
    summary = dict(summary or {})
    digest = hashlib.sha256(raw).hexdigest()
    # Prefer period/GSTIN from the PDF/JSON itself. UI "ALL"/FY must not wipe Table 3.1 mapping.
    file_period = normalize_gst_recon_period(
        summary.get("return_period") or infer_gst_period(file_name)
    )
    caller_period = normalize_gst_recon_period(return_period)
    period = file_period or caller_period
    batch_gstin = gst_text(gstin or summary.get("gstin", "") or infer_taxpayer_gstin_from_text(file_name)).upper()
    summary["return_period"] = period
    summary["gstin"] = batch_gstin
    summary["financial_year"] = gstr1_financial_year(period) or gst_portal_default_fy()
    if batch_gstin:
        gst_portal_set_context(batch_gstin, summary["financial_year"])
    connection = gst_recon_connection()
    try:
        existing = connection.execute(
            "SELECT id FROM gst_3b_import_batches WHERE file_digest=? AND return_period=? AND gstin=?",
            (digest, period, batch_gstin),
        ).fetchone()
        if not existing and period:
            # Repair older imports that stored blank return_period when UI filter was ALL.
            existing = connection.execute(
                "SELECT id FROM gst_3b_import_batches WHERE file_digest=? AND (return_period='' OR return_period IS NULL)",
                (digest,),
            ).fetchone()
        if existing:
            # Refresh stored summary so loader fixes (e.g. Table 3.1(a) outward) apply on re-import.
            connection.execute(
                "UPDATE gst_3b_import_batches SET return_period=?, gstin=?, financial_year=?, file_name=? WHERE id=?",
                (period, batch_gstin, summary["financial_year"], file_name, existing["id"]),
            )
            connection.execute(
                "UPDATE gst_3b_summary SET return_period=?, summary_json=?, imported_at=? WHERE batch_id=?",
                (period, json.dumps(summary, ensure_ascii=False), gst_recon_now(), existing["id"]),
            )
            connection.commit()
            gst_session_mark_imported("GSTR-3B", period)
            meta = gstr3b_rebuild_period_index() if rebuild_index else {
                "imported_periods": gstr3b_list_imported_periods(),
            }
            return {
                "duplicate": True,
                "batch_id": existing["id"],
                "message": "This GSTR-3B file was already imported for the same period.",
                "summary": summary,
                "return_period": period,
                "imported_periods": meta.get("imported_periods") or gstr3b_list_imported_periods(),
            }
        stamp = gst_recon_now()
        cursor = connection.execute(
            """INSERT INTO gst_3b_import_batches
            (gstin, financial_year, return_period, import_date, file_name, record_count, return_type, file_digest)
            VALUES (?,?,?,?,?,?, 'GSTR3B', ?)""",
            (batch_gstin, summary["financial_year"], period, stamp, file_name, 1, digest),
        )
        batch_id = cursor.lastrowid
        connection.execute(
            "INSERT INTO gst_3b_summary(batch_id, return_period, summary_json, imported_at) VALUES (?,?,?,?)",
            (batch_id, period, json.dumps(summary, ensure_ascii=False), stamp),
        )
        connection.commit()
    finally:
        connection.close()
    gst_recon_set_meta("gstr3b_last_import", {
        "batch_id": batch_id, "file_name": file_name, "return_period": period,
        "gstin": batch_gstin, "imported_at": stamp,
    })
    gst_session_mark_imported("GSTR-3B", period)
    meta = gstr3b_rebuild_period_index() if rebuild_index else {
        "imported_periods": gstr3b_list_imported_periods(),
    }
    return {
        "duplicate": False,
        "batch_id": batch_id,
        "summary": summary,
        "return_period": period,
        "imported_periods": meta.get("imported_periods") or gstr3b_list_imported_periods(),
    }


def gstr3b_load_summary(return_period=""):
    raw = gst_text(return_period)
    all_mode = is_gst_all_periods_selection(raw) or not normalize_gst_recon_period(raw)
    if all_mode:
        imported = gstr3b_list_imported_periods()
        if not imported:
            # Do not fall back to gst_recon_meta — that cache survived older clears.
            return None
        summaries = [gstr3b_load_summary_for_period(period) for period in imported]
        summaries = [item for item in summaries if item]
        if not summaries:
            return None
        aggregated = gstr3b_aggregate_summaries(summaries)
        aggregated["imported_periods"] = imported
        aggregated["periods"] = {
            period: tax_bucket_from((item.get("itc_claimed") or item.get("net_itc") or item))
            for period, item in zip(imported, summaries)
        }
        return aggregated

    period = normalize_gst_recon_period(return_period)
    connection = gst_recon_connection()
    try:
        row = connection.execute(
            "SELECT summary_json FROM gst_3b_summary WHERE return_period=? ORDER BY id DESC LIMIT 1",
            (period,),
        ).fetchone()
        if row:
            try:
                return json.loads(row["summary_json"])
            except json.JSONDecodeError:
                return None
        # Recover older rows saved with blank return_period (UI ALL wipe).
        for candidate in connection.execute(
            "SELECT return_period, summary_json FROM gst_3b_summary ORDER BY id DESC LIMIT 80"
        ).fetchall():
            try:
                data = json.loads(candidate["summary_json"] or "{}")
            except json.JSONDecodeError:
                continue
            source = gst_text(data.get("source_file") or "")
            candidate_period = normalize_gst_recon_period(
                data.get("return_period") or candidate["return_period"] or infer_gst_period(source)
            )
            if candidate_period == period or period in source:
                if not data.get("return_period"):
                    data["return_period"] = period
                return data
        # Specific month requested but not imported — do not return another month's meta.
        return None
    finally:
        connection.close()


def save_liability_summary(return_period, source, summary):
    period = normalize_gst_recon_period(return_period)
    stamp = gst_recon_now()
    connection = gst_recon_connection()
    try:
        connection.execute(
            "DELETE FROM gst_liability_summary WHERE return_period=? AND source=?",
            (period, gst_text(source)),
        )
        connection.execute(
            "INSERT INTO gst_liability_summary(return_period, source, summary_json, calculated_at) VALUES (?,?,?,?)",
            (period, gst_text(source), json.dumps(summary, ensure_ascii=False), stamp),
        )
        connection.commit()
    finally:
        connection.close()


def save_itc_claim_summary(return_period, summary):
    period = normalize_gst_recon_period(return_period)
    stamp = gst_recon_now()
    connection = gst_recon_connection()
    try:
        connection.execute("DELETE FROM gst_itc_claim_summary WHERE return_period=?", (period,))
        connection.execute(
            "INSERT INTO gst_itc_claim_summary(return_period, summary_json, calculated_at) VALUES (?,?,?)",
            (period, json.dumps(summary, ensure_ascii=False), stamp),
        )
        connection.commit()
    finally:
        connection.close()


def is_outward_nil_exempt_row(row):
    """Nil-rated / exempt outward: no output tax (matches GSTR-3B Table 3.1(c))."""
    if not isinstance(row, dict):
        return False
    tax = (
        abs(gst_number(row.get("igst")))
        + abs(gst_number(row.get("cgst")))
        + abs(gst_number(row.get("sgst")))
        + abs(gst_number(row.get("cess")))
    )
    if tax > 0.005:
        return False
    rate = gst_number(row.get("gst_rate"))
    # Zero-tax outward turnover (rate 0 / blank) is nil/exempt for Table 3.1(c).
    return abs(gst_number(row.get("taxable_value"))) > 0.005 or rate == 0


def outward_row_drill_fields(row):
    """Compact Books/Tally voucher fields for taxable / nil drill-down."""
    row = row or {}
    tax = round(
        gst_number(row.get("igst"))
        + gst_number(row.get("cgst"))
        + gst_number(row.get("sgst"))
        + gst_number(row.get("cess")),
        2,
    )
    nil = is_outward_nil_exempt_row(row)
    return {
        "invoice_no": gst_text(row.get("invoice_no") or row.get("voucher_number")),
        "invoice_date": gst_text(row.get("invoice_date") or row.get("date") or row.get("tally_entry_date")),
        "source_period": normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date") or row.get("date")),
        "party_name": gst_text(row.get("party_name") or row.get("party_ledger")),
        "gstin": gst_text(row.get("gstin")).upper(),
        "voucher_type": gst_text(row.get("voucher_type") or row.get("document_type")),
        "sales_ledger": gst_text(row.get("sales_ledger") or row.get("ledger") or row.get("item_name")),
        "hsn_code": gst_text(row.get("hsn_code") or row.get("hsn")),
        "gst_rate": gst_number(row.get("gst_rate") if row.get("gst_rate") is not None else row.get("rate")),
        "taxable_value": round(gst_number(row.get("taxable_value")), 2),
        "igst": round(gst_number(row.get("igst")), 2),
        "cgst": round(gst_number(row.get("cgst")), 2),
        "sgst": round(gst_number(row.get("sgst")), 2),
        "cess": round(gst_number(row.get("cess")), 2),
        "output_gst": tax,
        "classification": "Nil/Exempt (zero output tax)" if nil else "Taxable outward",
        "classification_basis": (
            "igst+cgst+sgst+cess = 0 and taxable_value > 0 → GSTR-3B Table 3.1(c) style"
            if nil
            else "output tax > 0 → GSTR-3B Table 3.1(a) style"
        ),
        "source": gst_text(row.get("source") or row.get("return_type") or "Tally"),
    }


def build_outward_classification_drilldown(return_period=""):
    """
    Drill-down for Outward Taxable vs Nil/Exempt classification mismatch.
    Books side: exact Tally (or GSTR-1) vouchers.
    Portal side: imported GSTR-3B Table 3.1(a)/(c) amounts + source file (no invented values).
    """
    raw_period = gst_text(return_period)
    all_mode = is_gst_all_periods_selection(raw_period) or not normalize_gst_recon_period(raw_period)
    selected = "" if all_mode else normalize_gst_recon_period(return_period)
    books = build_books_output_liability(selected)
    sales_rows_all = gst_recon_load_rows("TALLY_SALES") or []
    g1_rows_all = gst_recon_load_rows("GSTR-1") or []
    use_tally = str(books.get("books_source") or "").startswith("Tally")
    source_rows_all = sales_rows_all if use_tally else g1_rows_all
    if selected:
        source_rows = filter_gst_rows_by_period(source_rows_all, selected)
    else:
        source_rows = list(source_rows_all)
    taxable_rows, nil_rows = split_outward_taxable_and_nil_rows(source_rows)
    taxable_rows = sorted(taxable_rows, key=lambda r: (-abs(gst_number(r.get("taxable_value"))), gst_text(r.get("invoice_no"))))
    nil_rows = sorted(nil_rows, key=lambda r: (-abs(gst_number(r.get("taxable_value"))), gst_text(r.get("invoice_no"))))

    period_portal = []
    anomaly_notes = []
    for period, label in gst_fy_2025_26_periods():
        if selected and period != selected:
            continue
        summary = gstr3b_load_summary_for_period(period) or {}
        present = bool(summary)
        outward = tax_bucket_from(summary.get("outward_supplies")) if present else empty_tax_bucket()
        nil = tax_bucket_from(summary.get("outward_nil_exempt")) if present else empty_tax_bucket()
        b_month = build_books_output_liability(period) if all_mode or selected == period else books
        books_tax = tax_bucket_from(b_month.get("books_output"))
        books_nil = tax_bucket_from(b_month.get("books_nil_exempt"))
        books_gross = round(gst_number(books_tax.get("taxable_value")) + gst_number(books_nil.get("taxable_value")), 2)
        portal_tax = round(gst_number(outward.get("taxable_value")), 2)
        portal_nil = round(gst_number(nil.get("taxable_value")), 2)
        note = ""
        if present and abs(portal_nil) <= 0.005 and abs(portal_tax - books_gross) <= 1.0 and gst_number(books_nil.get("taxable_value")) > 1:
            note = (
                "GSTR-3B Table 3.1(a) taxable_value equals Books taxable+nil, while Table 3.1(c) is 0. "
                "Nil/Exempt turnover appears absorbed into 3.1(a) (parser miss or portal filing). "
                "Re-import this month's GSTR-3B PDF after verifying Table 3.1(c)/(e) on the PDF."
            )
            anomaly_notes.append(f"{label}: {note}")
        period_portal.append({
            "period": period,
            "period_label": label,
            "present": present,
            "source_file": gst_text(summary.get("source_file") or summary.get("file_name")),
            "table_31a_taxable_value": portal_tax,
            "table_31a_output_gst": round(gst_number(outward.get("output_gst")), 2),
            "table_31c_nil_exempt_taxable_value": portal_nil,
            "table_31c_output_gst": round(gst_number(nil.get("output_gst")), 2),
            "books_taxable_value": round(gst_number(books_tax.get("taxable_value")), 2),
            "books_nil_exempt_taxable_value": round(gst_number(books_nil.get("taxable_value")), 2),
            "books_gross_taxable_plus_nil": books_gross,
            "taxable_difference_books_minus_31a": round(gst_number(books_tax.get("taxable_value")) - portal_tax, 2),
            "nil_difference_books_minus_31c": round(gst_number(books_nil.get("taxable_value")) - portal_nil, 2),
            "anomaly_note": note,
            "portal_fields": {
                "outward_supplies": "GSTR-3B Table 3.1(a) Outward taxable supplies",
                "outward_nil_exempt": "GSTR-3B Table 3.1(c) Nil/Exempt (+ 3.1(e) Non-GST when present)",
            },
        })

    gstr3b = gstr3b_load_summary("ALL" if all_mode else selected) or {}
    return {
        "return_period": "ALL" if all_mode else selected,
        "period_mode": "fy_all" if all_mode else "month",
        "books_source": books.get("books_source"),
        "taxable": {
            "label": "Outward Taxable Supplies",
            "books_total": round(sum(gst_number(r.get("taxable_value")) for r in taxable_rows), 2),
            "books_count": len(taxable_rows),
            "books_rows": [outward_row_drill_fields(r) for r in taxable_rows],
            "portal_total": round(gst_number((gstr3b.get("outward_supplies") or {}).get("taxable_value")), 2),
            "portal_field": "outward_supplies.taxable_value ← GSTR-3B PDF/JSON Table 3.1(a)",
            "portal_source_files": sorted({
                gst_text(p.get("source_file")) for p in period_portal if p.get("source_file")
            }),
        },
        "nil_exempt": {
            "label": "Nil/Exempt Outward Supplies",
            "books_total": round(sum(gst_number(r.get("taxable_value")) for r in nil_rows), 2),
            "books_count": len(nil_rows),
            "books_rows": [outward_row_drill_fields(r) for r in nil_rows],
            "portal_total": round(gst_number((gstr3b.get("outward_nil_exempt") or {}).get("taxable_value")), 2),
            "portal_field": "outward_nil_exempt.taxable_value ← GSTR-3B PDF/JSON Table 3.1(c)/(e)",
            "portal_source_files": sorted({
                gst_text(p.get("source_file")) for p in period_portal if p.get("source_file")
            }),
        },
        "portal_period_breakdown": period_portal,
        "anomaly_notes": anomaly_notes,
        "findings": {
            "fy_mismatch_origin_month": next(
                (
                    p["period"] for p in period_portal
                    if abs(gst_number(p.get("taxable_difference_books_minus_31a"))) > 1000
                    or abs(gst_number(p.get("nil_difference_books_minus_31c"))) > 1000
                ),
                "",
            ),
            "summary": (
                anomaly_notes[0] if anomaly_notes else
                "No taxable↔nil absorption anomaly detected for the selected scope."
            ),
        },
    }


def split_outward_taxable_and_nil_rows(rows):
    """Split Books/Tally outward rows into taxable (3.1(a)) vs nil/exempt (3.1(c))."""
    taxable_rows = []
    nil_rows = []
    for row in rows or []:
        if is_outward_nil_exempt_row(row):
            nil_rows.append(row)
        else:
            taxable_rows.append(row)
    return taxable_rows, nil_rows


def build_books_output_liability(return_period=""):
    """
    Books Output Liability for the selected MMYYYY.
    Taxable outward (Table 3.1(a)) is kept separate from Nil/Exempt (Table 3.1(c)).
    Output IGST/CGST/SGST/CESS come from taxable outward only.
    """
    period = normalize_gst_recon_period(return_period)
    sales_rows_all = gst_recon_load_rows("TALLY_SALES")
    g1_rows_all = gst_recon_load_rows("GSTR-1")
    sales_rows = filter_gst_rows_by_period(sales_rows_all, period) if period else list(sales_rows_all or [])
    g1_rows = filter_gst_rows_by_period(g1_rows_all, period) if period else list(g1_rows_all or [])

    sales_taxable_rows, sales_nil_rows = split_outward_taxable_and_nil_rows(sales_rows)
    g1_taxable_rows, g1_nil_rows = split_outward_taxable_and_nil_rows(g1_rows)

    tally_taxable = tax_bucket_from(gstr1_tax_totals(sales_taxable_rows))
    tally_nil = tax_bucket_from(gstr1_tax_totals(sales_nil_rows))
    tally_gross = tax_bucket_from(gstr1_tax_totals(sales_rows))
    gstr1_taxable = tax_bucket_from(gstr1_tax_totals(g1_taxable_rows))
    gstr1_nil = tax_bucket_from(gstr1_tax_totals(g1_nil_rows))
    gstr1_gross = tax_bucket_from(gstr1_tax_totals(g1_rows))

    use_tally = bool(
        tally_taxable.get("output_gst")
        or tally_taxable.get("taxable_value")
        or tally_nil.get("taxable_value")
        or tally_gross.get("taxable_value")
    )
    books = tally_taxable if use_tally else gstr1_taxable
    books_nil = tally_nil if use_tally else gstr1_nil
    sync_meta = gst_recon_get_meta("tally_sales_sync", {}) or {}
    result = {
        "return_period": period or "ALL",
        "tally_output": tally_taxable,
        "tally_nil_exempt": tally_nil,
        "tally_gross_output": tally_gross,
        "gstr1_output": gstr1_taxable,
        "gstr1_nil_exempt": gstr1_nil,
        "gstr1_gross_output": gstr1_gross,
        "books_output": books,
        "books_nil_exempt": books_nil,
        "books_source": "Tally Sales" if use_tally else ("GSTR-1" if (gstr1_taxable.get("output_gst") or gstr1_taxable.get("taxable_value") or gstr1_nil.get("taxable_value")) else "None"),
        "tally_sales_count": len(sales_taxable_rows),
        "tally_nil_exempt_count": len(sales_nil_rows),
        "tally_sales_total_count": len(sales_rows_all or []),
        "gstr1_count": len(g1_taxable_rows),
        "difference_tally_vs_gstr1": tax_bucket_diff(gstr1_taxable, tally_taxable),
        "tally_sync": {
            "ok": bool(sync_meta.get("ok")),
            "count": gst_number(sync_meta.get("count")),
            "synced_at": gst_text(sync_meta.get("synced_at")),
            "company": gst_text(sync_meta.get("company")),
        },
    }
    if period and not use_tally and not (gstr1_taxable.get("output_gst") or gstr1_taxable.get("taxable_value") or gstr1_nil.get("taxable_value")):
        if len(sales_rows_all or []) == 0 and gst_number(sync_meta.get("count")) > 0:
            result["warning"] = (
                "Tally Sales rows are missing from the local store. "
                "Click Sync Tally Sales (or One Click GST Sync) and reconcile again."
            )
        elif len(sales_rows_all or []) > 0:
            result["warning"] = (
                f"No Tally Sales / GSTR-1 outward liability found for period {period}. "
                "Check the Period (MMYYYY) filter."
            )
        else:
            result["warning"] = "Sync Tally Sales before Books vs GSTR-3B comparison."
    save_liability_summary(period, "books_output", result)
    return result


def build_itc_claim_breakdown(return_period="", tolerance=1.0):
    period = normalize_gst_recon_period(return_period)
    ensure_tally_purchase_rows_for_itc(period)
    rows_2b = gst_recon_load_rows("GSTR-2B")
    rows_purchase = normalize_tally_purchase_itc_rows(gst_recon_load_rows("TALLY_PURCHASE"))
    if period:
        rows_2b = [row for row in rows_2b if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period]
        rows_purchase = [row for row in rows_purchase if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date")) == period]
    gstr3b = gstr3b_load_summary(period) or {}
    available = tax_bucket_from(gstr2b_tax_totals(rows_2b))
    # Books ITC Claimed = Tally GSTR-3B Net ITC Available (4C A−B).
    # Excludes VCHGSTSTATUSISUNCERTAIN=Yes; tax from Input Duties & Taxes amounts.
    tally_booked = build_tally_booked_itc_summary(rows_purchase, gstr3b_eligible_only=True)
    books_itc = tax_bucket_from({
        "igst": tally_booked["igst"],
        "cgst": tally_booked["cgst"],
        "sgst": tally_booked["sgst"],
        "cess": 0.0,
    })
    books_itc["cess"] = round(gst_number(tally_booked.get("cess")), 2)
    books_itc["output_gst"] = round(gst_number(tally_booked.get("net_itc")), 2)
    claimed = tax_bucket_from(gstr3b.get("itc_claimed") or gstr3b.get("net_itc") or gstr3b)
    reversed_itc = tax_bucket_from(gstr3b.get("itc_reversed"))
    ineligible = tax_bucket_from(gstr3b.get("itc_ineligible"))
    eligible = {
        key: round(max(0.0, gst_number(available.get(key)) - gst_number(reversed_itc.get(key)) - gst_number(ineligible.get(key))), 2)
        for key in ("taxable_value", *TAX_KEYS, "output_gst")
    }
    pending = {
        key: round(max(0.0, gst_number(available.get(key)) - gst_number(claimed.get(key))), 2)
        for key in ("taxable_value", *TAX_KEYS, "output_gst")
    }
    unused = {
        key: round(max(0.0, gst_number(claimed.get(key)) - gst_number(books_itc.get(key))), 2)
        if key != "taxable_value" else 0.0
        for key in ("taxable_value", *TAX_KEYS, "output_gst")
    }
    result = {
        "return_period": period,
        "available_itc": available,
        "claimed_itc": claimed,
        "eligible_itc": eligible,
        "ineligible_itc": ineligible,
        "pending_itc": pending,
        "reversed_itc": reversed_itc,
        "unused_itc": unused,
        "books_itc": books_itc,
        "tally_booked": tally_booked,
        "tolerance": tolerance,
    }
    save_itc_claim_summary(period, result)
    return result


def compute_gst_utilisation(output_liability, available_itc):
    liability = tax_bucket_from(output_liability)
    itc = tax_bucket_from(available_itc)
    remaining_itc = {key: gst_number(itc.get(key)) for key in TAX_KEYS}
    remaining_liability = {key: gst_number(liability.get(key)) for key in TAX_KEYS}
    utilised = empty_tax_bucket()
    cash = empty_tax_bucket()

    def use(credit_key, liability_key):
        take = min(remaining_itc[credit_key], remaining_liability[liability_key])
        if take <= 0:
            return
        remaining_itc[credit_key] = round(remaining_itc[credit_key] - take, 2)
        remaining_liability[liability_key] = round(remaining_liability[liability_key] - take, 2)
        utilised[credit_key] = round(utilised[credit_key] + take, 2)

    use("igst", "igst"); use("cgst", "igst"); use("sgst", "igst")
    use("cgst", "cgst"); use("igst", "cgst")
    use("sgst", "sgst"); use("igst", "sgst")
    use("cess", "cess")
    for key in TAX_KEYS:
        cash[key] = round(max(0.0, remaining_liability[key]), 2)
    utilised["output_gst"] = round(sum(utilised[key] for key in TAX_KEYS), 2)
    cash["output_gst"] = round(sum(cash[key] for key in TAX_KEYS), 2)
    return {
        "liability": liability,
        "available_itc": itc,
        "itc_utilised": utilised,
        "cash_required": cash,
        "remaining_itc": tax_bucket_from(remaining_itc),
    }


def build_net_gst_payable(output_liability, itc_utilised, interest=None, late_fee=None, cash_required=None):
    """
    Split cash components cleanly:
      net_gst_payable     = GST tax cash only (prefer utilisation cash_required; never interest/late fee)
      interest / late_fee = portal GSTR-3B amounts (passed through for callers)
      total_cash_payable  = tax + interest + late fee

    Prefer cash_required from compute_gst_utilisation so IGST↔CGST/SGST cross-utilisation
    is not double-counted as tax payable.
    """
    liability = tax_bucket_from(output_liability)
    utilised = tax_bucket_from(itc_utilised)
    interest = tax_bucket_from(interest)
    late_fee = tax_bucket_from(late_fee)
    if cash_required is not None:
        net = tax_bucket_from(cash_required)
    else:
        net = empty_tax_bucket()
        for key in TAX_KEYS:
            net[key] = round(max(0.0, liability[key] - utilised[key]), 2)
        net["output_gst"] = round(sum(net[key] for key in TAX_KEYS), 2)
    net["taxable_value"] = liability["taxable_value"]
    total_cash = empty_tax_bucket()
    for key in TAX_KEYS:
        total_cash[key] = round(gst_number(net.get(key)) + interest[key] + late_fee[key], 2)
    total_cash["output_gst"] = round(sum(total_cash[key] for key in TAX_KEYS), 2)
    interest_total = round(gst_number(interest.get("output_gst")), 2)
    late_total = round(gst_number(late_fee.get("output_gst")), 2)
    net_total = round(gst_number(net.get("output_gst")), 2)
    return {
        "output_tax": liability,
        "itc_utilised": utilised,
        "interest": interest,
        "late_fee": late_fee,
        "net_gst_payable": net,
        "net_gst_interest_payable": interest_total,
        "net_gst_late_fee_payable": late_total,
        "total_cash_payable": round(net_total + interest_total + late_total, 2),
        "total_cash_payable_bucket": total_cash,
    }


def gstr3b_portal_tax_payable(summary):
    """GST tax payable from one GSTR-3B summary: max(0, outward − claimed) overall. No interest/late fee."""
    summary = summary or {}
    outward = tax_bucket_from(summary.get("outward_supplies") or summary)
    claimed = tax_bucket_from(summary.get("itc_claimed") or summary.get("net_itc") or {})
    # Use overall totals (not head-wise max) so CGST surplus can offset SGST shortfall the way
    # the return's net tax payable is reviewed against the PDF.
    tax_total = round(max(0.0, gst_number(outward.get("output_gst")) - gst_number(claimed.get("output_gst"))), 2)
    tax = empty_tax_bucket()
    if tax_total <= 0:
        return tax
    # Allocate residual tax proportionally to heads that are short, for display only.
    shorts = {key: round(max(0.0, outward[key] - claimed[key]), 2) for key in TAX_KEYS}
    short_sum = round(sum(shorts.values()), 2)
    if short_sum <= 0:
        tax["cgst"] = round(tax_total / 2.0, 2)
        tax["sgst"] = round(tax_total - tax["cgst"], 2)
    else:
        allocated = 0.0
        for index, key in enumerate(TAX_KEYS):
            if index == len(TAX_KEYS) - 1:
                tax[key] = round(tax_total - allocated, 2)
            else:
                share = round(tax_total * (shorts[key] / short_sum), 2) if shorts[key] else 0.0
                tax[key] = share
                allocated = round(allocated + share, 2)
    tax["output_gst"] = tax_total
    return tax


def build_gstr3b_cash_payable_breakdown(return_period="", gstin="", financial_year=""):
    """
    Month-wise GST tax / interest / late fee / total cash from imported gst_3b_summary.
    Interest and late fee come only from portal GSTR-3B rows — never from Books vs 3B diffs.
    Returns an empty months list when no GSTR-3B is imported for the active GSTIN+FY.
    """
    raw_period = gst_text(return_period)
    all_mode = is_gst_all_periods_selection(raw_period) or not normalize_gst_recon_period(raw_period)
    selected = "" if all_mode else normalize_gst_recon_period(return_period)
    imported = set(gstr3b_list_imported_periods(gstin=gstin, financial_year=financial_year))
    months = []
    totals = {
        "gst_tax_payable": 0.0,
        "interest_payable": 0.0,
        "late_fee_payable": 0.0,
        "total_cash_payable": 0.0,
    }
    if not imported:
        return {
            "return_period": "ALL" if all_mode else selected,
            "period_mode": "fy_all" if all_mode else "month",
            "months": [],
            "totals": totals,
            "imported": False,
        }
    for period, label in gst_fy_2025_26_periods():
        if selected and period != selected:
            continue
        if period not in imported:
            # FY table: show missing months only when at least one month is imported.
            if all_mode:
                months.append({
                    "period": period,
                    "period_label": label,
                    "present": False,
                    "gst_tax_payable": 0.0,
                    "interest_payable": 0.0,
                    "late_fee_payable": 0.0,
                    "total_cash_payable": 0.0,
                    "interest": empty_tax_bucket(),
                    "late_fee": empty_tax_bucket(),
                    "gst_tax": empty_tax_bucket(),
                })
            continue
        summary = gstr3b_load_summary_for_period(period) or {}
        present = bool(summary)
        tax = gstr3b_portal_tax_payable(summary) if present else empty_tax_bucket()
        interest = tax_bucket_from(summary.get("interest")) if present else empty_tax_bucket()
        late_fee = tax_bucket_from(summary.get("late_fee")) if present else empty_tax_bucket()
        tax_amt = round(gst_number(tax.get("output_gst")), 2)
        interest_amt = round(gst_number(interest.get("output_gst")), 2)
        late_amt = round(gst_number(late_fee.get("output_gst")), 2)
        total_amt = round(tax_amt + interest_amt + late_amt, 2)
        row = {
            "period": period,
            "period_label": label,
            "present": present,
            "gst_tax_payable": tax_amt,
            "interest_payable": interest_amt,
            "late_fee_payable": late_amt,
            "total_cash_payable": total_amt,
            "interest": interest,
            "late_fee": late_fee,
            "gst_tax": tax,
        }
        months.append(row)
        if present:
            totals["gst_tax_payable"] = round(totals["gst_tax_payable"] + tax_amt, 2)
            totals["interest_payable"] = round(totals["interest_payable"] + interest_amt, 2)
            totals["late_fee_payable"] = round(totals["late_fee_payable"] + late_amt, 2)
            totals["total_cash_payable"] = round(totals["total_cash_payable"] + total_amt, 2)
    return {
        "return_period": "ALL" if all_mode else selected,
        "period_mode": "fy_all" if all_mode else "month",
        "months": months,
        "totals": totals,
        "imported": True,
    }


def reconcile_gstr3b_vs_books(return_period="", tolerance=1.0):
    raw_period = gst_text(return_period)
    all_mode = is_gst_all_periods_selection(raw_period) or not normalize_gst_recon_period(raw_period)
    period = "" if all_mode else normalize_gst_recon_period(return_period)
    imported_periods = gstr3b_list_imported_periods()
    if period:
        ok, _summary = gstr3b_imported_for_period(period)
        has_portal = bool(ok)
    else:
        has_portal = bool(imported_periods)
    if not has_portal:
        # No portal GSTR-3B for this scope — wipe derived caches so Clear stays permanent.
        gst_recon_delete_meta_keys(("gstr3b_dashboard", "GSTR-3B"))
        empty = {
            "imported": False,
            "status": "Not Imported",
            "return_period": "ALL" if all_mode else period,
            "imported_periods": [],
            "cards": {},
            "rows": [],
            "cash_payable_breakdown": {
                "return_period": "ALL" if all_mode else period,
                "period_mode": "fy_all" if all_mode else "month",
                "months": [],
                "totals": {
                    "gst_tax_payable": 0.0,
                    "interest_payable": 0.0,
                    "late_fee_payable": 0.0,
                    "total_cash_payable": 0.0,
                },
                "imported": False,
            },
            "itc": {},
            "utilisation": {},
            "payable": {},
            "gstr3b": {},
            "tolerance": tolerance,
        }
        return empty
    gstr3b = gstr3b_load_summary("ALL" if all_mode else period) or {}
    books = build_books_output_liability(period)
    # If local Tally Sales store is empty but a prior successful sync exists (or Tally is online),
    # refresh Books from Tally so Output Liability is not shown as zero.
    if not (books.get("books_output") or {}).get("output_gst") and not (books.get("books_output") or {}).get("taxable_value"):
        sync_meta = gst_recon_get_meta("tally_sales_sync", {}) or {}
        needs_sales = len(gst_recon_load_rows("TALLY_SALES") or []) == 0
        if needs_sales or gst_number(sync_meta.get("count")) > len(gst_recon_load_rows("TALLY_SALES") or []):
            try:
                synced = sync_tally_sales_vouchers()
                rows = synced.get("rows") or []
                if rows:
                    gst_recon_save_rows("TALLY_SALES", rows)
                    gst_recon_set_meta("tally_sales_sync", {
                        "company": synced.get("company", ""),
                        "count": len(rows),
                        "synced_at": gst_recon_now(),
                        "taxable_value": synced.get("taxable_value"),
                        "output_gst": synced.get("output_gst"),
                        "ok": True,
                        "igst": synced.get("igst"),
                        "cgst": synced.get("cgst"),
                        "sgst": synced.get("sgst"),
                    })
                    books = build_books_output_liability(period)
            except Exception as exc:
                books = build_books_output_liability(period)
                books["warning"] = books.get("warning") or f"Tally Sales sync failed: {exc}"
    itc = build_itc_claim_breakdown(period, tolerance)
    books_out = books["books_output"]
    books_nil = tax_bucket_from(books.get("books_nil_exempt"))
    portal_out = tax_bucket_from(gstr3b.get("outward_supplies"))
    portal_nil = tax_bucket_from(gstr3b.get("outward_nil_exempt"))
    portal_source = "GSTR-3B Table 3.1(a)"
    portal_nil_source = "GSTR-3B Table 3.1(c)/(e)"
    if not portal_out["output_gst"] and not portal_out["taxable_value"]:
        portal_out = tax_bucket_from(books.get("gstr1_output"))
        portal_source = "GSTR-1 (3B outward not in file)"
    # Never substitute GSTR-1 nil into the GSTR-3B column — that falsely inflates
    # Nil/Exempt portal figures when Table 3.1(c) was missing/zero on import.
    if not portal_nil.get("taxable_value"):
        portal_nil = empty_tax_bucket()
        portal_nil_source = "GSTR-3B Table 3.1(c)/(e) (none imported for scope)"
    liability_diff = tax_bucket_diff(portal_out, books_out)
    nil_diff = round(gst_number(books_nil.get("taxable_value")) - gst_number(portal_nil.get("taxable_value")), 2)
    claimed = itc["claimed_itc"]
    available = itc["available_itc"]
    utilisation = compute_gst_utilisation(books_out, claimed if claimed["output_gst"] else available)
    interest_bucket = tax_bucket_from(gstr3b.get("interest"))
    late_fee_bucket = tax_bucket_from(gstr3b.get("late_fee"))
    payable = build_net_gst_payable(
        books_out,
        utilisation["itc_utilised"],
        interest_bucket,
        late_fee_bucket,
        cash_required=utilisation["cash_required"],
    )
    # Portal tax payable only — never fold interest / late fee into Net Tax Payable.
    portal_tax_payable = round(max(0.0, portal_out.get("output_gst", 0) - claimed.get("output_gst", 0)), 2)
    interest_payable = round(gst_number(interest_bucket.get("output_gst")), 2)
    late_fee_payable = round(gst_number(late_fee_bucket.get("output_gst")), 2)
    net_gst_tax_payable = round(gst_number((payable.get("net_gst_payable") or {}).get("output_gst")), 2)
    total_cash_payable = round(net_gst_tax_payable + interest_payable + late_fee_payable, 2)
    cash_payable_breakdown = build_gstr3b_cash_payable_breakdown("ALL" if all_mode else period)
    rows = [
        {"particulars": "Outward Taxable Supplies", "books": books_out.get("taxable_value"), "gstr3b": portal_out.get("taxable_value"), "difference": liability_diff.get("taxable_value"),
         "drilldown_key": "taxable", "books_field": "Tally Sales taxable outward (zero-tax vouchers excluded)", "gstr3b_field": portal_source},
        {"particulars": "Nil/Exempt Outward Supplies", "books": books_nil.get("taxable_value"), "gstr3b": portal_nil.get("taxable_value"), "difference": nil_diff,
         "drilldown_key": "nil_exempt", "books_field": "Tally Sales nil/exempt (zero output tax)", "gstr3b_field": portal_nil_source},
        {"particulars": "Output IGST", "books": books_out.get("igst"), "gstr3b": portal_out.get("igst"), "difference": liability_diff.get("igst")},
        {"particulars": "Output CGST", "books": books_out.get("cgst"), "gstr3b": portal_out.get("cgst"), "difference": liability_diff.get("cgst")},
        {"particulars": "Output SGST", "books": books_out.get("sgst"), "gstr3b": portal_out.get("sgst"), "difference": liability_diff.get("sgst")},
        {"particulars": "Output CESS", "books": books_out.get("cess"), "gstr3b": portal_out.get("cess"), "difference": liability_diff.get("cess")},
        {"particulars": "ITC Claimed", "books": itc["books_itc"].get("output_gst"), "gstr3b": claimed.get("output_gst"),
         "difference": round(claimed.get("output_gst", 0) - itc["books_itc"].get("output_gst", 0), 2),
         "books_field": "tally_booked.net_itc (GSTR-3B 4C A−B; excludes VCHGSTSTATUSISUNCERTAIN=Yes)",
         "books_source": (itc.get("tally_booked") or {}).get("source"),
         "books_formula": (itc.get("tally_booked") or {}).get("formula"),
         "books_function": (itc.get("tally_booked") or {}).get("function"),
         "books_ledgers": (itc.get("tally_booked") or {}).get("ledgers"),
         "books_xml_tags": (itc.get("tally_booked") or {}).get("xml_tags"),
         "gstr3b_field": "itc_claimed.output_gst (Portal GSTR-3B Table 4 / net_itc)"},
        {"particulars": "Net GST Payable (Tax only)", "books": net_gst_tax_payable, "gstr3b": portal_tax_payable,
         "difference": round(net_gst_tax_payable - portal_tax_payable, 2)},
        {"particulars": "Net GST Interest Payable", "books": 0.0, "gstr3b": interest_payable,
         "difference": interest_payable},
        {"particulars": "Net GST Late Fee Payable", "books": 0.0, "gstr3b": late_fee_payable,
         "difference": late_fee_payable},
        {"particulars": "Total Cash Payable", "books": total_cash_payable,
         "gstr3b": round(portal_tax_payable + interest_payable + late_fee_payable, 2),
         "difference": round(
             total_cash_payable - (portal_tax_payable + interest_payable + late_fee_payable), 2
         )},
    ]
    # Turnover diffs (taxable + nil/exempt) + tax/ITC diffs — do not mix nil turnover into taxable.
    books_vs_3b_diff = (
        abs(liability_diff.get("taxable_value", 0))
        + abs(nil_diff)
        + abs(liability_diff.get("output_gst", 0))
        + abs(rows[6]["difference"])
    )
    dashboard = {
        "imported": True,
        "status": "Imported",
        "return_period": "ALL" if all_mode else period,
        "imported_periods": gstr3b_list_imported_periods(),
        "period_mode": "fy_all" if all_mode else "month",
        "cards": {
            "output_gst": books_out.get("output_gst", 0),
            "available_itc": available.get("output_gst", 0),
            "claimed_itc": claimed.get("output_gst", 0),
            "net_gst_payable": net_gst_tax_payable,
            "net_gst_interest_payable": interest_payable,
            "net_gst_late_fee_payable": late_fee_payable,
            "total_cash_payable": total_cash_payable,
            "cash_required": utilisation["cash_required"].get("output_gst", 0),
            "interest": interest_payable,
            "late_fee": late_fee_payable,
            "books_vs_3b_difference": round(books_vs_3b_diff, 2),
        },
        "cash_payable_breakdown": cash_payable_breakdown,
        "rows": rows,
        "books_liability": books,
        "books_source": books.get("books_source"),
        "books_itc_source": (itc.get("tally_booked") or {}).get("source"),
        "books_itc_field": "tally_booked.net_itc (GSTR-3B 4C A−B; VCHGSTSTATUSISUNCERTAIN)",
        "books_itc_formula": (itc.get("tally_booked") or {}).get("formula"),
        "books_itc_function": (itc.get("tally_booked") or {}).get("function"),
        "books_itc_ledgers": (itc.get("tally_booked") or {}).get("ledgers"),
        "books_itc_xml_tags": (itc.get("tally_booked") or {}).get("xml_tags"),
        "books_itc_uncertain_excluded": (itc.get("tally_booked") or {}).get("uncertain_excluded"),
        "books_itc_lines": (itc.get("tally_booked") or {}).get("lines"),
        "books_warning": books.get("warning") or "",
        "portal_outward": portal_out,
        "portal_nil_exempt": portal_nil,
        "portal_source": portal_source,
        "portal_nil_source": portal_nil_source,
        "outward_classification_drilldown": build_outward_classification_drilldown("ALL" if all_mode else period),
        "itc": itc,
        "utilisation": utilisation,
        "payable": payable,
        "gstr3b": gstr3b,
        "tolerance": tolerance,
    }
    gst_recon_set_meta("gstr3b_dashboard", gst_strip_gstr3b_dashboard_wire(dashboard, include_drilldown=False))
    save_liability_summary("ALL" if all_mode else period, "gstr3b_vs_books", dashboard)
    return dashboard


def make_gstr3b_export(dashboard, title="Books vs GSTR-3B", fmt="xlsx"):
    rows = (dashboard or {}).get("rows") or []
    if fmt == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["particulars", "books", "gstr3b", "difference"])
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in writer.fieldnames})
        return output.getvalue().encode("utf-8-sig")
    wb = Workbook()
    ws = wb.active
    ws.title = re.sub(r"[\\/*?:\[\]]", " ", gst_text(title))[:31] or "GSTR3B"
    ws.append(["Particulars", "Books", "GSTR-3B", "Difference"])
    for row in rows:
        ws.append([row.get("particulars"), row.get("books"), row.get("gstr3b"), row.get("difference")])
    cards = (dashboard or {}).get("cards") or {}
    ws2 = wb.create_sheet("Dashboard Cards")
    ws2.append(["Metric", "Value"])
    for key, value in cards.items():
        ws2.append([key, value])
    util = (dashboard or {}).get("utilisation") or {}
    ws3 = wb.create_sheet("Utilisation")
    ws3.append(["Type", "IGST", "CGST", "SGST", "CESS", "Total"])
    for label, key in (
        ("Liability", "liability"),
        ("ITC Utilised", "itc_utilised"),
        ("Cash Required", "cash_required"),
        ("Remaining ITC", "remaining_itc"),
    ):
        bucket = tax_bucket_from(util.get(key))
        ws3.append([label, bucket["igst"], bucket["cgst"], bucket["sgst"], bucket["cess"], bucket["output_gst"]])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



def run_gst_one_click_sync(tolerance=1.0, return_period=""):
    """Combined Phase 1 + Phase 2 + Phase 3 sync workflow."""
    steps = []
    def mark(step, label, percent):
        steps.append({"step": step, "label": label, "percent": percent})

    scope_period = "ALL" if is_gst_all_periods_selection(return_period) or not normalize_gst_recon_period(return_period) else normalize_gst_recon_period(return_period)

    mark(1, "Read GSTR-2B data", 8)
    rows_2b = gst_recon_load_rows("GSTR-2B")
    mark(2, "Read GSTR-1 data", 16)
    rows_g1 = gst_recon_load_rows("GSTR-1")
    mark(3, "Read GSTR-3B data", 24)
    gstr3b = gstr3b_load_summary() or {}
    mark(4, "Sync Tally Purchases", 36)
    purchase = sync_tally_purchase_vouchers()
    gst_recon_save_rows("TALLY_PURCHASE", purchase.get("rows", []))
    purchase_booked = build_tally_booked_itc_summary(
        normalize_tally_purchase_itc_rows(purchase.get("rows", []))
    )
    gst_recon_set_meta("tally_sync", {
        "company": purchase.get("company", ""),
        "count": purchase.get("count", 0),
        "synced_at": purchase.get("synced_at", ""),
        "ok": True,
        "net_itc": purchase_booked.get("net_itc"),
        "igst": purchase_booked.get("igst"),
        "cgst": purchase_booked.get("cgst"),
        "sgst": purchase_booked.get("sgst"),
        "cess": purchase_booked.get("cess"),
        "section_4c_source": purchase.get("section_4c_source"),
        "formula": purchase_booked.get("formula"),
        "function": purchase_booked.get("function"),
    })
    mark(5, "Sync Tally Sales", 48)
    sales = {"ok": False, "rows": [], "count": 0, "error": ""}
    sales_sync_ok = False
    try:
        sales = sync_tally_sales_vouchers()
        sales_sync_ok = bool(sales.get("ok"))
        if sales_sync_ok:
            gst_recon_save_rows("TALLY_SALES", sales.get("rows", []))
            gst_recon_set_meta("tally_sales_sync", {
                "company": sales.get("company", ""),
                "count": sales.get("count", 0),
                "synced_at": sales.get("synced_at", ""),
                "ok": True,
            })
        else:
            tally_log("one-click | sales sync returned not ok — preserving existing TALLY_SALES")
    except ValueError as exc:
        sales = {
            "ok": False,
            "rows": gst_recon_load_rows("TALLY_SALES"),
            "count": len(gst_recon_load_rows("TALLY_SALES")),
            "error": str(exc),
            "preserved_existing": True,
        }
        tally_log(f"one-click | sales sync FAILED — recon skipped | {exc}")
    mark(6, "Run Purchase Reconciliation", 58)
    purchase_results, purchase_counts = ([], {})
    purchase_doc_summary = {}
    if rows_2b and purchase.get("rows"):
        purchase_results, purchase_counts, purchase_doc_summary = reconcile_gstr2b_tally(
            rows_2b, purchase.get("rows", []), tolerance
        )
        gst_recon_save_results(purchase_results, "2b_tally")
    mark(7, "Run Sales Reconciliation", 68)
    sales_results, sales_counts = ([], {})
    sales_doc_summary = {}
    # Do not reconcile sales when Tally Sales sync failed.
    if sales_sync_ok and rows_g1 and sales.get("rows"):
        sales_results, sales_counts, sales_doc_summary = reconcile_gstr1_tally(
            rows_g1, sales.get("rows", []), tolerance
        )
        gstr1_save_reconciliation(sales_results)
    elif not sales_sync_ok:
        mark(7, "Sales Reconciliation skipped (Tally Sales sync failed)", 68)
    mark(8, "ITC Calculation", 78)
    itc = build_itc_dashboard(rows_2b, gstr3b, purchase.get("rows", []), tolerance)
    gst_recon_set_meta("itc_dashboard", itc)
    itc_claim = build_itc_claim_breakdown(scope_period, tolerance)
    mark(9, "Liability + Utilisation + Net Payable", 88)
    sales_rows_for_dash = sales.get("rows", []) if sales_sync_ok else gst_recon_load_rows("TALLY_SALES")
    sales_dash = build_sales_recon_dashboard(
        rows_g1,
        sales_rows_for_dash,
        sales_results or None if sales_sync_ok else None,
        tolerance,
        scope_period,
    )
    gst_recon_set_meta("sales_dashboard", sales_dash)
    gstr3b_dash = reconcile_gstr3b_vs_books(scope_period, tolerance)
    mark(10, "Save results", 95)
    overview = build_gst_recon_overview(scope_period, tolerance)
    gst_recon_set_meta("gst_recon_overview", overview)
    mark(11, "Refresh dashboard", 100)
    return {
        "steps": steps,
        "percent": 100,
        "purchase": {
            "rows": purchase_results,
            "counts": purchase_counts,
            "synced": purchase,
            "document_summary": purchase_doc_summary,
        },
        "sales": {
            "rows": sales_results,
            "counts": sales_counts,
            "synced": sales,
            "dashboard": sales_dash,
            "document_summary": sales_doc_summary or (sales_dash or {}).get("document_summary"),
        },
        "itc_dashboard": itc,
        "itc_claim": itc_claim,
        "gstr3b_dashboard": gstr3b_dash,
        "overview": overview,
    }



def make_gst_purchase_xml(rows, ledger_config, note_mode=False):
    messages = []
    purchase_ledger = gst_text(ledger_config.get("purchaseLedger")) or "Purchase A/c"
    purchase_ledgers = {
        rate: gst_text(ledger_config.get(f"purchaseLedger{rate}")) or f"{rate}% Purchase"
        for rate in (0, 5, 12, 18, 28)
    }
    tax_ledgers = {
        "igst": gst_text(ledger_config.get("igstLedger")) or "Input IGST",
        "cgst": gst_text(ledger_config.get("cgstLedger")) or "Input CGST",
        "sgst": gst_text(ledger_config.get("sgstLedger")) or "Input SGST",
        "cess": gst_text(ledger_config.get("cessLedger")) or "Input Cess",
    }
    round_ledger = gst_text(ledger_config.get("roundLedger")) or "Round Off"
    for row in rows:
        if not row.get("selected") or not (row.get("ready_for_purchase_note") if note_mode else row.get("ready_for_tally")):
            continue
        party = gst_text(row.get("party_ledger") or row.get("party_name"))
        if not party:
            raise ValueError(f"Select a Tally Party Ledger for invoice {row.get('invoice_no')}.")
        invoice_no = gst_text(row.get("original_invoice_no") or row.get("invoice_no"))
        original_date_text = gst_text(row.get("original_invoice_date") or row.get("invoice_date"))
        raw_date = gst_text(row.get("tally_entry_date") or row.get("invoice_date")).replace("/", "-")
        parsed_date = None
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%b-%y", "%d-%b-%Y"):
            try:
                parsed_date = datetime.strptime(raw_date, fmt)
                break
            except ValueError:
                pass
        if not parsed_date:
            raise ValueError(f"Invalid invoice date for {invoice_no}.")
        date = parsed_date.strftime("%Y%m%d")
        reference_date = None
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y", "%d-%b-%Y", "%d-%b-%y"):
            try:
                reference_date = datetime.strptime(original_date_text.replace("/", "-"), fmt).strftime("%Y%m%d")
                break
            except ValueError:
                pass
        invoice_value = gst_number(row.get("invoice_value"))
        taxable = gst_number(row.get("taxable_value"))
        if note_mode:
            invoice_value = abs(invoice_value)
            taxable = abs(taxable)
        if invoice_value <= 0 or taxable <= 0:
            raise ValueError(f"Invoice value/taxable value is missing for {invoice_no}.")
        document_type = gst_text(row.get("document_type"))
        note_hint = " ".join((
            document_type,
            gst_text(row.get("note_type")),
            gst_text(row.get("amendment_type")),
            gst_text(row.get("status")),
        ))
        # Purchase-side note mapping is the reverse of Sales: a supplier's
        # portal Credit Note reduces our payable and is a Tally Debit Note;
        # a supplier Debit Note/increase is a Tally Credit Note.
        is_reduction = bool(note_mode and re.search(
            r"credit\s*note|return|amendment[_\s-]*decrease|\bdecrease\b",
            note_hint, re.I,
        ))
        voucher_type = (
            ("Debit Note" if is_reduction else "Credit Note")
            if note_mode else "Purchase"
        )
        sign = -1 if is_reduction else 1
        expense_ledger = gst_text(row.get("expense_ledger"))
        raw_allocations = [] if expense_ledger else (row.get("sales_allocations") or row.get("items") or [])
        allocations = []
        for source_allocation in raw_allocations:
            allocation = dict(source_allocation or {})
            if allocation.get("rate") in (None, ""):
                allocation["rate"] = allocation.get("gst_rate", row.get("gst_rate"))
            if not gst_text(allocation.get("item_name")):
                allocation["item_name"] = gst_text(allocation.get("name") or row.get("item_name") or row.get("stock_item")) or "Items"
            if not gst_text(allocation.get("hsn")):
                allocation["hsn"] = allocation.get("hsn_code") or row.get("hsn_code")
            allocations.append(allocation)
        # Purchase/GSTR-2 must post as an Item Invoice.  Portal rows often
        # carry rate-wise ``items`` but no browser-created sales_allocations;
        # normalize those items here so they cannot silently fall back to an
        # Accounting Voucher / purchase-ledger-only entry.
        if not expense_ledger and not allocations:
            allocations = [{
                "item_name": gst_text(row.get("item_name") or row.get("stock_item")) or f"{row.get('gst_rate') or 0}% Items",
                "rate": row.get("gst_rate"), "hsn": row.get("hsn_code"),
                "quantity": row.get("quantity") or 1, "unit": row.get("unit") or row.get("uqc") or "Pcs",
                "taxable_value": taxable, "igst": row.get("igst"), "cgst": row.get("cgst"),
                "sgst": row.get("sgst"), "cess": row.get("cess"),
            }]
        row_rate = min((0, 5, 12, 18, 28), key=lambda rate: abs(rate - (100 * sum(gst_number(row.get(key)) for key in ("igst", "cgst", "sgst", "cess")) / taxable if taxable else 0)))
        configured_purchase = gst_text(ledger_config.get(f"purchaseLedger{row_rate}")) or gst_text(row.get("purchase_ledger"))
        if not expense_ledger and not configured_purchase and not allocations:
            raise ValueError(f"Purchase Ledger Required for invoice {invoice_no} ({row_rate}%).")
        posting_ledger = expense_ledger or configured_purchase or purchase_ledgers.get(row_rate, purchase_ledger)
        # Tally XML polarity is debit=negative + IsDeemedPositive Yes and
        # credit=positive + IsDeemedPositive No.  A normal Purchase therefore
        # credits the supplier and debits purchase/Input GST immediately.
        party_amount = sign * invoice_value
        entries = [(party, party_amount, "Yes" if party_amount < 0 else "No")]
        if not allocations:
            posting_amount = sign * -taxable
            entries.append((posting_ledger, posting_amount, "Yes" if posting_amount < 0 else "No"))
        # When item allocations are present Tally books their ACCOUNTINGALLOCATIONS,
        # not row.taxable_value.  The balancing/round-off amount must therefore use
        # the exact inventory allocation total; otherwise Debit and Credit differ.
        allocation_taxable_total = round(sum(
            abs(gst_number(allocation.get("taxable_value"))) if note_mode
            else gst_number(allocation.get("taxable_value"))
            for allocation in allocations
        ), 2)
        component_total = allocation_taxable_total if allocations else taxable
        for field, ledger in tax_ledgers.items():
            tax = gst_number(row.get(field))
            if note_mode:
                tax = abs(tax)
            if tax:
                tax_amount = sign * -tax
                entries.append((ledger, tax_amount, "Yes" if tax_amount < 0 else "No"))
                component_total += tax
        difference = round(invoice_value - component_total, 2)
        if difference:
            round_amount = sign * -difference
            entries.append((round_ledger, round_amount, "Yes" if round_amount < 0 else "No"))
        # Invoice-view vouchers must use LEDGERENTRIES.LIST.  Tally accepted
        # ALLLEDGERENTRIES.LIST syntactically but silently discarded those
        # ledgers while retaining the inventory rows, leaving every Purchase
        # voucher with only its stock-item debit and therefore an import
        # exception (credit/debit mismatch).
        ledger_xml = "".join(
            f"<LEDGERENTRIES.LIST><LEDGERNAME>{xml_escape(name)}</LEDGERNAME>"
            f"<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE>"
            f"<ISPARTYLEDGER>{'Yes' if index == 0 else 'No'}</ISPARTYLEDGER>"
            f"<ISLASTDEEMEDPOSITIVE>{positive}</ISLASTDEEMEDPOSITIVE>"
            f"<AMOUNT>{value:.2f}</AMOUNT></LEDGERENTRIES.LIST>"
            for index, (name, value, positive) in enumerate(entries)
        )
        inventory_xml = ""
        for allocation in allocations:
            allocation_taxable = gst_number(allocation.get("taxable_value"))
            if note_mode:
                allocation_taxable = abs(allocation_taxable)
            if not allocation_taxable:
                continue
            item_name = gst_rate_stock_item_name(allocation)
            quantity = gst_number(allocation.get("quantity")) or 1.0
            unit = gst_text(allocation.get("unit")) or "Pcs"
            item_rate = round(allocation_taxable / quantity, 4) if quantity else allocation_taxable
            amount = sign * -allocation_taxable
            positive = "Yes" if amount < 0 else "No"
            allocation_rate = min((0, 5, 12, 18, 28), key=lambda rate: abs(rate - gst_number(allocation.get("rate"))))
            allocation_ledger = expense_ledger or purchase_ledgers.get(allocation_rate, posting_ledger)
            inventory_xml += (
                f'<ALLINVENTORYENTRIES.LIST><STOCKITEMNAME>{xml_escape(item_name)}</STOCKITEMNAME>'
                f'<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE><RATE>{item_rate:.4f}/{xml_escape(unit)}</RATE>'
                f'<AMOUNT>{amount:.2f}</AMOUNT><ACTUALQTY>{quantity:g} {xml_escape(unit)}</ACTUALQTY>'
                f'<BILLEDQTY>{quantity:g} {xml_escape(unit)}</BILLEDQTY>'
                f'<ACCOUNTINGALLOCATIONS.LIST><LEDGERNAME>{xml_escape(allocation_ledger)}</LEDGERNAME>'
                f'<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE><AMOUNT>{amount:.2f}</AMOUNT>'
                f'</ACCOUNTINGALLOCATIONS.LIST></ALLINVENTORYENTRIES.LIST>'
            )
        # Each retry gets a fresh REMOTEID.  A rejected Import Exception can
        # retain the earlier id inside Tally; reusing it may make a corrected
        # voucher collide with that failed object.  Duplicate protection is
        # performed before send from GSTIN/invoice/date/value matching.
        remote_id = str(uuid.uuid4())
        # Purchase voucher No. and Supplier Invoice No. both carry the
        # supplier's original invoice number, as requested.  DATE remains
        # the GSTR-2B-period posting date and REFERENCEDATE is the original
        # supplier invoice date.
        voucher_number = invoice_no
        party_gstin = re.sub(r"\s+", "", gst_text(row.get("gstin")).upper())
        party_state = GST_STATE_CODES.get(party_gstin[:2], "") if re.fullmatch(r"\d{2}[A-Z0-9]{13}", party_gstin) else ""
        party_detail_xml = (
            f'<BASICBUYERNAME>{xml_escape(party)}</BASICBUYERNAME>'
            f'<CONSIGNEEMAILINGNAME>{xml_escape(party)}</CONSIGNEEMAILINGNAME>'
            f'<CONSIGNEECOUNTRYNAME>India</CONSIGNEECOUNTRYNAME>'
            f'{"<PARTYGSTIN>" + xml_escape(party_gstin) + "</PARTYGSTIN>" if party_gstin else ""}'
            f'{"<CONSIGNEEGSTIN>" + xml_escape(party_gstin) + "</CONSIGNEEGSTIN>" if party_gstin else ""}'
            f'{"<STATENAME>" + xml_escape(party_state) + "</STATENAME>" if party_state else ""}'
            f'{"<CONSIGNEESTATENAME>" + xml_escape(party_state) + "</CONSIGNEESTATENAME>" if party_state else ""}'
            f'{"<PLACEOFSUPPLY>" + xml_escape(party_state) + "</PLACEOFSUPPLY>" if party_state else ""}'
        )
        narration = (
            f"{voucher_type} {invoice_no} imported from reviewed GSTR-2/Purchase data. "
            f"Original invoice date: {original_date_text}. GSTR-2B period: {row.get('gstr2b_period') or 'Not available'}. "
            f"Supplier GSTIN: {row.get('gstin')}."
        )
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER REMOTEID="{remote_id}" '
            f'VCHTYPE="{xml_escape(voucher_type)}" ACTION="Create" OBJVIEW="Invoice Voucher View"><DATE>{date}</DATE>'
            f'<VOUCHERTYPENAME>{xml_escape(voucher_type)}</VOUCHERTYPENAME><VOUCHERNUMBER>{voucher_number}</VOUCHERNUMBER>'
            f'<REFERENCE>{xml_escape(invoice_no)}</REFERENCE>{f"<REFERENCEDATE>{reference_date}</REFERENCEDATE>" if reference_date else ""}<PARTYLEDGERNAME>{xml_escape(party)}</PARTYLEDGERNAME>'
            f'{party_detail_xml}'
            f'<PERSISTEDVIEW>{"Invoice Voucher View" if inventory_xml else "Accounting Voucher View"}</PERSISTEDVIEW>'
            f'{"<ISINVOICE>Yes</ISINVOICE><OBJVIEW>Invoice Voucher View</OBJVIEW>" if inventory_xml else ""}'
            f'<NARRATION>{xml_escape(narration)}</NARRATION>'
            f'{ledger_xml}{inventory_xml}</VOUCHER></TALLYMESSAGE>'
        )
    if not messages:
        raise ValueError("Select at least one reviewed Purchase entry.")
    return ("<?xml version=\"1.0\" encoding=\"UTF-8\"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>"
            "</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>"
            f"<REQUESTDATA>{''.join(messages)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>").encode("utf-8")


def gst_rate_for_values(taxable, igst=0, cgst=0, sgst=0, cess=0):
    taxable = gst_number(taxable)
    if taxable <= 0:
        return 0
    calculated = 100 * sum(gst_number(value) for value in (igst, cgst, sgst, cess)) / taxable
    return min((0, 5, 12, 18, 28), key=lambda rate: abs(rate - calculated))


def adjust_gst_sales_rows(rows, reductions, additions=None):
    additions = additions or {}
    adjusted = json.loads(json.dumps(rows))
    buckets = {rate: [] for rate in (0, 5, 12, 18, 28)}
    for row in adjusted:
        # Missing selection must never mean selected. The browser explicitly
        # sends the user's checked rows; defaulting to True selected notes and
        # other non-visible records after an amendment.
        row["selected"] = bool(row.get("selected", False))
        is_note = any(word in gst_text(row.get("document_type")).lower()
                      for word in ("note", "refund", "return"))
        allocations = []
        source_items = row.get("items") or [{
            "taxable_value": row.get("taxable_value", 0), "igst": row.get("igst", 0),
            "cgst": row.get("cgst", 0), "sgst": row.get("sgst", 0), "cess": row.get("cess", 0),
        }]
        grouped = {}
        for item in source_items:
            taxable = gst_number(item.get("taxable_value"))
            rate = gst_rate_for_values(taxable, item.get("igst"), item.get("cgst"), item.get("sgst"), item.get("cess"))
            if taxable <= 0:
                continue
            hsn = gst_text(item.get("hsn"))
            item_name = gst_text(item.get("item_name") or item.get("name")) or (
                f"HSN {hsn} Items" if hsn else f"{rate}% Items"
            )
            group_key = (rate, hsn, item_name)
            target = grouped.setdefault(group_key, {
                "rate": rate, "hsn": hsn, "item_name": item_name, "quantity": 0.0,
                "item_rate": gst_number(item.get("rate")), "unit": gst_text(item.get("unit") or item.get("uqc")) or "Pcs",
                "taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0,
            })
            target["quantity"] = round(target["quantity"] + gst_number(item.get("quantity")), 3)
            for field in ("taxable_value", "igst", "cgst", "sgst", "cess"):
                target[field] = round(target[field] + gst_number(item.get(field)), 2)
        allocations.extend(grouped.values())
        row["sales_allocations"] = allocations
        for allocation in allocations:
            # Distribute only among the invoices explicitly selected by the
            # user. Every selected invoice line at this GST rate receives its
            # proportional share; unselected invoices remain unchanged.
            if not is_note and row["selected"] and allocation["rate"] in buckets:
                buckets[allocation["rate"]].append(allocation)
        row["is_sales_note"] = is_note

    applied = {}
    for rate, allocations in buckets.items():
        requested = max(0.0, gst_number(reductions.get(str(rate), reductions.get(rate, 0))))
        added = max(0.0, gst_number(additions.get(str(rate), additions.get(rate, 0))))
        available = round(sum(item["taxable_value"] for item in allocations), 2)
        if requested > available:
            raise ValueError(f"{rate}% less amount cannot exceed taxable sales ₹{available:.2f}.")
        if added > 0 and not allocations:
            raise ValueError(f"{rate}% taxable add needs at least one {rate}% sales invoice.")
        remaining = round(requested - added, 2)
        remaining_available = available
        for index, allocation in enumerate(allocations):
            old_taxable = allocation["taxable_value"]
            share = remaining if index == len(allocations) - 1 else round(
                remaining * old_taxable / remaining_available, 2
            ) if remaining_available else 0.0
            remaining = round(remaining - share, 2)
            remaining_available = round(remaining_available - old_taxable, 2)
            new_taxable = round(old_taxable - share, 2)
            interstate = gst_number(allocation.get("igst")) > 0
            allocation["less_amount"] = round(max(share, 0), 2)
            allocation["add_amount"] = round(max(-share, 0), 2)
            allocation["taxable_value"] = new_taxable
            allocation["igst"] = round(new_taxable * rate / 100, 2) if interstate else 0.0
            allocation["cgst"] = round(new_taxable * rate / 200, 2) if not interstate else 0.0
            allocation["sgst"] = round(new_taxable * rate / 200, 2) if not interstate else 0.0
        applied[str(rate)] = {"requested": requested, "added": added, "available": available}

    for row in adjusted:
        allocations = row.get("sales_allocations") or []
        for field in ("taxable_value", "igst", "cgst", "sgst", "cess"):
            row[field] = round(sum(gst_number(item.get(field)) for item in allocations), 2)
        row["invoice_value"] = round(
            gst_number(row.get("taxable_value"))
            + gst_number(row.get("igst"))
            + gst_number(row.get("cgst"))
            + gst_number(row.get("sgst"))
            + gst_number(row.get("cess")),
            2,
        )
        row["ready_for_sales_tally"] = bool(
            not row.get("is_sales_note") and allocations and row.get("invoice_no")
            and gst_number(row.get("invoice_value")) > 0.005
        )
        row["ready_for_note_tally"] = bool(
            row.get("is_sales_note") and allocations and row.get("invoice_no")
        )
    return adjusted, applied


def is_tally_pure_sales_invoice_row(row):
    """Sales vouchers only — never Credit Note / Debit Note."""
    doc = gst_text(row.get("document_type")).lower()
    vtype = gst_text(row.get("voucher_type")).lower()
    if any(token in doc for token in ("credit note", "debit note", "refund", "return")):
        return False
    if any(token in vtype for token in ("credit note", "debit note")):
        return False
    if "sales" in vtype or doc in {"", "sales invoice", "invoice", "b2b", "b2c"}:
        return True
    return "sales" in doc


def sales_row_exists_in_tally(row, tally_sales_rows, tolerance=1.0):
    """
    Confirm a GSTR-1 Sales invoice already exists in Tally using real voucher data:
    Sales voucher type + voucher/invoice number + date + party + amount.
    """
    tolerance = max(1.0, gst_number(tolerance))
    inv = gst_text(row.get("invoice_no_norm") or normalize_invoice_number(row.get("invoice_no")))
    if not inv or inv == "0":
        return None
    date = normalize_invoice_date_key(row.get("invoice_date"))
    party = normalize_party_key(row.get("party_ledger") or row.get("party_name"))
    amount = abs(gst_number(row.get("invoice_value")))
    for candidate in tally_sales_rows or []:
        if not is_tally_pure_sales_invoice_row(candidate):
            continue
        c_inv = gst_text(
            candidate.get("invoice_no_norm")
            or normalize_invoice_number(candidate.get("invoice_no") or candidate.get("voucher_number"))
        )
        if c_inv != inv:
            continue
        c_date = normalize_invoice_date_key(
            candidate.get("invoice_date") or candidate.get("tally_entry_date")
        )
        if date and c_date and date != c_date:
            continue
        c_party = normalize_party_key(candidate.get("party_ledger") or candidate.get("party_name"))
        if party and c_party and party != c_party:
            continue
        c_amount = abs(gst_number(candidate.get("invoice_value")))
        if abs(c_amount - amount) > tolerance:
            continue
        return candidate
    return None


def filter_tally_sales_invoices_for_period(tally_rows, return_period=""):
    period = normalize_gst_recon_period(return_period)
    rows = [row for row in (tally_rows or []) if is_tally_pure_sales_invoice_row(row)]
    if not period:
        return rows
    return [
        row for row in rows
        if normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date") or row.get("tally_entry_date"))
        == period
    ]


def build_sales_existence_report(rows, tally_sales_rows, tolerance=1.0, return_period=""):
    """Classify selected GSTR-1 Sales rows as already in Tally vs missing."""
    scoped_tally = filter_tally_sales_invoices_for_period(tally_sales_rows, return_period)
    already = []
    missing = []
    report = []
    for row in rows or []:
        item = dict(row or {})
        if not item.get("selected"):
            continue
        if any(token in gst_text(item.get("document_type")).lower() for token in ("note", "refund", "return")):
            continue
        match = sales_row_exists_in_tally(item, scoped_tally, tolerance=tolerance)
        entry = {
            "invoice_no": gst_text(item.get("invoice_no")),
            "invoice_date": gst_text(item.get("invoice_date")),
            "party": gst_text(item.get("party_ledger") or item.get("party_name")),
            "amount": round(abs(gst_number(item.get("invoice_value"))), 2),
            "found_in_tally": bool(match),
            "matching_tally_voucher_no": gst_text((match or {}).get("voucher_number") or (match or {}).get("invoice_no")),
            "status": "ALREADY EXISTS" if match else "MISSING IN TALLY",
        }
        report.append(entry)
        if match:
            item["tally_status"] = "Already in Tally"
            item["tally_voucher_no"] = entry["matching_tally_voucher_no"]
            already.append(item)
        else:
            item["tally_status"] = "MISSING IN TALLY"
            item["ready_for_sales_tally"] = True
            missing.append(item)
    return {
        "already": already,
        "missing": missing,
        "report": report,
        "tally_sales_count": len(scoped_tally),
        "period": normalize_gst_recon_period(return_period) or "ALL",
    }


def extract_tally_missing_master(details):
    """Parse LINEERROR / detail text for missing ledger/master names."""
    texts = details if isinstance(details, (list, tuple)) else [details]
    for text in texts:
        message = gst_text(text)
        if not message:
            continue
        match = re.search(r"Ledger\s+'([^']+)'\s+does\s+not\s+exist", message, re.I)
        if match:
            return {"kind": "ledger", "name": match.group(1), "message": message}
        match = re.search(
            r"(Stock\s+Item|Godown|Voucher\s+Type|Unit)\s+'([^']+)'\s+does\s+not\s+exist",
            message,
            re.I,
        )
        if match:
            return {"kind": match.group(1).strip().lower(), "name": match.group(2), "message": message}
        if "does not exist" in message.lower():
            return {"kind": "master", "name": "", "message": message}
    return None


def send_one_gst_sales_voucher_to_tally(row, ledger_config, return_period="", tolerance=1.0, tally_rows=None):
    """Send exactly one Sales voucher and capture that voucher's own Tally response."""
    item = dict(row or {})
    item["selected"] = True
    item["ready_for_sales_tally"] = True
    invoice_no = gst_text(item.get("invoice_no"))
    party = gst_text(item.get("party_ledger") or item.get("party_name"))
    amount = round(abs(gst_number(item.get("invoice_value"))), 2)
    period = normalize_gst_recon_period(return_period) or normalize_gst_recon_period(
        item.get("source_period") or item.get("invoice_date")
    )
    if tally_rows is None:
        try:
            tally_pack = sync_tally_sales_vouchers()
            tally_rows = list(tally_pack.get("rows") or [])
            gst_recon_save_rows("TALLY_SALES", tally_rows)
        except Exception as exc:
            return {
                "invoice_no": invoice_no,
                "invoice_date": gst_text(item.get("invoice_date")),
                "party": party,
                "amount": amount,
                "send_result": "FAILED",
                "status": "TALLY SYNC FAILED",
                "created": 0,
                "altered": 0,
                "errors": 1,
                "exceptions": 0,
                "lineerror": str(exc),
                "found_in_tally_after": False,
                "matching_tally_voucher_no": "",
                "missing_master": None,
            }
    scoped = filter_tally_sales_invoices_for_period(tally_rows, period)
    existing = sales_row_exists_in_tally(item, scoped, tolerance=tolerance)
    if existing:
        return {
            "invoice_no": invoice_no,
            "invoice_date": gst_text(item.get("invoice_date")),
            "party": party,
            "amount": amount,
            "send_result": "SKIPPED",
            "status": "ALREADY EXISTS",
            "created": 0,
            "altered": 0,
            "errors": 0,
            "exceptions": 0,
            "lineerror": "",
            "found_in_tally_after": True,
            "matching_tally_voucher_no": gst_text(existing.get("voucher_number") or existing.get("invoice_no")),
            "missing_master": None,
            "tally_sales_count": len(scoped),
        }
    raw = make_gst_sales_xml([item], ledger_config or {}, fresh_remote_id=True)
    try:
        # Use the configured local/remote Connector so a phone browser can
        # submit an invoice to the computer where TallyPrime is running.
        tally_response = tally_post(raw, timeout=120, purpose="sales-invoice-photo")
    except Exception as exc:
        return {
            "invoice_no": invoice_no,
            "invoice_date": gst_text(item.get("invoice_date")),
            "party": party,
            "amount": amount,
            "send_result": "FAILED",
            "status": "CONNECTION FAILED",
            "created": 0,
            "altered": 0,
            "errors": 1,
            "exceptions": 0,
            "lineerror": str(exc),
            "found_in_tally_after": False,
            "matching_tally_voucher_no": "",
            "missing_master": None,
        }
    (DATA_DIR / "tally_last_sales_request.xml").write_bytes(raw)
    (DATA_DIR / "tally_last_sales_response.xml").write_text(
        tally_response, encoding="utf-8", errors="replace"
    )
    result = tally_import_result(tally_response)
    missing_master = extract_tally_missing_master(result.get("details") or [])
    try:
        post_pack = sync_tally_sales_vouchers()
        post_rows = list(post_pack.get("rows") or [])
        gst_recon_save_rows("TALLY_SALES", post_rows)
    except Exception:
        post_rows = tally_rows
    post_scoped = filter_tally_sales_invoices_for_period(post_rows, period)
    match = sales_row_exists_in_tally(item, post_scoped, tolerance=tolerance)
    if missing_master:
        status, send_result = "MASTER MISSING", "FAILED"
    elif match:
        status = "CREATED" if result.get("created") else "VERIFIED IN TALLY"
        send_result = "SUCCESS"
    else:
        status, send_result = "MISSING IN TALLY", "FAILED"
    lineerror = next((detail for detail in (result.get("details") or []) if detail), "")
    return {
        "invoice_no": invoice_no,
        "invoice_date": gst_text(item.get("invoice_date")),
        "party": party,
        "amount": amount,
        "send_result": send_result,
        "status": status,
        "created": result.get("created", 0),
        "altered": result.get("altered", 0),
        "errors": result.get("errors", 0),
        "exceptions": result.get("exceptions", 0),
        "lineerror": lineerror,
        "details": result.get("details") or [],
        "found_in_tally_after": bool(match),
        "matching_tally_voucher_no": gst_text((match or {}).get("voucher_number") or (match or {}).get("invoice_no")),
        "missing_master": missing_master,
        "tally_sales_count": len(post_scoped),
        "raw_response": result.get("raw_response") or "",
    }


def make_gst_sales_xml(rows, ledger_config, fresh_remote_id=False):
    sales_ledgers = {
        0: gst_text(ledger_config.get("salesLedger0")) or "Sales 0%",
        5: gst_text(ledger_config.get("salesLedger5")) or "Sales 5%",
        12: gst_text(ledger_config.get("salesLedger12")) or "Sales 12%",
        18: gst_text(ledger_config.get("salesLedger18")) or "Sales 18%",
        28: gst_text(ledger_config.get("salesLedger28")) or "Sales 28%",
    }
    tax_ledgers = {
        "igst": gst_text(ledger_config.get("igstLedger")) or "Output IGST",
        "cgst": gst_text(ledger_config.get("cgstLedger")) or "Output CGST",
        "sgst": gst_text(ledger_config.get("sgstLedger")) or "Output SGST",
        "cess": gst_text(ledger_config.get("cessLedger")) or "Output Cess",
    }
    round_ledger = gst_text(ledger_config.get("roundLedger")) or "Round Off"
    messages = []
    for row in rows:
        if not row.get("selected") or not row.get("ready_for_sales_tally"):
            continue
        party = gst_party_ledger(row)
        invoice_no = gst_text(row.get("invoice_no"))
        raw_date = gst_text(row.get("invoice_date")).replace("/", "-")
        parsed_date = None
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y", "%d/%m/%Y"):
            try:
                parsed_date = datetime.strptime(raw_date, fmt)
                break
            except ValueError:
                pass
        if not parsed_date:
            raise ValueError(f"Invalid invoice date for {invoice_no}.")
        date = parsed_date.strftime("%Y%m%d")
        allocations = row.get("sales_allocations") or []
        invoice_value = gst_number(row.get("invoice_value"))
        ledger_entries = [(party, -invoice_value, "Yes", "Yes")]
        component_total = 0.0
        inventory_xml = []
        for allocation in allocations:
            taxable = gst_number(allocation.get("taxable_value"))
            if taxable:
                gst_rate = int(gst_number(allocation.get("rate")))
                sales_ledger = sales_ledgers.get(gst_rate, "Sales")
                item_name = gst_rate_stock_item_name(allocation)
                quantity = gst_number(allocation.get("quantity")) or 1.0
                unit = gst_text(allocation.get("unit")) or "Pcs"
                item_rate = round(taxable / quantity, 4) if quantity else taxable
                gst_nature_xml = (
                    '<GSTOVRDNNATURE>Sales Nil Rated</GSTOVRDNNATURE>'
                    if gst_rate == 0 else ''
                )
                inventory_xml.append(
                    f'<ALLINVENTORYENTRIES.LIST><STOCKITEMNAME>{xml_escape(item_name)}</STOCKITEMNAME>'
                    f'<ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE><RATE>{item_rate:.4f}/{xml_escape(unit)}</RATE>'
                    f'<AMOUNT>{taxable:.2f}</AMOUNT><ACTUALQTY>{quantity:g} {xml_escape(unit)}</ACTUALQTY>'
                    f'<BILLEDQTY>{quantity:g} {xml_escape(unit)}</BILLEDQTY>'
                    f'<ACCOUNTINGALLOCATIONS.LIST><LEDGERNAME>{xml_escape(sales_ledger)}</LEDGERNAME>'
                    f'<ISDEEMEDPOSITIVE>No</ISDEEMEDPOSITIVE>{gst_nature_xml}'
                    f'<AMOUNT>{taxable:.2f}</AMOUNT>'
                    f'</ACCOUNTINGALLOCATIONS.LIST></ALLINVENTORYENTRIES.LIST>'
                )
                component_total += taxable
        for field, ledger in tax_ledgers.items():
            value = round(sum(gst_number(item.get(field)) for item in allocations), 2)
            if value:
                ledger_entries.append((ledger, value, "No", "No"))
                component_total += value
        difference = round(invoice_value - component_total, 2)
        if difference:
            ledger_entries.append((round_ledger, difference, "No" if difference > 0 else "Yes", "No"))
        ledger_xml = "".join(
            f"<LEDGERENTRIES.LIST><LEDGERNAME>{xml_escape(name)}</LEDGERNAME>"
            f"<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE><ISPARTYLEDGER>{is_party}</ISPARTYLEDGER>"
            f"<ISLASTDEEMEDPOSITIVE>{positive}</ISLASTDEEMEDPOSITIVE>"
            f"<AMOUNT>{value:.2f}</AMOUNT></LEDGERENTRIES.LIST>"
            for name, value, positive, is_party in ledger_entries
        )
        # Verified-missing invoices get a fresh REMOTEID so a prior failed import
        # REMOTEID cannot falsely ALTER / EXCEPTION as "already exists".
        remote_seed = f"bank2tally-sales:{invoice_no}:{date}:{invoice_value:.2f}"
        # A retry must not reuse the REMOTEID of an earlier rejected import.
        # Tally may otherwise answer CREATED=0/ERRORS=0 and silently ignore it.
        remote_id = (
            str(uuid.uuid4()) if fresh_remote_id
            else str(uuid.uuid5(uuid.NAMESPACE_URL, remote_seed))
        )
        narration = "Being goods sold for cash." if party.lower() == "cash" else "Being goods sold on credit."
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER REMOTEID="{remote_id}" VCHTYPE="Sales" '
            f'ACTION="Create" OBJVIEW="Invoice Voucher View">'
            f'<DATE>{date}</DATE><VOUCHERTYPENAME>Sales</VOUCHERTYPENAME><VOUCHERNUMBER>{xml_escape(invoice_no)}</VOUCHERNUMBER>'
            f'<REFERENCE>{xml_escape(invoice_no)}</REFERENCE><PARTYLEDGERNAME>{xml_escape(party)}</PARTYLEDGERNAME>'
            f'<PERSISTEDVIEW>Invoice Voucher View</PERSISTEDVIEW><ISINVOICE>Yes</ISINVOICE>'
            f'<OBJVIEW>Invoice Voucher View</OBJVIEW>'
            f'<NARRATION>{xml_escape(narration)}</NARRATION>{ledger_xml}{"".join(inventory_xml)}'
            f'</VOUCHER></TALLYMESSAGE>'
        )
    if not messages:
        raise ValueError("Select at least one reviewed Sales Invoice.")
    return ("<?xml version=\"1.0\" encoding=\"UTF-8\"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>"
            "</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>"
            f"<REQUESTDATA>{''.join(messages)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>").encode("utf-8")


def make_gst_note_xml(rows, ledger_config, voucher_type):
    voucher_type = gst_text(voucher_type)
    if voucher_type not in {"Credit Note", "Debit Note", "Journal"}:
        raise ValueError("Choose Credit Note, Debit Note or Journal.")
    sales_ledgers = {
        0: gst_text(ledger_config.get("salesLedger0")) or "Sales 0%",
        5: gst_text(ledger_config.get("salesLedger5")) or "Sales 5%",
        12: gst_text(ledger_config.get("salesLedger12")) or "Sales 12%",
        18: gst_text(ledger_config.get("salesLedger18")) or "Sales 18%",
        28: gst_text(ledger_config.get("salesLedger28")) or "Sales 28%",
    }
    tax_ledgers = {
        "igst": gst_text(ledger_config.get("igstLedger")) or "Output IGST",
        "cgst": gst_text(ledger_config.get("cgstLedger")) or "Output CGST",
        "sgst": gst_text(ledger_config.get("sgstLedger")) or "Output SGST",
        "cess": gst_text(ledger_config.get("cessLedger")) or "Output Cess",
    }
    round_ledger = gst_text(ledger_config.get("roundLedger")) or "Round Off"
    tally_voucher_type = gst_text(ledger_config.get("tallyVoucherType")) or voucher_type
    reverse = voucher_type in {"Credit Note", "Journal"}
    messages = []
    for row in rows:
        if not row.get("selected"):
            continue
        party = gst_party_ledger(row)
        if not party:
            raise ValueError(f"Party ledger missing for note {gst_text(row.get('invoice_no'))}.")
        note_no = gst_text(row.get("invoice_no"))
        if not note_no:
            raise ValueError("Credit/Debit Note number is missing.")
        raw_date = gst_text(row.get("invoice_date")).replace("/", "-")
        parsed_date = None
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y"):
            try:
                parsed_date = datetime.strptime(raw_date, fmt)
                break
            except ValueError:
                pass
        if not parsed_date:
            raise ValueError(f"Invalid note date for {note_no}.")
        date = parsed_date.strftime("%Y%m%d")
        invoice_value = abs(gst_number(row.get("invoice_value")))
        if invoice_value < 0.01:
            raise ValueError(f"Note value is zero for {note_no}.")
        party_amount = invoice_value if reverse else -invoice_value
        party_positive = "No" if reverse else "Yes"
        counter_sign = -1 if reverse else 1
        entries = [(party, party_amount, party_positive, "Yes")]
        component_total = 0.0
        inventory_xml = []
        for allocation in row.get("sales_allocations") or []:
            taxable = abs(gst_number(allocation.get("taxable_value")))
            if taxable:
                gst_rate = int(gst_number(allocation.get("rate")))
                sales_ledger = sales_ledgers.get(gst_rate, "Sales")
                if allocation.get("use_ledger_entry"):
                    income_ledger = gst_text(allocation.get("item_name"))
                    if not income_ledger:
                        raise ValueError(f"Select the Tally income ledger for {note_no}.")
                    entries.append((income_ledger, counter_sign * taxable,
                                    "Yes" if reverse else "No", "No"))
                elif voucher_type == "Journal":
                    entries.append((sales_ledger, counter_sign * taxable,
                                    "Yes" if reverse else "No", "No"))
                else:
                    item_name = gst_rate_stock_item_name(allocation)
                    quantity = abs(gst_number(allocation.get("quantity"))) or 1.0
                    unit = gst_text(allocation.get("unit")) or "Pcs"
                    item_rate = round(taxable / quantity, 4) if quantity else taxable
                    item_positive = "Yes" if reverse else "No"
                    item_amount = counter_sign * taxable
                    inventory_xml.append(
                        f'<ALLINVENTORYENTRIES.LIST><STOCKITEMNAME>{xml_escape(item_name)}</STOCKITEMNAME>'
                        f'<ISDEEMEDPOSITIVE>{item_positive}</ISDEEMEDPOSITIVE>'
                        f'<RATE>{item_rate:.4f}/{xml_escape(unit)}</RATE><AMOUNT>{item_amount:.2f}</AMOUNT>'
                        f'<ACTUALQTY>{quantity:g} {xml_escape(unit)}</ACTUALQTY>'
                        f'<BILLEDQTY>{quantity:g} {xml_escape(unit)}</BILLEDQTY>'
                        f'<ACCOUNTINGALLOCATIONS.LIST><LEDGERNAME>{xml_escape(sales_ledger)}</LEDGERNAME>'
                        f'<ISDEEMEDPOSITIVE>{item_positive}</ISDEEMEDPOSITIVE>'
                        f'<AMOUNT>{item_amount:.2f}</AMOUNT></ACCOUNTINGALLOCATIONS.LIST>'
                        f'</ALLINVENTORYENTRIES.LIST>'
                    )
                component_total += taxable
        for field, ledger in tax_ledgers.items():
            value = abs(round(sum(gst_number(item.get(field)) for item in row.get("sales_allocations") or []), 2))
            if value:
                entries.append((ledger, counter_sign * value, "Yes" if reverse else "No", "No"))
                component_total += value
        difference = round(invoice_value - component_total, 2)
        if difference:
            entries.append((round_ledger, counter_sign * difference, "Yes" if reverse else "No", "No"))
        ledger_xml = "".join(
            f"<LEDGERENTRIES.LIST><LEDGERNAME>{xml_escape(name)}</LEDGERNAME>"
            f"<ISDEEMEDPOSITIVE>{positive}</ISDEEMEDPOSITIVE><ISPARTYLEDGER>{is_party}</ISPARTYLEDGER>"
            f"<ISLASTDEEMEDPOSITIVE>{positive}</ISLASTDEEMEDPOSITIVE><AMOUNT>{value:.2f}</AMOUNT>"
            f"</LEDGERENTRIES.LIST>" for name, value, positive, is_party in entries
        )
        remote_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"bank2tally-note:{voucher_type}:{note_no}:{date}:{invoice_value:.2f}"))
        invoice_mode = voucher_type != "Journal"
        view = "Invoice Voucher View" if invoice_mode else "Accounting Voucher View"
        narration = (
            "Being goods returned by customer."
            if voucher_type == "Credit Note"
            else ("Being additional sales value debited to customer." if voucher_type == "Debit Note"
                  else "Being sales adjustment recorded.")
        )
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><VOUCHER REMOTEID="{remote_id}" VCHTYPE="{xml_escape(tally_voucher_type)}" '
            f'ACTION="Create" OBJVIEW="{view}">'
            f'<DATE>{date}</DATE><VOUCHERTYPENAME>{xml_escape(tally_voucher_type)}</VOUCHERTYPENAME>'
            f'<VOUCHERNUMBER>{xml_escape(note_no)}</VOUCHERNUMBER><REFERENCE>{xml_escape(note_no)}</REFERENCE>'
            f'<PARTYLEDGERNAME>{xml_escape(party)}</PARTYLEDGERNAME><PERSISTEDVIEW>{view}</PERSISTEDVIEW>'
            f'<ISINVOICE>{"Yes" if invoice_mode else "No"}</ISINVOICE><OBJVIEW>{view}</OBJVIEW>'
            f'<NARRATION>{xml_escape(narration)}</NARRATION>'
            f'{ledger_xml}{"".join(inventory_xml)}</VOUCHER></TALLYMESSAGE>'
        )
    if not messages:
        raise ValueError("Select at least one Credit/Debit Note or Sales Return.")
    return ("<?xml version=\"1.0\" encoding=\"UTF-8\"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>"
            "</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>Vouchers</REPORTNAME></REQUESTDESC>"
            f"<REQUESTDATA>{''.join(messages)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>").encode("utf-8")


def tally_import_result(tally_response):
    def count(tag):
        match = re.search(rf"<{tag}>\s*(\d+)\s*</{tag}>", tally_response, re.I)
        return int(match.group(1)) if match else 0

    details = []
    for tag in ("LINEERROR", "ERROR", "MESSAGE", "DESC"):
        for value in re.findall(rf"<{tag}(?:\s[^>]*)?>(.*?)</{tag}>", tally_response, re.I | re.S):
            message = re.sub(r"<[^>]+>", " ", html.unescape(value))
            message = re.sub(r"\s+", " ", message).strip()
            if message and not message.isdigit() and message not in details:
                details.append(message)
    created = count("CREATED")
    errors = count("ERRORS")
    exceptions = count("EXCEPTIONS")
    return {
        "created": created,
        "altered": count("ALTERED"),
        "ignored": count("IGNORED"),
        "errors": errors,
        "exceptions": exceptions,
        "details": details[:20],
        "raw_response": tally_response[:4000] if tally_response else "",
    }


def parse_bulk_voucher_file(name, raw):
    """Read an optional customer workbook without imposing a fixed template."""
    suffix = Path(name).suffix.lower()
    records = []
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        workbook_source = io.BytesIO(raw)
        conversion_dir = None
        if suffix == ".xls":
            # openpyxl cannot read the legacy binary XLS format. Bank2Tally is a
            # Windows/Tally application, so use the locally installed Excel to
            # make a temporary XLSX copy without modifying the customer's file.
            conversion_dir = tempfile.TemporaryDirectory(prefix="bank2tally_xls_")
            conversion_path = Path(conversion_dir.name)
            source_path = conversion_path / "source.xls"
            target_path = conversion_path / "converted.xlsx"
            script_path = conversion_path / "convert.ps1"
            source_path.write_bytes(raw)
            script_path.write_text(
                "param([string]$Source,[string]$Target)\n"
                "$excel = $null\n$book = $null\n"
                "try {\n"
                "  $excel = New-Object -ComObject Excel.Application\n"
                "  $excel.Visible = $false\n$excel.DisplayAlerts = $false\n"
                "  $book = $excel.Workbooks.Open($Source, 0, $true)\n"
                "  $book.SaveAs($Target, 51)\n"
                "} finally {\n"
                "  if ($book -ne $null) { $book.Close($false) }\n"
                "  if ($excel -ne $null) { $excel.Quit() }\n"
                "}\n",
                encoding="utf-8",
            )
            try:
                completed = subprocess.run(
                    ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File",
                     str(script_path), str(source_path), str(target_path)],
                    capture_output=True, text=True, timeout=90, check=False,
                    creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
                )
                if completed.returncode != 0 or not target_path.exists():
                    detail = (completed.stderr or completed.stdout or "Excel conversion failed").strip()
                    raise ValueError(
                        "Could not read the old .xls file. Open it in Excel, Save As .xlsx, then upload again. "
                        + detail[:180]
                    )
                workbook_source = io.BytesIO(target_path.read_bytes())
            except (OSError, subprocess.SubprocessError) as exc:
                raise ValueError(
                    "Could not read the old .xls file. Open it in Excel, Save As .xlsx, then upload again."
                ) from exc
        workbook = load_workbook(workbook_source, data_only=True, read_only=True)
        # Customer workbooks often keep a summary/list on the first sheet and
        # the actual voucher rows on another sheet. Pick the sheet whose first
        # rows contain the strongest voucher-column signature.
        def sheet_score(candidate):
            sample = list(candidate.iter_rows(min_row=1, max_row=min(candidate.max_row, 20), values_only=True))
            joined = " ".join(gst_text(value).lower() for row in sample for value in row)
            return sum(1 for word in ("invoice", "credit note", "voucher", "date", "item", "amount", "taxable") if word in joined)
        sheet = max(workbook.worksheets, key=sheet_score)
        values = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if conversion_dir is not None:
            conversion_dir.cleanup()
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        values = list(csv.reader(io.StringIO(text)))
    if not values:
        return records
    header_index = 0
    for index, row in enumerate(values[:20]):
        joined = " ".join(gst_text(value).lower() for value in row)
        if any(word in joined for word in ("party", "customer", "ledger", "amount", "invoice")):
            header_index = index
            break
    headers = [re.sub(r"[^a-z0-9]+", " ", gst_text(value).lower()).strip() for value in values[header_index]]
    def find(*needles):
        for index, header in enumerate(headers):
            if any(needle in header for needle in needles):
                return index
        return -1
    columns = {
        "party": find("party", "customer", "person", "ledger name"),
        "amount": find("invoice value", "total amount", "amount", "value"),
        "date": find("invoice date", "date"),
        "invoice_no": find("invoice no", "voucher no", "bill no", "credit note", "debit note"),
        "item": find("item name", "stock item", "item"),
        "rate": find("gst rate", "tax rate"),
        "hsn": find("hsn code", "hsn", "sac code", "sac"),
        "quantity": find("quantity", "qty"),
    }
    # Detailed GSTR-1 reports commonly put one invoice on a header row and its
    # remaining HSN lines immediately below it with invoice/party cells blank or
    # zero. Preserve every HSN as a separate allocation instead of dropping the
    # continuation lines.
    is_gstr1_detail = all(token in headers for token in ("desc", "gstin", "invoice date", "invoice no", "hsn code", "taxable amount"))
    if is_gstr1_detail:
        def exact(label):
            try:
                return headers.index(label)
            except ValueError:
                return -1

        pos = {
            "party": exact("desc"), "gstin": exact("gstin"),
            "date": exact("invoice date"), "invoice_no": exact("invoice no"),
            "invoice_value": exact("invoice value"), "hsn": exact("hsn code"),
            "quantity": exact("quantity"), "taxable": exact("taxable amount"),
            "cess": exact("cess"),
        }
        # The percentage/amount sub-headings are on the next row; in this
        # report layout SGST, CGST and IGST each occupy two adjacent columns.
        pos["sgst_pct"] = exact("sgst")
        pos["sgst"] = pos["sgst_pct"] + 1
        pos["cgst_pct"] = exact("cgst")
        pos["cgst"] = pos["cgst_pct"] + 1
        pos["igst_pct"] = exact("igst")
        pos["igst"] = pos["igst_pct"] + 1

        def at(row, key):
            index = pos.get(key, -1)
            return row[index] if index >= 0 and index < len(row) else ""

        grouped = []
        current = None
        for row in values[header_index + 2:]:
            if not row or not any(gst_text(value) for value in row):
                continue
            invoice_no = gst_text(at(row, "invoice_no"))
            party = gst_text(at(row, "party"))
            is_new_invoice = bool(invoice_no and invoice_no != "0" and party and party != "0")
            if is_new_invoice:
                date_value = at(row, "date")
                if isinstance(date_value, datetime):
                    date_value = date_value.strftime("%Y-%m-%d")
                current = {
                    "party": party,
                    "gstin": gst_text(at(row, "gstin")),
                    "amount": gst_number(at(row, "invoice_value")),
                    "date": gst_text(date_value),
                    "invoice_no": invoice_no,
                    "item": "",
                    "rate": 0.0,
                    "hsn": "",
                    "quantity": 0.0,
                    "sales_allocations": [],
                }
                grouped.append(current)
            if current is None:
                continue
            hsn_match = re.search(r"\b\d{4,8}\b", gst_text(at(row, "hsn")))
            taxable = gst_number(at(row, "taxable"))
            if not hsn_match or abs(taxable) <= 0.0001:
                continue
            sgst_pct = gst_number(at(row, "sgst_pct"))
            cgst_pct = gst_number(at(row, "cgst_pct"))
            igst_pct = gst_number(at(row, "igst_pct"))
            rate = round(igst_pct if igst_pct else sgst_pct + cgst_pct, 4)
            allocation = {
                "hsn": hsn_match.group(0),
                "quantity": gst_number(at(row, "quantity")),
                "taxable_value": taxable,
                "rate": rate,
                "igst": gst_number(at(row, "igst")),
                "cgst": gst_number(at(row, "cgst")),
                "sgst": gst_number(at(row, "sgst")),
                "cess": gst_number(at(row, "cess")),
                "unit": "Pcs",
            }
            current["sales_allocations"].append(allocation)

        for record in grouped:
            allocations = record["sales_allocations"]
            if allocations:
                record["rate"] = allocations[0]["rate"]
                record["hsn"] = allocations[0]["hsn"]
                record["quantity"] = sum(gst_number(item.get("quantity")) for item in allocations)
        return [record for record in grouped if record["sales_allocations"]]
    # Some accounting exports leave the Party heading blank (for example the
    # column between Date and Items). Infer that column from its position.
    if columns["party"] < 0 and columns["date"] >= 0 and columns["item"] > columns["date"] + 1:
        columns["party"] = columns["date"] + 1
    for row in values[header_index + 1:]:
        if not row or not any(gst_text(value) for value in row):
            continue
        def cell(key):
            pos = columns[key]
            return row[pos] if pos >= 0 and pos < len(row) else ""
        amount = gst_number(cell("amount"))
        if not amount and not gst_text(cell("party")):
            continue
        date_value = cell("date")
        if isinstance(date_value, datetime):
            date_value = date_value.strftime("%Y-%m-%d")
        records.append({
            "party": gst_text(cell("party")), "amount": amount,
            "date": gst_text(date_value), "invoice_no": gst_text(cell("invoice_no")),
            "item": gst_text(cell("item")), "rate": gst_number(cell("rate")),
            "hsn": gst_text(cell("hsn")), "quantity": gst_number(cell("quantity")),
        })
    return records


def gst_rate_stock_item_name(allocation):
    """Keep one Tally Stock Item per GST rate + display item + HSN."""
    rate = int(gst_number(allocation.get("rate")))
    base_name = gst_text(allocation.get("item_name")) or "Items"
    # A bulk entry may already contain the rate prefix.  Add it only once;
    # ``preserve_item_name`` means preserve the user's base label, not collapse
    # all HSNs into one Tally Stock Item.
    if not re.match(rf"^\s*{rate}\s*%\s+", base_name, re.I):
        base_name = f"{rate}% {base_name}"

    hsn_candidates = re.findall(r"\b\d{4,8}\b", gst_text(allocation.get("hsn")))
    hsn = hsn_candidates[0] if hsn_candidates else ""
    if not hsn:
        return base_name
    # Avoid duplicating the suffix when a reviewed/imported item already has it.
    if re.search(rf"(?:^|\s)HSN\s*[-:#]?\s*{re.escape(hsn)}\s*$", base_name, re.I):
        return base_name
    return f"{base_name} - HSN {hsn}"


def gst_stock_item_taxability_xml(rate):
    """Classify a 0% Stock Item explicitly as Nil Rated in Tally."""
    if abs(gst_number(rate)) > 0.000001:
        return ""
    return (
        '<GSTDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM>'
        '<CALCULATIONTYPE>On Value</CALCULATIONTYPE>'
        '<TAXABILITY>Nil Rated</TAXABILITY>'
        '<SRCOFGSTDETAILS>Specify Details Here</SRCOFGSTDETAILS>'
        '</GSTDETAILS.LIST>'
    )


def ensure_gst_party_ledgers(rows, ledger_config=None):
    ledger_config = ledger_config or {}
    party_parent = gst_text(ledger_config.get("partyParent")) or "Sundry Debtors"
    purchase_party_mode = party_parent.strip().lower() == "sundry creditors"
    cache = sync_tally()
    ledger_items = cache.get("ledgers", [])
    stock_items = cache.get("items", [])
    stock_groups = cache.get("stock_groups", [])
    company_state = gst_text(cache.get("company_state")) or "Assam"
    company_country = gst_text(cache.get("company_country")) or "India"
    existing_by_name = {
        gst_text(item.get("name")).lower(): item for item in ledger_items if gst_text(item.get("name"))
    }
    existing_by_gstin = {}
    for item in ledger_items:
        gstin = re.sub(r"\s+", "", gst_text(item.get("gstin")).upper())
        if gstin and gstin not in existing_by_gstin:
            existing_by_gstin[gstin] = item
    parties = {}
    party_alters = {}
    mappings = {}
    for row in rows:
        raw_party = gst_text(row.get("party_ledger") or row.get("party_name"))
        raw_party_key = raw_party.lower()
        party = gst_party_ledger(row)
        party_key = party.lower()
        gstin = re.sub(r"\s+", "", gst_text(row.get("gstin")).upper())
        if party_key == "cash":
            existing_cash = existing_by_name.get("cash")
            mappings[party_key] = existing_cash["name"] if existing_cash else "Cash"
            if raw_party_key:
                mappings[raw_party_key] = mappings[party_key]
            if not existing_cash and party_key not in parties:
                parties[party_key] = {
                    "name": "Cash", "gstin": "", "state": company_state,
                    "country": company_country, "registration_type": "Unregistered/Consumer",
                    "parent": "Cash-in-Hand",
                }
            continue
        if gstin and gstin in existing_by_gstin:
            mappings[party_key] = existing_by_gstin[gstin]["name"]
            continue
        if party_key in existing_by_name:
            existing_party = existing_by_name[party_key]
            existing_gstin = re.sub(r"\s+", "", gst_text(existing_party.get("gstin")).upper())
            # Purchase rule: GSTIN is the primary identity.  Do not attach a
            # registered supplier invoice to a same-name ledger whose GSTIN is
            # blank/different; create a distinct GSTIN-specific creditor.
            if not (purchase_party_mode and gstin and existing_gstin != gstin):
                mappings[party_key] = existing_party["name"]
                if (not purchase_party_mode and not gstin and
                        gst_text(existing_party.get("state")).lower() in {"", "not applicable"}):
                    party_alters[party_key] = {
                        "name": existing_party["name"],
                        "parent": existing_party.get("parent") or party_parent,
                        "state": company_state, "country": company_country,
                        "registration_type": "Unregistered/Consumer",
                    }
                continue
            party = f"{party} - {gstin}"
            party_key = party.lower()
        if party_key not in parties:
            valid_gstin = bool(re.fullmatch(r"\d{2}[A-Z0-9]{13}", gstin))
            party_state = GST_STATE_CODES.get(gstin[:2], company_state) if valid_gstin else company_state
            parties[party_key] = {
                "name": party,
                "gstin": gstin if valid_gstin else "",
                "state": party_state,
                "country": company_country,
                "registration_type": "Regular" if valid_gstin else "Unregistered/Consumer",
                "parent": party_parent,
            }
            # Resolve both the displayed/source party name and the generated
            # GSTIN-specific ledger name to the newly-created creditor.
            mappings[raw_party_key] = party
    sales_mappings = {}
    sales_to_create = {}
    for rate in (0, 5, 12, 18, 28):
        field = f"salesLedger{rate}"
        requested = gst_text(ledger_config.get(field)) or f"Sales {rate}%"
        rate_pattern = re.compile(rf"(?<!\d){rate}\s*%")
        rate_ledgers = [
            item for item in ledger_items
            if gst_text(item.get("parent")).lower() == "sales accounts"
            and rate_pattern.search(gst_text(item.get("name")))
            and "sale" in gst_text(item.get("name")).lower()
        ]
        # Prefer the user's Tally convention ("12% Sales") over an older
        # auto-created duplicate ("Sales 12%"), then use the requested name.
        resolved = None
        if rate == 0:
            # Use Tally's configured Nil Rated/Exempt sales ledger.  A plain
            # Sales 0% ledger without GST details makes Tally discard the
            # stock item's accounting allocation as an Import Exception.
            resolved = next((
                item for item in ledger_items
                if gst_text(item.get("parent")).lower() == "sales accounts"
                and gst_text(item.get("taxability")).lower() == "nil rated"
            ), None)
        resolved = resolved or next((
            item for item in rate_ledgers
            if re.match(rf"^\s*{rate}\s*%\s*sales?\b", gst_text(item.get("name")), re.I)
        ), None)
        resolved = resolved or existing_by_name.get(requested.lower())
        resolved = resolved or (rate_ledgers[0] if rate_ledgers else None)
        if resolved:
            sales_mappings[field] = resolved["name"]
        else:
            sales_mappings[field] = requested
            sales_to_create[requested.lower()] = requested
    purchase_mappings = {}
    purchase_to_create = {}
    for rate in (0, 5, 12, 18, 28):
        field = f"purchaseLedger{rate}"
        requested = gst_text(ledger_config.get(field)) or f"{rate}% Purchase"
        rate_pattern = re.compile(rf"(?<!\d){rate}\s*%")
        rate_ledgers = [item for item in ledger_items
                        if gst_text(item.get("parent")).lower() == "purchase accounts"
                        and rate_pattern.search(gst_text(item.get("name")))
                        and "purchase" in gst_text(item.get("name")).lower()]
        resolved = next((item for item in rate_ledgers
                         if re.match(rf"^\s*{rate}\s*%\s*purchase\b", gst_text(item.get("name")), re.I)), None)
        resolved = resolved or existing_by_name.get(requested.lower()) or (rate_ledgers[0] if rate_ledgers else None)
        if resolved:
            purchase_mappings[field] = resolved["name"]
        else:
            purchase_mappings[field] = requested
            purchase_to_create[requested.lower()] = requested
    tax_specs = {
        "igstLedger": ("Output IGST", "IGST", "Integrated Tax"),
        "cgstLedger": ("Output CGST", "CGST", "Central Tax"),
        "sgstLedger": ("Output SGST", "SGST", "State Tax"),
        "cessLedger": ("Output Cess", "CESS", "Cess"),
    }
    tax_mappings = {}
    tax_to_create = {}
    for field, (default_name, token, duty_head) in tax_specs.items():
        requested = gst_text(ledger_config.get(field)) or default_name
        resolved = existing_by_name.get(requested.lower())
        if not resolved:
            resolved = next((
                item for item in ledger_items
                if gst_text(item.get("parent")).lower() == "duties & taxes"
                and token.lower() in gst_text(item.get("name")).lower()
                and "output" in gst_text(item.get("name")).lower()
            ), None)
        if resolved:
            tax_mappings[field] = resolved["name"]
        else:
            tax_mappings[field] = requested
            tax_to_create[requested.lower()] = (requested, duty_head)
    round_requested = gst_text(ledger_config.get("roundLedger")) or "Round Off"
    round_resolved = existing_by_name.get(round_requested.lower()) or next((
        item for item in ledger_items
        if "round" in gst_text(item.get("name")).lower()
        and "off" in gst_text(item.get("name")).lower()
    ), None)
    tax_mappings["roundLedger"] = round_resolved["name"] if round_resolved else round_requested
    round_to_create = None if round_resolved else round_requested
    existing_stock_by_name = {
        gst_text(item.get("name")).lower(): item
        for item in stock_items if gst_text(item.get("name"))
    }
    existing_stock_names = set(existing_stock_by_name)
    expense_mappings = {}
    expenses_to_create = {}
    for row in rows:
        requested = gst_text(row.get("expense_ledger"))
        if not requested:
            continue
        key = requested.lower()
        existing = existing_by_name.get(key)
        if existing:
            expense_mappings[key] = existing["name"]
        else:
            expense_mappings[key] = requested
            expenses_to_create[key] = requested
    stock_to_create = {}
    stock_to_alter = {}
    primary_stock_parent = "&#4; Primary"
    for row in rows:
        for allocation in row.get("sales_allocations") or []:
            rate = int(gst_number(allocation.get("rate")))
            item_name = gst_rate_stock_item_name(allocation)
            item_key = item_name.lower()
            hsn_candidates = re.findall(r"\b\d{4,8}\b", gst_text(allocation.get("hsn")))
            source_hsn = hsn_candidates[0] if hsn_candidates else ""
            if item_key not in existing_stock_names and item_key not in stock_to_create:
                stock_to_create[item_key] = {
                    "name": item_name,
                    "hsn": source_hsn,
                    "unit": gst_text(allocation.get("unit")) or "Pcs",
                    "rate": rate,
                }
            elif item_key in existing_stock_by_name:
                existing_item = existing_stock_by_name[item_key]
                existing_parent = html.unescape(gst_text(existing_item.get("parent"))).replace("\x04", "").strip()
                needs_hsn = bool(source_hsn and not gst_text(existing_item.get("hsn")))
                needs_primary = existing_parent.lower() != "primary"
                needs_nil_rated = rate == 0
                if (needs_hsn or needs_primary or needs_nil_rated) and item_key not in stock_to_alter:
                    stock_to_alter[item_key] = {
                        "name": existing_item["name"],
                        "hsn": gst_text(existing_item.get("hsn")) or source_hsn,
                        "unit": gst_text(existing_item.get("unit")) or gst_text(allocation.get("unit")) or "Pcs",
                        "rate": rate,
                    }
    messages = []
    for party in parties.values():
        gstin = party["gstin"]
        state = party["state"]
        country = party["country"]
        registration_type = party["registration_type"]
        gst_details = (
            f'<LEDGSTREGDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM>'
            f'<STATE>{xml_escape(state)}</STATE><PLACEOFSUPPLY>{xml_escape(state)}</PLACEOFSUPPLY>'
            f'<GSTREGISTRATIONTYPE>{xml_escape(registration_type)}</GSTREGISTRATIONTYPE>'
            f'{"<GSTIN>" + xml_escape(gstin) + "</GSTIN>" if gstin else ""}'
            f'</LEDGSTREGDETAILS.LIST>'
            f'<LEDMAILINGDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM>'
            f'<MAILINGNAME>{xml_escape(party["name"])}</MAILINGNAME>'
            f'<STATE>{xml_escape(state)}</STATE><COUNTRY>{xml_escape(country)}</COUNTRY>'
            f'</LEDMAILINGDETAILS.LIST>'
        )
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(party["name"])}" ACTION="Create">'
            f'<NAME>{xml_escape(party["name"])}</NAME><PARENT>{xml_escape(party.get("parent") or party_parent)}</PARENT>'
            f'<ISBILLWISEON>Yes</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
            f'<COUNTRYNAME>{xml_escape(country)}</COUNTRYNAME>'
            f'<COUNTRYOFRESIDENCE>{xml_escape(country)}</COUNTRYOFRESIDENCE>'
            f'<LEDSTATENAME>{xml_escape(state)}</LEDSTATENAME>'
            f'<GSTREGISTRATIONTYPE>{xml_escape(registration_type)}</GSTREGISTRATIONTYPE>'
            f'{"<PARTYGSTIN>" + xml_escape(gstin) + "</PARTYGSTIN>" if gstin else ""}'
            f'{gst_details}'
            f'</LEDGER></TALLYMESSAGE>'
        )
    for party in party_alters.values():
        state = party["state"]
        country = party["country"]
        registration_type = party["registration_type"]
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(party["name"])}" ACTION="Alter">'
            f'<NAME>{xml_escape(party["name"])}</NAME><PARENT>{xml_escape(party["parent"])}</PARENT>'
            f'<COUNTRYNAME>{xml_escape(country)}</COUNTRYNAME><COUNTRYOFRESIDENCE>{xml_escape(country)}</COUNTRYOFRESIDENCE>'
            f'<LEDSTATENAME>{xml_escape(state)}</LEDSTATENAME><GSTREGISTRATIONTYPE>{xml_escape(registration_type)}</GSTREGISTRATIONTYPE>'
            f'<LEDGSTREGDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM><STATE>{xml_escape(state)}</STATE>'
            f'<PLACEOFSUPPLY>{xml_escape(state)}</PLACEOFSUPPLY><GSTREGISTRATIONTYPE>{xml_escape(registration_type)}</GSTREGISTRATIONTYPE></LEDGSTREGDETAILS.LIST>'
            f'<LEDMAILINGDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM><MAILINGNAME>{xml_escape(party["name"])}</MAILINGNAME>'
            f'<STATE>{xml_escape(state)}</STATE><COUNTRY>{xml_escape(country)}</COUNTRY></LEDMAILINGDETAILS.LIST>'
            f'</LEDGER></TALLYMESSAGE>'
        )
    for ledger_name in sales_to_create.values():
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(ledger_name)}" ACTION="Create">'
            f'<NAME>{xml_escape(ledger_name)}</NAME><PARENT>Sales Accounts</PARENT>'
            f'<ISBILLWISEON>No</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
            f'</LEDGER></TALLYMESSAGE>'
        )
    for ledger_name in purchase_to_create.values():
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(ledger_name)}" ACTION="Create">'
            f'<NAME>{xml_escape(ledger_name)}</NAME><PARENT>Purchase Accounts</PARENT>'
            f'<ISBILLWISEON>No</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
            f'</LEDGER></TALLYMESSAGE>'
        )
    for ledger_name, duty_head in tax_to_create.values():
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(ledger_name)}" ACTION="Create">'
            f'<NAME>{xml_escape(ledger_name)}</NAME><PARENT>Duties &amp; Taxes</PARENT>'
            f'<TAXTYPE>GST</TAXTYPE><GSTDUTYHEAD>{xml_escape(duty_head)}</GSTDUTYHEAD>'
            f'<ISBILLWISEON>No</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
            f'</LEDGER></TALLYMESSAGE>'
        )
    if round_to_create:
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(round_to_create)}" ACTION="Create">'
            f'<NAME>{xml_escape(round_to_create)}</NAME><PARENT>Indirect Expenses</PARENT>'
            f'<ISBILLWISEON>No</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
            f'</LEDGER></TALLYMESSAGE>'
        )
    for ledger_name in expenses_to_create.values():
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><LEDGER NAME="{xml_escape(ledger_name)}" ACTION="Create">'
            f'<NAME>{xml_escape(ledger_name)}</NAME><PARENT>Indirect Expenses</PARENT>'
            f'<ISBILLWISEON>No</ISBILLWISEON><AFFECTSSTOCK>No</AFFECTSSTOCK>'
            f'</LEDGER></TALLYMESSAGE>'
        )
    for item in stock_to_create.values():
        hsn_xml = (
            f'<HSNDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM>'
            f'<HSNCODE>{xml_escape(item["hsn"])}</HSNCODE>'
            f'<HSN>{xml_escape(item["hsn"])}</HSN><SRCOFHSNDETAILS>Specify Details Here</SRCOFHSNDETAILS>'
            f'</HSNDETAILS.LIST>'
            if item["hsn"] else ""
        )
        taxability_xml = gst_stock_item_taxability_xml(item["rate"])
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><STOCKITEM NAME="{xml_escape(item["name"])}" ACTION="Create">'
            f'<NAME>{xml_escape(item["name"])}</NAME><PARENT TYPE="String">{primary_stock_parent}</PARENT>'
            f'<BASEUNITS>{xml_escape(item["unit"])}</BASEUNITS><GSTAPPLICABLE>Applicable</GSTAPPLICABLE>'
            f'<GSTTYPEOFSUPPLY>Goods</GSTTYPEOFSUPPLY>{hsn_xml}{taxability_xml}'
            f'</STOCKITEM></TALLYMESSAGE>'
        )
    for item in stock_to_alter.values():
        hsn_xml = (
            f'<HSNDETAILS.LIST><APPLICABLEFROM>20250401</APPLICABLEFROM>'
            f'<HSNCODE>{xml_escape(item["hsn"])}</HSNCODE>'
            f'<HSN>{xml_escape(item["hsn"])}</HSN><SRCOFHSNDETAILS>Specify Details Here</SRCOFHSNDETAILS>'
            f'</HSNDETAILS.LIST>'
        )
        taxability_xml = gst_stock_item_taxability_xml(item["rate"])
        messages.append(
            f'<TALLYMESSAGE xmlns:UDF="TallyUDF"><STOCKITEM NAME="{xml_escape(item["name"])}" ACTION="Alter">'
            f'<NAME>{xml_escape(item["name"])}</NAME><PARENT TYPE="String">{primary_stock_parent}</PARENT>'
            f'<GSTAPPLICABLE>Applicable</GSTAPPLICABLE><GSTTYPEOFSUPPLY>Goods</GSTTYPEOFSUPPLY>'
            f'{hsn_xml}{taxability_xml}</STOCKITEM></TALLYMESSAGE>'
        )
    if not messages:
        return {
            "created": 0, "existing": len(existing_by_name), "names": [],
            "mappings": mappings, "salesLedgers": sales_mappings, "purchaseLedgers": purchase_mappings, "taxLedgers": tax_mappings,
            "stockItemsCreated": 0, "stockItemsUpdated": 0, "expenseMappings": expense_mappings,
        }
    raw = (
        '<?xml version="1.0" encoding="UTF-8"?><ENVELOPE><HEADER><TALLYREQUEST>Import Data</TALLYREQUEST>'
        '</HEADER><BODY><IMPORTDATA><REQUESTDESC><REPORTNAME>All Masters</REPORTNAME></REQUESTDESC>'
        f'<REQUESTDATA>{"".join(messages)}</REQUESTDATA></IMPORTDATA></BODY></ENVELOPE>'
    ).encode("utf-8")
    result = tally_import_result(tally_post(raw, timeout=90))
    hard_errors = result.get("errors", 0)
    if hard_errors or result["ignored"]:
        detail = f" First reason: {result['details'][0]}" if result["details"] else ""
        raise ValueError(
            f"Tally could not create the required Ledger/Stock Item masters. Errors: {hard_errors}, "
            f"Ignored: {result['ignored']}.{detail}"
        )
    for party_key, party in parties.items():
        mappings[party_key] = party["name"]
    return {
        "created": result["created"] + result["altered"],
        "existing": len(existing_by_name),
        "names": [party["name"] for party in parties.values()] + list(sales_to_create.values()) + list(purchase_to_create.values()),
        "mappings": mappings,
        "salesLedgers": sales_mappings,
        "purchaseLedgers": purchase_mappings,
        "taxLedgers": tax_mappings,
        "expenseMappings": expense_mappings,
        "stockItemsCreated": len(stock_to_create),
        "stockItemsUpdated": len(stock_to_alter),
    }


# Browser refresh/tab close while a long Phase 1 request is still writing a
# response must not crash the request thread with WinError 10054 noise.
CLIENT_DISCONNECT_ERRORS = (ConnectionResetError, BrokenPipeError, ConnectionAbortedError)


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)
    def log_message(self, fmt, *args):
        return
    def end_headers(self):
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()
    def send_json(self, status, payload):
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        try:
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(raw)))
            self.end_headers()
            self.wfile.write(raw)
        except CLIENT_DISCONNECT_ERRORS:
            # Client already gone (refresh / close / aborted fetch). Response is dropped.
            return False
        return True
    def do_GET(self):
        # Long-poll endpoint: the Windows Connector calls this endpoint; no
        # request is ever made from Render to the customer's computer.
        if self.path.startswith("/api/connector/poll"):
            from urllib.parse import parse_qs, urlparse
            connector_id = parse_qs(urlparse(self.path).query).get("id", [""])[0]
            if not connector_authorized(self.headers, connector_id):
                self.send_json(401, {"error": "Connector authentication failed."})
                return
            try:
                command = CONNECTOR_COMMANDS.get(timeout=25)
            except queue.Empty:
                command = None
            self.send_json(200, {"command": command})
            return
        super().do_GET()
    def do_POST(self):
        global ACCOUNT_LOGGED_IN
        try:
            payload = json.loads(self.rfile.read(int(self.headers.get("Content-Length", "0"))))
            if self.path == "/api/connector/result":
                connector_id = str(payload.get("connector_id", ""))
                if not connector_authorized(self.headers, connector_id):
                    self.send_json(401, {"error": "Connector authentication failed."})
                    return
                command_id = str(payload.get("command_id", ""))
                with CONNECTOR_LOCK:
                    pending = CONNECTOR_PENDING.pop(command_id, None)
                if not pending:
                    self.send_json(404, {"error": "Command expired or unknown."})
                    return
                pending["result"] = payload.get("result") or {}
                pending["event"].set()
                self.send_json(200, {"accepted": True})
                return
            if self.path == "/api/shutdown":
                self.send_json(200, {"closed": True})
                threading.Thread(target=self.server.shutdown, daemon=True).start()
                return
            if self.path == "/api/account/status":
                account = load_account()
                self.send_json(200, {
                    "registered": bool(account), "logged_in": ACCOUNT_LOGGED_IN,
                    "account": public_account(account) if ACCOUNT_LOGGED_IN else {},
                })
                return
            if self.path == "/api/account/register":
                if load_account():
                    raise ValueError("An account is already registered on this computer.")
                account = save_account(
                    payload.get("customerName", ""), payload.get("mobile", ""),
                    payload.get("businessName", ""), str(payload.get("pin", "")),
                )
                ACCOUNT_LOGGED_IN = True
                gst_session_reset_portal_imports()
                save_login_session(True)
                self.send_json(200, {"registered": True, "logged_in": True, "account": public_account(account)})
                return
            if self.path == "/api/account/login":
                account = load_account()
                if not account:
                    raise ValueError("Create the customer account first.")
                if not verify_pin(account, str(payload.get("pin", ""))):
                    raise ValueError("Incorrect Login PIN.")
                ACCOUNT_LOGGED_IN = True
                gst_session_reset_portal_imports()
                save_login_session(True)
                self.send_json(200, {"registered": True, "logged_in": True, "account": public_account(account)})
                return
            if self.path == "/api/account/logout":
                ACCOUNT_LOGGED_IN = False
                gst_session_reset_portal_imports()
                clear_login_session()
                self.send_json(200, {"logged_in": False})
                return
            if not ACCOUNT_LOGGED_IN:
                # Try restoring a session file if the server process was replaced mid-UI.
                if not restore_login_session():
                    self.send_json(401, {"error": "Login required. Enter your PIN, then Connect Tally again."})
                    return
            if self.path == "/api/license/status":
                self.send_json(200, LICENSE_STORE.status())
                return
            if self.path == "/api/license/activate":
                account = load_account() or {}
                status = LICENSE_STORE.activate(payload.get("licenseKey", ""), account.get("mobile", ""))
                self.send_json(200, {"message": "License activated successfully.", **status})
                return
            if self.path == "/api/hsn/list":
                rows = list_hsn_master(payload.get("query", ""), payload.get("limit", 1000))
                self.send_json(200, {"rows": rows, "count": len(rows)})
                return
            if self.path == "/api/hsn/save":
                save_hsn_record(payload)
                rows = list_hsn_master(payload.get("hsn_code", ""), 100)
                self.send_json(200, {"message": "HSN item saved.", "rows": rows})
                return
            if self.path == "/api/hsn/import":
                import base64
                file = payload.get("file") or {}
                saved, skipped = import_hsn_master(
                    file.get("name", "HSN Master.xlsx"), base64.b64decode(file.get("data", ""))
                )
                self.send_json(200, {"message": "HSN Master imported.", "saved": saved, "skipped": skipped})
                return
            if self.path == "/api/tally/sync":
                cache = sync_tally()
                self.send_json(200, {
                    "connected": True, "company": cache["company"],
                    "ledgers": cache["ledgers"], "items": cache["items"],
                    "stock_groups": cache["stock_groups"],
                    "voucher_types": cache["voucher_types"], "synced_at": cache["synced_at"],
                    "counts": {
                        "ledgers": len(cache["ledgers"]), "items": len(cache["items"]),
                        "stock_groups": len(cache["stock_groups"]),
                        "voucher_types": len(cache["voucher_types"]),
                    },
                })
                return
            if self.path == "/api/gst/tally-monthly":
                self.send_json(200, sync_tally_gst_monthly())
                return
            if self.path == "/api/tally/resolve-ledgers":
                if not TALLY_CACHE.get("connected"):
                    sync_tally()
                names = payload.get("names", [])
                resolved = []
                seen = set()
                for name in names:
                    cleaned = clean_ledger_name(name)
                    if not cleaned or cleaned in seen:
                        continue
                    seen.add(cleaned)
                    resolved.append({
                        "source": cleaned,
                        "ledger": resolve_tally_ledger_name(cleaned),
                    })
                self.send_json(200, {"resolved": resolved, "company": TALLY_CACHE.get("company", "")})
                return
            if self.path == "/api/tally/match-ledgers":
                if not TALLY_CACHE.get("connected"):
                    sync_tally()
                matches = []
                for index, text in enumerate(payload.get("particulars", [])):
                    ledger, score = match_tally_ledger(text)
                    resolved = resolve_tally_ledger_name(text) if ledger else ""
                    matches.append({
                        "index": index,
                        "ledger": resolved or (clean_ledger_name(ledger.get("name", "")) if ledger else ""),
                        "score": score, "matched": bool(ledger),
                    })
                self.send_json(200, {"company": TALLY_CACHE.get("company", ""), "matches": matches})
                return
            if self.path == "/api/company/parse":
                import base64
                all_rows, files = [], []
                for item in payload.get("files", []):
                    name = item.get("name", "Company statement")
                    raw = base64.b64decode(item.get("data", ""))
                    suffix = Path(name).suffix.lower()
                    if suffix == ".pdf":
                        parsed, meta = parse_company_pdf(name, raw, item.get("password", ""))
                    elif suffix in {".jpeg", ".jpg", ".png"}:
                        parsed, meta = parse_company_image(name, raw)
                    else:
                        parsed, meta = parse_file(name, raw, payload.get("partyLedger", "Party Ledger"))
                    all_rows.extend(parsed)
                    files.append({"name": name, **meta})
                self.send_json(200, {"rows": all_rows, "files": files, "count": len(all_rows)})
                return
            if self.path == "/api/company/reconcile":
                statement_rows = payload.get("statementRows", [])
                party_ledger = gst_text(payload.get("partyLedger", ""))
                if not party_ledger:
                    raise ValueError("Select the matching Tally party ledger.")
                tally_rows = sync_tally_vouchers(party_ledger)
                results = reconcile_company_rows(statement_rows, tally_rows, payload.get("tolerance", 1))
                counts = {}
                for row in results:
                    counts[row["match_status"]] = counts.get(row["match_status"], 0) + 1
                self.send_json(200, {"rows": results, "counts": counts, "tally_rows": len(tally_rows),
                                     "party_ledger": party_ledger})
                return
            if self.path == "/api/company/tally/send":
                raw = make_company_statement_xml(
                    payload.get("rows", []), gst_text(payload.get("partyLedger", "")),
                    gst_text(payload.get("counterLedger", "")),
                )
                if not payload.get("partyLedger") or not payload.get("counterLedger"):
                    raise ValueError("Party Ledger and Counter Ledger are required.")
                tally_response = tally_post(raw, timeout=30)
                def company_count(tag):
                    match = re.search(rf"<{tag}>\s*(\d+)\s*</{tag}>", tally_response, re.I)
                    return int(match.group(1)) if match else 0
                errors = company_count("ERRORS") + company_count("EXCEPTIONS")
                if errors:
                    raise ValueError(f"Tally reported {errors} error(s). No success is assumed.")
                self.send_json(200, {"created": company_count("CREATED"),
                                     "altered": company_count("ALTERED"),
                                     "ignored": company_count("IGNORED")})
                return
            if self.path == "/api/gst/import":
                import base64
                all_rows = []
                return_type = gst_text(payload.get("returnType", "GST"))
                filing_gstin = gst_text(payload.get("gstin", "")).upper()
                financial_year = gst_text(payload.get("financialYear") or payload.get("financial_year") or "")
                for item in payload.get("files", []):
                    item_name = item.get("name", "GST file")
                    item_raw = base64.b64decode(item.get("data", ""))
                    if Path(item_name).suffix.lower() == ".mbk":
                        password = gst_text(payload.get("backupPassword"))
                        try:
                            with zipfile.ZipFile(io.BytesIO(item_raw)) as marg_archive:
                                encrypted = any(member.flag_bits & 0x1 for member in marg_archive.infolist())
                                if encrypted and not password:
                                    self.send_json(400, {
                                        "error": "This MARG backup is password-protected. Enter the Backup Password and import again.",
                                        "password_required": True,
                                    })
                                    return
                                if encrypted:
                                    first_file = next((member for member in marg_archive.infolist() if not member.is_dir()), None)
                                    if first_file:
                                        try:
                                            marg_archive.read(first_file, pwd=password.encode("utf-8"))
                                        except RuntimeError:
                                            self.send_json(400, {
                                                "error": "Incorrect MARG Backup Password.",
                                                "password_required": True,
                                            })
                                            return
                        except zipfile.BadZipFile:
                            raise ValueError("The selected MARG backup file is damaged or unsupported.")
                        raise ValueError("MARG backup opened successfully, but its internal sales database needs the MARG data reader before invoices can be imported.")
                    parsed = parse_gst_file(item_name, item_raw)
                    file_gstin = filing_gstin or infer_taxpayer_gstin_from_text(item_name)
                    for row in parsed:
                        if normalize_portal_return_type(return_type) in {"GSTR-2B", "GSTR-2A", "GSTR-1"}:
                            fy = financial_year or gstr1_financial_year(
                                row.get("source_period") or row.get("gstr2b_period") or infer_gst_period(item_name)
                            ) or gst_portal_default_fy()
                            all_rows.append(
                                gst_stamp_portal_row(
                                    row,
                                    taxpayer_gstin=file_gstin,
                                    financial_year=fy,
                                    return_type=normalize_portal_return_type(return_type) or return_type,
                                )
                            )
                        else:
                            all_rows.append(row)
                if not all_rows:
                    raise ValueError("No sales invoice rows were found. For MARG Backup / Sales Register, use an invoice-level or HSN-wise sales register Excel file.")
                if filing_gstin or any(row.get("taxpayer_gstin") for row in all_rows):
                    gst_portal_set_context(
                        filing_gstin or next((row.get("taxpayer_gstin") for row in all_rows if row.get("taxpayer_gstin")), ""),
                        financial_year or gst_portal_default_fy(),
                    )
                self.send_json(200, {"return_type": return_type,
                                     "rows": all_rows, "summary": gst_summary(all_rows),
                                     "gstin": gst_portal_get_context().get("gstin", ""),
                                     "financial_year": gst_portal_get_context().get("financial_year", "")})
                return
            if self.path == "/api/gst/financial-import":
                import base64
                reports = [
                    parse_gst_financial_file(
                        item.get("name", "GST report"), base64.b64decode(item.get("data", "")),
                        payload.get("reportType", ""),
                    ) for item in payload.get("files", [])
                ]
                self.send_json(200, {"reports": reports, "count": len(reports)})
                return
            if self.path == "/api/gst/gstr3b-import":
                import base64
                totals = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
                reverse_charge = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
                net_itc = {"taxable_value": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}
                periods = {}
                net_periods = {}
                outward_periods = {}
                import_reports = []
                import_errors = []
                last_summary = {}
                for item in payload.get("files", []):
                    name = item.get("name", "GSTR-3B file")
                    try:
                        raw = base64.b64decode(item.get("data", ""))
                        parsed = parse_gstr3b_full(name, raw)
                        file_period = (
                            parsed.get("return_period")
                            or infer_gst_period(name)
                            or ""
                        )
                        # Rebuild FY meta once after the batch — not after every PDF.
                        report = gstr3b_save_import_batch(
                            name, raw, parsed,
                            gstin=payload.get("gstin", "") or parsed.get("gstin", ""),
                            # Prefer PDF period; UI ALL/FY must not blank Table 3.1 mapping.
                            return_period=file_period or payload.get("returnPeriod", ""),
                            rebuild_index=False,
                        )
                        import_reports.append(report)
                        summary = report.get("summary") or parsed
                        last_summary = summary
                        for key in totals:
                            totals[key] += gst_number(parsed.get(key))
                            net_itc[key] += gst_number((parsed.get("net_itc") or parsed).get(key))
                            reverse_charge[key] += gst_number((parsed.get("reverse_charge") or {}).get(key))
                        period = normalize_gst_recon_period(
                            report.get("return_period") or file_period or summary.get("return_period")
                        )
                        if period:
                            claimed = tax_bucket_from(parsed.get("itc_claimed") or parsed.get("net_itc") or parsed)
                            outward = tax_bucket_from(parsed.get("outward_supplies"))
                            periods[period] = {key: round(gst_number(parsed.get(key)), 2) for key in totals}
                            net_periods[period] = {key: claimed.get(key, 0) for key in totals}
                            outward_periods[period] = outward
                            gst_session_mark_imported("GSTR-3B", period)
                    except Exception as exc:
                        import_errors.append({"file": name, "error": str(exc)})
                        import_reports.append({
                            "file": name,
                            "duplicate": False,
                            "error": str(exc),
                            "message": str(exc),
                        })
                if not last_summary and not periods:
                    message = "; ".join(
                        f"{item.get('file')}: {item.get('error')}" for item in import_errors
                    ) or "No GSTR-3B PDF could be imported."
                    self.send_json(400, {"error": message, "imports": import_reports, "errors": import_errors})
                    return
                meta = gstr3b_rebuild_period_index()
                imported_periods = meta.get("imported_periods") or gstr3b_list_imported_periods()
                self.send_json(200, {
                    "kind": "GSTR-3B",
                    "totals": {key: round(value, 2) for key, value in totals.items()},
                    "net_itc": {key: round(value, 2) for key, value in net_itc.items()},
                    "periods": meta.get("periods") or periods,
                    "net_periods": meta.get("net_periods") or net_periods,
                    "outward_periods": meta.get("outward_periods") or outward_periods,
                    "nil_periods": meta.get("nil_periods") or {},
                    "imported_periods": imported_periods,
                    "reverse_charge": {key: round(value, 2) for key, value in reverse_charge.items()},
                    "imports": import_reports,
                    "errors": import_errors,
                    "summary": meta if len(imported_periods) > 1 else (last_summary or meta or {}),
                    "duplicate": bool(import_reports) and all(r.get("duplicate") for r in import_reports if not r.get("error")),
                    "return_period": "ALL" if len(imported_periods) > 1 else (imported_periods[0] if imported_periods else ""),
                })
                return
            if self.path == "/api/bulk-vouchers/parse":
                import base64
                rows = []
                for item in payload.get("files", []):
                    rows.extend(parse_bulk_voucher_file(
                        item.get("name", "bulk.xlsx"), base64.b64decode(item.get("data", ""))
                    ))
                self.send_json(200, {"rows": rows, "count": len(rows)})
                return
            if self.path == "/api/gst/reconcile":
                results, counts, category_counts, dashboard = reconcile_gst_rows(
                    payload.get("gstr2a", []), payload.get("gstr2b", []), payload.get("tolerance", 1)
                )
                self.send_json(200, {
                    "rows": results,
                    "counts": counts,
                    "category_counts": category_counts,
                    "dashboard": dashboard,
                    "total": len(results),
                })
                return
            if self.path == "/api/gst/recon/load":
                ctx = gst_portal_resolve_context(
                    payload.get("gstin", ""),
                    payload.get("financialYear") or payload.get("financial_year") or "",
                )
                gstin = ctx.get("gstin")
                fy = ctx.get("financial_year") or gst_portal_default_fy()
                mode = gst_text(payload.get("mode") or "summary").lower()
                include = payload.get("include") or payload.get("datasets") or []
                if isinstance(include, str):
                    include = [part.strip() for part in include.split(",") if part.strip()]
                # summary (default): counts + light meta. full / include list: load row arrays.
                want_full = mode in {"full", "all"} or bool(include)
                sales_dash = gst_strip_dashboard_rows(gst_recon_get_meta("sales_dashboard", {}) or {})
                g3_dash = gst_strip_gstr3b_dashboard_wire(
                    gst_recon_get_meta("gstr3b_dashboard", {}) or {}, include_drilldown=False
                )
                overview_meta = gst_recon_overview_wire(gst_recon_get_meta("gst_recon_overview", {}) or {})
                payload_out = {
                    "portal_context": ctx,
                    "mode": "full" if want_full and mode in {"full", "all"} and not include else (
                        "datasets" if include else "summary"
                    ),
                    "counts": {
                        "gstr2b": gst_recon_count_rows("GSTR-2B", gstin=gstin, financial_year=fy),
                        "gstr2a": gst_recon_count_rows("GSTR-2A", gstin=gstin, financial_year=fy),
                        "tally_purchase": gst_recon_count_rows("TALLY_PURCHASE"),
                        "gstr1": gstr1_count_invoices(gstin=gstin, financial_year=fy),
                        "tally_sales": gst_recon_count_rows("TALLY_SALES"),
                        "results": gst_recon_count_results("2b_tally"),
                        "gstr1_results": gstr1_count_reconciliation(),
                    },
                    "gstr3b": gstr3b_load_summary() or {},
                    "tally_sync": gst_recon_get_meta("tally_sync", {}),
                    "tally_sales_sync": gst_recon_get_meta("tally_sales_sync", {}),
                    "sales_dashboard": sales_dash,
                    "gstr3b_dashboard": g3_dash,
                    "overview": overview_meta,
                }
                if want_full:
                    bundle = gst_recon_load_dataset_bundle(
                        include if include else ["all"], gstin=gstin, financial_year=fy
                    )
                    payload_out.update(bundle)
                else:
                    # Keep keys present so older clients don't crash on undefined.
                    payload_out.update({
                        "gstr2b": [],
                        "gstr2a": [],
                        "tally_purchase": [],
                        "gstr1": [],
                        "tally_sales": [],
                        "results": [],
                        "gstr1_results": [],
                    })
                self.send_json(200, payload_out)
                return
            if self.path == "/api/gst/recon/datasets":
                include = payload.get("include") or payload.get("datasets") or []
                if isinstance(include, str):
                    include = [part.strip() for part in include.split(",") if part.strip()]
                if not include:
                    self.send_json(400, {"error": "include dataset list is required."})
                    return
                bundle = gst_recon_load_dataset_bundle(
                    include,
                    gstin=payload.get("gstin", ""),
                    financial_year=payload.get("financialYear") or payload.get("financial_year") or "",
                )
                self.send_json(200, bundle)
                return
            if self.path == "/api/gst/recon/save":
                dataset_key = gst_text(payload.get("datasetKey", ""))
                rows = payload.get("rows")
                save_gstin = gst_text(payload.get("gstin", "")).upper()
                save_fy = gst_text(payload.get("financialYear") or payload.get("financial_year") or "")
                if rows is not None and dataset_key:
                    if dataset_key in GST_PORTAL_DATASETS and not save_gstin:
                        # Infer filing GSTIN from filenames embedded in rows / prior context.
                        save_gstin = gst_text(
                            next(
                                (
                                    row.get("taxpayer_gstin") or row.get("filing_gstin")
                                    for row in (rows or [])
                                    if row.get("taxpayer_gstin") or row.get("filing_gstin")
                                ),
                                "",
                            )
                        ).upper() or (gst_portal_get_context().get("gstin") or "")
                    gst_recon_save_rows(dataset_key, rows, gstin=save_gstin, financial_year=save_fy)
                    if dataset_key == "GSTR-2B":
                        selected = normalize_gst_recon_period(payload.get("returnPeriod", ""))
                        if selected:
                            gst_session_mark_imported("GSTR-2B", selected)
                if payload.get("gstr3b") is not None:
                    gst_recon_set_meta("GSTR-3B", payload.get("gstr3b"))
                if payload.get("tally_sync") is not None:
                    gst_recon_set_meta("tally_sync", payload.get("tally_sync"))
                if payload.get("tally_sales_sync") is not None:
                    gst_recon_set_meta("tally_sales_sync", payload.get("tally_sales_sync"))
                if payload.get("results") is not None:
                    gst_recon_save_results(payload.get("results"), "2b_tally")
                if payload.get("gstr1_results") is not None:
                    gstr1_save_reconciliation(payload.get("gstr1_results"), payload.get("returnPeriod", ""))
                if payload.get("sales_dashboard") is not None:
                    gst_recon_set_meta(
                        "sales_dashboard",
                        gst_strip_dashboard_rows(payload.get("sales_dashboard") or {}),
                    )
                if payload.get("gstr3b_dashboard") is not None:
                    dash = payload.get("gstr3b_dashboard") or {}
                    if not dash or dash.get("imported") is False or not (
                        dash.get("imported_periods") or dash.get("cards") or dash.get("rows")
                    ):
                        gst_recon_delete_meta_keys(("gstr3b_dashboard",))
                    else:
                        gst_recon_set_meta(
                            "gstr3b_dashboard",
                            gst_strip_gstr3b_dashboard_wire(dash, include_drilldown=False),
                        )
                self.send_json(200, {"saved": True})
                return
            if self.path == "/api/gst/recon/clear":
                return_type = normalize_portal_return_type(
                    payload.get("returnType") or payload.get("return_type") or ""
                )
                gstin = gst_text(payload.get("gstin", "")).upper()
                financial_year = gst_text(payload.get("financialYear") or payload.get("financial_year") or "2025-26")
                if return_type:
                    cleared = gst_recon_clear_portal_return(
                        return_type, gstin=gstin, financial_year=financial_year
                    )
                else:
                    cleared = gst_recon_clear_all()
                overview = build_gst_recon_overview(
                    payload.get("returnPeriod", ""),
                    payload.get("tolerance", 1),
                    gstin=cleared.get("gstin") or gstin,
                    financial_year=cleared.get("financial_year") or financial_year,
                )
                self.send_json(200, {
                    **cleared,
                    "overview": overview,
                    "import_status": overview.get("import_status"),
                })
                return
            if self.path == "/api/gst/recon/payment/status":
                self.send_json(
                    200,
                    gst_payment_build_status(
                        gstin=payload.get("gstin", ""),
                        financial_year=payload.get("financialYear") or payload.get("financial_year") or "",
                    ),
                )
                return
            if self.path == "/api/gst/recon/payment/load":
                data_type = normalize_gst_payment_data_type(
                    payload.get("dataType") or payload.get("data_type") or ""
                )
                gstin = payload.get("gstin", "")
                fy = payload.get("financialYear") or payload.get("financial_year") or ""
                status = gst_payment_build_status(gstin=gstin, financial_year=fy)
                rows = gst_payment_load_rows(data_type, gstin=gstin, financial_year=fy) if data_type else {}
                payload_out = {
                    **status,
                    "rows_by_type": {
                        key: gst_payment_load_rows(key, gstin=status.get("gstin"), financial_year=status.get("financial_year"))
                        for key in GST_PAYMENT_DATA_TYPES
                    },
                }
                if data_type:
                    payload_out["data_type"] = data_type
                    payload_out["rows"] = rows if isinstance(rows, list) else []
                self.send_json(200, payload_out)
                return
            if self.path == "/api/gst/recon/payment/clear":
                cleared = gst_payment_clear_dataset(
                    payload.get("dataType") or payload.get("data_type") or "",
                    gstin=payload.get("gstin", ""),
                    financial_year=payload.get("financialYear") or payload.get("financial_year") or "",
                )
                self.send_json(200, cleared)
                return
            if self.path == "/api/gst/recon/payment/import":
                import base64
                data_type = normalize_gst_payment_data_type(
                    payload.get("dataType") or payload.get("data_type") or payload.get("reportType") or ""
                )
                files = payload.get("files") or []
                if not data_type:
                    self.send_json(400, {
                        "ok": False,
                        "error": "Payment & Ledger import failed",
                        "detail": "dataType must be challan_history, cash_ledger, credit_ledger (or GST_PAYMENT_LIST / GST_CASH_LEDGER / GST_ITC_LEDGER).",
                    })
                    return
                if not files:
                    self.send_json(400, {
                        "ok": False,
                        "error": "Payment & Ledger import failed",
                        "detail": "Select a GST Payment / Ledger file to import.",
                    })
                    return
                try:
                    file_items = []
                    for item in files:
                        file_items.append({
                            "name": item.get("name", "GST payment file"),
                            "raw": base64.b64decode(item.get("data", "")),
                        })
                    last = gst_payment_save_imports(
                        data_type,
                        file_items,
                        gstin=payload.get("gstin", ""),
                        financial_year=payload.get("financialYear") or payload.get("financial_year") or "",
                    )
                except Exception as exc:
                    self.send_json(400, {
                        "ok": False,
                        "error": f"{GST_PAYMENT_TYPE_LABELS.get(data_type, data_type)} import failed",
                        "detail": str(exc),
                        "data_type": data_type,
                    })
                    return
                meta = (last or {}).get("meta") or {}
                validation = (last or {}).get("validation") or {}
                merge_stats = (last or {}).get("merge_stats") or {}
                self.send_json(200, {
                    **(last or {}),
                    "ok": True,
                    "data_type": (last or {}).get("data_type") or data_type,
                    "rows_imported": (last or {}).get("record_count") or (last or {}).get("row_count") or 0,
                    "gstin": (last or {}).get("gstin") or meta.get("detected_gstin") or "",
                    "from_date": meta.get("from_date") or validation.get("from_date") or "",
                    "to_date": meta.get("to_date") or validation.get("to_date") or "",
                    "financial_year": (last or {}).get("financial_year") or "",
                    "merge_stats": merge_stats,
                    "count": len(file_items),
                })
                return
            if self.path == "/api/gst/recon/session-start":
                reset = gst_session_reset_portal_imports()
                overview = build_gst_recon_overview(
                    payload.get("returnPeriod", ""),
                    payload.get("tolerance", 1),
                )
                self.send_json(200, {**reset, "overview": overview})
                return
            if self.path == "/api/gst/tally/purchase-sync":
                synced = sync_tally_purchase_vouchers()
                gst_recon_save_rows("TALLY_PURCHASE", synced.get("rows", []))
                gst_recon_set_meta("tally_sync", {
                    "company": synced.get("company", ""),
                    "count": synced.get("count", 0),
                    "synced_at": synced.get("synced_at", ""),
                })
                self.send_json(200, synced)
                return
            if self.path == "/api/gst/tally/sales-sync":
                existing_sales = gst_recon_load_rows("TALLY_SALES")
                try:
                    synced = sync_tally_sales_vouchers()
                except ValueError as exc:
                    tally_log(f"sales-sync API | FAILED preserved_existing={len(existing_sales)} | {exc}")
                    gst_recon_set_meta("tally_sales_sync", {
                        **(gst_recon_get_meta("tally_sales_sync", {}) or {}),
                        "ok": False,
                        "error": str(exc),
                        "preserved_existing": True,
                        "existing_count": len(existing_sales),
                        "failed_at": gst_recon_now(),
                    })
                    self.send_json(400, {
                        "ok": False,
                        "error": str(exc),
                        "preserved_existing": True,
                        "existing_count": len(existing_sales),
                        "url": TALLY_HTTP_URL,
                        "port_reachable": tally_port_reachable()[0],
                        "rows": existing_sales,
                        "count": len(existing_sales),
                        "recon_blocked": True,
                    })
                    return
                if not synced.get("ok"):
                    tally_log("sales-sync API | not ok — preserving existing sales rows")
                    self.send_json(400, {
                        "ok": False,
                        "error": synced.get("error") or "Tally Sales sync failed.",
                        "preserved_existing": True,
                        "existing_count": len(existing_sales),
                        "rows": existing_sales,
                        "count": len(existing_sales),
                        "recon_blocked": True,
                        "connection": synced.get("connection"),
                    })
                    return
                gst_recon_save_rows("TALLY_SALES", synced.get("rows", []))
                gst_recon_set_meta("tally_sales_sync", {
                    "company": synced.get("company", ""),
                    "count": synced.get("count", 0),
                    "synced_at": synced.get("synced_at", ""),
                    "taxable_value": synced.get("taxable_value", 0),
                    "output_gst": synced.get("output_gst", 0),
                    "ok": True,
                })
                self.send_json(200, synced)
                return
            if self.path == "/api/gst/tally/sales-connection-test":
                result = tally_test_connection(timeout=15)
                self.send_json(200 if result.get("ok") else 400, result)
                return
            if self.path == "/api/gst/recon/gstr1-import":
                import base64
                all_rows = []
                import_reports = []
                for item in payload.get("files", []):
                    name = item.get("name", "GSTR-1 file")
                    raw = base64.b64decode(item.get("data", ""))
                    rows = parse_gstr1_file(name, raw)
                    report = gstr1_save_import_batch(
                        name, raw, rows,
                        gstin=payload.get("gstin", ""),
                        return_period=payload.get("returnPeriod", ""),
                    )
                    import_reports.append(report)
                    if not report.get("duplicate"):
                        all_rows = report.get("rows") or all_rows
                if not import_reports:
                    raise ValueError("Select at least one GSTR-1 JSON, Excel or CSV file.")
                if all(report.get("duplicate") for report in import_reports) and not gst_recon_load_rows("GSTR-1"):
                    raise ValueError(import_reports[0].get("message") or "Duplicate GSTR-1 import.")
                rows = gstr1_load_invoices(payload.get("returnPeriod", ""))
                section_summary = build_gstr1_section_totals(rows if not payload.get("returnPeriod") else gstr1_load_invoices(payload.get("returnPeriod", "")))
                # Prefer section summary of the latest non-duplicate import when present.
                for report in reversed(import_reports):
                    if report.get("section_summary") and not report.get("duplicate"):
                        section_summary = report["section_summary"]
                        break
                self.send_json(200, {
                    "return_type": "GSTR-1",
                    "rows": rows,
                    "summary": gst_summary(rows),
                    "section_summary": section_summary,
                    "imports": import_reports,
                    "duplicate": all(report.get("duplicate") for report in import_reports),
                })
                return
            if self.path == "/api/gst/recon/document-summary":
                portal_rows = payload.get("portal") or payload.get("rows") or []
                tally_rows = payload.get("tally") or []
                portal_label = gst_text(payload.get("portal_label") or payload.get("source") or "Portal") or "Portal"
                tally_label = gst_text(payload.get("tally_label") or "Tally") or "Tally"
                if tally_rows:
                    pack = build_signed_reconciliation_pack(
                        portal_rows, tally_rows, portal_label, tally_label, payload.get("tolerance", 1)
                    )
                else:
                    portal_summary = build_signed_document_type_summary(portal_rows, portal_label)
                    pack = {
                        "portal": portal_summary,
                        "tally": build_signed_document_type_summary([], tally_label),
                        "by_type": [],
                        "net": {
                            "portal_net_gst": ((portal_summary or {}).get("net") or {}).get("net_gst", 0),
                            "tally_net_gst": 0.0,
                            "difference": 0.0,
                            "matched": False,
                            "formula": "Invoices + Debit Notes - Credit Notes ± Amendments",
                        },
                    }
                self.send_json(200, pack)
                return
            if self.path == "/api/gst/recon/2b-tally":
                rows_2b = payload.get("gstr2b") or gst_recon_load_rows("GSTR-2B")
                rows_tally = payload.get("tally_purchase") or gst_recon_load_rows("TALLY_PURCHASE")
                if not rows_2b:
                    raise ValueError("Import GSTR-2B before reconciling with Tally.")
                if not rows_tally:
                    raise ValueError("Sync Tally Purchase vouchers before reconciliation.")
                results, counts, document_summary = reconcile_gstr2b_tally(
                    rows_2b, rows_tally, payload.get("tolerance", 1)
                )
                gst_recon_save_results(results, "2b_tally")
                self.send_json(200, {
                    "rows": results,
                    "counts": counts,
                    "total": len(results),
                    "document_summary": document_summary,
                })
                return
            if self.path == "/api/gst/recon/gstr1-tally":
                return_period = payload.get("returnPeriod", "")
                all_mode = is_gst_all_periods_selection(return_period) or not normalize_gst_recon_period(return_period)
                if all_mode:
                    rows_g1 = gstr1_load_invoices() or payload.get("gstr1") or []
                else:
                    rows_g1 = payload.get("gstr1") or gstr1_load_invoices(return_period)
                rows_tally = payload.get("tally_sales") or gst_recon_load_rows("TALLY_SALES")
                rows_tally = repair_stored_tally_sales_igst(
                    rows_tally, persist=not bool(payload.get("tally_sales"))
                ).get("rows") or rows_tally
                sales_meta = gst_recon_get_meta("tally_sales_sync", {}) or {}
                if not rows_g1:
                    raise ValueError("Import GSTR-1 before reconciling with Tally Sales.")
                if sales_meta.get("ok") is False:
                    raise ValueError(
                        "Tally Sales sync failed earlier. Fix the TallyPrime connection and "
                        "Sync Tally Sales successfully before reconciliation. "
                        "GSTR-1 portal totals were preserved."
                    )
                if not rows_tally:
                    raise ValueError(
                        "Sync Tally Sales vouchers before reconciliation. "
                        "Reconciliation is blocked until Tally Sales data is available."
                    )
                results, counts, document_summary = reconcile_gstr1_tally(
                    rows_g1, rows_tally, payload.get("tolerance", 1), return_period
                )
                gstr1_save_reconciliation(results, return_period)
                dashboard = build_sales_recon_dashboard(
                    rows_g1, rows_tally, results, payload.get("tolerance", 1), return_period
                )
                if document_summary:
                    dashboard["document_summary"] = document_summary
                gst_recon_set_meta("sales_dashboard", dashboard)
                self.send_json(200, {
                    "rows": results, "counts": counts, "total": len(results),
                    "dashboard": dashboard, "document_summary": document_summary,
                })
                return
            if self.path == "/api/gst/recon/sales-dashboard":
                return_period = payload.get("returnPeriod", "")
                all_mode = is_gst_all_periods_selection(return_period) or not normalize_gst_recon_period(return_period)
                include_rows = bool(payload.get("includeRows") or payload.get("include_rows"))
                # Prefer SQLite sources — avoid round-tripping multi-MB client arrays.
                use_client_rows = bool(payload.get("useClientRows") or payload.get("use_client_rows"))
                if all_mode:
                    rows_g1 = gstr1_load_invoices() or ((payload.get("gstr1") or []) if use_client_rows else [])
                else:
                    rows_g1 = (
                        (payload.get("gstr1") if use_client_rows else None)
                        or gstr1_load_invoices(return_period)
                    )
                rows_tally = (
                    (payload.get("tally_sales") if use_client_rows else None)
                    or gst_recon_load_rows("TALLY_SALES")
                )
                # IGST mapping only — align stored Tally Sales IGST to official GSTR-1 Output IGST.
                repaired = repair_stored_tally_sales_igst(
                    rows_tally, persist=not (use_client_rows and bool(payload.get("tally_sales")))
                )
                rows_tally = repaired.get("rows") or rows_tally
                results = payload.get("results") if use_client_rows else None
                if not results:
                    # Month recon cache only; ALL / FY always rebuilds from rows.
                    if normalize_gst_recon_period(return_period):
                        results = gstr1_load_reconciliation(return_period) or None
                    else:
                        results = None
                dashboard = build_sales_recon_dashboard(
                    rows_g1, rows_tally, results, payload.get("tolerance", 1), return_period
                )
                dashboard["igst_audit"] = {
                    "xml_tag": repaired.get("igst_xml_tag") or "ALLLEDGERENTRIES.LIST/LEDGERNAME+AMOUNT",
                    "raw_mapped_output_igst": gst_number((repaired.get("totals") or {}).get("igst")),
                    "final_value_displayed_in_ui": gst_number(((dashboard.get("output_summary") or {}).get("tally") or {}).get("igst")),
                    "rows_repaired": repaired.get("changed", 0),
                }
                # Persist cards without embedding thousands of voucher rows in meta.
                gst_recon_set_meta("sales_dashboard", gst_strip_dashboard_rows(dashboard))
                wire = gst_strip_dashboard_rows(dashboard)
                wire["igst_audit"] = dashboard.get("igst_audit")
                if include_rows:
                    wire["rows"] = dashboard.get("rows") or []
                self.send_json(200, wire)
                return
            if self.path == "/api/gst/recon/overview":
                overview = build_gst_recon_overview(
                    payload.get("returnPeriod", ""),
                    payload.get("tolerance", 1),
                )
                wire = gst_recon_overview_wire(overview)
                gst_recon_set_meta("gst_recon_overview", wire)
                self.send_json(200, wire)
                return
            if self.path == "/api/gst/recon/review-action":
                results = payload.get("results") or gstr1_load_reconciliation(payload.get("returnPeriod", ""))
                updated = apply_gstr1_review_action(
                    results,
                    payload.get("invoiceNo", ""),
                    payload.get("gstin", ""),
                    payload.get("action", ""),
                    payload.get("note", ""),
                )
                gstr1_save_reconciliation(updated, payload.get("returnPeriod", ""))
                self.send_json(200, {"rows": updated, "total": len(updated)})
                return
            if self.path == "/api/gst/recon/one-click-sync":
                result = run_gst_one_click_sync(
                    payload.get("tolerance", 1),
                    payload.get("returnPeriod", "") or payload.get("return_period", "") or "ALL",
                )
                self.send_json(200, result)
                return
            if self.path == "/api/gst/recon/sales-export":
                rows = payload.get("rows")
                if rows is None:
                    rows = gstr1_load_reconciliation(payload.get("returnPeriod", ""))
                status_filter = gst_text(payload.get("statusFilter", ""))
                if status_filter:
                    if status_filter == "Mismatch":
                        mismatch = {
                            "Value Difference", "Tax Difference", "Date Difference", "GSTIN Difference",
                            "Invoice Number Difference", "Possible Match", "Duplicate",
                        }
                        rows = [row for row in rows if row.get("status") in mismatch]
                    else:
                        rows = [row for row in rows if gst_text(row.get("status")) == status_filter]
                title = gst_text(payload.get("title")) or "GSTR-1 vs Tally"
                fmt = gst_text(payload.get("format") or "xlsx").lower()
                if fmt == "csv":
                    output = io.StringIO()
                    fields = [
                        "status", "gstin", "party_name", "invoice_no", "invoice_date", "voucher_type",
                        "gstr1_taxable", "tally_taxable", "gstr1_igst", "tally_igst",
                        "gstr1_cgst", "tally_cgst", "gstr1_sgst", "tally_sgst",
                        "gstr1_cess", "tally_cess", "total_difference", "document_type", "review_action",
                    ]
                    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
                    writer.writeheader()
                    for row in rows or []:
                        writer.writerow({field: row.get(field, "") for field in fields})
                    raw = output.getvalue().encode("utf-8-sig")
                    self.send_response(200)
                    self.send_header("Content-Type", "text/csv; charset=utf-8")
                    self.send_header("Content-Disposition", f'attachment; filename="{title}.csv"')
                    self.send_header("Content-Length", str(len(raw)))
                    self.end_headers()
                    self.wfile.write(raw)
                    return
                raw = make_gstr1_recon_export(rows, title)
                safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", title).strip("_") or "GSTR1_Tally"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.xlsx"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return

            if self.path == "/api/gst/recon/gstr3b-dashboard":
                period = payload.get("returnPeriod", "")
                include_drilldown = bool(
                    payload.get("includeDrilldown") or payload.get("include_drilldown")
                )
                dashboard = reconcile_gstr3b_vs_books(period, payload.get("tolerance", 1))
                # Keep full dashboard in meta for export; wire response drops voucher lists.
                wire = gst_strip_gstr3b_dashboard_wire(dashboard, include_drilldown=include_drilldown)
                self.send_json(200, wire)
                return
            if self.path == "/api/gst/recon/gstr3b-drilldown":
                period = payload.get("returnPeriod", "")
                dashboard = reconcile_gstr3b_vs_books(period, payload.get("tolerance", 1))
                self.send_json(200, {
                    "return_period": dashboard.get("return_period"),
                    "outward_classification_drilldown": dashboard.get("outward_classification_drilldown") or {},
                })
                return
            if self.path == "/api/gst/recon/liability":
                result = build_books_output_liability(payload.get("returnPeriod", ""))
                self.send_json(200, result)
                return
            if self.path == "/api/gst/recon/itc-claim":
                result = build_itc_claim_breakdown(
                    payload.get("returnPeriod", ""), payload.get("tolerance", 1)
                )
                self.send_json(200, result)
                return
            if self.path == "/api/gst/recon/utilisation":
                period = payload.get("returnPeriod", "")
                books = build_books_output_liability(period)
                itc = build_itc_claim_breakdown(period, payload.get("tolerance", 1))
                claimed = itc.get("claimed_itc") or {}
                available = itc.get("available_itc") or {}
                util = compute_gst_utilisation(
                    books.get("books_output"),
                    claimed if tax_bucket_from(claimed).get("output_gst") else available,
                )
                self.send_json(200, util)
                return
            if self.path == "/api/gst/recon/net-payable":
                period = payload.get("returnPeriod", "")
                dash = reconcile_gstr3b_vs_books(period, payload.get("tolerance", 1))
                self.send_json(200, dash.get("payable") or {})
                return
            if self.path == "/api/gst/recon/gstr3b-export":
                dashboard = payload.get("dashboard") or gst_recon_get_meta("gstr3b_dashboard", {})
                if not dashboard:
                    dashboard = reconcile_gstr3b_vs_books(
                        payload.get("returnPeriod", ""), payload.get("tolerance", 1)
                    )
                title = gst_text(payload.get("title")) or "Books_vs_GSTR3B"
                fmt = gst_text(payload.get("format") or "xlsx").lower()
                report = gst_text(payload.get("report") or "books_vs_3b")
                if report == "liability":
                    title = "Liability_Summary"
                    rows = []
                    books = (dashboard.get("books_liability") or {}).get("books_output") or {}
                    portal = dashboard.get("portal_outward") or {}
                    for key, label in (
                        ("taxable_value", "Outward Taxable"),
                        ("igst", "Output IGST"),
                        ("cgst", "Output CGST"),
                        ("sgst", "Output SGST"),
                        ("cess", "Output CESS"),
                        ("output_gst", "Total Output Tax"),
                    ):
                        rows.append({
                            "particulars": label,
                            "books": books.get(key, 0),
                            "gstr3b": portal.get(key, 0),
                            "difference": round(gst_number(books.get(key)) - gst_number(portal.get(key)), 2),
                        })
                    dashboard = {**dashboard, "rows": rows}
                elif report == "itc":
                    title = "ITC_Summary"
                    itc = dashboard.get("itc") or {}
                    rows = []
                    for label, key in (
                        ("Available ITC", "available_itc"),
                        ("Claimed ITC", "claimed_itc"),
                        ("Eligible ITC", "eligible_itc"),
                        ("Ineligible ITC", "ineligible_itc"),
                        ("Pending ITC", "pending_itc"),
                        ("Reversed ITC", "reversed_itc"),
                        ("Unused ITC", "unused_itc"),
                    ):
                        bucket = tax_bucket_from(itc.get(key))
                        rows.append({
                            "particulars": label,
                            "books": bucket.get("output_gst"),
                            "gstr3b": bucket.get("output_gst"),
                            "difference": 0,
                            "igst": bucket.get("igst"),
                            "cgst": bucket.get("cgst"),
                            "sgst": bucket.get("sgst"),
                            "cess": bucket.get("cess"),
                        })
                    dashboard = {**dashboard, "rows": rows}
                elif report == "cash":
                    title = "Cash_Ledger_Summary"
                    util = dashboard.get("utilisation") or {}
                    cash = tax_bucket_from(util.get("cash_required"))
                    rem = tax_bucket_from(util.get("remaining_itc"))
                    rows = [
                        {"particulars": "Cash Required", "books": cash.get("output_gst"), "gstr3b": cash.get("output_gst"), "difference": 0},
                        {"particulars": "Remaining ITC", "books": rem.get("output_gst"), "gstr3b": rem.get("output_gst"), "difference": 0},
                    ]
                    dashboard = {**dashboard, "rows": rows}
                elif report == "tax_diff":
                    title = "Tax_Difference"
                raw = make_gstr3b_export(dashboard, title, fmt)
                safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", title).strip("_") or "GSTR3B"
                self.send_response(200)
                if fmt == "csv":
                    self.send_header("Content-Type", "text/csv; charset=utf-8")
                    self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.csv"')
                else:
                    self.send_header(
                        "Content-Type",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.xlsx"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return

            if self.path == "/api/gst/recon/itc-dashboard":
                rows_2b = payload.get("gstr2b") or gst_recon_load_rows("GSTR-2B")
                rows_tally = payload.get("tally_purchase") or gst_recon_load_rows("TALLY_PURCHASE")
                gstr3b = payload.get("gstr3b") or gstr3b_load_summary() or {}
                dashboard = build_itc_dashboard(
                    rows_2b, gstr3b, rows_tally, payload.get("tolerance", 1)
                )
                gst_recon_set_meta("itc_dashboard", dashboard)
                self.send_json(200, dashboard)
                return
            if self.path == "/api/gst/recon/itc-difference":
                rows_2b = payload.get("gstr2b") or gst_recon_load_rows("GSTR-2B")
                rows_tally = payload.get("tally_purchase") or gst_recon_load_rows("TALLY_PURCHASE")
                recon = build_itc_available_difference_recon(
                    rows_2b, rows_tally, payload.get("tolerance", 1)
                )
                gst_recon_set_meta("itc_difference_recon", {
                    "summary": recon.get("summary"),
                    "counts": recon.get("counts"),
                    "generated_at": gst_recon_now(),
                })
                self.send_json(200, recon)
                return
            if self.path == "/api/gst/recon/itc-difference-export":
                rows_2b = payload.get("gstr2b") or gst_recon_load_rows("GSTR-2B")
                rows_tally = payload.get("tally_purchase") or gst_recon_load_rows("TALLY_PURCHASE")
                recon = payload.get("recon") or build_itc_available_difference_recon(
                    rows_2b, rows_tally, payload.get("tolerance", 1)
                )
                title = gst_text(payload.get("title") or "Available ITC Difference")
                raw = make_itc_difference_export(recon, title)
                safe_name = re.sub(r"[^\w\-]+", "_", title)[:80] or "Available_ITC_Difference"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.xlsx"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return
            if self.path == "/api/gst/recon/gstr1-difference":
                return_period = payload.get("returnPeriod", "")
                all_mode = is_gst_all_periods_selection(return_period) or not normalize_gst_recon_period(return_period)
                if all_mode:
                    rows_g1 = gstr1_load_invoices() or payload.get("gstr1") or []
                else:
                    rows_g1 = payload.get("gstr1") or gstr1_load_invoices(return_period)
                rows_tally = payload.get("tally_sales") or gst_recon_load_rows("TALLY_SALES")
                rows_tally = repair_stored_tally_sales_igst(
                    rows_tally, persist=not bool(payload.get("tally_sales"))
                ).get("rows") or rows_tally
                recon = build_gstr1_output_difference_recon(
                    rows_g1, rows_tally, payload.get("tolerance", 1), return_period
                )
                gst_recon_set_meta("gstr1_difference_recon", {
                    "summary": recon.get("summary"),
                    "counts": recon.get("counts"),
                    "generated_at": gst_recon_now(),
                })
                self.send_json(200, recon)
                return
            if self.path == "/api/gst/recon/gstr1-difference-export":
                return_period = payload.get("returnPeriod", "")
                all_mode = is_gst_all_periods_selection(return_period) or not normalize_gst_recon_period(return_period)
                if all_mode:
                    rows_g1 = gstr1_load_invoices() or payload.get("gstr1") or []
                else:
                    rows_g1 = payload.get("gstr1") or gstr1_load_invoices(return_period)
                rows_tally = payload.get("tally_sales") or gst_recon_load_rows("TALLY_SALES")
                rows_tally = repair_stored_tally_sales_igst(
                    rows_tally, persist=not bool(payload.get("tally_sales"))
                ).get("rows") or rows_tally
                recon = payload.get("recon") or build_gstr1_output_difference_recon(
                    rows_g1, rows_tally, payload.get("tolerance", 1), return_period
                )
                title = gst_text(payload.get("title") or "GSTR-1 vs Tally Difference")
                fmt = gst_text(payload.get("format") or "xlsx").lower()
                if fmt == "csv":
                    import csv
                    buffer = io.StringIO()
                    writer = csv.writer(buffer)
                    writer.writerow([
                        "Status", "GSTIN", "Party Name", "Invoice Number", "Invoice Date", "Voucher Type",
                        "GSTR-1 Taxable", "Tally Taxable", "Difference",
                        "GSTR-1 IGST", "Tally IGST", "GSTR-1 CGST", "Tally CGST",
                        "GSTR-1 SGST", "Tally SGST", "CESS", "Reason",
                    ])
                    for row in (recon.get("rows") or []):
                        writer.writerow([
                            row.get("status"), row.get("gstin"), row.get("party_name"),
                            row.get("invoice_no"), row.get("invoice_date"), row.get("voucher_type"),
                            row.get("gstr1_taxable"), row.get("tally_taxable"), row.get("difference"),
                            row.get("gstr1_igst"), row.get("tally_igst"),
                            row.get("gstr1_cgst"), row.get("tally_cgst"),
                            row.get("gstr1_sgst"), row.get("tally_sgst"),
                            row.get("cess"), row.get("reason"),
                        ])
                    raw = buffer.getvalue().encode("utf-8-sig")
                    safe_name = re.sub(r"[^\w\-]+", "_", title)[:80] or "GSTR1_Tally_Difference"
                    self.send_response(200)
                    self.send_header("Content-Type", "text/csv; charset=utf-8")
                    self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.csv"')
                    self.send_header("Content-Length", str(len(raw)))
                    self.end_headers()
                    self.wfile.write(raw)
                    return
                raw = make_gstr1_difference_export(recon, title)
                safe_name = re.sub(r"[^\w\-]+", "_", title)[:80] or "GSTR1_Tally_Difference"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.xlsx"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return
            if self.path == "/api/gst/sales/adjust":
                results, applied = adjust_gst_sales_rows(
                    payload.get("rows", []), payload.get("reductions", {}), payload.get("additions", {})
                )
                self.send_json(200, {"rows": results, "applied": applied, "summary": gst_summary(results)})
                return
            if self.path == "/api/gst/party-ledgers/ensure":
                result = ensure_gst_party_ledgers(payload.get("rows", []), payload.get("ledgers", {}))
                self.send_json(200, result)
                return
            if self.path == "/api/gst/sales/tally/send":
                selected_rows = [dict(row or {}) for row in (payload.get("rows") or []) if row]
                for row in selected_rows:
                    row["selected"] = True
                    row["ready_for_sales_tally"] = True
                tolerance = gst_number(payload.get("tolerance", 1))
                return_period = gst_text(payload.get("returnPeriod") or payload.get("return_period") or "")
                # Infer July-style period from selected invoice dates when caller omits it.
                if not return_period and selected_rows:
                    periods = {
                        normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date"))
                        for row in selected_rows
                    }
                    periods.discard("")
                    if len(periods) == 1:
                        return_period = next(iter(periods))
                # Always verify against live Tally Sales vouchers — never trust
                # local sent-history or raw EXCEPTIONS as "Already Exists".
                try:
                    tally_pack = sync_tally_sales_vouchers()
                    tally_rows = list(tally_pack.get("rows") or [])
                    gst_recon_save_rows("TALLY_SALES", tally_rows)
                    gst_recon_set_meta("tally_sales_sync", {
                        "ok": True,
                        "synced_at": datetime.now().isoformat(timespec="seconds"),
                        "company": (tally_pack.get("company") or TALLY_CACHE.get("company") or ""),
                        "count": len(tally_rows),
                    })
                except Exception as exc:
                    # Fall back to last synced sales rows if live sync fails.
                    try:
                        tally_rows = [ensure_gst_invoice_fields(row) for row in gst_recon_load_rows("TALLY_SALES")]
                    except Exception:
                        tally_rows = []
                    if not tally_rows:
                        raise ValueError(
                            "Could not read Sales vouchers from Tally to verify duplicates. "
                            f"{exc}"
                        ) from exc
                existence = build_sales_existence_report(
                    selected_rows, tally_rows, tolerance=tolerance, return_period=return_period
                )
                already = existence["already"]
                missing = existence["missing"]
                skipped_report = existence["report"]
                if not missing:
                    self.send_json(200, {
                        "message": "No missing Sales invoices to create — all selected vouchers already exist in Tally.",
                        "created": 0,
                        "altered": 0,
                        "ignored": 0,
                        "errors": 0,
                        "exceptions": 0,
                        "already_exists_count": len(already),
                        "missing_count": 0,
                        "missing_sent_count": 0,
                        "created_confirmed": 0,
                        "still_missing_count": 0,
                        "tally_sales_count": existence["tally_sales_count"],
                        "tally_sales_count_before": existence["tally_sales_count"],
                        "tally_sales_count_after": existence["tally_sales_count"],
                        "period": existence["period"],
                        "skipped_report": skipped_report,
                        "failed_report": [],
                        "invoice_results": [],
                        "details": [],
                        "missing_only_mode": True,
                    })
                    return
                # CRITICAL: never batch-send all missing vouchers in one XML.
                # One voucher per Tally request so CREATED/ERRORS/EXCEPTIONS/LINEERROR
                # belong to that invoice only.
                ledger_config = payload.get("ledgers", {}) or {}
                invoice_results = []
                created_confirmed = 0
                still_missing = 0
                master_missing = 0
                created_total = 0
                altered_total = 0
                errors_total = 0
                exceptions_total = 0
                details = []
                failed_report = []
                live_tally_rows = tally_rows
                for row in missing:
                    one = send_one_gst_sales_voucher_to_tally(
                        row,
                        ledger_config,
                        return_period=return_period,
                        tolerance=tolerance,
                        tally_rows=live_tally_rows,
                    )
                    invoice_results.append(one)
                    created_total += int(one.get("created") or 0)
                    altered_total += int(one.get("altered") or 0)
                    errors_total += int(one.get("errors") or 0)
                    exceptions_total += int(one.get("exceptions") or 0)
                    if one.get("lineerror"):
                        details.append(
                            f"{one.get('invoice_no')}: {one.get('lineerror')}"
                        )
                    if one.get("found_in_tally_after"):
                        created_confirmed += 1
                        # Keep in-memory Tally list warm for subsequent existence checks.
                        if one.get("matching_tally_voucher_no"):
                            live_tally_rows = list(live_tally_rows or []) + [{
                                "invoice_no": one.get("invoice_no"),
                                "voucher_number": one.get("matching_tally_voucher_no"),
                                "invoice_date": one.get("invoice_date"),
                                "party_ledger": one.get("party"),
                                "party_name": one.get("party"),
                                "invoice_value": one.get("amount"),
                                "document_type": "Sales Invoice",
                                "voucher_type": "Sales",
                                "source_period": return_period,
                            }]
                    else:
                        still_missing += 1
                        if one.get("status") == "MASTER MISSING":
                            master_missing += 1
                        failed_report.append({
                            "invoice_no": one.get("invoice_no"),
                            "invoice_date": one.get("invoice_date"),
                            "party": one.get("party"),
                            "amount": one.get("amount"),
                            "found_in_tally": False,
                            "matching_tally_voucher_no": "",
                            "status": one.get("status") or "MISSING IN TALLY",
                            "tally_detail": one.get("lineerror") or "",
                            "missing_master": one.get("missing_master"),
                        })
                tally_after = invoice_results[-1].get("tally_sales_count") if invoice_results else existence["tally_sales_count"]
                self.send_json(200, {
                    "message": (
                        f"Verified Tally Sales first. Skipped {len(already)} already in Tally. "
                        f"Attempted {len(missing)} missing one-by-one. "
                        f"Confirmed created/present after send: {created_confirmed}."
                    ),
                    "created": created_total,
                    "altered": altered_total,
                    "ignored": 0,
                    "errors": errors_total,
                    "exceptions": exceptions_total,
                    "details": details,
                    "raw_response": "",
                    "already_exists_count": len(already),
                    "missing_count": len(missing),
                    "missing_sent_count": len(missing),
                    "created_confirmed": created_confirmed,
                    "still_missing_count": still_missing,
                    "master_missing_count": master_missing,
                    "tally_sales_count_before": existence["tally_sales_count"],
                    "tally_sales_count_after": tally_after,
                    "period": existence["period"],
                    "skipped_report": skipped_report,
                    "failed_report": failed_report,
                    "invoice_results": invoice_results,
                    "missing_only_mode": True,
                })
                return
            if self.path == "/api/gst/sales/tally/verify":
                selected_rows = [dict(row or {}) for row in (payload.get("rows") or []) if row]
                for row in selected_rows:
                    row["selected"] = True
                tolerance = gst_number(payload.get("tolerance", 1))
                return_period = gst_text(payload.get("returnPeriod") or payload.get("return_period") or "")
                if not return_period and selected_rows:
                    periods = {
                        normalize_gst_recon_period(row.get("source_period") or row.get("invoice_date"))
                        for row in selected_rows
                    }
                    periods.discard("")
                    if len(periods) == 1:
                        return_period = next(iter(periods))
                try:
                    tally_pack = sync_tally_sales_vouchers()
                    tally_rows = list(tally_pack.get("rows") or [])
                    gst_recon_save_rows("TALLY_SALES", tally_rows)
                except Exception as exc:
                    try:
                        tally_rows = [ensure_gst_invoice_fields(row) for row in gst_recon_load_rows("TALLY_SALES")]
                    except Exception:
                        tally_rows = []
                    if not tally_rows:
                        raise ValueError(f"Could not verify Sales against Tally. {exc}") from exc
                existence = build_sales_existence_report(
                    selected_rows, tally_rows, tolerance=tolerance, return_period=return_period
                )
                self.send_json(200, {
                    "message": "Sales existence verified against Tally.",
                    "already_exists_count": len(existence["already"]),
                    "missing_count": len(existence["missing"]),
                    "tally_sales_count": existence["tally_sales_count"],
                    "period": existence["period"],
                    "skipped_report": existence["report"],
                    "missing_rows": existence["missing"],
                    "already_rows": existence["already"],
                    "missing_sales_vouchers": existence["missing"],
                })
                return
            if self.path == "/api/gst/sales/tally/send-bulk-fast":
                rows = [dict(row or {}) for row in (payload.get("rows") or []) if row]
                if not rows:
                    raise ValueError("Select at least one missing Sales invoice.")
                for row in rows:
                    row["selected"] = True
                    row["ready_for_sales_tally"] = True
                ledger_config = payload.get("ledgers", {}) or {}
                return_period = gst_text(payload.get("returnPeriod") or payload.get("return_period") or "")
                tolerance = gst_number(payload.get("tolerance", 1))

                # Resolve/create party, rate-wise Sales, GST, Round-Off and
                # HSN-specific Stock Item masters once for the whole batch.
                ensured = ensure_gst_party_ledgers(rows, ledger_config)
                party_mappings = ensured.get("mappings") or {}
                for row in rows:
                    party_key = gst_party_ledger(row).lower()
                    if party_mappings.get(party_key):
                        row["party_ledger"] = party_mappings[party_key]
                resolved_ledgers = dict(ledger_config)
                resolved_ledgers.update(ensured.get("salesLedgers") or {})
                resolved_ledgers.update(ensured.get("taxLedgers") or {})

                # Keep very large months responsive by sending bounded batches.
                # For example, 6,920 invoices become 28 requests instead of one
                # multi-megabyte request that can stall TallyPrime.
                batch_size = 250
                import_results = []
                request_parts = []
                response_parts = []
                for batch_no, start in enumerate(range(0, len(rows), batch_size), 1):
                    batch_rows = rows[start:start + batch_size]
                    raw = make_gst_sales_xml(batch_rows, resolved_ledgers, fresh_remote_id=True)
                    tally_response = tally_post(
                        raw, timeout=300,
                        purpose=f"bulk-missing-sales-import batch {batch_no}",
                    )
                    request_parts.append(raw.decode("utf-8", errors="replace"))
                    response_parts.append(tally_response)
                    import_results.append(tally_import_result(tally_response))

                result = {
                    key: sum(int(item.get(key) or 0) for item in import_results)
                    for key in ("created", "altered", "deleted", "ignored", "errors", "cancelled", "exceptions")
                }
                result["details"] = []
                for item in import_results:
                    for detail in item.get("details") or []:
                        if detail not in result["details"]:
                            result["details"].append(detail)
                result["details"] = result["details"][:20]
                result["batches"] = len(import_results)
                # Keep the voucher import exchange separate from the live
                # verification export, which uses the legacy filenames.
                (DATA_DIR / "tally_last_sales_import_request.xml").write_text(
                    "\n<!-- NEXT BATCH -->\n".join(request_parts), encoding="utf-8", errors="replace"
                )
                (DATA_DIR / "tally_last_sales_import_response.xml").write_text(
                    "\n<!-- NEXT BATCH -->\n".join(response_parts), encoding="utf-8", errors="replace"
                )
                (DATA_DIR / "tally_last_sales_request.xml").write_text(
                    request_parts[-1], encoding="utf-8", errors="replace"
                )
                (DATA_DIR / "tally_last_sales_response.xml").write_text(
                    response_parts[-1], encoding="utf-8", errors="replace"
                )

                # One final live read only. If Tally is temporarily too busy to
                # export, keep the import response and mark verification pending.
                post_rows = []
                verify_error = ""
                try:
                    post_pack = sync_tally_sales_vouchers()
                    post_rows = list(post_pack.get("rows") or [])
                    gst_recon_save_rows("TALLY_SALES", post_rows)
                except Exception as exc:
                    verify_error = str(exc)
                missing_after = []
                confirmed = 0
                if post_rows:
                    report = build_sales_existence_report(
                        rows, post_rows, tolerance=tolerance, return_period=return_period
                    )
                    missing_after = report["missing"]
                    confirmed = len(report["already"])
                elif result.get("errors") or result.get("exceptions"):
                    missing_after = rows

                # Tally's import counters are authoritative when every sent
                # voucher was created without errors/exceptions.  On very large
                # exports the secondary matcher can miss a few vouchers because
                # of display/reference normalization; never offer a duplicate
                # retry for vouchers Tally has already confirmed as CREATED.
                import_fully_created = (
                    int(result.get("created") or 0) == len(rows)
                    and not int(result.get("errors") or 0)
                    and not int(result.get("exceptions") or 0)
                )
                if import_fully_created:
                    missing_after = []
                    confirmed = len(rows)

                details = result.get("details") or []
                self.send_json(200, {
                    **result,
                    "message": (
                        f"Bulk request sent for {len(rows)} missing Sales vouchers. "
                        f"Tally reported created={result.get('created', 0)}, "
                        f"altered={result.get('altered', 0)}, errors={result.get('errors', 0)}."
                    ),
                    "missing_sent_count": len(rows),
                    "created_confirmed": confirmed or int(result.get("created") or 0),
                    "still_missing_count": len(missing_after),
                    "master_missing_count": 1 if extract_tally_missing_master(details) else 0,
                    "missing_rows": missing_after,
                    "failed_report": [
                        {
                            "invoice_no": gst_text(row.get("invoice_no")),
                            "party": gst_party_ledger(row),
                            "amount": abs(gst_number(row.get("invoice_value"))),
                            "status": "MISSING AFTER BULK SEND",
                        }
                        for row in missing_after
                    ],
                    "invoice_results": [],
                    "verification_pending": bool(verify_error),
                    "verification_error": verify_error,
                    "bulk_fast_mode": True,
                })
                return
            if self.path == "/api/gst/sales/tally/send-one":
                row = dict(payload.get("row") or {})
                if not row:
                    rows = payload.get("rows") or []
                    row = dict(rows[0] or {}) if rows else {}
                if not gst_text(row.get("invoice_no")):
                    raise ValueError("send-one requires exactly one Sales invoice row.")
                result = send_one_gst_sales_voucher_to_tally(
                    row,
                    payload.get("ledgers", {}),
                    return_period=gst_text(payload.get("returnPeriod") or payload.get("return_period") or ""),
                    tolerance=gst_number(payload.get("tolerance", 1)),
                )
                self.send_json(200, {
                    "message": f"Single Sales voucher {result.get('invoice_no')} processed.",
                    **result,
                })
                return
            if self.path == "/api/gst/notes/tally/send":
                raw = make_gst_note_xml(
                    payload.get("rows", []),
                    {**payload.get("ledgers", {}), "tallyVoucherType": payload.get("tallyVoucherType", "")},
                    payload.get("voucherType", "")
                )
                request = urllib.request.Request(
                    "http://127.0.0.1:9000", data=raw,
                    headers={"Content-Type": "text/xml; charset=utf-8"}, method="POST",
                )
                try:
                    with urllib.request.urlopen(request, timeout=180) as response:
                        tally_response = response.read().decode("utf-8", errors="replace")
                except (urllib.error.URLError, TimeoutError) as exc:
                    raise ValueError(
                        "TallyPrime connection failed. Open TallyPrime, load the correct company, "
                        "and enable the HTTP Server on port 9000."
                    ) from exc
                (DATA_DIR / "tally_last_note_request.xml").write_bytes(raw)
                (DATA_DIR / "tally_last_note_response.xml").write_text(
                    tally_response, encoding="utf-8", errors="replace"
                )
                result = tally_import_result(tally_response)
                self.send_json(200, {"message": "Reviewed notes sent to Tally.", **result})
                return
            if self.path == "/api/gst/tally/send":
                selected_rows = list(payload.get("rows") or [])
                tolerance = gst_number(payload.get("tolerance", 1))
                try:
                    live_purchase = sync_tally_purchase_vouchers()
                    live_rows = live_purchase.get("rows", [])
                    gst_recon_save_rows("TALLY_PURCHASE", live_rows)
                    tally_rows = [ensure_gst_invoice_fields(row) for row in live_rows]
                except Exception:
                    tally_rows = [ensure_gst_invoice_fields(row) for row in gst_recon_load_rows("TALLY_PURCHASE")]
                sendable = []
                skipped = []
                failed = []
                for row in selected_rows:
                    item = dict(row or {})
                    if not item.get("selected"):
                        continue
                    # Require purchase ledger mapping when rate-based ledger missing.
                    taxable = gst_number(item.get("taxable_value"))
                    if taxable <= 0:
                        failed.append({
                            "invoice_no": item.get("invoice_no"),
                            "gstin": item.get("gstin"),
                            "error": "Taxable value missing — Review Required",
                            "status": "Tally Entry Failed",
                        })
                        continue
                    existing = purchase_row_already_in_tally(item, tally_rows, tolerance=tolerance)
                    if existing:
                        skipped.append({
                            "invoice_no": item.get("invoice_no"),
                            "gstin": item.get("gstin"),
                            "status": "Already in Tally",
                            "tally_voucher_no": gst_text(
                                existing.get("voucher_no") or existing.get("voucher_number") or existing.get("invoice_no")
                            ),
                            "error": "Duplicate purchase already exists in Tally — voucher not recreated.",
                        })
                        continue
                    if not item.get("ready_for_tally"):
                        failed.append({
                            "invoice_no": item.get("invoice_no"),
                            "gstin": item.get("gstin"),
                            "error": item.get("itc_status") or item.get("status") or "Not ready for Tally",
                            "status": "Review Required",
                        })
                        continue
                    if not gst_text(item.get("party_ledger") or item.get("party_name")):
                        failed.append({
                            "invoice_no": item.get("invoice_no"),
                            "gstin": item.get("gstin"),
                            "error": "Purchase party ledger required",
                            "status": "Review Required",
                        })
                        continue
                    taxable_rate = gst_number(item.get("taxable_value"))
                    rate_guess = 0
                    if taxable_rate:
                        rate_pct = 100 * sum(gst_number(item.get(k)) for k in ("igst", "cgst", "sgst", "cess")) / taxable_rate
                        rate_guess = min((0, 5, 12, 18, 28), key=lambda rate: abs(rate - rate_pct))
                    ledgers = payload.get("ledgers") or {}
                    purchase_ledger_name = (
                        gst_text(item.get("expense_ledger"))
                        or gst_text(item.get("purchase_ledger"))
                        or gst_text(ledgers.get(f"purchaseLedger{rate_guess}"))
                    )
                    if not purchase_ledger_name and not (item.get("sales_allocations") or []):
                        failed.append({
                            "invoice_no": item.get("invoice_no"),
                            "gstin": item.get("gstin"),
                            "error": "Purchase Ledger Required",
                            "status": "Review Required",
                        })
                        continue
                    item["selected"] = True
                    sendable.append(item)
                created = 0
                ignored = 0
                sent_count = 0
                if sendable:
                    def gst_tally_count(tag):
                        match = re.search(rf"<{tag}>\s*(\d+)\s*</{tag}>", tally_response, re.I)
                        return int(match.group(1)) if match else 0
                    # A single multi-thousand-voucher request can be imported
                    # fully while Tally never returns its final HTTP response.
                    # Send bounded batches so every accepted group receives a
                    # reliable result and the UI never labels 3,177 created
                    # vouchers as Failed because one giant response timed out.
                    batch_size = 100
                    for batch_start in range(0, len(sendable), batch_size):
                        batch = sendable[batch_start:batch_start + batch_size]
                        raw = make_gst_purchase_xml(batch, payload.get("ledgers", {}))
                        request = urllib.request.Request(
                            "http://127.0.0.1:9000", data=raw,
                            headers={"Content-Type": "text/xml; charset=utf-8"}, method="POST",
                        )
                        try:
                            with urllib.request.urlopen(request, timeout=180) as response:
                                tally_response = response.read().decode("utf-8", errors="replace")
                        except (urllib.error.URLError, TimeoutError) as exc:
                            first_invoice = gst_text(batch[0].get("invoice_no")) if batch else ""
                            last_invoice = gst_text(batch[-1].get("invoice_no")) if batch else ""
                            raise ValueError(
                                f"Tally did not return the result for Purchase batch "
                                f"{batch_start // batch_size + 1} ({first_invoice} to {last_invoice}). "
                                f"Accepted before this batch: {sent_count}. Keep Tally open and retry; "
                                f"existing vouchers will be verified and will not be resent."
                            ) from exc
                        batch_created = gst_tally_count("CREATED")
                        batch_ignored = gst_tally_count("IGNORED")
                        batch_errors = gst_tally_count("ERRORS") + gst_tally_count("EXCEPTIONS")
                        created += batch_created
                        ignored += batch_ignored
                        if batch_errors and batch_created == 0:
                            for item in batch:
                                failed.append({
                                    "invoice_no": item.get("invoice_no"),
                                    "gstin": item.get("gstin"),
                                    "error": f"Tally reported {batch_errors} error(s) for this batch",
                                    "status": "Tally Entry Failed",
                                })
                            continue
                        if batch_created or batch_ignored:
                            sent_count += len(batch)
                            for item in batch:
                                item["tally_status"] = "Sent to Tally"
                                item["purchase_booked"] = True
                self.send_json(200, {
                    "ok": True,
                    "message": (
                        f"Sent {created} Purchase voucher(s) to Tally. "
                        f"Skipped duplicates: {len(skipped)}. Failed/review: {len(failed)}."
                    ),
                    "created": created,
                    "ignored": ignored,
                    "skipped": skipped,
                    "failed": failed,
                    "sent_count": sent_count,
                })
                return
            if self.path == "/api/gst/purchase-notes/tally/send":
                rows = payload.get("rows", []) or []
                totals = {"created": 0, "altered": 0, "ignored": 0, "errors": 0, "exceptions": 0}
                per_invoice = []
                for source in rows:
                    item = dict(source or {})
                    item["selected"] = True
                    item["ready_for_purchase_note"] = True
                    invoice_no = gst_text(item.get("original_invoice_no") or item.get("invoice_no"))
                    try:
                        raw = make_gst_purchase_xml([item], payload.get("ledgers", {}), note_mode=True)
                        tally_response = tally_post(raw, timeout=60)
                        result = tally_import_result(tally_response)
                        for key in totals:
                            totals[key] += int(result.get(key, 0) or 0)
                        successful = bool(result.get("created") or result.get("altered"))
                        details = result.get("details") or []
                        if not successful and not result.get("ignored") and not result.get("errors") and not result.get("exceptions"):
                            totals["errors"] += 1
                            details = details or ["Tally returned no Created/Altered result for this voucher."]
                        per_invoice.append({
                            "invoice_no": invoice_no,
                            "gstin": item.get("gstin"),
                            "status": "Created" if result.get("created") else ("Already Exists" if result.get("altered") else "Failed"),
                            "created": result.get("created", 0),
                            "altered": result.get("altered", 0),
                            "ignored": result.get("ignored", 0),
                            "error": details[0] if details else "",
                        })
                    except Exception as exc:
                        totals["errors"] += 1
                        per_invoice.append({
                            "invoice_no": invoice_no, "gstin": item.get("gstin"),
                            "status": "Failed", "created": 0, "altered": 0,
                            "ignored": 0, "error": str(exc),
                        })
                self.send_json(200, {
                    "ok": totals["errors"] == 0 and totals["exceptions"] == 0,
                    "message": "Purchase notes/amendments processed invoice by invoice.",
                    **totals,
                    "per_invoice": per_invoice,
                    "failed": [row for row in per_invoice if row.get("status") == "Failed"],
                })
                return
            if self.path == "/api/parse":
                import base64
                bank_ledger = payload.get("bankLedger") or "Bank"
                all_rows, files = [], []
                for item in payload.get("files", []):
                    raw = base64.b64decode(item["data"])
                    digest = file_digest(raw)
                    duplicate = LICENSE_STORE.already_processed(digest)
                    suffix = Path(item["name"]).suffix.lower()
                    estimated = 1
                    if suffix == ".pdf" and item.get("password"):
                        estimated = credit_cost(item["name"], raw, 0, item.get("password", ""))
                    if not duplicate and not LICENSE_STORE.can_charge(estimated):
                        raise ValueError(f"Not enough credits. This file needs at least {estimated} credit(s).")
                    rows, meta = parse_file(item["name"], raw, bank_ledger, item.get("password", ""))
                    if meta.get("password_required"):
                        files.append({"name": item["name"], "rows": 0, **meta})
                        continue
                    if meta.get("mapping_required"):
                        token = uuid.uuid4().hex
                        PENDING_CHARGES[token] = {
                            "digest": digest, "filename": item["name"],
                            "pdf_cost": credit_cost(item["name"], raw, 0, item.get("password", "")) if suffix == ".pdf" else 0,
                        }
                        meta["charge_token"] = token
                        files.append({"name": item["name"], "rows": 0, **meta})
                        continue
                    cost = credit_cost(item["name"], raw, len(rows), item.get("password", ""))
                    charge = LICENSE_STORE.charge(cost, digest, item["name"])
                    all_rows.extend(rows)
                    files.append({"name": item["name"], "rows": len(rows), "credit": charge, **meta})
                self.send_json(200, {"rows": all_rows, "files": files, "license": LICENSE_STORE.status()})
                return
            if self.path == "/api/map":
                grid = payload.get("grid", [])
                mapping = payload.get("mapping", {})
                header_row = int(payload.get("headerRow", 0))
                bank_ledger = payload.get("bankLedger") or "Bank"
                effective_header = effective_mapping_header_row(grid, mapping, header_row)
                rows = apply_mapping(grid, mapping, effective_header, bank_ledger, classify)
                token = payload.get("chargeToken", "")
                pending = PENDING_CHARGES.get(token)
                if not pending:
                    raise ValueError("Upload session expired. Please choose the file again.")
                cost = pending["pdf_cost"] or max(1, (len(rows) + 24) // 25)
                charge = LICENSE_STORE.charge(cost, pending["digest"], pending["filename"])
                PENDING_CHARGES.pop(token, None)
                if payload.get("save", True):
                    fingerprint = format_fingerprint(payload.get("filename", "unknown"), grid, header_row)
                    profiles = load_profiles(MAPPINGS_PATH)
                    profiles[fingerprint] = {
                        "name": payload.get("profileName") or payload.get("filename") or "Saved mapping",
                        "suffix": Path(payload.get("filename", "")).suffix.lower(),
                        "headers": [str(value or "").strip().lower() for value in grid[header_row]],
                        "mapping": mapping,
                    }
                    save_profiles(MAPPINGS_PATH, profiles)
                self.send_json(200, {"rows": rows, "format": "Mapped format", "credit": charge, "license": LICENSE_STORE.status()})
                return
            if self.path == "/api/tally/send":
                rows = payload.get("rows", [])
                for row in rows:
                    row["bank_ledger"] = clean_ledger_name(row.get("bank_ledger", ""))
                    row["counter_ledger"] = clean_ledger_name(row.get("counter_ledger", ""))
                sync_tally()
                resolve_row_ledgers(rows)
                batch_id = uuid.uuid4().hex
                def tally_count(tag):
                    match = re.search(rf"<{tag}>\s*(\d+)\s*</{tag}>", current_response, re.I)
                    return int(match.group(1)) if match else 0

                # Large bank statements can exceed TallyPrime's response time when
                # sent as one XML envelope. Send smaller envelopes, but retain one
                # history batch so the complete import can still be undone together.
                totals = {name: 0 for name in ("CREATED", "ALTERED", "IGNORED", "ERRORS", "EXCEPTIONS")}
                records = []
                error_details = []
                ready_rows = [row for row in rows if str(row.get("approval", row.get("status", "Ready"))).lower() == "ready"]
                send_rows = ready_rows or rows
                ledger_notes = ensure_bank_counter_ledgers(send_rows)
                resolve_row_ledgers(send_rows)
                for note in ledger_notes:
                    if note not in error_details:
                        error_details.append(note)
                for offset in range(0, len(send_rows), 100):
                    chunk = send_rows[offset:offset + 100]
                    raw, chunk_records = make_tally_xml(chunk, batch_id=batch_id, return_records=True)
                    if offset == 0:
                        (DATA_DIR / "tally_last_bank_request.xml").write_bytes(raw)
                    current_response = tally_post(raw, timeout=75)
                    if offset == 0:
                        (DATA_DIR / "tally_last_bank_response.xml").write_text(
                            current_response, encoding="utf-8", errors="replace"
                        )
                    records.extend(chunk_records)
                    chunk_result = tally_import_result(current_response)
                    for detail in chunk_result["details"]:
                        if detail not in error_details:
                            error_details.append(detail)
                    for name in totals:
                        totals[name] += tally_count(name)

                created = totals["CREATED"]
                altered = totals["ALTERED"]
                ignored = totals["IGNORED"]
                errors = totals["ERRORS"]
                exceptions = totals["EXCEPTIONS"]
                missing_ledgers = extract_missing_ledgers_from_details(error_details)
                missing_cf = {name.casefold() for name in missing_ledgers}
                if created or altered:
                    successful_records = [
                        record for record in records
                        if clean_ledger_name(record.get("counter_ledger", "")).casefold() not in missing_cf
                    ] if missing_cf else records
                    dates = [record["date"] for record in successful_records if record.get("date")]
                    history = load_import_history()
                    history.append({
                        "batch_id": batch_id,
                        "bank_ledger": next((row.get("bank_ledger", "") for row in rows if row.get("bank_ledger")), ""),
                        "from_date": min(dates) if dates else "",
                        "to_date": max(dates) if dates else "",
                        "count": len(successful_records),
                        "created_at": datetime.now().isoformat(timespec="seconds"),
                        "undone": False,
                        "partial": bool(errors or exceptions),
                        "records": successful_records,
                    })
                    save_import_history(history[-50:])
                self.send_json(200, {
                    "message": "TallyPrime response received.",
                    "created": created,
                    "altered": altered,
                    "ignored": ignored,
                    "errors": errors,
                    "exceptions": exceptions,
                    "details": error_details[:8],
                    "missing_ledgers": missing_ledgers,
                    "ledger_notes": [note for note in ledger_notes if note not in error_details[:8]],
                })
                return
            if self.path == "/api/tally/history":
                history = load_import_history()
                self.send_json(200, {"batches": [{
                    "batch_id": item.get("batch_id", ""),
                    "bank_ledger": item.get("bank_ledger", ""),
                    "from_date": item.get("from_date", ""),
                    "to_date": item.get("to_date", ""),
                    "count": item.get("count", 0),
                    "created_at": item.get("created_at", ""),
                    "undone": bool(item.get("undone")),
                } for item in reversed(history) if not item.get("undone")]})
                return
            if self.path == "/api/tally/undo":
                account = load_account() or {}
                if not verify_pin(account, str(payload.get("pin", ""))):
                    raise ValueError("Incorrect Login PIN.")
                batch_id = str(payload.get("batchId", ""))
                history = load_import_history()
                batch = next((item for item in history if item.get("batch_id") == batch_id and not item.get("undone")), None)
                if not batch:
                    raise ValueError("Import batch was not found or was already undone.")
                raw = make_tally_delete_xml(batch.get("records", []))
                request = urllib.request.Request(
                    "http://127.0.0.1:9000", data=raw,
                    headers={"Content-Type": "text/xml; charset=utf-8"}, method="POST",
                )
                try:
                    with urllib.request.urlopen(request, timeout=30) as response:
                        tally_response = response.read().decode("utf-8", errors="replace")
                except (urllib.error.URLError, TimeoutError) as exc:
                    raise ValueError(
                        "TallyPrime connection failed. Open TallyPrime, load the correct company, "
                        "and enable the HTTP Server on port 9000."
                    ) from exc
                def undo_count(tag):
                    match = re.search(rf"<{tag}>\s*(\d+)\s*</{tag}>", tally_response, re.I)
                    return int(match.group(1)) if match else 0
                deleted, errors = undo_count("DELETED"), undo_count("ERRORS") + undo_count("EXCEPTIONS")
                if errors:
                    raise ValueError(f"Tally reported {errors} error(s). Nothing was marked as undone.")
                if not deleted:
                    raise ValueError("Tally did not confirm any deleted vouchers. The batch remains available.")
                batch["undone"] = True
                batch["undone_at"] = datetime.now().isoformat(timespec="seconds")
                batch["deleted"] = deleted
                save_import_history(history)
                self.send_json(200, {"message": "Bank2Tally import undone.", "deleted": deleted})
                return
            if self.path in {"/api/export/xlsx", "/api/export/xml"}:
                rows = payload.get("rows", [])
                if self.path.endswith("xlsx"):
                    raw, mime, name = make_xlsx(rows), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "Bank2Tally_Review.xlsx"
                else:
                    raw, mime, name = make_tally_xml(rows), "application/xml", "Bank2Tally_Import.xml"
                self.send_response(200)
                self.send_header("Content-Type", mime)
                self.send_header("Content-Disposition", f'attachment; filename="{name}"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return
            if self.path == "/api/gst/preview-xlsx":
                rows = payload.get("rows", [])
                title = gst_text(payload.get("title")) or "GSTR-2 Preview"
                raw = make_gst_summary_xlsx(payload.get("summaryRows", []), title) if payload.get("summaryRows") is not None else make_gst_preview_xlsx(rows, title)
                safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", title).strip("_") or "GSTR2_Preview"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{safe_name}.xlsx"')
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                self.wfile.write(raw)
                return
            self.send_json(404, {
                "ok": False,
                "error": "Not found",
                "detail": f"No API route for {self.path}. Restart Bank2Tally if this endpoint was added recently.",
                "path": self.path,
                "method": "POST",
            })
        except CLIENT_DISCONNECT_ERRORS:
            # Client disconnected before/while the handler finished. No response possible.
            return
        except Exception as exc:
            # Log the real backend failure first. A later client-disconnect while
            # sending the JSON error body must not replace or hide this error.
            try:
                (DATA_DIR / "last_app_error.txt").write_text(
                    f"{datetime.now().isoformat(timespec='seconds')}\n"
                    f"Path: {self.path}\n"
                    f"Error: {exc}\n\n{traceback.format_exc()}",
                    encoding="utf-8",
                )
            except OSError:
                pass
            if not isinstance(exc, ValueError):
                traceback.print_exc()
            self.send_json(
                400,
                {"error": str(exc) or f"{type(exc).__name__}: statement processing failed."},
            )


def main():
    port = PORT
    while True:
        try:
            server = ThreadingHTTPServer((HOST, port), Handler)
            break
        except OSError:
            port += 1
            if port > PORT + 20:
                raise RuntimeError("Bank2Tally could not find a free local port.")
    browser_host = "127.0.0.1" if HOST in {"0.0.0.0", "::"} else HOST
    url = f"http://{browser_host}:{port}"
    print(f"Bank2Tally is running at {url}")
    print("Close this window to stop the application.")
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
